"""Build farada_5y_v1.xlsx from the WIP by wiring the Working-Capital drivers and
Cash-Flow sections of the ProForma (and downstream CF/BS in later phases).

Idempotent: always rebuilds the target from the WIP source, so re-running after a
code change reproduces the workbook exactly. Run from repo root:

    .venv/bin/python scripts/build_farada_5y.py

Plan: docs/superpowers/specs/2026-06-29-farada-5y-cashflow-wiring.md
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter as gc

MODELING = Path("clients/farada/modeling")
SRC = MODELING / "FaradaIC - 5Y plan - WIP.xlsx"
DST = MODELING / "farada_5y_v2.xlsx"

PF = "ProForma"
INP = " Inputs"  # note leading space

# Month columns on ProForma: C(3) .. BJ(62) = 60 months.
C0, C1 = 3, 62


def col(c: int) -> str:
    return gc(c)


# ---------------------------------------------------------------------------
# Task 1 — set Payroll payable days to 14 (both scenario columns L & M).
# ---------------------------------------------------------------------------
def task1_payroll_days(wb) -> None:
    ws = wb[INP]
    for c in (12, 13):  # L, M
        ws.cell(188, c).value = 14
    # leave Receivable days (184) and Payable days (187) at 30.


# Inputs driver scalars (resolved via OFFSET on the Inputs sheet).
DSO = f"' Inputs'!$J$184/30"
DPO = f"' Inputs'!$J$187/30"
PAYDAYS = f"' Inputs'!$J$188/30"

# Arrears-billed revenue/recognised row -> cash-in row (ProForma). Each cash line
# lags its RECOGNISED revenue by DSO. Subscription (202-204) is billed annually in
# advance => Task 4b. Overage is lagged at the subtotal off the recognised total
# (row 64), whose own formula already handles the ramp delay (Inputs J78).
CASHIN_MAP = {
    188: 47, 189: 48, 190: 49,        # Components #1
    192: 51, 193: 52, 194: 53,        # Components #2
    198: 57, 199: 58, 200: 59,        # SaaS #3 — Hardware (device, billed in arrears)
    205: 64,                          # SaaS #3 — Overage (recognised total, lagged)
}


def _lag(ws, cash_row: int, src_row: int, ratio: str) -> None:
    """cash[m] = accrual[m] - ratio*(accrual[m] - accrual[m-1]); month-1 prior = 0."""
    for c in range(C0, C1 + 1):
        cur = f"{col(c)}{src_row}"
        if c == C0:
            f = f"={cur}-{ratio}*{cur}"
        else:
            prev = f"{col(c - 1)}{src_row}"
            f = f"={cur}-{ratio}*({cur}-{prev})"
        ws.cell(cash_row, c).value = f


def _balance(ws, bal_row: int, expr_fn, ratio: str) -> None:
    """Closing payable/receivable balance = ratio * accrual expression."""
    for c in range(C0, C1 + 1):
        ws.cell(bal_row, c).value = f"={ratio}*{expr_fn(c)}"


def _cash_from_balance(ws, cash_row: int, accrual_fn, bal_row: int, opening="0") -> None:
    """Cash = accrual − Δ(working-capital balance row). The deviation is tracked by
    the WC driver balance, not recomputed from the P&L line. month-1 prior = opening."""
    for c in range(C0, C1 + 1):
        if c == C0:
            delta = f"({col(c)}{bal_row}-{opening})"
        else:
            delta = f"({col(c)}{bal_row}-{col(c - 1)}{bal_row})"
        ws.cell(cash_row, c).value = f"={accrual_fn(c)}-{delta}"


def _sum(ws, row: int, lo: int, hi: int) -> None:
    for c in range(C0, C1 + 1):
        ws.cell(row, c).value = f"=SUM({col(c)}{lo}:{col(c)}{hi})"


def _add(ws, row: int, parts: list[int]) -> None:
    for c in range(C0, C1 + 1):
        ws.cell(row, c).value = "=" + "+".join(f"{col(c)}{p}" for p in parts)


# ---------------------------------------------------------------------------
# Task 2 — AR balance (166) + cash in from clients (186-208).
# ---------------------------------------------------------------------------
def task2_ar_cashin(wb) -> None:
    ws = wb[PF]
    # AR closing balance = DSO/30 * arrears-billed revenue (Revenue 45 minus the
    # annually-prepaid Subscription 60, which is deferred revenue, not a receivable).
    for c in range(C0, C1 + 1):
        ws.cell(166, c).value = f"={DSO}*({col(c)}45-{col(c)}60)"
    # Cash-in rows lag their matching recognised revenue line (incl. overage subtotal 205).
    for cash_row, src_row in CASHIN_MAP.items():
        _lag(ws, cash_row, src_row, DSO)
    # Clear stray leftovers in the Subscription (202-204) and Overage bundle (206-208)
    # detail rows — they carry a `=-#REF!` / junk in the tail months. Subscription is
    # repopulated in Task 4b; overage is summarised at the 205 subtotal (no bundle split).
    for r in (202, 203, 204, 206, 207, 208):
        for c in range(C0, C1 + 1):
            ws.cell(r, c).value = None
    # Subtotals and totals. Subscription (201<-202:204) stays blank until Task 4b.
    _sum(ws, 187, 188, 190)   # Components #1
    _sum(ws, 191, 192, 194)   # Components #2
    _sum(ws, 197, 198, 200)   # Hardware
    _sum(ws, 201, 202, 204)   # Subscription (0 until Task 4b)
    _add(ws, 196, [197, 201, 205])  # SaaS #3 (205 overage set above)
    # Total cash in from clients — balance-driven: cash = Revenue − ΔReceivables(166)
    # + ΔDeferred revenue(183). The timing deviation comes from the WC drivers, not a
    # recomputed P&L delta. (Detail rows 187-208 remain a by-product collection memo.)
    for c in range(C0, C1 + 1):
        if c == C0:
            dev = f"({col(c)}166-' Inputs'!$J$193)-({col(c)}183-' Inputs'!$J$196)"
        else:
            dev = f"({col(c)}166-{col(c - 1)}166)-({col(c)}183-{col(c - 1)}183)"
        ws.cell(186, c).value = f"={col(c)}45-({dev})"


# Supplier (non-payroll) cost buckets: (payable-balance row, cash-paid row, cost expr).
# Payroll lives in the personnel section (Task 4), so these are all "excl. Payroll".
def _cost_cos(c):  return f"{col(c)}69"
def _cost_sm(c):   return f"({col(c)}90-{col(c)}91)"
def _cost_ga(c):   return f"({col(c)}99-{col(c)}100)"
def _cost_rd(c):   return f"({col(c)}110-{col(c)}111)"

SUPPLIERS = [
    (174, 211, _cost_cos),  # CoS
    (170, 212, _cost_sm),   # S&M excl. payroll
    (172, 213, _cost_ga),   # G&A excl. payroll
    (176, 214, _cost_rd),   # R&D excl. payroll
]


# ---------------------------------------------------------------------------
# Task 3 — supplier payables (168-176) + cash paid to suppliers (210-214).
# ---------------------------------------------------------------------------
def task3_supplier_payables(wb) -> None:
    ws = wb[PF]
    for bal_row, cash_row, expr in SUPPLIERS:
        _balance(ws, bal_row, expr, DPO)               # closing payable = DPO/30 * cost
        _cash_from_balance(ws, cash_row, expr, bal_row)  # cash = cost − Δ(payable balance)
    _add(ws, 168, [170, 172, 174, 176])      # Trade payables (excl. payroll)
    _sum(ws, 210, 211, 214)                  # Cash Paid to Suppliers


# Personnel payroll buckets: (payable-balance row, payment row, payroll expr).
# CoS has no payroll in this model (COGS is materials/usage only) -> 0.
PERSONNEL = [
    (179, 217, lambda c: "0"),            # Payroll CoS (none)
    (180, 218, lambda c: f"{col(c)}91"),  # Payroll S&M
    (181, 219, lambda c: f"{col(c)}100"), # Payroll G&A
    (182, 220, lambda c: f"{col(c)}111"), # Payroll R&D (Germany + Serbia)
]


# ---------------------------------------------------------------------------
# Task 4 — personnel payables (178-182) + payments to personnel (216-220).
# ---------------------------------------------------------------------------
def task4_personnel_payables(wb) -> None:
    ws = wb[PF]
    for bal_row, pay_row, expr in PERSONNEL:
        _balance(ws, bal_row, expr, PAYDAYS)               # payable = 14/30 * payroll
        _cash_from_balance(ws, pay_row, expr, bal_row)     # paid = payroll − Δ(payable balance)
    _sum(ws, 178, 179, 182)                    # Account payables to personnel
    _sum(ws, 216, 217, 220)                    # Payments to Personnel


# Subscription recognised-revenue rows (Bundle S/M/L) -> annual-billing cash rows.
SUBSCRIPTION = {202: 61, 203: 62, 204: 63}
RENEWAL_LAGS = [0, 12, 24, 36, 48]  # annual re-billing across the 60-month horizon
DEFERRED_ROW = 183                  # repurposed blank separator -> Deferred revenue


def _drec(x: int, rr: int) -> str:
    """Monthly increment of recognised subscription (new ARR) in column x."""
    if x == C0:
        return f"{col(x)}{rr}"
    return f"({col(x)}{rr}-{col(x - 1)}{rr})"


# ---------------------------------------------------------------------------
# Task 4b — SaaS subscription billed annually in advance + deferred revenue.
# ---------------------------------------------------------------------------
def task4b_subscription_deferred(wb) -> None:
    ws = wb[PF]
    # Each cohort prepays 12 months at signup and re-bills every 12 months, so the
    # billing in month c = 12 × Σ ΔARR at lags 0,12,24,36,48 (months still on the grid).
    for cash_row, rec_row in SUBSCRIPTION.items():
        for c in range(C0, C1 + 1):
            terms = [_drec(c - k, rec_row) for k in RENEWAL_LAGS if c - k >= C0]
            ws.cell(cash_row, c).value = "=12*(" + "+".join(terms) + ")"
    # 201 = SUM(202:204) already wired in Task 2; it now picks up the billings.
    # Deferred revenue (row 183) = prior + billings(201) − recognised(60); opening J196.
    ws.cell(DEFERRED_ROW, 1).value = "Deferred revenue"
    for c in range(C0, C1 + 1):
        if c == C0:
            f = f"=' Inputs'!$J$196+{col(c)}201-{col(c)}60"
        else:
            f = f"={col(c - 1)}{DEFERRED_ROW}+{col(c)}201-{col(c)}60"
        ws.cell(DEFERRED_ROW, c).value = f


# ---------------------------------------------------------------------------
# Task 5b — clear the stray #REF! depreciation so the P&L is clean. The balance
# sheet is NOT built this iteration, so PP&E depreciation is left at 0 (no D&A).
# ---------------------------------------------------------------------------
def task5b_clear_depreciation(wb) -> None:
    ws = wb[PF]
    for c in range(C0, C1 + 1):
        ws.cell(122, c).value = 0  # Depreciation (PP&E) -> 0; row 121 D&A (=C122) follows


# ---------------------------------------------------------------------------
# Task 5 — complete (direct-method) Cash Flow statement on the CF sheet.
# Direct method needs no depreciation: CFO = cash receipts − cash payments.
# ---------------------------------------------------------------------------
def task5_cf_statement(wb) -> None:
    cf = wb["CF"]
    for c in range(C0, C1 + 1):
        L = col(c)
        d = f"ProForma!{L}2"  # month date
        # Operating
        cf.cell(3, c).value = f"=ProForma!{L}186"            # cash received from customers
        cf.cell(4, c).value = f"=-ProForma!{L}210"           # cash paid to suppliers
        cf.cell(5, c).value = f"=-ProForma!{L}216"           # payment for personnel
        cf.cell(6, c).value = f"=-ProForma!{L}124"           # bank charges / finance costs
        cf.cell(7, c).value = 0                              # VAT (not modelled)
        cf.cell(8, c).value = f"=ProForma!{L}233"            # corporate tax (already negative)
        cf.cell(9, c).value = 0
        cf.cell(10, c).value = f"={L}3+{L}4+{L}5+{L}6+{L}7+{L}8+{L}9"
        # Investing — CAPEX from the input schedule (no BS needed)
        cf.cell(11, c).value = (
            f"=-IF(AND({d}>=' Inputs'!$G$177,{d}<=' Inputs'!$H$177),' Inputs'!$J$177,0)")
        cf.cell(12, c).value = 0                             # R&D (not capitalised)
        cf.cell(13, c).value = 0
        cf.cell(14, c).value = f"={L}11+{L}12+{L}13"
        # Financing — equity round, grants, debt draw from the funding inputs
        cf.cell(15, c).value = f"=IF({d}=' Inputs'!$J$11,' Inputs'!$J$10,0)"
        cf.cell(16, c).value = (
            f"=IF(AND({d}>=' Inputs'!$G$30,{d}<=' Inputs'!$H$30),' Inputs'!$J$30,0)")
        cf.cell(17, c).value = f"=IF({d}=' Inputs'!$J$19,' Inputs'!$J$18,0)"
        cf.cell(18, c).value = 0
        cf.cell(19, c).value = 0                             # dividends (none)
        cf.cell(20, c).value = f"={L}15+{L}16+{L}17+{L}18+{L}19"
        # Roll-up
        cf.cell(22, c).value = f"={L}10+{L}14+{L}20"
        cf.cell(23, c).value = "=' Inputs'!$J$192" if c == C0 else f"={col(c - 1)}24"
        cf.cell(24, c).value = f"={L}23+{L}22"


def build() -> Path:
    wb = openpyxl.load_workbook(SRC, data_only=False)
    task1_payroll_days(wb)
    task2_ar_cashin(wb)
    task3_supplier_payables(wb)
    task4_personnel_payables(wb)
    task4b_subscription_deferred(wb)
    task5b_clear_depreciation(wb)
    task5_cf_statement(wb)
    # Force Excel/LibreOffice to recompute on open (openpyxl writes no cached values).
    wb.calculation.fullCalcOnLoad = True
    wb.save(DST)
    return DST


if __name__ == "__main__":
    out = build()
    print(f"wrote {out}")
