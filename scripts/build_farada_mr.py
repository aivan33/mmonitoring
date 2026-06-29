#!/usr/bin/env python3
"""Build Farada's monthly MR workbook by populating a new month column.

The MR (master accounting workbook) is a formula-driven Excel file:
  - `P&L Mapping` VLOOKUPs per-account values out of the `BWA` sheet, tagged
    with a leaf label; the front `P&L` then SUMs that mapping by leaf.
  - The front `CF` SUMIFs the `ControllingReport BWA` sheet by mapping code.
  - The front `BS` builds off the `Balance Sheet` / `Trial balance` sheets.

So the monthly job has two parts:
  1. DATA  — paste the new month's column into the German data sheets (this
     module), keyed by DATEV account number so nothing is dropped.
  2. FRONT — extend the front P&L/CF/BS formula columns one month to the right
     (handled separately; the front lags the data by a month in the source file).

This module does part (1) and self-validates by reproducing a known month
(April) from its raw sources and diffing against the delivered workbook.
"""
from __future__ import annotations
import argparse
from pathlib import Path
import openpyxl
import re
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter, column_index_from_string

RAW = Path("clients/farada/raw/accounting")
MR_DIR = Path("clients/farada/raw")

# Month index helpers (1=Jan ... 12=Dec) -> column in each sheet.
# Raw German Jahresübersicht sheets: account in col 2, Jan in col 4 (so month m -> col 3+m).
RAW_ACC_COL = 2
def raw_month_col(m: int) -> int:
    return 3 + m

# MR data sheets that are a FAITHFUL account-keyed paste of a raw German export.
# month m -> col (jan_col - 1 + m).
#
# Only BWA qualifies. Deliberately excluded:
#   - ControllingReport BWA : derived, =VLOOKUP(... 'CR-Upload' ...) — extend the
#     formula, never paste values.
#   - CR-Upload (CF data)   : ~93% raw Controlling-BWA but carries MANUAL CF
#     reclassifications each month (e.g. acct 15900 €606k, sign-flipped 14000)
#     that are not mechanically derivable — must be booked by hand.
#   - Trial balance / Balance Sheet : bespoke S/H (Soll/Haben) layouts, not yet
#     validated; treated as manual for now.
DATA_SHEETS = {
    "BWA": dict(
        raw="BWA - Jahresübersicht {mm}-2026.xlsx",
        acc_col=2, jan_col=4,           # Jan=4 ... Apr=7, May=8
    ),
}


def _num(x):
    return float(x) if isinstance(x, (int, float)) else None


def _raw_account_values(raw_path: Path, month: int) -> dict[str, float]:
    """account number -> month value, from a raw German Jahresübersicht."""
    ws = openpyxl.load_workbook(raw_path, data_only=True).worksheets[0]
    col = raw_month_col(month)
    out: dict[str, float] = {}
    for r in range(3, ws.max_row + 1):
        acc = ws.cell(r, RAW_ACC_COL).value
        if acc not in (None, ""):
            out[str(acc).strip()] = _num(ws.cell(r, col).value) or 0.0
    return out


def _mr_account_rows(ws, acc_col: int) -> dict[str, int]:
    """account number -> row, for an MR data sheet (account-level rows only)."""
    out: dict[str, int] = {}
    for r in range(3, ws.max_row + 1):
        acc = ws.cell(r, acc_col).value
        if acc not in (None, ""):
            out.setdefault(str(acc).strip(), r)
    return out


def populate_sheet(wb, sheet: str, spec: dict, month: int, mm: str, write: bool):
    """Fill `sheet`'s month column from its raw source. Returns a report dict."""
    ws = wb[sheet]
    raw_path = RAW / f"{mm}-2026" / spec["raw"].format(mm=mm)
    raw_vals = _raw_account_values(raw_path, month)
    mr_rows = _mr_account_rows(ws, spec["acc_col"])
    mr_col = spec["jan_col"] - 1 + month

    written, unmapped, changed = 0, [], 0
    for acc, val in raw_vals.items():
        row = mr_rows.get(acc)
        if row is None:
            if abs(val) > 0.005:
                unmapped.append((acc, val))
            continue
        if write:
            prev = _num(ws.cell(row, mr_col).value)
            if prev is None or abs((prev or 0) - val) > 0.005:
                changed += 1
            ws.cell(row, mr_col).value = val
        written += 1
    return dict(sheet=sheet, raw=str(raw_path.name), written=written,
               unmapped=unmapped, changed=changed, mr_col=mr_col)


def golden_april(mr_path: Path) -> int:
    """Reproduce April from raw 04 and diff against the delivered workbook.

    Returns the number of account-level mismatches (0 = perfect reproduction).
    """
    wb = openpyxl.load_workbook(mr_path, data_only=True)
    mism = 0
    for sheet, spec in DATA_SHEETS.items():
        ws = wb[sheet]
        raw_vals = _raw_account_values(
            RAW / "04-2026" / spec["raw"].format(mm="04"), month=4)
        mr_rows = _mr_account_rows(ws, spec["acc_col"])
        col = spec["jan_col"] - 1 + 4
        sheet_mism = 0
        for acc, val in raw_vals.items():
            row = mr_rows.get(acc)
            if row is None:
                continue
            cur = _num(ws.cell(row, col).value) or 0.0
            if abs(cur - val) > 0.01:
                sheet_mism += 1
                if sheet_mism <= 5:
                    print(f"    {sheet} acct {acc}: MR={cur} raw04={val}")
        print(f"  {sheet}: {len(raw_vals)} raw accounts, {sheet_mism} April mismatches")
        mism += sheet_mism
    return mism


# --- Serbia: row-label-matched copy of the template's month column ----------
SERBIA_TPL = "FaradaIC Serbia_ 01_2026_new template_NVS_{stamp}.xlsx"
SERBIA_STAMP = {"04": "20260522", "05": "20260622"}
SERBIA_PL_ROWS = range(5, 16)   # R&D Payroll .. Depreciation (label rows)


def populate_serbia(wb, month: int, mm: str):
    """Fill the MR Serbia sheet's month column from the template, by label.

    MR Serbia: Jan=col3 .. May=col7.  Template: Jan=col4 .. May=col8.
    Returns (written, restatements) — restatements = prior months the template
    revised vs what the MR already holds.
    """
    mr = wb["Serbia"]
    tpl_path = (RAW / f"{mm}-2026" / "Serbia"
                / SERBIA_TPL.format(stamp=SERBIA_STAMP[mm]))
    tpl = openpyxl.load_workbook(tpl_path, data_only=True)["FaradaIC Serbia"]
    mr_col = 2 + month        # May -> 7
    tpl_col = 3 + month       # May -> 8

    written, restated = 0, []
    for r in SERBIA_PL_ROWS:
        mr_lab = str(mr.cell(r, 2).value or "").strip()
        tpl_lab = str(tpl.cell(r, 3).value or "").strip()
        # labels are abbreviated identically; compare on a common prefix
        if mr_lab[:12].lower() != tpl_lab[:12].lower():
            raise ValueError(f"Serbia row {r} label mismatch: MR={mr_lab!r} tpl={tpl_lab!r}")
        val = _num(tpl.cell(r, tpl_col).value) or 0.0
        mr.cell(r, mr_col).value = val
        written += 1
        # restatement check on prior months (Jan..month-1)
        for pm in range(1, month):
            a = _num(mr.cell(r, 2 + pm).value) or 0.0
            b = _num(tpl.cell(r, 3 + pm).value) or 0.0
            if abs(a - b) > 0.5:
                restated.append((mr_lab, pm, a, b))
    return written, restated


# --- New-account mapping rows (resolves the May unmapped-account flags) -------
# P&L Mapping month columns 2..13 (Jan..Dec) each VLOOKUP the matching BWA month
# column; for mapping col c the BWA index is c+1 (verified against acct 83360).
def _vlookup(row: int, c: int) -> str:
    return f"=VLOOKUP($A{row},BWA!$B:$P,{c + 1},FALSE)"


# May's new accounts. Each must be added to BOTH the BWA sheet (the data the
# P&L Mapping VLOOKUPs) and P&L Mapping (the leaf/Type routing).
NEW_ACCOUNTS = [
    # 83380: tax-exempt 3rd-country sales -> revenue 'Other' (spare row 19, which
    # the template already tags Revenue/Other; front Sales 'Other' sums 19:21).
    dict(acct=83380, name="Tax-exempt sales 3rd country",
         bwa_row=157, pm_row=19, capitalise=False),
    # 31004: Fremdleistungen (R&D), Fraunhofer -> CAPITALISED like sibling 31014:
    # tracked on spare row 150 (outside every opex sum) + subtracted in the Check.
    dict(acct=31004, name="Fremdleistungen (R&D)",
         bwa_row=158, pm_row=150, capitalise=True),
    # --- tiny new accounts, mapped to their functional bucket via a free row ---
    # 49203 Telephone (G&A) -> G&A Office Expenses (exact).
    dict(acct=49203, name="Telephone (G&A)",
         bwa_row=159, pm_row=96, capitalise=False),
    # 46680 Employee mileage -> G&A Travel & Representative (row labelled
    # "travel costs employees" — exact).
    dict(acct=46680, name="Employee mileage reimbursement",
         bwa_row=160, pm_row=111, capitalise=False),
    # 46502 Entertainment (S&M) -> S&M; Travel range is full, so the only free
    # S&M row is 'Other marketing expenses' (r73): S&M total correct, sub-line is
    # a catch-all. NOTE the compromise.
    dict(acct=46502, name="Entertainment expenses (S&M)",
         bwa_row=161, pm_row=73, capitalise=False),
    # 37360 (Skonti received, -4.38) intentionally NOT mapped: COGS range is full
    # and discounts-received treatment is a judgement call. Left flagged.
]


def add_mapping_rows(wb, mm: str):
    """Wire May's new accounts into the BWA + P&L Mapping chain (nothing dropped)."""
    bwa = wb["BWA"]
    pm = wb["P&L Mapping"]
    pl = wb["P&L"]
    raw_path = RAW / f"{mm}-2026" / "BWA - Jahresübersicht {mm}-2026.xlsx".format(mm=mm)
    raw = openpyxl.load_workbook(raw_path, data_only=True).worksheets[0]
    # raw account -> {raw_col: value} for the 12 month columns (Jan col4 .. Dec col15)
    raw_rows = {str(raw.cell(r, 2).value).strip(): r
                for r in range(3, raw.max_row + 1) if raw.cell(r, 2).value not in (None, "")}

    done = []
    for spec in NEW_ACCOUNTS:
        acc = str(spec["acct"])
        # 1) BWA sheet row: account in col 2, monthly values (cols 4..15) from raw
        br = spec["bwa_row"]
        bwa.cell(br, 2).value = spec["acct"]
        bwa.cell(br, 3).value = spec["name"]
        rr = raw_rows.get(acc)
        if rr:
            for col in range(4, 16):          # Jan..Dec
                v = _num(raw.cell(rr, col).value)
                if v is not None:
                    bwa.cell(br, col).value = v
        # 2) P&L Mapping row: VLOOKUP the BWA month columns
        pr = spec["pm_row"]
        pm.cell(pr, 1).value = spec["acct"]
        if spec["capitalise"]:
            pm.cell(pr, 18).value = "R&D"     # mirror 31014; no leaf -> not in opex sums
        for c in range(2, 14):
            pm.cell(pr, c).value = _vlookup(pr, c)
        # 3) capitalised accounts: subtract in the front-P&L 'Check' row 84
        if spec["capitalise"]:
            for col in (16, 17, 18):
                f = pl.cell(84, col).value
                ml = get_column_letter(col_to_mapping_letter(col))
                if isinstance(f, str) and f"{ml}146" in f and f"{ml}{pr}" not in f:
                    pl.cell(84, col).value = f.replace(
                        f"'P&L Mapping'!{ml}146",
                        f"'P&L Mapping'!{ml}146-'P&L Mapping'!{ml}{pr}")
        kind = ("capitalised" if spec["capitalise"]
                else "revenue" if pr == 19 else "expense")
        done.append((spec["acct"], kind, pr))
    return done


def col_to_mapping_letter(front_col: int) -> int:
    """Front P&L month col -> the P&L Mapping month col it references.

    Front col 16 (Mar) references mapping col D(4); each +1 front col -> +1
    mapping col.  So mapping_col = front_col - 12.
    """
    return front_col - 12


# --- Front statements: extend a formula column to the right (relative refs) --
def extend_formula_column(ws, src_col: int, dst_col: int):
    """Copy every formula in src_col to dst_col, translating relative refs."""
    n = 0
    for r in range(1, ws.max_row + 1):
        cell = ws.cell(r, src_col)
        f = cell.value
        if isinstance(f, str) and f.startswith("="):
            src = f"{get_column_letter(src_col)}{r}"
            dst = f"{get_column_letter(dst_col)}{r}"
            ws.cell(r, dst_col).value = Translator(f, origin=src).translate_formula(dst)
            ws.cell(r, dst_col).number_format = cell.number_format
            n += 1
    return n


# --- Balance Sheet (Germany-only, mirrors the BWA->Mapping->front pattern) ----
# raw BS file (Kontennachweis): account col1, EUR col3, Geschäftsjahr col4.
# 'Balance Sheet' data sheet: account col2; each month = 3 cols (EUR/FinYr/PY),
#   Jan EUR=4 -> month m EUR col = 1 + 3*m, FinYr = +1.  (May EUR=16, FinYr=17.)
def bs_eur_col(month: int) -> int:
    return 1 + 3 * month


def populate_balance_sheet(wb, mm: str, month: int):
    """Paste a month's German BS balances into the 'Balance Sheet' data sheet."""
    bsd = wb["Balance Sheet"]
    raw = openpyxl.load_workbook(
        RAW / f"{mm}-2026" / f"BS {mm}-2026.xlsx", data_only=True).worksheets[0]
    raw_vals = {}
    for r in range(3, raw.max_row + 1):
        a = raw.cell(r, 1).value
        if a not in (None, "") and str(a).strip() not in ("", "Konto"):
            raw_vals[str(a).strip()] = (_num(raw.cell(r, 3).value), _num(raw.cell(r, 4).value))
    rows = {}
    for r in range(9, bsd.max_row + 1):
        a = bsd.cell(r, 2).value
        if a not in (None, ""):
            rows.setdefault(str(a).strip(), r)
    eur_col = bs_eur_col(month)
    written, unmapped = 0, []
    for acc, (eur, finyr) in raw_vals.items():
        row = rows.get(acc)
        if row is None:
            if (eur or finyr):
                unmapped.append((acc, eur, finyr))
            continue
        bsd.cell(row, eur_col).value = eur
        bsd.cell(row, eur_col + 1).value = finyr
        written += 1

    # Manual subtotal rows: a tagged row with no account number holds the sum of
    # the contiguous untagged account rows above it (e.g. Cash = the 4 bank
    # accounts, PP&E = 200 0 + 210 0). Recompute them into the FinYr column.
    for r in range(9, bsd.max_row + 1):
        if bsd.cell(r, 1).value in (None, "") or bsd.cell(r, 2).value not in (None, ""):
            continue
        total, rr = 0.0, r - 1
        while rr >= 9 and bsd.cell(rr, 2).value not in (None, "") \
                and bsd.cell(rr, 1).value in (None, ""):
            total += (_num(bsd.cell(rr, eur_col).value) or 0)
            total += (_num(bsd.cell(rr, eur_col + 1).value) or 0)
            rr -= 1
        if rr < r - 1:                      # found at least one untagged account
            bsd.cell(r, eur_col + 1).value = total
    return written, unmapped


def _shift_bs_cols(formula: str, step: int) -> str:
    """Shift the absolute 'Balance Sheet'!$X$1:$X$168 column refs by `step`."""
    def repl(m):
        n1 = get_column_letter(column_index_from_string(m.group(1)) + step)
        n2 = get_column_letter(column_index_from_string(m.group(2)) + step)
        return f"'Balance Sheet'!${n1}$1:${n2}$168"
    return re.sub(r"'Balance Sheet'!\$([A-Z]+)\$1:\$([A-Z]+)\$168", repl, formula)


def extend_bs_front(wb):
    """Build front-BS April (col18) + May (col19) from March (col17).

    SUMIF rows use ABSOLUTE Balance-Sheet column refs -> shift by +3/+6 explicitly
    (the row's own tag ref A{r} stays). Subtotal rows use relative front refs ->
    Translator shifts them correctly.
    """
    bs = wb["BS"]
    n = 0
    for r in range(1, bs.max_row + 1):
        f = bs.cell(r, 17).value          # March
        if not isinstance(f, str) or not f.startswith("="):
            continue
        for dst, step in ((18, 3), (19, 6)):
            if "'Balance Sheet'" in f:
                bs.cell(r, dst).value = _shift_bs_cols(f, step)
            else:
                bs.cell(r, dst).value = Translator(
                    f, origin=f"Q{r}").translate_formula(f"{get_column_letter(dst)}{r}")
            bs.cell(r, dst).number_format = bs.cell(r, 17).number_format
            n += 1
    return n


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--golden", action="store_true", help="reproduce April and diff")
    ap.add_argument("--month", type=int, default=5)
    ap.add_argument("--base", default=str(MR_DIR / "mr_2026-04.xlsx"))
    ap.add_argument("--out", default=str(MR_DIR / "mr_2026-05.xlsx"))
    args = ap.parse_args()

    if args.golden:
        print("GOLDEN: reproducing April data-sheet columns from raw 04...")
        n = golden_april(Path(args.base))
        print(f"\nTotal April account-level mismatches: {n}")
        raise SystemExit(0 if n == 0 else 1)

    mm = f"{args.month:02d}"
    print(f"Building {args.out} — populating month {mm} data columns from raw {mm}...")
    wb = openpyxl.load_workbook(args.base)  # keep formulas

    # 1) German P&L data: paste BWA month column (validated, account-keyed).
    for sheet, spec in DATA_SHEETS.items():
        rep = populate_sheet(wb, sheet, spec, args.month, mm, write=True)
        print(f"  {sheet}: wrote {rep['written']} accounts into col {rep['mr_col']} "
              f"({rep['changed']} changed); unmapped={len(rep['unmapped'])}")
        for acc, val in rep["unmapped"]:
            print(f"      UNMAPPED account {acc} = {val:.2f}  (raw has it, MR has no row)")

    # 2) Serbia P&L: label-matched copy of the template month column.
    written, restated = populate_serbia(wb, args.month, mm)
    print(f"  Serbia: wrote {written} P&L rows into col {2 + args.month}")
    for lab, pm, a, b in restated:
        print(f"      RESTATED Serbia {lab!r} month {pm}: MR={a:.2f} -> template={b:.2f}")

    # 3) Front P&L: extend the formula column from the last built month (March,
    #    col 16) into April (17) and May (18). Relative refs shift to the
    #    matching P&L Mapping / Serbia columns, which are now populated.
    pl = wb["P&L"]
    for dst in (17, 18):
        n = extend_formula_column(pl, src_col=16, dst_col=dst)
        print(f"  Front P&L: extended {n} formulas into col {dst} "
              f"({'Apr' if dst == 17 else 'May'} 2026)")
    print("  NOTE: CF + BS fronts NOT extended — they need CR-Upload (manual CF "
          "reclassifications) and the Trial balance / Balance Sheet data layers first.")

    # 4) Resolve May's new accounts: 83380 -> revenue 'Other'; 31004 -> capitalised
    for acct, kind, pr in add_mapping_rows(wb, mm):
        print(f"  Mapping: {acct} -> {kind} (BWA + P&L Mapping row {pr})")

    # 5) Balance Sheet (Germany-only): paste Apr + May German balances, then
    #    extend the front BS formulas to cover both months.
    for pm_mm, pm_month in (("04", 4), (str(args.month).zfill(2), args.month)):
        w, un = populate_balance_sheet(wb, pm_mm, pm_month)
        print(f"  Balance Sheet {pm_mm}: pasted {w} accounts; unmapped={len(un)}")
        for acc, e, fy in un:
            print(f"      UNMAPPED BS account {acc} = EUR {e or 0:.2f} / FinYr {fy or 0:.2f}")
    nbs = extend_bs_front(wb)
    print(f"  Front BS: extended {nbs} formulas into Apr (col18) + May (col19)")

    wb.save(args.out)
    print(f"Saved {args.out}")


if __name__ == "__main__":
    main()
