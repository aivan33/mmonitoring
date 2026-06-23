"""Restructure the BS & CF statements to the reference structure (sub-groups + the missing lines),
leaving lines we don't model BLANK but DEFINED. Thin statements → rebuild from a declarative layout:
existing wired lines keep their full per-column formulas (matched by label), new lines are blank,
totals are explicit sums of the real value rows. The yearly sheets (BS_Y/CF_Y) reference these by
row, so their refs are remapped old→new. Gate: BS check=0 preserved (blank lines are 0).
"""
from __future__ import annotations

import re

from openpyxl.worksheet.formula import ArrayFormula

FIRST, LAST = 3, 62  # month columns C..BJ


def _ft(c):
    v = c.value
    return v.text if isinstance(v, ArrayFormula) else v


def _snapshot(ws):
    """label → list of (value, style) across all columns, for the current rows."""
    out = {}
    for r in range(1, ws.max_row + 1):
        lbl = ws.cell(r, 1).value
        if isinstance(lbl, str) and lbl.strip():
            out[lbl.strip()] = [(ws.cell(r, c).value, ws.cell(r, c)._style) for c in range(1, LAST + 1)]
    return out


def _remap_year(wb, ysheet, src, old2new):
    pat = re.compile(rf"({src}!\$?[A-Z]{{1,2}}\$?)(\d+)")
    for row in wb[ysheet].iter_rows():
        for c in row:
            t = _ft(c)
            if isinstance(t, str) and f"{src}!" in t:
                c.value = pat.sub(lambda m: f"{m.group(1)}{old2new.get(int(m.group(2)), int(m.group(2)))}", t)


# BS layout: ('keep', current-label[, new-label]) · ('new', label) · ('sub', label) · ('hdr', label)
#            · ('asset_total'|'el_total'|'check'|'title'|'ccy'|'blank')   (value rows tracked for totals)
BS_LAYOUT = [
    ("title", "Balance Sheet — Monthly (EUR)"), ("ccy",), ("blank",),
    ("hdr", "ASSETS"),
    ("sub", "  Non-current assets"),
    ("keep", "Intangible fixed assets (R&D)"), ("keep", "Tangible fixed assets (PP&E)"),
    ("new", "    Business equipment"),
    ("sub", "  Current assets"),
    ("keep", "Cash & cash equivalents"), ("keep", "Trade receivable", "Trade receivables"),
    ("new", "    Prepaid expenses"), ("new", "    Other receivables"),
    ("sub", "  Inventory"),
    ("new", "    Raw materials"), ("new", "    Work-in-progress"), ("new", "    Finished goods"),
    ("asset_total", "TOTAL ASSETS"), ("blank",),
    ("hdr", "EQUITY & LIABILITIES"),
    ("sub", "  Equity"),
    ("keep", "Share capital"), ("keep", "Retained earnings"),
    ("sub", "  Long-term liabilities"),
    ("keep", "Loan facility financing"), ("new", "    Grants"),
    ("sub", "  Current liabilities"),
    ("keep", "Trade payables"), ("keep", "Personnel & social security payables"),
    ("keep", "Deferred revenue"), ("keep", "Tax payables"), ("new", "    Other payables"),
    ("el_total", "TOTAL EQUITY & LIABILITIES"), ("check", "check (Assets − E&L)"), ("blank",),
    ("hdr", "KPIs"),
    ("new", "  Current ratio"), ("new", "  Quick ratio"), ("new", "  Cash ratio"),
]


def restructure_bs(wb):
    from openpyxl.utils import get_column_letter
    bs = wb["BS"]
    snap = _snapshot(bs)
    old_row = {bs.cell(r, 1).value.strip(): r for r in range(1, bs.max_row + 1)
               if isinstance(bs.cell(r, 1).value, str) and bs.cell(r, 1).value.strip()}
    hdr_st = snap["ASSETS"][0][1]
    sub_st = snap["Intangible fixed assets (R&D)"][0][1]
    tot_st = snap["TOTAL ASSETS"]                      # full per-column total style

    # --- plan pass: assign rows, track asset/el value rows + total/check rows + old→new ---
    plan, nr, o2n, asset_rows, el_rows = [], 0, {2: None}, [], []
    ta_row = tel_row = None
    section = None
    for item in BS_LAYOUT:
        nr += 1
        plan.append((nr, item))
        k = item[0]
        if k == "hdr":
            section = {"ASSETS": "A", "EQUITY & LIABILITIES": "E", "KPIs": "K"}.get(item[1], section)
        if k in ("keep", "new") and section in ("A", "E"):
            (asset_rows if section == "A" else el_rows).append(nr)
        if k == "ccy":
            o2n[2] = nr
        elif k == "keep":
            o2n[old_row[item[1]]] = nr
        elif k == "asset_total":
            o2n[old_row[item[1]]] = nr; ta_row = nr
        elif k == "el_total":
            o2n[old_row[item[1]]] = nr; tel_row = nr
        elif k == "check":
            o2n[old_row[item[1]]] = nr
    last = nr

    # --- clear + write ---
    maxr = bs.max_row
    for r in range(1, maxr + 1):
        for c in range(1, LAST + 1):
            cell = bs.cell(r, c); cell.value = None; cell.style = "Normal"
    for nr, item in plan:
        k = item[0]
        if k == "title" or k == "hdr":
            bs.cell(nr, 1, item[1])._style = hdr_st
        elif k == "ccy":
            bs.cell(nr, 1, "Currency: EUR")._style = sub_st
            bs.cell(nr, 3, "=IS!C2")
        elif k == "sub" or k == "new":
            bs.cell(nr, 1, item[1])._style = sub_st
        elif k == "keep":
            for c, (val, st) in enumerate(snap[item[1]], start=1):
                bs.cell(nr, c).value = val
                bs.cell(nr, c)._style = st
            if len(item) == 3:
                bs.cell(nr, 1, item[2])._style = snap[item[1]][0][1]
        elif k in ("asset_total", "el_total", "check"):
            bs.cell(nr, 1, item[1])._style = hdr_st
            for c in range(FIRST, LAST + 1):
                x = get_column_letter(c)
                if k == "check":
                    f = f"={x}{ta_row}-{x}{tel_row}"
                else:
                    rows = asset_rows if k == "asset_total" else el_rows
                    f = "=" + "+".join(f"{x}{rr}" for rr in rows)
                bs.cell(nr, c, f)._style = tot_st[c - 1][1]
    if maxr > last:
        bs.delete_rows(last + 1, maxr - last)

    # rebuild BS_Y to MIRROR the new BS — each value/total/check row = BS December snapshot per year
    bsy = wb["BS_Y"]
    valrow = next(r for r in range(1, bsy.max_row + 1)
                  if isinstance(_ft(bsy.cell(r, 3)), str) and "BS!" in _ft(bsy.cell(r, 3)))
    deccols = [re.search(r"BS!\$?([A-Z]+)", _ft(bsy.cell(valrow, c))).group(1) for c in range(3, 8)]
    yhdr = [(bsy.cell(2, c).value, bsy.cell(2, c)._style) for c in range(1, 9)]
    title_st = bsy.cell(1, 1)._style
    ymax = bsy.max_row
    for r in range(1, ymax + 1):
        for c in range(1, 9):
            cell = bsy.cell(r, c); cell.value = None; cell.style = "Normal"
    bsy.cell(1, 1, "Balance Sheet — Yearly (calendar year-end; 2026 = Dec)")._style = title_st
    for c, (v, st) in enumerate(yhdr, start=1):
        bsy.cell(2, c).value = v
        bsy.cell(2, c)._style = st
    for nr, item in plan:
        if nr <= 2:
            continue
        k = item[0]
        if k in ("hdr", "sub", "new"):
            bsy.cell(nr, 1, item[1])._style = (hdr_st if k == "hdr" else sub_st)
        elif k == "keep":
            bsy.cell(nr, 1, item[2] if len(item) == 3 else item[1])._style = snap[item[1]][0][1]
            for y, dc in enumerate(deccols):
                bsy.cell(nr, 3 + y, f"=BS!{dc}{nr}")._style = snap[item[1]][0][1]
        elif k in ("asset_total", "el_total", "check"):
            bsy.cell(nr, 1, item[1])._style = hdr_st
            for y, dc in enumerate(deccols):
                bsy.cell(nr, 3 + y, f"=BS!{dc}{nr}")._style = tot_st[2][1]
    if ymax > last:
        bsy.delete_rows(last + 1, ymax - last)
    print(f"  BS: restructured + BS_Y mirrored ({last} rows, {len(asset_rows)} assets / {len(el_rows)} E&L)")


CF_LAYOUT = [
    ("title", "Cash Flow Statement — Monthly (Direct method, EUR)"), ("ccy",), ("blank",),
    ("sub", "Operating activities"),
    ("op", "Cash received from customers"), ("op", "Cash paid to suppliers"),
    ("op", "Payment for personnel and social security"), ("op", "Bank charges paid"),
    ("op_new", "  Recovery/(repayment) of VAT"), ("op", "Corporate and other taxes, net"),
    ("op", "Movement in deferred revenue"), ("op_new", "  Other (operating)"),
    ("op_total", "Cash Flow from Operating Activities"), ("blank",),
    ("sub", "Investing activities"),
    ("inv", "CAPEX"), ("inv", "R&D (capitalised)"), ("inv_new", "  Other (investing)"),
    ("inv_total", "Cash Flow from Investing Activities"), ("blank",),
    ("sub", "Financing activities"),
    ("fin", "Capital Increase"), ("fin", "Loan facility financing"), ("fin", "Grants"),
    ("fin_new", "  Other payments / proceeds, net"), ("fin_new", "  Distribution of dividends"),
    ("fin_total", "Cash Flow from Financing Activities"), ("blank",),
    ("excess", "Excess Cash for the Period"), ("begin", "Beginning Cash Balance"),
    ("end", "Ending Cash Balance"), ("pct", "% Change in cash"), ("blank",),
    ("burn", "Net Cash Burn (Operating + Investing)"), ("new", "  Average Monthly Burn 2026"),
]
YEAR_RANGES = [("C", "H"), ("I", "T"), ("U", "AF"), ("AG", "AR"), ("AS", "BD")]  # calendar-year month spans


def restructure_cf(wb):
    from openpyxl.utils import get_column_letter
    cf = wb["CF"]
    snap = _snapshot(cf)
    old_row = {cf.cell(r, 1).value.strip(): r for r in range(1, cf.max_row + 1)
               if isinstance(cf.cell(r, 1).value, str) and cf.cell(r, 1).value.strip()}
    sub_st = snap["Cash received from customers"][0][1]
    tot_st = snap["Cash Flow from Operating Activities"]
    open_ref = re.search(r"' Inputs'!\$J\$\d+", snap["Beginning Cash Balance"][2][0]).group(0)

    # plan
    plan, nr, o2n = [], 0, {2: None}
    op, inv, fin = [], [], []
    rows = {}
    for item in CF_LAYOUT:
        nr += 1
        plan.append((nr, item))
        k = item[0]
        if k in ("op", "op_new"):
            op.append(nr)
        elif k in ("inv", "inv_new"):
            inv.append(nr)
        elif k in ("fin", "fin_new"):
            fin.append(nr)
        if k == "ccy":
            o2n[2] = nr
        elif k in ("op", "inv", "fin"):
            o2n[old_row[item[1]]] = nr
        elif k in ("op_total", "inv_total", "fin_total", "excess", "begin", "end", "pct"):
            o2n[old_row[item[1]]] = nr
            rows[k] = nr
    last = nr

    maxr = cf.max_row
    for r in range(1, maxr + 1):
        for c in range(1, LAST + 1):
            cell = cf.cell(r, c); cell.value = None; cell.style = "Normal"
    for nr, item in plan:
        k = item[0]
        if k in ("title", "sub"):
            cf.cell(nr, 1, item[1])._style = (tot_st[0][1] if k == "title" else sub_st)
        elif k == "ccy":
            cf.cell(nr, 1, "Currency: EUR")._style = sub_st
            cf.cell(nr, 3, "=IS!C2")
        elif k in ("op_new", "inv_new", "fin_new", "new"):
            cf.cell(nr, 1, item[1])._style = sub_st
        elif k in ("op", "inv", "fin"):
            for c, (val, st) in enumerate(snap[item[1]], start=1):
                cf.cell(nr, c).value = val
                cf.cell(nr, c)._style = st
        elif k in ("op_total", "inv_total", "fin_total", "excess", "begin", "end", "pct", "burn"):
            cf.cell(nr, 1, item[1])._style = tot_st[0][1]
            for c in range(FIRST, LAST + 1):
                x, px = get_column_letter(c), get_column_letter(c - 1)
                if k == "op_total":
                    f = f"=SUM({x}{op[0]}:{x}{op[-1]})"
                elif k == "inv_total":
                    f = f"=SUM({x}{inv[0]}:{x}{inv[-1]})"
                elif k == "fin_total":
                    f = f"=SUM({x}{fin[0]}:{x}{fin[-1]})"
                elif k == "excess":
                    f = f"={x}{rows['op_total']}+{x}{rows['inv_total']}+{x}{rows['fin_total']}"
                elif k == "begin":
                    f = f"={open_ref}" if c == FIRST else f"={px}{rows['end']}"
                elif k == "end":
                    f = f"={x}{rows['begin']}+{x}{rows['excess']}"
                elif k == "pct":
                    f = f"=IF({x}{rows['begin']}=0,0,{x}{rows['end']}/{x}{rows['begin']}-1)"
                elif k == "burn":
                    f = f"={x}{rows['op_total']}+{x}{rows['inv_total']}"
                cf.cell(nr, c, f)._style = tot_st[c - 1][1]
    if maxr > last:
        cf.delete_rows(last + 1, maxr - last)

    # remap BS's =CF! cash ref to the new ending row, then rebuild CF_Y to mirror CF
    _remap_year(wb, "BS", "CF", o2n)
    cfy = wb["CF_Y"]
    yhdr = [(cfy.cell(2, c).value, cfy.cell(2, c)._style) for c in range(1, 9)]
    title_st = cfy.cell(1, 1)._style
    ymax = cfy.max_row
    for r in range(1, ymax + 1):
        for c in range(1, 9):
            cell = cfy.cell(r, c); cell.value = None; cell.style = "Normal"
    cfy.cell(1, 1, "Cash Flow — Yearly (calendar; 2026 = Jul–Dec)")._style = title_st
    for c, (v, st) in enumerate(yhdr, start=1):
        cfy.cell(2, c).value = v
        cfy.cell(2, c)._style = st
    for nr, item in plan:
        if nr <= 2:
            continue
        k = item[0]
        if k in ("sub", "op_new", "inv_new", "fin_new", "new"):
            cfy.cell(nr, 1, item[1])._style = sub_st
        elif k in ("op", "inv", "fin", "op_total", "inv_total", "fin_total", "excess", "burn"):
            cfy.cell(nr, 1, item[1])._style = (sub_st if k in ("op", "inv", "fin") else tot_st[0][1])
            for y, (a, b) in enumerate(YEAR_RANGES):
                cfy.cell(nr, 3 + y, f"=SUM(CF!{a}{nr}:{b}{nr})")._style = sub_st
        elif k == "begin":
            cfy.cell(nr, 1, item[1])._style = tot_st[0][1]
            for y, (a, b) in enumerate(YEAR_RANGES):
                cfy.cell(nr, 3 + y, f"=CF!{a}{nr}")._style = sub_st
        elif k == "end":
            cfy.cell(nr, 1, item[1])._style = tot_st[0][1]
            for y in range(len(YEAR_RANGES)):
                x = get_column_letter(3 + y)
                cfy.cell(nr, 3 + y, f"={x}{rows['begin']}+{x}{rows['excess']}")._style = sub_st
        elif k == "pct":
            cfy.cell(nr, 1, item[1])._style = tot_st[0][1]
            for y in range(len(YEAR_RANGES)):
                x = get_column_letter(3 + y)
                cfy.cell(nr, 3 + y, f"=IF({x}{rows['begin']}=0,0,{x}{rows['end']}/{x}{rows['begin']}-1)")._style = sub_st
    if ymax > last:
        cfy.delete_rows(last + 1, ymax - last)
    print(f"  CF: restructured + CF_Y mirrored ({last} rows)")
