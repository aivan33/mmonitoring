"""Standalone investor unit-economics file for the FaradaIC sensor (hardware only).

Cost of Sales is built bottom-up from the WAFER: 1 wafer = N sensors, so
chip cost/unit = wafer cost / sensors-per-wafer. This is why a 1-unit order and
a 1,000-unit order cost the same (both inside one wafer's economics); unit cost
only steps down at the wafer-pricing / packaging-lot / scale thresholds.

No product lines — pure unit economics. Flow per volume band:
    volume -> chip+packaging+testing+ASIC -> unit cost -> ASP -> gross profit/margin

Founder inputs still to confirm are filled yellow and tagged "← confirm".
Run:  .venv/bin/python clients/farada/one_offs/build_unit_economics.py
"""
from __future__ import annotations

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

DST = "clients/farada/modeling/farada_unit_economics.xlsx"

YELLOW = PatternFill("solid", fgColor="FFFEF2CB")   # founder input cell
BLUE = PatternFill("solid", fgColor="FFD5EBF4")     # output (ASP / GP / GM)
GREY = PatternFill("solid", fgColor="FFD8D8D8")     # section / table header
COSTF = PatternFill("solid", fgColor="FFF2F2F2")    # cost build-up band
THIN = Side(style="thin", color="FFBFBFBF")
BORD = Border(THIN, THIN, THIN, THIN)
EUR = "€#,##0.00"
PCT = "0.0%"
INTF = "#,##0"


def build():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Unit Economics"
    ws.sheet_view.showGridLines = False

    def put(coord, val, *, fill=None, bold=False, fmt=None, align=None,
            border=False, size=10, italic=False, color=None):
        c = ws[coord]
        c.value = val
        c.font = Font(bold=bold, size=size, italic=italic, color=color)
        if fill: c.fill = fill
        if fmt: c.number_format = fmt
        if border: c.border = BORD
        wrap = align == "wrapcenter"
        horiz = "center" if wrap else align
        c.alignment = Alignment(horizontal=horiz, vertical="center", wrap_text=wrap)
        return c

    # ---------------------------------------------------------------- title
    put("A1", "FaradaIC — Sensor Unit Economics", bold=True, size=15)
    put("A2", "Hardware only. Cost of Sales built bottom-up from the wafer; "
              "Cost of Sales steps down at the wafer / packaging-lot / scale thresholds.",
        italic=True, size=9, color="FF666666")

    # ---------------------------------------------------------------- inputs
    put("A4", "INPUTS  (yellow = founder input)", bold=True, fill=GREY)
    ws.merge_cells("A4:C4")
    rows = [
        ("Sensors per wafer", 3000, INTF, False),
        ("Wafer cost — standard (€/wafer)", 2880, "€#,##0", True),
        ("Wafer cost — at scale (€/wafer)", 1440, "€#,##0", True),
        ("Yield (good die %)", 1.00, PCT, True),
        ("Packaging €/unit — < lot threshold", 0.30, EUR, False),
        ("Packaging €/unit — ≥ lot threshold", 0.10, EUR, False),
        ("Testing €/unit — standard", 0.50, EUR, False),
        ("Testing €/unit — at scale", 0.01, EUR, False),
        ("ASIC / readout €/unit — discrete", 1.04, EUR, False),
        ("ASIC / readout €/unit — own ASIC (scale)", 0.32, EUR, False),
        ("Packaging lot threshold (units)", 100000, INTF, False),
        ("Scale threshold — chip/test/ASIC (units)", 1000000, INTF, False),
    ]
    r0 = 5
    for i, (lbl, val, fmt, flag) in enumerate(rows):
        r = r0 + i
        put(f"A{r}", lbl, size=10)
        c = put(f"B{r}", val, fill=YELLOW, fmt=fmt, align="center", border=True, bold=True)
        if flag:
            put(f"C{r}", "← confirm with Ryan", italic=True, size=8, color="FFC00000")
    # named refs
    SPW, WSTD, WSCL, YLD = "$B$5", "$B$6", "$B$7", "$B$8"
    PKL, PKH, TSL, TSH = "$B$9", "$B$10", "$B$11", "$B$12"
    ASL, ASH, LOT, SCL = "$B$13", "$B$14", "$B$15", "$B$16"

    # ASP ladder (editable)
    put("D4", "ASP LADDER  (€/unit by order size)", bold=True, fill=GREY)
    ws.merge_cells("D4:E4")
    put("D5", "≥ units", bold=True, align="center", border=True)
    put("E5", "€/unit", bold=True, align="center", border=True)
    ladder = [(1, 125), (100, 75), (1000, 49), (10000, 29),
              (100000, 19), (100001, 15), (1000000, 5)]
    for i, (thr, price) in enumerate(ladder):
        r = 6 + i
        put(f"D{r}", thr, fmt=INTF, align="center", border=True)
        put(f"E{r}", price, fmt="€#,##0", fill=YELLOW, align="center", border=True, bold=True)
    # ladder price cells E6..E12, thresholds D6..D12
    def asp(vol_cell):
        # nested IF with the HIGHEST threshold outermost: IF(v>=D12,E12,IF(v>=D11,...E6))
        expr = "$E$6"
        for i in range(1, len(ladder)):
            expr = f"IF({vol_cell}>=$D${6+i},$E${6+i},{expr})"
        return "=" + expr

    # ---------------------------------------------------------------- table
    T = 19
    heads = ["Annual\nvolume (units)", "Wafers", "Chip\n€/u", "Packaging\n€/u",
             "Testing\n€/u", "ASIC\n€/u", "UNIT COST\n€/u", "ASP\n€/u",
             "Gross profit\n€/u", "Gross\nmargin %"]
    put(f"A{T-1}", "UNIT ECONOMICS  →  cost build-up flows into ASP and gross margin",
        bold=True, fill=GREY)
    ws.merge_cells(f"A{T-1}:J{T-1}")
    for j, h in enumerate(heads):
        c = put(f"{get_column_letter(1+j)}{T}", h, bold=True, fill=GREY,
                align="wrapcenter", border=True, size=9)

    bands = [1, 100, 1000, 3000, 10000, 100000, 1000000, 5000000, 10000000]
    for i, v in enumerate(bands):
        r = T + 1 + i
        A = f"A{r}"
        put(A, v, fmt=INTF, align="center", border=True)
        put(f"B{r}", f"=ROUNDUP(A{r}/{SPW},0)", fmt=INTF, align="center", border=True)
        # chip = wafer cost / sensors per wafer / yield, stepped at scale threshold
        put(f"C{r}", f"=IF(A{r}>={SCL},{WSCL},{WSTD})/{SPW}/{YLD}",
            fmt=EUR, fill=COSTF, border=True)
        put(f"D{r}", f"=IF(A{r}>={LOT},{PKH},{PKL})", fmt=EUR, fill=COSTF, border=True)
        put(f"E{r}", f"=IF(A{r}>={SCL},{TSH},{TSL})", fmt=EUR, fill=COSTF, border=True)
        put(f"F{r}", f"=IF(A{r}>={SCL},{ASH},{ASL})", fmt=EUR, fill=COSTF, border=True)
        put(f"G{r}", f"=C{r}+D{r}+E{r}+F{r}", fmt=EUR, border=True, bold=True)
        put(f"H{r}", asp(A), fmt=EUR, fill=BLUE, border=True, bold=True)
        put(f"I{r}", f"=H{r}-G{r}", fmt=EUR, fill=BLUE, border=True)
        put(f"J{r}", f"=IF(H{r}=0,0,I{r}/H{r})", fmt=PCT, fill=BLUE, border=True, bold=True)

    # ---------------------------------------------------------- thresholds note
    n = T + len(bands) + 3
    put(f"A{n}", "CRITICAL THRESHOLDS — where Cost of Sales steps down", bold=True, fill=GREY)
    ws.merge_cells(f"A{n}:J{n}")
    notes = [
        "1 wafer = 3,000 sensors → below one wafer, unit cost is flat (1 unit and 1,000 "
        "units share the same wafer economics).",
        "≥ 100,000 units (packaging lot): packaging €0.30 → €0.10.",
        "≥ 1,000,000 units (scale): chip halves (cheaper wafers), testing €0.50 → €0.01, "
        "own ASIC €1.04 → €0.32.",
        "TODO (need Ryan): finer steps — e.g. a ~10k point and a 10M mass-scale wafer price; "
        "yield %; confirm wafer costs.",
    ]
    for i, t in enumerate(notes):
        put(f"A{n+1+i}", ("• " + t), size=9, italic=(i == 3),
            color=("FFC00000" if i == 3 else None))
        ws.merge_cells(f"A{n+1+i}:J{n+1+i}")

    # widths
    for col, w in {"A": 34, "B": 11, "C": 10, "D": 11, "E": 10, "F": 9,
                   "G": 12, "H": 10, "I": 13, "J": 10}.items():
        ws.column_dimensions[col].width = w
    ws.row_dimensions[T].height = 30

    wb.save(DST)
    print(f"Saved {DST}")


if __name__ == "__main__":
    build()
