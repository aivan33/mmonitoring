"""Build farada_model_v4 from farada_model_v3.5: split the ProForma (calculation engine)
from a real Income Statement.

User feedback on v3.5: the ProForma was being used AS the income statement. It should be the
calc engine (drivers + value lines + the capex/depreciation schedule); the *statement* and its
margins/KPIs belong on a separate Income Statement. So v4:

  - `IS`   — new monthly Income Statement: a clean CONTIGUOUS full mirror that PULLS value lines
             from the ProForma (`=ProForma!<col><row>`) and COMPUTES the margins on itself.
  - `IS_Y` — yearly Income Statement on CALENDAR years (2026 = Jul–Dec partial … 2030 full;
             2031 H1 dropped), full statement incl. EBITDA + below; flows = SUM of IS monthly
             columns per year; margins recomputed.
  - ProForma `% margin/KPI rows` (56-61, 117, 121, 122, 126, 133) are blanked — they live on IS.

Reads v3.5 (preserved), writes v4. Idempotent. No recalc engine → verify_model_v4.py checks
structure. Run:  .venv/bin/python clients/farada/one_offs/build_model_v4.py
"""
from __future__ import annotations

from copy import copy

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.formula import ArrayFormula

SRC = "clients/farada/modeling/farada_model_v3.5.xlsx"
DST = "clients/farada/modeling/farada_model_v4.xlsx"
FIRST, LAST = 3, 62                       # ProForma/IS month cols C..BJ (Jul-2026..Jun-2031)

# Statement rows in the ProForma P&L (drivers/capex-schedule excluded).
PL_SRC = list(range(4, 62)) + list(range(85, 134))
# %-margin rows -> (numerator src row, denominator src row); recomputed on the statement.
PCT_SRC = {56: (44, 4), 57: (45, 5), 58: (46, 9), 59: (47, 14), 60: (48, 15), 61: (52, 19),
           117: (116, 4), 122: (116, 4), 126: (125, 4), 133: (132, 4)}
PCT_SPECIAL = {121: (116, 120, 4)}        # EBITDA margin incl. grant = (EBITDA+grant)/rev
MARGIN_ROWS = list(PCT_SRC) + list(PCT_SPECIAL)
# calendar year -> (first ProForma col, last col). 2026 = Jul-Dec (partial); 2031 H1 dropped.
CAL_YEARS = [("2026", 3, 8), ("2027", 9, 20), ("2028", 21, 32), ("2029", 33, 44), ("2030", 45, 56)]


def _ft(cell):
    v = cell.value
    return v.text if isinstance(v, ArrayFormula) else v


def statement_spec(ws):
    """Ordered statement lines assigned to CONTIGUOUS rows (single-row spacers between blocks);
    returns (items, src->is_row map). Each item: {is_row, src, kind in VAL|PCT|HEADER, label}."""
    items, m, is_r, prev = [], {}, 4, None
    for src in PL_SRC:
        a, cf = ws.cell(src, 1).value, _ft(ws.cell(src, 3))
        if a is None and cf is None:
            continue
        if prev is not None and src > prev + 1:
            is_r += 1                                   # collapse any gap to ONE spacer row
        kind = "HEADER" if cf is None else ("PCT" if "%" in ws.cell(src, 3).number_format else "VAL")
        items.append({"is_row": is_r, "src": src, "kind": kind, "label": a})
        m[src] = is_r
        prev, is_r = src, is_r + 1
    return items, m


def _pct_formula(col, src, m):
    if src in PCT_SPECIAL:
        n1, n2, den = PCT_SPECIAL[src]
        return f"=IF({col}{m[den]}=0,0,({col}{m[n1]}+{col}{m[n2]})/{col}{m[den]})"
    num, den = PCT_SRC[src]
    return f"=IF({col}{m[den]}=0,0,{col}{m[num]}/{col}{m[den]})"


def build_is(wb):
    ws = wb["ProForma"]
    items, m = statement_spec(ws)
    if "IS" in wb.sheetnames:
        del wb["IS"]
    s = wb.create_sheet("IS", index=wb.sheetnames.index("ProForma") + 1)
    s.sheet_view.showGridLines = False
    s.freeze_panes = "C3"
    s.column_dimensions["A"].width = ws.column_dimensions["A"].width or 42
    s["A1"] = "Income Statement — Monthly (EUR)"
    s["A1"]._style = copy(ws.cell(116, 1)._style)
    s.cell(2, 1, ws.cell(2, 1).value)._style = copy(ws.cell(2, 1)._style)   # "Currency: EUR"
    for c in range(FIRST, LAST + 1):
        s.cell(2, c, f"=ProForma!{get_column_letter(c)}2")._style = copy(ws.cell(2, c)._style)
    for it in items:
        r, src, kind = it["is_row"], it["src"], it["kind"]
        s.cell(r, 1, it["label"])._style = copy(ws.cell(src, 1)._style)
        if kind == "HEADER":
            for c in range(2, LAST + 1):
                s.cell(r, c)._style = copy(ws.cell(src, c)._style)
            continue
        for c in range(FIRST, LAST + 1):
            col = get_column_letter(c)
            f = f"=ProForma!{col}{src}" if kind == "VAL" else _pct_formula(col, src, m)
            s.cell(r, c, f)._style = copy(ws.cell(src, c)._style)
    print(f"  IS: {len(items)} statement rows (monthly, pulls from ProForma)")
    return items, m


def build_is_yearly(wb, items, m):
    s = wb["IS"]
    if "IS_Y" in wb.sheetnames:
        del wb["IS_Y"]
    y = wb.create_sheet("IS_Y", index=wb.sheetnames.index("IS") + 1)
    y.sheet_view.showGridLines = False
    y.freeze_panes = "C3"
    y.column_dimensions["A"].width = s.column_dimensions["A"].width or 42
    y["A1"] = "Income Statement — Yearly (calendar years; 2026 = Jul–Dec)"
    y["A1"]._style = copy(s["A1"]._style)
    ycols = [get_column_letter(3 + k) for k in range(len(CAL_YEARS))]
    for k, (lbl, _, _) in enumerate(CAL_YEARS):
        c = y[f"{ycols[k]}2"]
        c.value = lbl
        c._style = copy(s.cell(2, 3)._style)
        c.number_format = "General"
        y.column_dimensions[ycols[k]].width = 15
    for it in items:
        r, src, kind = it["is_row"], it["src"], it["kind"]
        y.cell(r, 1, s.cell(r, 1).value)._style = copy(s.cell(r, 1)._style)
        if kind == "HEADER":
            for yc in ycols:
                y[f"{yc}{r}"]._style = copy(s.cell(r, 3)._style)
            continue
        for k, (lbl, c0, c1) in enumerate(CAL_YEARS):
            yc = ycols[k]
            if kind == "VAL":
                f = f"=SUM(IS!{get_column_letter(c0)}{r}:{get_column_letter(c1)}{r})"
            else:
                f = _pct_formula(yc, src, m)
            y.cell(r, 3 + k, f)._style = copy(s.cell(r, 3)._style)
    print(f"  IS_Y: {len(CAL_YEARS)} calendar years (2026 partial..2030); 2031 H1 dropped")


def strip_proforma_margins(wb):
    ws = wb["ProForma"]
    for r in MARGIN_ROWS:
        for c in range(1, LAST + 1):
            ws.cell(r, c).value = None
    print(f"  ProForma: blanked {len(MARGIN_ROWS)} % margin rows (moved to IS)")


def build():
    wb = openpyxl.load_workbook(SRC)
    items, m = build_is(wb)              # build statement (reads ProForma incl. margins) ...
    build_is_yearly(wb, items, m)
    strip_proforma_margins(wb)           # ... then strip ProForma margins (IS pulls VALUES)
    wb.save(DST)
    print(f"Saved {DST}")


if __name__ == "__main__":
    build()
