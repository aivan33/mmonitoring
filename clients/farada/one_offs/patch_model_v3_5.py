"""Patch the hand-built farada_model_v3.5.xlsx IN PLACE (idempotent).

v3.5 is the source of truth (user hand-added the HR sheet + re-sorted the inputs into groups;
no builder reproduces it). We extend it by patching only the cells we own, so the user's hand
work is preserved. A one-time pristine backup is written before the first patch (the file is
gitignored — no git safety net). Safe to re-run.

Phase 1 (this file):
  - Salary indexation input — HR escalates every salary by ' Inputs'!$J$250, but that cell is
    empty (=> no indexation). Add the input in its LOGICAL GROUP (end of OPEX, row 137 — a blank
    row, so no row-shift / no $J$NN ref breakage) and repoint HR's 2,340 refs J250 -> J137.
    (Payroll itself is already wired: ProForma OPEX pulls HR subtotal rows O16/O48/O66/O67.)

Run from repo root:  .venv/bin/python clients/farada/one_offs/patch_model_v3_5.py
"""
from __future__ import annotations

import os
import shutil
from copy import copy

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.formula import ArrayFormula

P = "clients/farada/modeling/farada_model_v3.5.xlsx"
BAK = "clients/farada/modeling/farada_model_v3.5.prepatch.xlsx"
FIRST, LAST = 3, 62        # ProForma month columns C..BJ (Jul-2026..Jun-2031)

SAL_IDX_ROW = 137          # salary indexation — blank row at the end of the OPEX group
STYLE_ROW = 33             # "Annual price indexation" (a %-rate input) — style template
SAL_IDX_RATE = 0.03        # 3% default (matches price indexation J33); ← user to confirm

# --- Phase 2: below-EBITDA P&L + explicit D&A schedule (mockup values) -------
# ProForma rows (appended after EBITDA margin r117; r118-119, r134 = spacers).
R_GRANT, R_EBM_I, R_EBM_X = 120, 121, 122
R_DA, R_DEP, R_EBIT, R_EBIT_M = 123, 124, 125, 126
R_FIN_I, R_FIN_C, R_FIN_N, R_PBT, R_TAX, R_NP, R_NP_M = 127, 128, 129, 130, 131, 132, 133
R_SCHED, R_CAPEX, R_OPEN, R_DEPS, R_CLOSE = 135, 136, 137, 138, 139
# Inputs rows (new "BELOW EBITDA" group, after OPEX/notes — empty zone, no shift).
I_HDR, I_CAPEX, I_LIFE, I_OPEN, I_GRANT, I_FIN, I_TAX = 152, 153, 154, 155, 156, 157, 158


def _ft(cell):
    v = cell.value
    return v.text if isinstance(v, ArrayFormula) else v


def add_salary_indexation(wb) -> None:
    inp, hr = wb[" Inputs"], wb["HR"]
    # 1) the input row, styled like the price-indexation rate input (C/D/J/L, no dates).
    for col in (3, 4, 10, 12):
        inp.cell(SAL_IDX_ROW, col)._style = copy(inp.cell(STYLE_ROW, col)._style)
    inp.cell(SAL_IDX_ROW, 3, "Annual salary indexation  ← confirm (defaulted to 3%)")
    inp.cell(SAL_IDX_ROW, 4, "%")
    inp.cell(SAL_IDX_ROW, 10, f"=OFFSET(K{SAL_IDX_ROW},0,$D$2)")
    inp.cell(SAL_IDX_ROW, 12, SAL_IDX_RATE)
    # 2) repoint HR's salary-escalation refs J250 -> J137 (idempotent: no-op once replaced).
    n = 0
    for row in hr.iter_rows():
        for cell in row:
            t = _ft(cell)
            if isinstance(t, str) and "$J$250" in t:
                cell.value = t.replace("$J$250", f"$J${SAL_IDX_ROW}")
                n += 1
    print(f"  salary indexation: input at Inputs!{openpyxl.utils.get_column_letter(12)}{SAL_IDX_ROW}"
          f" = {SAL_IDX_RATE:.0%}; repointed {n} HR cells J250 -> J{SAL_IDX_ROW}")


def add_below_ebitda_and_da(wb) -> None:
    """Complete the P&L to Net Profit (mirrors the reference `is`) with an explicitly
    BUILT D&A: a capex → PP&E → straight-line-depreciation schedule, not a typed input.
    Styles are copied from existing ProForma rows (no naked rows)."""
    ws, inp = wb["ProForma"], wb[" Inputs"]

    # ---- Inputs: new grouped block (after OPEX/notes). Mockup values, flagged. ----
    g, h = inp.cell(113, 7).value, inp.cell(113, 8).value      # OPEX date window (Jul26-Jun31)
    s_hdr = inp.cell(110, 3)._style                            # section-header style
    s_date = {c: inp.cell(113, c)._style for c in (3, 4, 7, 8, 10, 12)}   # date-gated input
    s_rate = {c: inp.cell(33, c)._style for c in (3, 4, 10, 12)}          # scalar/rate input
    inp.cell(I_HDR, 3, "BELOW EBITDA — D&A, FINANCE & TAX  (mockup ← confirm)")._style = copy(s_hdr)

    def date_input(r, label, unit, val):
        for c in (3, 4, 7, 8, 10, 12):
            inp.cell(r, c)._style = copy(s_date[c])
        inp.cell(r, 3, label); inp.cell(r, 4, unit)
        inp.cell(r, 7, g); inp.cell(r, 8, h)
        inp.cell(r, 10, f"=OFFSET(K{r},0,$D$2)"); inp.cell(r, 12, val)

    def scalar_input(r, label, unit, val, numfmt=None):
        for c in (3, 4, 10, 12):
            inp.cell(r, c)._style = copy(s_rate[c])
        inp.cell(r, 3, label); inp.cell(r, 4, unit)
        inp.cell(r, 10, f"=OFFSET(K{r},0,$D$2)")
        cell = inp.cell(r, 12, val)
        if numfmt:
            cell.number_format = numfmt

    date_input(I_CAPEX, "Capex – PP&E (monthly)", "EUR/mo", 10000)
    scalar_input(I_LIFE, "PP&E useful life", "years", 5, numfmt="#,##0")
    scalar_input(I_OPEN, "Opening PP&E (NBV)", "EUR", 0, numfmt="€#,##0")
    date_input(I_GRANT, "Grant financing (monthly)", "EUR/mo", 0)
    date_input(I_FIN, "Finance costs (monthly)", "EUR/mo", 1000)
    scalar_input(I_TAX, "Corporate tax rate", "%", 0.25)

    # ---- ProForma: harvest role styles from existing rows (no naked cells) ----
    tot = {c: ws.cell(116, c)._style for c in range(1, LAST + 1)}   # headline (EBITDA)
    mgn = {c: ws.cell(117, c)._style for c in range(1, LAST + 1)}   # margin %
    leaf = {c: ws.cell(89, c)._style for c in range(1, LAST + 1)}   # plain leaf
    hdr = {c: ws.cell(85, c)._style for c in range(1, LAST + 1)}    # section band
    ROLE = {R_GRANT: leaf, R_EBM_I: mgn, R_EBM_X: mgn, R_DA: leaf, R_DEP: leaf,
            R_EBIT: tot, R_EBIT_M: mgn, R_FIN_I: leaf, R_FIN_C: leaf, R_FIN_N: leaf,
            R_PBT: tot, R_TAX: leaf, R_NP: tot, R_NP_M: mgn, R_SCHED: hdr,
            R_CAPEX: leaf, R_OPEN: leaf, R_DEPS: leaf, R_CLOSE: leaf}
    LABEL = {R_GRANT: "Grant financing", R_EBM_I: "EBITDA margin (incl. grant)",
             R_EBM_X: "EBITDA margin excl. grant", R_DA: "Depreciation & amortisation",
             R_DEP: "   Depreciation (PP&E)", R_EBIT: "EBIT", R_EBIT_M: "EBIT margin",
             R_FIN_I: "Finance income", R_FIN_C: "Finance costs",
             R_FIN_N: "Finance (costs), net", R_PBT: "Profit / (loss) before income tax",
             R_TAX: "Income tax (expense)", R_NP: "Net profit / (loss) for the period",
             R_NP_M: "Profit margin", R_SCHED: "CAPEX & DEPRECIATION (D&A build)",
             R_CAPEX: "  Capex – PP&E", R_OPEN: "  Opening PP&E (NBV)",
             R_DEPS: "  Depreciation (PP&E)", R_CLOSE: "  Closing PP&E (NBV)"}

    def fml(r, L, prev):
        if r == R_GRANT:  return f"=IF(AND({L}2>=' Inputs'!$G${I_GRANT},{L}2<=' Inputs'!$H${I_GRANT}),' Inputs'!$J${I_GRANT},0)"
        if r == R_EBM_I:  return f"=IF({L}4=0,0,({L}116+{L}{R_GRANT})/{L}4)"
        if r == R_EBM_X:  return f"=IF({L}4=0,0,{L}116/{L}4)"
        if r == R_DA:     return f"={L}{R_DEP}"
        if r == R_DEP:    return f"={L}{R_DEPS}"
        if r == R_EBIT:   return f"={L}116+{L}{R_GRANT}-{L}{R_DA}"
        if r == R_EBIT_M: return f"=IF({L}4=0,0,{L}{R_EBIT}/{L}4)"
        if r == R_FIN_I:  return "0"
        if r == R_FIN_C:  return f"=IF(AND({L}2>=' Inputs'!$G${I_FIN},{L}2<=' Inputs'!$H${I_FIN}),' Inputs'!$J${I_FIN},0)"
        if r == R_FIN_N:  return f"={L}{R_FIN_I}-{L}{R_FIN_C}"
        if r == R_PBT:    return f"={L}{R_EBIT}+{L}{R_FIN_N}"
        if r == R_TAX:    return f"=-MAX(0,{L}{R_PBT})*' Inputs'!$J${I_TAX}"
        if r == R_NP:     return f"={L}{R_PBT}+{L}{R_TAX}"
        if r == R_NP_M:   return f"=IF({L}4=0,0,{L}{R_NP}/{L}4)"
        if r == R_CAPEX:  return f"=IF(AND({L}2>=' Inputs'!$G${I_CAPEX},{L}2<=' Inputs'!$H${I_CAPEX}),' Inputs'!$J${I_CAPEX},0)"
        if r == R_OPEN:   return (f"=' Inputs'!$J${I_OPEN}" if L == get_column_letter(FIRST) else f"={prev}{R_CLOSE}")
        if r == R_DEPS:   return f"=MIN(({L}{R_OPEN}+{L}{R_CAPEX})/(' Inputs'!$J${I_LIFE}*12),{L}{R_OPEN}+{L}{R_CAPEX})"
        if r == R_CLOSE:  return f"={L}{R_OPEN}+{L}{R_CAPEX}-{L}{R_DEPS}"
        return None

    for r, style in ROLE.items():
        ws.cell(r, 1, LABEL[r])._style = copy(style[1])
        if r == R_SCHED:                                  # section band: style across, no data
            for c in range(2, LAST + 1):
                ws.cell(r, c)._style = copy(style[c])
            continue
        for c in range(FIRST, LAST + 1):
            L, prev = get_column_letter(c), get_column_letter(c - 1)
            ws.cell(r, c, fml(r, L, prev))._style = copy(style[c])
    print(f"  below-EBITDA P&L (rows {R_GRANT}-{R_NP_M}) + D&A capex schedule "
          f"(rows {R_SCHED}-{R_CLOSE}); inputs {I_CAPEX}-{I_TAX}")


# --- Phase 3: yearly P&L (IS_Y) -------------------------------------------------
PL_ROWS = list(range(4, 62)) + list(range(85, 134))   # income-statement rows (skip drivers)
# %-margin rows -> (numerator row, denominator row), recomputed yearly (not summed).
PCT_MAP = {56: (44, 4), 57: (45, 5), 58: (46, 9), 59: (47, 14), 60: (48, 15), 61: (52, 19),
           117: (116, 4), 122: (116, 4), 126: (125, 4), 133: (132, 4)}
FY_LABELS = ["FY1  Jul'26–Jun'27", "FY2  Jul'27–Jun'28", "FY3  Jul'28–Jun'29",
             "FY4  Jul'29–Jun'30", "FY5  Jul'30–Jun'31"]


def add_yearly_pl(wb) -> None:
    """Yearly P&L `IS_Y`: each flow line = SUM of its 12 monthly ProForma cells per
    fiscal year (Jul-Jun); margins recomputed from IS_Y's own lines (reference `isy`
    pattern). Mirrors the ProForma row numbers 1:1 so the recompute refs are trivial."""
    ws = wb["ProForma"]
    if "IS_Y" in wb.sheetnames:                       # idempotent
        del wb["IS_Y"]
    y = wb.create_sheet("IS_Y", index=wb.sheetnames.index("ProForma") + 1)
    y.sheet_view.showGridLines = False
    ycols = [get_column_letter(3 + k) for k in range(5)]    # C..G = FY1..FY5

    # title + year header (styled from ProForma analogues)
    y["A1"]._style = copy(ws.cell(1, 1)._style)
    y["A1"] = "Income Statement — Yearly (fiscal years Jul–Jun)"
    for k, yc in enumerate(ycols):
        c = y[f"{yc}2"]
        c._style = copy(ws.cell(2, 3)._style)
        c.value = FY_LABELS[k]
        c.number_format = "General"
    y.freeze_panes = "C3"
    y.column_dimensions["A"].width = ws.column_dimensions["A"].width or 34
    for yc in ycols:
        y.column_dimensions[yc].width = 15

    for r in PL_ROWS:
        a, cval = ws.cell(r, 1).value, _ft(ws.cell(r, 3))
        if a is None and cval is None:
            continue
        y.cell(r, 1, a)._style = copy(ws.cell(r, 1)._style)
        if cval is None:                              # section header (band, no data)
            for k, yc in enumerate(ycols):
                y[f"{yc}{r}"]._style = copy(ws.cell(r, 3 + k)._style)
            continue
        is_pct = "%" in ws.cell(r, 3).number_format
        for k, yc in enumerate(ycols):
            cell = y[f"{yc}{r}"]
            cell._style = copy(ws.cell(r, 3)._style)  # carry money/% format -> no naked rows
            if is_pct:
                if r == 121:
                    cell.value = f"=IF({yc}4=0,0,({yc}116+{yc}120)/{yc}4)"
                elif r in PCT_MAP:
                    num, den = PCT_MAP[r]
                    cell.value = f"=IF({yc}{den}=0,0,{yc}{num}/{yc}{den})"
                else:
                    raise ValueError(f"unmapped % row {r} ({a!r})")
            else:
                s = get_column_letter(FIRST + 12 * k)
                e = get_column_letter(FIRST + 12 * k + 11)
                cell.value = f"=SUM(ProForma!{s}{r}:{e}{r})"
    print(f"  yearly P&L IS_Y: {len([r for r in PL_ROWS if ws.cell(r,1).value or _ft(ws.cell(r,3))])} "
          f"rows over 5 fiscal years (C..G)")


def patch() -> None:
    if not os.path.exists(BAK):
        shutil.copy2(P, BAK)
        print(f"  wrote pristine backup -> {BAK}")
    else:
        print(f"  backup already exists ({BAK}) — left untouched")
    wb = openpyxl.load_workbook(P)
    add_salary_indexation(wb)
    add_below_ebitda_and_da(wb)
    add_yearly_pl(wb)
    wb.save(P)
    print(f"Saved {P}")


if __name__ == "__main__":
    patch()
