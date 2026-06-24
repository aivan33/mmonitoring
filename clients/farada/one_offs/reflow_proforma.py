"""Reflow the ProForma into the skill-outline order: VOLUMES & DRIVERS first, then Revenue → COGS
→ OPEX → below-EBITDA → working-capital/financing rolls. Same permutation+remap discipline as the
Inputs reflow, but the ProForma has INTERNAL relative refs (=C5+C9) and is referenced by the
statements (=ProForma!C44), so the remap is context-aware:

  • on ProForma: rewrite BARE internal row-refs (sheet-qualified refs ' Inputs'!/HR!/Revenue_Inputs!
    are left, including sheet-qualified ranges like Revenue_Inputs!$D$2:$D$31);
  • on IS/CF/BS(+_Y): rewrite ProForma!<cell> row-refs only.

Comprehensive gate: every relocated formula must reference the SAME logical rows after the move —
the multiset of referenced-row LABELS is preserved (catches any missed/mis-remapped ref). With no
recalc engine this label-isomorphism + the balance oracle are the value guarantee. Run after the
Inputs reflow (build_model_v7 calls it). Rewrites farada_model_v7.xlsx in place.
"""
from __future__ import annotations

import re
from copy import copy

from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.formula import ArrayFormula

SHEET = "ProForma"
STMTS = ("IS", "CF", "BS", "IS_Y", "CF_Y", "BS_Y")
NCOLS = 62  # A..BJ (months)

# DRIVERS reordered (D2): header · sensors · run-rate · CoS/sensor (spw,yield,chip,pkg,sensor-test,
# final-test,ASIC) · measurements · secondary (blended ASP/cost, capacity, util, clients, expansion)
DRIVERS = [63, 64, 65, 66, 67, 82, 83, 69, 70, 71, 72, 73, 68, 74, 75, 76, 77, 78, 79, 80, 81]
# block order: HEAD · DRIVERS · REVENUE · COGS · OPEX · BELOW · ROLLS (ranges as (s,e); lists kept verbatim)
BLOCKS = [(1, 2), DRIVERS, (4, 22), (24, 41), (85, 114), (120, 139), (143, 155)]

# a cell or range, optionally sheet-qualified (group 'q' present ⇒ external, leave it)
TOKEN = re.compile(r"(?P<q>'[^']*'!|[A-Za-z_][A-Za-z0-9_]*!)?"
                   r"(?P<ref>\$?[A-Z]{1,3}\$?\d+(?::\$?[A-Z]{1,3}\$?\d+)?)")
PFREF = re.compile(r"ProForma!(\$?[A-Z]{1,3}\$?)(\d+)(?::(\$?[A-Z]{1,3}\$?)(\d+))?")


def _ft(c):
    v = c.value
    return v.text if isinstance(v, ArrayFormula) else v


def _internal_rows(formula):
    """Rows referenced by BARE (ProForma-internal) refs in a formula."""
    out = set()
    if not isinstance(formula, str):
        return out
    for m in TOKEN.finditer(formula):
        if m.group("q") is None:
            out.update(int(d) for d in re.findall(r"\d+", m.group("ref")))
    return out


def _remap_internal(formula, o2n):
    if not isinstance(formula, str):
        return formula
    def repl(m):
        if m.group("q") is not None:
            return m.group(0)                       # external ref — leave
        ref = re.sub(r"\d+", lambda d: str(o2n.get(int(d.group()), int(d.group()))), m.group("ref"))
        return m.group(0)[: -len(m.group("ref"))] + ref
    return TOKEN.sub(repl, formula)


def _remap_pfrefs(formula, o2n):
    if not isinstance(formula, str):
        return formula
    def repl(m):
        s = f"ProForma!{m.group(1)}{o2n.get(int(m.group(2)), int(m.group(2)))}"
        if m.group(3):                              # range end (same sheet, bare col+row)
            s += f":{m.group(3)}{o2n.get(int(m.group(4)), int(m.group(4)))}"
        return s
    return PFREF.sub(repl, formula)


def reflow(wb):
    pf = wb[SHEET]
    snap = {r: [(pf.cell(r, c).value, pf.cell(r, c)._style) for c in range(1, NCOLS + 1)]
            for r in range(1, pf.max_row + 1)}
    label_old = {r: pf.cell(r, 1).value for r in snap}

    # plan: assign new rows block-by-block with a spacer before each (except HEAD)
    o2n, plan, nr = {}, [], 0
    for i, blk in enumerate(BLOCKS):
        if i:
            nr += 1; plan.append((nr, None))
        rows = range(blk[0], blk[1] + 1) if isinstance(blk, tuple) else blk
        for old in rows:
            nr += 1; o2n[old] = nr; plan.append((nr, old))
    last = nr

    # gate part 1: every internally-referenced ProForma row must be relocated
    refd = set()
    for r in snap:
        refd |= _internal_rows(_ft(pf.cell(r, 3)))
    for sn in STMTS:
        for row in wb[sn].iter_rows():
            for c in row:
                t = _ft(c)
                if isinstance(t, str):
                    refd |= {int(m.group(2)) for m in PFREF.finditer(t)}
    missing = sorted((refd & set(snap)) - set(o2n))
    assert not missing, f"referenced ProForma rows dropped by the reflow: {missing}"

    # reset the working region clean, then write the plan, remapping each moved formula's internals
    maxr, maxc = pf.max_row, max(NCOLS, pf.max_column)
    for r in range(1, maxr + 1):
        for c in range(1, maxc + 1):
            cell = pf.cell(r, c); cell.value = None; cell.style = "Normal"
    for nr, old in plan:
        if old is None:
            continue
        for c, (val, st) in enumerate(snap[old], start=1):
            cell = pf.cell(nr, c)
            cell.value = _remap_internal(val, o2n) if isinstance(val, str) and val.startswith("=") else val
            cell._style = st
    if maxr > last:
        pf.delete_rows(last + 1, maxr - last)

    # remap ProForma!<cell> refs in the statements
    for sn in STMTS:
        for row in wb[sn].iter_rows():
            for c in row:
                t = _ft(c)
                if isinstance(t, str) and "ProForma!" in t:
                    c.value = _remap_pfrefs(t, o2n)

    # gate part 2: each relocated formula references the same logical rows (by label multiset)
    def labelset(formula, rowlabel):
        return sorted(str(rowlabel.get(r)) for r in _internal_rows(formula))
    label_new = {r: pf.cell(r, 1).value for r in range(1, pf.max_row + 1)}
    bad = []
    for old, new in o2n.items():
        old_f, new_f = _ft_snap(snap, old), _ft(pf.cell(new, 3))
        if isinstance(old_f, str) and old_f.startswith("="):
            if labelset(old_f, label_old) != labelset(new_f, label_new):
                bad.append((old, new, label_old.get(old)))
    assert not bad, f"formula ref-labels changed at: {bad[:6]}"
    print(f"  PF gate ✓ {len(o2n)} rows relocated; ref-label multisets preserved on all formulas")
    return o2n, last


def _ft_snap(snap, row):
    v = snap[row][2][0]  # col C value
    return v.text if isinstance(v, ArrayFormula) else v


PF_SECTIONS = [
    ("WC DRIVERS & RATIOS", ["Receivable days (DSO)", "Payable days (DPO)",
                             "Current ratio", "Quick ratio", "Cash ratio"]),
    ("CASH FLOW", ["Operating activities", "Investing activities", "Financing activities"]),
    ("TAXATION", ["Tax expense (P&L)", "Tax payable (BS)"]),
    ("FUNDING", ["Equity round", "Debt draw", "Grants"]),
]


def add_proforma_sections(wb):
    """Complete the ProForma's skill-outline lower sections. The existing rolls ARE the balance
    sheet → relabel that header 'BALANCE SHEET (rolls)', then APPEND the remaining named sections
    (WC drivers & ratios · Cash Flow · Taxation · Funding) as blank-but-defined placeholders. Append
    only — no row-shift, no remap (nothing references rows past the rolls)."""
    pf = wb[SHEET]
    band_st = label_st = None
    for r in range(1, pf.max_row + 1):
        v = pf.cell(r, 1).value
        if isinstance(v, str) and "WORKING CAPITAL & FINANCING ROLLS" in v:
            band_st = pf.cell(r, 1)._style
            label_st = pf.cell(r + 1, 1)._style
            pf.cell(r, 1, "BALANCE SHEET (rolls)")._style = band_st
    nr = max(r for r in range(1, pf.max_row + 1) if pf.cell(r, 1).value is not None)
    for title, lines in PF_SECTIONS:
        nr += 2                                       # blank spacer + header
        pf.cell(nr, 1, title)._style = band_st
        for ln in lines:
            nr += 1
            pf.cell(nr, 1, "  " + ln)._style = label_st
    print("  PF sections: BALANCE SHEET + appended WC / Cash Flow / Taxation / Funding (blank-but-defined)")


def fix_run_rate(wb, FIRST=3, LAST=62):
    """D1 — replace the frozen `Total run-rate (sensors/yr) = SUM(C5:N7)` (identical in every column)
    with a real **LTM trailing-12-months** run-rate: for month m, Σ of the 3 sensor rows over the
    window [max(first, m-11) … m]. Drives the 6-point cost curve off realised scale (early months
    partial → lower volume → higher unit cost, which is correct). Label-based (survives the reorder)."""
    pf = wb[SHEET]
    L = {pf.cell(r, 1).value.strip(): r for r in range(1, pf.max_row + 1)
         if isinstance(pf.cell(r, 1).value, str) and pf.cell(r, 1).value.strip()}
    rr = L["Total run-rate (sensors/yr)"]
    s1, s3 = L["Sensors Line 1 (monthly)"], L["Sensors Line 3 (monthly)"]
    for c in range(FIRST, LAST + 1):
        a, x = get_column_letter(max(FIRST, c - 11)), get_column_letter(c)
        pf.cell(rr, c, f"=SUM({a}{s1}:{x}{s3})")
    print(f"  D1: run-rate → LTM trailing-12 over sensor rows {s1}-{s3}")


def style_subtotals(wb, LAST=62):
    """Bold the ProForma sum/subtotal lines (a row whose formula is purely +-joined internal cell
    refs or a SUM range) so totals stand out from their indented leaf children — readability (E)."""
    pf = wb[SHEET]
    sub = re.compile(r"^=(SUM\([A-Z]+\d+:[A-Z]+\d+\)|[A-Z]+\d+(\+[A-Z]+\d+)+)$")
    n = 0
    for r in range(1, pf.max_row + 1):
        f = _ft(pf.cell(r, 3))
        if isinstance(f, str) and sub.match(f.replace(" ", "")):
            for c in [1] + list(range(3, LAST + 1)):
                cell = pf.cell(r, c)
                fo = cell.font
                cell.font = Font(name=fo.name, size=fo.size, bold=True, color=fo.color,
                                 italic=fo.italic)
            n += 1
    print(f"  style: bolded {n} ProForma subtotal lines")


def _line3_bundle_inputs(inp):
    """Resolve the Line-3 per-bundle Input rows by label (robust to the reflow): returns dicts of
    sensors / included / list-price / discount rows + the avg-meas row, all for S/M/L."""
    def bundles(prefix):
        hdr = next(r for r in range(1, inp.max_row + 1) if isinstance(inp.cell(r, 3).value, str)
                   and inp.cell(r, 3).value.strip().startswith(prefix))
        out, r = [], hdr + 1
        while len(out) < 3 and r <= inp.max_row:
            if isinstance(inp.cell(r, 10).value, str) and "OFFSET" in inp.cell(r, 10).value:
                out.append(r)
            r += 1
        return out
    avg = next(r for r in range(1, inp.max_row + 1) if isinstance(inp.cell(r, 3).value, str)
               and inp.cell(r, 3).value.strip().startswith("Avg measurements"))
    return dict(sens=bundles("Line 3 — sensors per bundle"), incl=bundles("Line 3 — included measurements"),
                price=bundles("Line 3 — overage price"), disc=bundles("Line 3 — plan tier discount"), avg=avg)


def add_installed_base(wb, FIRST=3, LAST=62):
    """V2 — explicit per-bundle Installed base (cumulative sensors) driver rows: IB_b(c) = prior +
    phased new bundles(c) × sensors/bundle_b. This is the single accumulator the subscription / overage
    / measurements read as `IB × rate` — removing the per-line re-accumulation that bred the stale-ref
    and double-count bugs. Inserts 3 rows just before the measurements block + remaps. Value-neutral
    (new driver rows, not yet referenced). Runs after the ProForma reflow, before the SaaS rebuild."""
    pf, inp = wb[SHEET], wb[" Inputs"]
    R = next(r for r in range(1, pf.max_row + 1)
             if isinstance(pf.cell(r, 1).value, str) and "Measurements Line 3" in pf.cell(r, 1).value)
    I = _line3_bundle_inputs(inp)
    lbl_st = pf.cell(R, 1)._style
    val_st = {c: pf.cell(R, c)._style for c in range(FIRST, LAST + 1)}
    pf.insert_rows(R, 3)                                       # 3 IB rows BEFORE the measurements total
    o2n = {r: (r if r < R else r + 3) for r in range(1, pf.max_row + 4)}
    for row in pf.iter_rows():
        for cell in row:
            t = _ft(cell)
            if isinstance(t, str) and t.startswith("="):
                cell.value = _remap_internal(t, o2n)
    for sn in STMTS:
        for row in wb[sn].iter_rows():
            for cell in row:
                t = _ft(cell)
                if isinstance(t, str) and "ProForma!" in t:
                    cell.value = _remap_pfrefs(t, o2n)
    names = ["  Installed base — Bundle S", "  Installed base — Bundle M", "  Installed base — Bundle L"]
    for i, (br, sr, nm) in enumerate(zip([12, 13, 14], I["sens"], names)):
        ib = R + i
        pf.cell(ib, 1, nm)._style = lbl_st
        for c in range(FIRST, LAST + 1):
            x, p = get_column_letter(c), get_column_letter(c - 1)
            add = f"{_phase(c, br, FIRST)}*' Inputs'!$J${sr}"
            pf.cell(ib, c, f"={add}" if c == FIRST else f"={p}{ib}+{add}")._style = val_st[c]
    print(f"  V2: installed-base rows {R}-{R + 2} (cumulative sensors per bundle)")


def add_measurement_children(wb, FIRST=3, LAST=62):
    """V4 — measurements off the installed base. Included = Σ IB_b × included_b /12; Overage (gross) =
    Σ IB_b × MAX(0, avg − included_b)/12 — both LEVELS (IB is already cumulative → NO re-accumulation).
    The displayed Overage = the gross initially; add_overage_delay ramp-delays it. Total = Included +
    Overage (clean =C+C). Fixes the old stale-$J$71 bug (Included == Overage, ~2× cloud COGS). Input
    rows resolved by label. Inserts 3 rows (Included, Overage, Overage-gross helper)."""
    pf, inp = wb[SHEET], wb[" Inputs"]
    R = next(r for r in range(1, pf.max_row + 1)
             if isinstance(pf.cell(r, 1).value, str) and "Measurements Line 3" in pf.cell(r, 1).value)
    I = _line3_bundle_inputs(inp)
    IB = [next(r for r in range(1, pf.max_row + 1) if isinstance(pf.cell(r, 1).value, str)
               and pf.cell(r, 1).value.strip() == f"Installed base — Bundle {b}") for b in ("S", "M", "L")]
    lbl_st = pf.cell(R, 1)._style
    val_st = {c: pf.cell(R, c)._style for c in range(FIRST, LAST + 1)}
    pf.insert_rows(R + 1, 3)
    o2n = {r: (r if r <= R else r + 3) for r in range(1, pf.max_row + 4)}
    for row in pf.iter_rows():
        for cell in row:
            t = _ft(cell)
            if isinstance(t, str) and t.startswith("="):
                cell.value = _remap_internal(t, o2n)
    for sn in STMTS:
        for row in wb[sn].iter_rows():
            for cell in row:
                t = _ft(cell)
                if isinstance(t, str) and "ProForma!" in t:
                    cell.value = _remap_pfrefs(t, o2n)
    INC, OVR, GRS, avg = R + 1, R + 2, R + 3, I["avg"]
    pf.cell(INC, 1, "    Included (subscription)")._style = lbl_st
    pf.cell(OVR, 1, "    Overage (beyond subscription)")._style = lbl_st
    pf.cell(GRS, 1, "    Overage — gross (pre-ramp, calc)")._style = lbl_st
    for c in range(FIRST, LAST + 1):
        x = get_column_letter(c)
        inc = "+".join(f"{x}{ib}*' Inputs'!$J${ir}/12" for ib, ir in zip(IB, I["incl"]))
        grs = "+".join(f"{x}{ib}*MAX(0,' Inputs'!$J${avg}-' Inputs'!$J${ir})/12" for ib, ir in zip(IB, I["incl"]))
        pf.cell(INC, c, f"={inc}")._style = val_st[c]
        pf.cell(GRS, c, f"={grs}")._style = val_st[c]
        pf.cell(OVR, c, f"={x}{GRS}")._style = val_st[c]            # undelayed; add_overage_delay ramp-delays
        pf.cell(R, c, f"={x}{INC}+{x}{OVR}")._style = val_st[c]     # total = Included + Overage (clean)
    print(f"  V4: measurements off IB — Included {INC}, Overage {OVR} (gross {GRS}); total = clean sum")


def wire_yield_inputs(wb, FIRST=3, LAST=62):
    """V1 — point the ProForma yield row at the staged Yield Inputs (was a hardcoded IF curve with
    literal 4000000/0.95/… breakpoints) and the sensors-per-wafer row at its Input (was literal 4000).
    Cascades high→low run-rate threshold off ' Inputs'!$F/$J rungs, exactly like the cost-of-sales
    curves. Chip = wafer ÷ spw ÷ yield unchanged. Post-reflow, by label."""
    pf, inp = wb[SHEET], wb[" Inputs"]
    Lp = {pf.cell(r, 1).value.strip(): r for r in range(1, pf.max_row + 1)
          if isinstance(pf.cell(r, 1).value, str) and pf.cell(r, 1).value.strip()}
    def irow(prefix):
        return next(r for r in range(1, inp.max_row + 1) if isinstance(inp.cell(r, 3).value, str)
                    and inp.cell(r, 3).value.strip().startswith(prefix))
    spw_in = irow("Sensors per wafer")
    rungs = [irow(t) for t in ("Yield @ 1 /yr", "Yield @ 10,000", "Yield @ 100,000",
                               "Yield @ 1,000,000", "Yield @ 4,000,000")]
    rr, spw_row, yld_row = (Lp["Total run-rate (sensors/yr)"], Lp["Sensors per wafer"],
                            Lp["Yield (staged by run-rate)"])
    for c in range(FIRST, LAST + 1):
        x = get_column_letter(c)
        pf.cell(spw_row, c, f"=' Inputs'!$J${spw_in}")
        casc = f"' Inputs'!$J${rungs[0]}"                       # base rung (@1) = fallback
        for rg in rungs[1:]:
            casc = f"IF({x}{rr}>=' Inputs'!$F${rg},' Inputs'!$J${rg},{casc})"
        pf.cell(yld_row, c, f"={casc}")
    print("  V1: ProForma yield cascades off staged Inputs; spw → Input (no literals)")


def _phase(c, brow, FIRST=3):
    """Quarter-bookings phasing for ProForma column c and a Revenue_Inputs bundle row — the same
    INT/MOD spread the hardware/overage lines use (col C → quarter B, threshold 1; +1 per month)."""
    q = get_column_letter(2 + (c - FIRST) // 3)          # Revenue_Inputs quarter column (B = 2026 Q3)
    k = (c - FIRST) % 3 + 1                               # month-in-quarter threshold
    return f"(INT(Revenue_Inputs!{q}${brow}/3)+IF(MOD(Revenue_Inputs!{q}${brow},3)>={k},1,0))"


def add_subscription_lines(wb, FIRST=3, LAST=62):
    """D5b — add the Subscription (recurring) revenue block + a billings memo to the SaaS #3 group.
    Subscription accrues on the cumulative installed base at the discounted plan rate
    (included × list × (1−tier_discount)/12, per bundle) — like overage, but on the included quota.
    Billings is the per-period ANNUAL amount collected upfront (drives the deferred-revenue roll, D5e).
    SaaS #3 = Hardware + Subscription + Overage; the billings memo is excluded. Inserts 5 rows + remaps.
    Bundle inputs: sensors J53/54/55 · included J58/59/60 · list(=overage) J63/64/65 · discount J68/69/70."""
    pf = wb[SHEET]
    saas = next(r for r in range(1, pf.max_row + 1) if isinstance(pf.cell(r, 1).value, str)
                and "Hardware-enabled SaaS" in pf.cell(r, 1).value)
    overage = next(r for r in range(1, pf.max_row + 1) if isinstance(pf.cell(r, 1).value, str)
                   and "SaaS (overage" in pf.cell(r, 1).value)
    hw = saas + 1                                         # hardware subtotal (directly under SaaS #3)
    sub_st = pf.cell(overage, 1)._style                  # overage subtotal label style
    kid_st = pf.cell(overage + 1, 1)._style              # overage child label style
    sv = {c: pf.cell(overage, c)._style for c in range(FIRST, LAST + 1)}
    kv = {c: pf.cell(overage + 1, c)._style for c in range(FIRST, LAST + 1)}

    N = 5
    pf.insert_rows(overage, N)                           # insert the subscription block BEFORE overage
    o2n = {r: (r if r < overage else r + N) for r in range(1, pf.max_row + N + 1)}
    for row in pf.iter_rows():
        for cell in row:
            t = _ft(cell)
            if isinstance(t, str) and t.startswith("="):
                cell.value = _remap_internal(t, o2n)
    for sn in STMTS:
        for row in wb[sn].iter_rows():
            for cell in row:
                t = _ft(cell)
                if isinstance(t, str) and "ProForma!" in t:
                    cell.value = _remap_pfrefs(t, o2n)

    SUB, S_S, S_M, S_L, BILL = overage, overage + 1, overage + 2, overage + 3, overage + 4
    OVER = overage + N                                   # overage subtotal, shifted down
    pf.cell(SUB, 1, "  Subscription (recurring)")._style = sub_st
    pf.cell(S_S, 1, "    Bundle S")._style = kid_st
    pf.cell(S_M, 1, "    Bundle M")._style = kid_st
    pf.cell(S_L, 1, "    Bundle L")._style = kid_st
    pf.cell(BILL, 1, "  Subscription billings (annual, upfront — memo)")._style = sub_st
    BUNDLES = [(S_S, 12, 53, 58, 63, 68), (S_M, 13, 54, 59, 64, 69), (S_L, 14, 55, 60, 65, 70)]
    rate = lambda ir, pr, dr: f"' Inputs'!$J${ir}*' Inputs'!$J${pr}*(1-' Inputs'!$J${dr})"
    for c in range(FIRST, LAST + 1):
        x, p = get_column_letter(c), get_column_letter(c - 1)
        for (row, brow, sr, ir, pr, dr) in BUNDLES:      # subscription children — cumulative installed base
            add = f"{_phase(c, brow, FIRST)}*' Inputs'!$J${sr}*{rate(ir, pr, dr)}/12"
            pf.cell(row, c, f"={add}" if c == FIRST else f"={p}{row}+{add}")._style = kv[c]
        pf.cell(SUB, c, f"={x}{S_S}+{x}{S_M}+{x}{S_L}")._style = sv[c]
        bill = "+".join(f"{_phase(c, brow, FIRST)}*' Inputs'!$J${sr}*{rate(ir, pr, dr)}"
                        for (_, brow, sr, ir, pr, dr) in BUNDLES)        # per-period annual (upfront)
        pf.cell(BILL, c, f"={bill}")._style = sv[c]
        pf.cell(saas, c, f"={x}{hw}+{x}{SUB}+{x}{OVER}")._style = pf.cell(saas, c)._style
    print(f"  D5b: subscription block (rev {SUB}, children {S_S}-{S_L}, billings {BILL}); SaaS#3 = HW+Sub+Overage")


def rewire_saas_off_ib(wb, FIRST=3, LAST=62):
    """V3 — point the SaaS revenue lines at the installed base (levels), removing the per-line cohort
    accumulation: Subscription_b = IB_b × included_b × list_b × (1−disc_b)/12; Overage_b (gross) = IB_b
    × MAX(0, avg − included_b) × list_b /12; Subscription billings_b = ΔIB_b × included_b × list_b ×
    (1−disc_b) (the new annual plans sold). Value-equivalent to the old cohort sums, but clean (no
    double-accumulate). Subtotals already sum the children. Runs after add_subscription_lines."""
    pf, inp = wb[SHEET], wb[" Inputs"]
    L = {pf.cell(r, 1).value.strip(): r for r in range(1, pf.max_row + 1)
         if isinstance(pf.cell(r, 1).value, str) and pf.cell(r, 1).value.strip()}
    I = _line3_bundle_inputs(inp)
    IB = [L[f"Installed base — Bundle {b}"] for b in ("S", "M", "L")]
    sub, ov = L["Subscription (recurring)"], L["SaaS (overage, recurring)"]
    bill, avg = L["Subscription billings (annual, upfront — memo)"], I["avg"]
    rate = lambda ir, pr, dr: f"' Inputs'!$J${ir}*' Inputs'!$J${pr}*(1-' Inputs'!$J${dr})"
    for c in range(FIRST, LAST + 1):
        x, p = get_column_letter(c), get_column_letter(c - 1)
        billterms = []
        for i, (ib, ir, pr, dr) in enumerate(zip(IB, I["incl"], I["price"], I["disc"])):
            pf.cell(sub + 1 + i, c, f"={x}{ib}*{rate(ir, pr, dr)}/12")                      # subscription
            pf.cell(ov + 1 + i, c, f"={x}{ib}*MAX(0,' Inputs'!$J${avg}-' Inputs'!$J${ir})*' Inputs'!$J${pr}/12")  # overage gross
            dib = f"{x}{ib}" if c == FIRST else f"({x}{ib}-{p}{ib})"                          # ΔIB = new sensors
            billterms.append(f"{dib}*{rate(ir, pr, dr)}")
        pf.cell(bill, c, "=" + "+".join(billterms))
    print("  V3: subscription / overage / billings rewired off the installed base (IB × rate)")


def cloud_cogs_measurement_driven(wb, FIRST=3, LAST=62):
    """D5d — replace the SaaS-COGS gross-margin plug (overage × (1−GM-target)) with a real
    measurement-driven cost: cloud COGS = total Line-3 measurements × cloud_cost/measurement. The
    measurements total already equals installed_base × avg/12 (D4), so this = installed × avg × cloud/12.
    GM becomes an OUTPUT (subscription + overage − cloud), not a plug. Supersedes the placeholder."""
    pf, inp = wb[SHEET], wb[" Inputs"]
    L = {pf.cell(r, 1).value.strip(): r for r in range(1, pf.max_row + 1)
         if isinstance(pf.cell(r, 1).value, str) and pf.cell(r, 1).value.strip()}
    usage, meas = L["Usage (cloud / compute)"], L["Measurements Line 3 (monthly)"]
    inrow = lambda pfx: next(r for r in range(1, inp.max_row + 1) if isinstance(inp.cell(r, 3).value, str)
                             and inp.cell(r, 3).value.strip().startswith(pfx))
    cloud, gm = inrow("Cloud / compute per measurement"), inrow("SaaS gross margin target")
    for c in range(FIRST, LAST + 1):
        pf.cell(usage, c, f"={get_column_letter(c)}{meas}*' Inputs'!$J${cloud}")
    inp.cell(cloud, 15, "← drives SaaS/cloud COGS = total measurements × this (SaaS GM is an output)")
    inp.cell(gm, 15, "← SUPERSEDED by measurement-driven cloud COGS (D5d); kept for reference")
    print(f"  D5d: cloud COGS = measurements(row {meas}) × cloud_cost(J{cloud}); 80% GM plug retired")


def rework_wc_rolls(wb, FIRST=3, LAST=62):
    """D5e — re-wire the working-capital rolls to the split SaaS streams (the 3-statement seam):
      • Trade receivables = (components + hardware) × (1−prepay) + overage, spread over DSO.
        Subscription is EXCLUDED (billed annually upfront → deferred, no receivable).
      • Deferred revenue = running balance: prior + subscription billings − subscription revenue
        (replaces the =overage×SAAS_ANN×6 proxy). The CF Δ-deferred then resolves subscription cash to
        the upfront billings. AP is unchanged (COGS TOTAL already includes the cloud COGS)."""
    pf, inp = wb[SHEET], wb[" Inputs"]
    L = {pf.cell(r, 1).value.strip(): r for r in range(1, pf.max_row + 1)
         if isinstance(pf.cell(r, 1).value, str) and pf.cell(r, 1).value.strip()}
    ar, dfr = L["Trade receivables (AR)"], L["Deferred revenue (SaaS annual)"]
    comp1, comp2, hwdev = (L["Components #1 - Low Volume"], L["Components #2 - High Volume"],
                           L["Hardware (device, cost + markup)"])
    sub, bill, ov = (L["Subscription (recurring)"],
                     L["Subscription billings (annual, upfront — memo)"], L["SaaS (overage, recurring)"])
    jrow = lambda pfx: next(r for r in range(1, inp.max_row + 1) if isinstance(inp.cell(r, 3).value, str)
                            and inp.cell(r, 3).value.strip().startswith(pfx))
    PREPAY, DSO, OBDEF = jrow("Hardware prepayment"), jrow("Receivable days"), jrow("Opening deferred revenue")
    for c in range(FIRST, LAST + 1):
        x, p = get_column_letter(c), get_column_letter(c - 1)
        pf.cell(ar, c, f"=(({x}{comp1}+{x}{comp2}+{x}{hwdev})*(1-' Inputs'!$J${PREPAY})"
                       f"+{x}{ov})/30*' Inputs'!$J${DSO}")
        pf.cell(dfr, c, (f"=' Inputs'!$J${OBDEF}+{x}{bill}-{x}{sub}" if c == FIRST
                         else f"={p}{dfr}+{x}{bill}-{x}{sub}"))
    print(f"  D5e: AR=(comp+hw)(1−prepay)+overage; deferred=running(billings−subscription)")


def add_overage_delay(wb, FIRST=3, LAST=62):
    """OD — a client doesn't overuse credits from month 1; overage ramps in `delay` months after a
    cohort starts. A UNIFORM per-cohort delay = an exact right-shift of the (cumulative) overage
    aggregate by `delay` months: overage_delayed(c) = overage_undelayed(c−delay). Implemented with
    OFFSET (as the scenario selector does) on the undelayed children, guarded so the first `delay`
    months are 0 (no #REF!). Applies to BOTH overage revenue (→ SaaS#3/AR/IS) and the overage
    MEASUREMENT child (→ cloud COGS). Hardware (one-time) and Subscription/Included (month 1) untouched.
    No row inserts — the undelayed children/child stay as the engine; only the consumed rows shift."""
    pf = wb[SHEET]
    L = {pf.cell(r, 1).value.strip(): r for r in range(1, pf.max_row + 1)
         if isinstance(pf.cell(r, 1).value, str) and pf.cell(r, 1).value.strip()}
    ov = L["SaaS (overage, recurring)"]
    ovm, grs = L["Overage (beyond subscription)"], L["Overage — gross (pre-ramp, calc)"]
    inp = wb[" Inputs"]
    drow = next(r for r in range(1, inp.max_row + 1) if isinstance(inp.cell(r, 3).value, str)
                and inp.cell(r, 3).value.strip().startswith("Overage ramp delay"))
    DLY = f"' Inputs'!$J${drow}"
    for c in range(FIRST, LAST + 1):
        x, m = get_column_letter(c), c - FIRST                    # m = 0-based month index
        shift = lambda row: f"OFFSET({x}{row},0,-{DLY})"          # value `delay` columns to the left
        # overage revenue subtotal → delayed sum of the (undelayed) children, 0 during the ramp
        kids = "+".join(shift(ov + i) for i in (1, 2, 3))
        pf.cell(ov, c, f"=IF({m}<{DLY},0,{kids})")
        # measurement Overage (displayed) = ramp-delayed gross; the TOTAL stays a clean Included+Overage
        pf.cell(ovm, c, f"=IF({m}<{DLY},0,{shift(grs)})")
    print(f"  OD: overage revenue (row {ov}) + measurement overage (row {ovm}) ramp-delayed by {DLY}")


def build_cupffee_cf(wb, FIRST=3, LAST=62):
    """RB1 — rebuild the ProForma CASH FLOW as a Cupffee-style BY-COMPONENT direct method (replaces the
    earlier clone). Each line = accrual ± Δ(working-capital balance), itemised by category using
    Farada's per-category AP buckets (COGS/S&M/G&A/R&D) + payroll payable. Value-NEUTRAL vs the prior
    lumped CF (the category lines sum to the same operating total). The CF statement then PULLS each
    line (=ProForma!) and SUMS the subtotals in-statement (Cupffee-style, not 100% bare refs); BS cash =
    CF ending. Re-appends blank TAXATION/FUNDING (CB4/RB2 fill). Runs after restructure_cf."""
    pf, cf, iss, inp = wb[SHEET], wb["CF"], wb["IS"], wb[" Inputs"]
    Lp = {pf.cell(r, 1).value.strip(): r for r in range(1, pf.max_row + 1)
          if isinstance(pf.cell(r, 1).value, str) and pf.cell(r, 1).value.strip()}
    istax = next(r for r in range(1, iss.max_row + 1)
                 if isinstance(iss.cell(r, 1).value, str) and "Income tax (expense)" in iss.cell(r, 1).value)
    def jr(prefix):
        r = next(rr for rr in range(1, inp.max_row + 1) if isinstance(inp.cell(rr, 3).value, str)
                 and inp.cell(rr, 3).value.strip().startswith(prefix))
        return f"' Inputs'!$J${r}"
    REV, AR, DEFr = Lp["Revenue"], Lp["Trade receivables (AR)"], Lp["Deferred revenue (SaaS annual)"]
    COGS, SM, GA, RD = Lp["COGS TOTAL"], Lp["S&M"], Lp["G&A"], Lp["R&D"]
    SMp, GAp, RDp = SM + 1, GA + 1, RD + 1
    APc, APs, APg, APr = (Lp["Trade payables — COGS"], Lp["Trade payables — S&M"],
                          Lp["Trade payables — G&A"], Lp["Trade payables — R&D"])
    PAY, TAXP, CAPEX = Lp["Payroll payable"], Lp["Tax payable"], Lp["Capex – PP&E"]
    SC, DEBT, GRANT, FINc = Lp["Share capital"], Lp["Debt"], Lp["Grant financing"], Lp["Finance costs"]
    OBcash, OBar, OBdef, OBsc, OBdebt = (jr("Opening cash"), jr("Opening AR"), jr("Opening deferred"),
                                         jr("Opening share capital"), jr("Opening debt"))
    band_st = pf.cell(Lp["CASH FLOW"], 1)._style
    lbl_st = pf.cell(AR, 1)._style
    valc = {c: pf.cell(AR, c)._style for c in range(1, LAST + 1)}
    st_lbl, st_val = cf.cell(5, 1)._style, {c: cf.cell(5, c)._style for c in range(1, LAST + 1)}
    st_tot, st_totv = cf.cell(13, 1)._style, {c: cf.cell(13, c)._style for c in range(1, LAST + 1)}
    st_sub = cf.cell(4, 1)._style

    def Dr(roll, base, x, p, first):                          # Δ(balance): close − (opening | prior col)
        return f"({x}{roll}-{base if first else f'{p}{roll}'})"
    O = "0"
    spec = [  # (key, label, kind, payload)  kind: band/sub/blank/line/subtotal/roll
        ("_b", "CASH FLOW", "band", None), ("_x0", "", "blank", None),
        ("_oh", "Operating activities", "sub", None),
        ("inflow", "  Cash inflow / clients", "line", lambda x, p, f: f"={x}{REV}-{Dr(AR, OBar, x, p, f)}"),
        ("deferred", "  Movement in deferred revenue", "line", lambda x, p, f: f"={Dr(DEFr, OBdef, x, p, f)}"),
        ("sup_cogs", "  Suppliers — Direct (COGS)", "line", lambda x, p, f: f"=-({x}{COGS}-{Dr(APc, O, x, p, f)})"),
        ("sup_sm", "  Suppliers — S&M", "line", lambda x, p, f: f"=-(({x}{SM}-{x}{SMp})-{Dr(APs, O, x, p, f)})"),
        ("sup_ga", "  Suppliers — G&A", "line", lambda x, p, f: f"=-(({x}{GA}-{x}{GAp})-{Dr(APg, O, x, p, f)})"),
        ("sup_rd", "  Suppliers — R&D", "line", lambda x, p, f: f"=-(({x}{RD}-{x}{RDp})-{Dr(APr, O, x, p, f)})"),
        ("suppliers", "  Cash outflow — Suppliers", "subtotal", ["sup_cogs", "sup_sm", "sup_ga", "sup_rd"]),
        ("per_sm", "  Personnel — S&M", "line", lambda x, p, f: f"=-{x}{SMp}"),
        ("per_ga", "  Personnel — G&A", "line", lambda x, p, f: f"=-{x}{GAp}"),
        ("per_rd", "  Personnel — R&D", "line", lambda x, p, f: f"=-{x}{RDp}"),
        ("mv_pay", "  Movement in payroll payable", "line", lambda x, p, f: f"={Dr(PAY, O, x, p, f)}"),
        ("personnel", "  Payments to personnel", "subtotal", ["per_sm", "per_ga", "per_rd", "mv_pay"]),
        ("taxes", "  Corporate & other taxes, net", "line", lambda x, p, f: f"=IS!{x}{istax}+{Dr(TAXP, O, x, p, f)}"),
        ("bank", "  Bank charges paid", "line", lambda x, p, f: f"=-{x}{FINc}"),
        ("op", "Cash Flow from Operating Activities", "subtotal",
         ["inflow", "deferred", "suppliers", "personnel", "taxes", "bank"]),
        ("_x1", "", "blank", None), ("_ih", "Investing activities", "sub", None),
        ("capex", "  CAPEX", "line", lambda x, p, f: f"=-{x}{CAPEX}"),
        ("rdcap", "  R&D (capitalised)", "line", lambda x, p, f: "0"),
        ("inv", "Cash Flow from Investing Activities", "subtotal", ["capex", "rdcap"]),
        ("_x2", "", "blank", None), ("_fh", "Financing activities", "sub", None),
        ("equity", "  Capital Increase", "line", lambda x, p, f: f"={Dr(SC, OBsc, x, p, f)}"),
        ("debt", "  Loan facility financing", "line", lambda x, p, f: f"={Dr(DEBT, OBdebt, x, p, f)}"),
        ("grants", "  Grants", "line", lambda x, p, f: f"={x}{GRANT}"),
        ("fin", "Cash Flow from Financing Activities", "subtotal", ["equity", "debt", "grants"]),
        ("_x3", "", "blank", None),
        ("excess", "Excess Cash for the Period", "subtotal", ["op", "inv", "fin"]),
        ("begin", "Beginning Cash Balance", "roll", "begin"),
        ("ending", "Ending Cash Balance", "roll", "ending"),
    ]
    cfh = Lp["CASH FLOW"]
    for r in range(cfh, pf.max_row + 1):                       # clear ProForma CASH FLOW+TAX+FUNDING
        for c in range(1, LAST + 1):
            pf.cell(r, c).value = None; pf.cell(r, c).style = "Normal"
    cf_max = max(r for r in range(1, cf.max_row + 1) if cf.cell(r, 1).value not in (None, ""))
    for r in range(3, cf_max + 1):                            # clear CF statement (keep title rows 1-2)
        for c in range(1, LAST + 1):
            cf.cell(r, c).value = None; cf.cell(r, c).style = "Normal"

    def lay(sheet, start, label_styles):                      # write labels, return key→row
        rowmap, r = {}, start
        for key, label, kind, _ in spec:
            if kind == "blank":
                r += 1; continue
            if sheet is cf and kind == "band":
                continue                                       # statement keeps its own title; no band row
            rowmap[key] = r
            sheet.cell(r, 1, label)._style = label_styles[kind]
            r += 1
        return rowmap
    pfrow = lay(pf, cfh, {"band": band_st, "sub": lbl_st, "line": lbl_st, "subtotal": lbl_st, "roll": lbl_st})
    strow = lay(cf, 4, {"sub": st_sub, "line": st_lbl, "subtotal": st_tot, "roll": st_tot})

    for c in range(FIRST, LAST + 1):
        x, p, first = get_column_letter(c), get_column_letter(c - 1), c == FIRST
        for key, label, kind, payload in spec:
            if kind in ("blank", "band", "sub"):
                continue
            pr, sr = pfrow[key], strow[key]
            if kind == "line":
                pf.cell(pr, c, payload(x, p, first))._style = valc[c]
                cf.cell(sr, c, f"=ProForma!{x}{pr}")._style = st_val[c]
            elif kind == "subtotal":
                pf.cell(pr, c, "=" + "+".join(f"{x}{pfrow[k]}" for k in payload))._style = valc[c]
                cf.cell(sr, c, "=" + "+".join(f"{x}{strow[k]}" for k in payload))._style = st_totv[c]
            elif kind == "roll":
                for sh, rm, vs in ((pf, pfrow, valc), (cf, strow, st_totv)):
                    R = rm[key]
                    if payload == "begin":
                        sh.cell(R, c, f"={OBcash}" if first else f"={p}{rm['ending']}")._style = vs[c]
                    else:
                        sh.cell(R, c, f"={x}{rm['begin']}+{x}{rm['excess']}")._style = vs[c]
    bs = wb["BS"]
    for rr in range(1, bs.max_row + 1):                        # BS cash → CF ending (statement-driven)
        if isinstance(bs.cell(rr, 1).value, str) and "Cash & cash equivalents" in bs.cell(rr, 1).value:
            for c in range(FIRST, LAST + 1):
                bs.cell(rr, c).value = f"=CF!{get_column_letter(c)}{strow['ending']}"

    nr = max(pfrow.values())                                  # re-append blank TAX + FUNDING (RB2/CB4 fill)
    for title, lines in [("TAXATION", ["Taxable profit before utilisation", "Utilisation of tax loss",
                                       "Taxable profit after utilisation", "Total taxation",
                                       "Tax-loss control — opening", "Additions", "Utilisation",
                                       "Tax-loss control — closing", "Corporate tax"]),
                         ("FUNDING", ["Equity round", "Debt draw", "Grants"])]:
        nr += 2
        pf.cell(nr, 1, title)._style = band_st
        for ln in lines:
            nr += 1
            pf.cell(nr, 1, "  " + ln)._style = lbl_st
    print(f"  RB1: CASH FLOW rebuilt Cupffee by-component (ProForma {cfh}-{max(pfrow.values())}); CF stmt pull+sum; BS cash→CF")


def populate_wc_ratios(wb, FIRST=3, LAST=62):
    """CB3 — fill the WC DRIVERS & RATIOS section in the ProForma (was blank). Realised DSO/DPO and
    the liquidity ratios, computed off the BS rolls + the cloned cash. Current assets = cash + AR
    (Farada has no inventory/prepaid); current liabilities = AP + payroll + deferred + tax payable.
    The BS statement's ratio rows then pull from here. Runs after relocate_cf_to_proforma."""
    pf = wb[SHEET]
    L = {pf.cell(r, 1).value.strip(): r for r in range(1, pf.max_row + 1)
         if isinstance(pf.cell(r, 1).value, str) and pf.cell(r, 1).value.strip()}
    dso, dpo, cur, qk, csh = (L["Receivable days (DSO)"], L["Payable days (DPO)"],
                              L["Current ratio"], L["Quick ratio"], L["Cash ratio"])
    ar, ap, pay, dfr, tax = (L["Trade receivables (AR)"], L["Trade payables (total)"], L["Payroll payable"],
                             L["Deferred revenue (SaaS annual)"], L["Tax payable"])
    rev, cogs, cash = L["Revenue"], L["COGS TOTAL"], L["Ending Cash Balance"]
    base_st = pf.cell(ar, FIRST)._style                       # a roll value-cell style to harvest
    for c in range(FIRST, LAST + 1):
        x = get_column_letter(c)
        CA, CL = f"({x}{cash}+{x}{ar})", f"({x}{ap}+{x}{pay}+{x}{dfr}+{x}{tax})"
        rows = {dso: (f"=IF({x}{rev}=0,0,{x}{ar}/{x}{rev}*30)", "#,##0"),
                dpo: (f"=IF({x}{cogs}=0,0,{x}{ap}/{x}{cogs}*30)", "#,##0"),
                cur: (f"=IF({CL}=0,0,{CA}/{CL})", "0.0"),
                qk:  (f"=IF({CL}=0,0,{CA}/{CL})", "0.0"),     # no inventory → quick = current
                csh: (f"=IF({CL}=0,0,{x}{cash}/{CL})", "0.0")}
        for r, (f, fmt) in rows.items():
            cell = pf.cell(r, c, f)
            cell._style = base_st
            cell.number_format = fmt
    # BS statement ratio rows → pull from the ProForma WC ratios
    bs = wb["BS"]
    for r in range(1, bs.max_row + 1):
        lbl = bs.cell(r, 1).value
        if isinstance(lbl, str):
            tgt = {"Current ratio": cur, "Quick ratio": qk, "Cash ratio": csh}.get(lbl.strip())
            if tgt:
                for c in range(FIRST, LAST + 1):
                    bs.cell(r, c, f"=ProForma!{get_column_letter(c)}{tgt}").number_format = "0.0"
    print(f"  CB3: WC ratios (DSO/DPO/current/quick/cash) computed in ProForma; BS pulls them")


def populate_tax_funding(wb, FIRST=3, LAST=62):
    """RB2/CB4 — TAXATION as a Cupffee tax-loss carryforward + FUNDING thin refs. Taxable profit (IS
    PBT) is reduced by utilisation of carried-forward losses via a tax-loss control account (opening →
    additions when loss → utilisation against profit → closing); Total taxation = MAX(after×rate, 0);
    Corporate tax = −Total taxation. The IS 'Income tax (expense)' now references this corporate tax
    (was the inline −MAX(0,PBT)×rate) — so the model only taxes profit net of accumulated losses.
    FUNDING presents the financing movements. Runs after build_cupffee_cf."""
    pf, iss, inp = wb[SHEET], wb["IS"], wb[" Inputs"]
    L = {pf.cell(r, 1).value.strip(): r for r in range(1, pf.max_row + 1)
         if isinstance(pf.cell(r, 1).value, str) and pf.cell(r, 1).value.strip()}
    istax = next(r for r in range(1, iss.max_row + 1)
                 if isinstance(iss.cell(r, 1).value, str) and "Income tax (expense)" in iss.cell(r, 1).value)
    pbt = next(r for r in range(1, iss.max_row + 1)
               if isinstance(iss.cell(r, 1).value, str) and "before income tax" in iss.cell(r, 1).value)
    rate = next(f"' Inputs'!$J${r}" for r in range(1, inp.max_row + 1)
                if isinstance(inp.cell(r, 3).value, str) and "Corporate tax rate" in inp.cell(r, 3).value)
    base_st = pf.cell(L["Trade receivables (AR)"], FIRST)._style
    tax_hdr = L["TAXATION"]
    cf_grants = next(r for r in range(1, tax_hdr)
                     if isinstance(pf.cell(r, 1).value, str) and pf.cell(r, 1).value.strip() == "Grants")
    bf, ut, af, tt = (L["Taxable profit before utilisation"], L["Utilisation of tax loss"],
                      L["Taxable profit after utilisation"], L["Total taxation"])
    op_, ad, uc, cl, ct = (L["Tax-loss control — opening"], L["Additions"], L["Utilisation"],
                           L["Tax-loss control — closing"], L["Corporate tax"])
    for c in range(FIRST, LAST + 1):
        x, p, first = get_column_letter(c), get_column_letter(c - 1), c == FIRST
        F = {bf: f"=IS!{x}{pbt}",
             uc: f"=-MIN({x}{op_},MAX(IS!{x}{pbt},0))",          # utilise losses against positive profit
             ut: f"={x}{uc}",
             af: f"={x}{bf}+{x}{ut}",
             tt: f"=MAX({x}{af}*{rate},0)",
             op_: ("0" if first else f"={p}{cl}"),               # opening loss balance = prior closing
             ad: f"=-MIN(IS!{x}{pbt},0)",                        # add losses (positive) when PBT < 0
             cl: f"={x}{op_}+{x}{ad}+{x}{uc}",
             ct: f"=-{x}{tt}",
             L["Equity round"]: f"={x}{L['Capital Increase']}",
             L["Debt draw"]: f"={x}{L['Loan facility financing']}",
             L["Grants"]: f"={x}{cf_grants}"}
        for r, f in F.items():
            pf.cell(r, c, f)._style = base_st
        iss.cell(istax, c).value = f"=ProForma!{x}{ct}"          # IS tax → carryforward corporate tax
    print("  RB2: TAXATION = tax-loss carryforward; IS tax → ProForma corporate tax; FUNDING refs")
