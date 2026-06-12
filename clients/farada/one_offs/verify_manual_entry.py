"""Manual-entry verification for the v3 rolling budget.

Simulates the user's headline scenario — a discrete +€1,000,000 CAPEX funded by a
+€1,000,000 loan, entered **on the Balance Sheet** in May — and asserts the
indirect-derived CF + cash-as-plug behave correctly:

  * In-month (May) cash impact ≈ 0 (the loan funds the capex).
  * Cash stays unchanged vs baseline in EVERY later month too — i.e. the manual
    capex is NOT depreciated (per the v3 rule: manual entries don't model useful
    life). A residual drift => depreciation is leaking onto the manual addition.
  * Balance Check stays ≈ 0 every month.

Usage: uv run python clients/farada/one_offs/verify_manual_entry.py
Exits non-zero on failure.
"""

from __future__ import annotations

import importlib.util
import logging
import shutil
import sys
from pathlib import Path

import openpyxl

HERE = Path(__file__).resolve().parent
WORKBOOK = HERE.parent / "reference" / "rolling_budget_v3.xlsx"
TMP = Path("/tmp/rb_manual_entry_test.xlsx")
TOL = 1.0
AMOUNT = 1_000_000
ENTRY_COL = "I"  # May (E=Jan … H=Apr, I=May)


def _gen():
    spec = importlib.util.spec_from_file_location(
        "build_rolling_budget", HERE / "build_rolling_budget.py")
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


def _recalc(path: Path):
    import formulas
    logging.getLogger("formulas").setLevel(logging.ERROR)
    xl = formulas.ExcelModel().loads(str(path)).finish()
    sol = xl.calculate()

    def get(col, row):
        suffix = f"PRO FORMA'!{col}{row}"
        for k, v in sol.items():
            if k.upper().endswith(suffix.upper()):
                try:
                    return float(v.value[0, 0])
                except Exception:
                    try:
                        return float(v.value)
                    except Exception:
                        return v.value
        return None
    return get


COLS = ["E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P"]
MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def _apply(edits: list[tuple[int, float]]) -> None:
    shutil.copy(WORKBOOK, TMP)
    wb = openpyxl.load_workbook(TMP)
    ws = wb["Pro Forma"]
    for row, amt in edits:
        cur = ws[f"{ENTRY_COL}{row}"].value
        body = str(cur)[1:] if str(cur).startswith("=") else str(cur)
        ws[f"{ENTRY_COL}{row}"] = f"={body}+{amt}"
    wb.save(TMP)
    wb.close()


def _run(name, edits, base, *, expect_cash_neutral):
    """Return list of failure strings; print a per-month line."""
    _apply(edits)
    test = _recalc(TMP)
    print(f"\n[{name}] edits in May: " +
          ", ".join(f"row {r} +{a:,.0f}" for r, a in edits))
    fails = []
    for col, mon in zip(COLS, MONTHS):
        dcash = test(col, R.BS_CASH) - base(col, R.BS_CASH)
        bal = test(col, R.BAL_CHECK)
        if bal is None or abs(bal) > 1.5:
            fails.append(f"{name}/{mon}: BalChk={bal} (BS must stay balanced)")
        if expect_cash_neutral and abs(dcash) > TOL:
            fails.append(f"{name}/{mon}: Δcash={dcash:,.1f} (expected neutral)")
    worst = max(abs(test(c, R.BAL_CHECK) or 0) for c in COLS)
    print(f"   max |BalChk| = {worst:,.1f}  ({'OK' if worst <= 1.5 else 'BROKEN'})")
    return fails


# Module-level R so helpers can reference it.
R = _gen().R


def main() -> int:
    if not WORKBOOK.exists():
        print(f"workbook not found: {WORKBOOK} (run the generator first)", file=sys.stderr)
        return 2
    base = _recalc(WORKBOOK)

    failures = []
    # Scenario 1: capex funded by loan — cash neutral, no depreciation/interest on manual.
    failures += _run("capex+loan", [(R.PPE_NBV, AMOUNT), (R.BS_LOAN, AMOUNT)],
                     base, expect_cash_neutral=True)
    # Scenario 2: a single asset edit (Prepaid) must still keep the BS balanced —
    # the CF must capture ΔPrepaid (else assets move with no offset).
    failures += _run("prepaid-only", [(R.BS_PREPAID, AMOUNT)],
                     base, expect_cash_neutral=False)

    if failures:
        print("\nFAIL:")
        for f in failures:
            print(f"  - {f}")
        return 1
    print("\nOK — all manual-entry scenarios keep the balance sheet balanced.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
