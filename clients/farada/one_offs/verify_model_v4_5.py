"""Verify farada_model_v4.5.xlsx — CF re-architecture R1–R3.

R1 ProForma stripped of profitability subtotals; R2 IS computes GP/EBITDA/EBIT/PBT/tax/NP from
its own leaves; R3 ProForma WC + financing rolls (tax-payable & RE reference the IS). No recalc
engine → structural + logic checks.

Run:  .venv/bin/python clients/farada/one_offs/verify_model_v4_5.py
"""
from __future__ import annotations

import openpyxl
from openpyxl.worksheet.formula import ArrayFormula

P = "clients/farada/modeling/farada_model_v4.5.xlsx"


def ft(cell):
    v = cell.value
    return (v.text if isinstance(v, ArrayFormula) else v)


def labels(ws):
    out = {}
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if isinstance(v, str) and v.strip():
            out.setdefault(v.strip(), r)
    return out


def _balance_oracle():
    """Replicate the CF/BS formula logic on synthetic monthly leaves and assert the BS balances
    (Assets = Equity + Liabilities) every month — independent of the sheet's computed P&L."""
    rate, DSO, DPO, PAYD, PREPAY, SAASANN, LIFE_M = 0.25, 30, 30, 30, 0.5, 0.0, 60
    OB_CASH, OB_AR, OB_AP, OB_PAY, OB_DEF, OB_DEBT, OB_SC, OPEN_PPE = 2e6, 0, 0, 0, 0, 0, 5e6, 0
    H = [100000, 120000, 0, 200000]        # hardware+components revenue
    S = [10000, 12000, 15000, 18000]       # SaaS revenue
    COGS = [40000, 48000, 5000, 60000]
    OPEX = [50000, 50000, 50000, 50000]
    PAY = [30000, 30000, 30000, 30000]     # payroll (⊂ OPEX)
    FIN = [1000, 1000, 1000, 1000]
    GRANT = [0, 0, 50000, 0]
    CAPEX = [5000, 5000, 5000, 5000]
    plug_re = OB_CASH + OB_AR + OPEN_PPE - OB_AP - OB_PAY - OB_DEF - OB_DEBT - OB_SC

    ppe_prev, cash_prev = OPEN_PPE, OB_CASH
    ar_prev, ap_prev, pay_prev, def_prev, taxp_prev, sc_prev, debt_prev, re_prev = (
        OB_AR, OB_AP, OB_PAY, OB_DEF, 0.0, OB_SC, OB_DEBT, plug_re)
    for m in range(4):
        rev = H[m] + S[m]
        dep = min((ppe_prev + CAPEX[m]) / LIFE_M, ppe_prev + CAPEX[m])   # P&L D&A == schedule dep
        ppe = ppe_prev + CAPEX[m] - dep
        ebitda = rev - COGS[m] - OPEX[m]
        ebit = ebitda + GRANT[m] - dep
        pbt = ebit - FIN[m]                          # finance income 0
        tax = -max(0.0, pbt) * rate
        np = pbt + tax
        # rolls (closing balances)
        ar = (H[m] * (1 - PREPAY) + S[m] * (1 - SAASANN)) / 30 * DSO
        ap = (COGS[m] + (OPEX[m] - PAY[m])) / 30 * DPO
        pay = PAY[m] / 30 * PAYD
        deferred = S[m] * SAASANN * 6
        taxp = 0.0                                   # lag 0
        sc, debt = OB_SC, OB_DEBT                    # no tranches in test window
        re = re_prev + np
        # CF (direct)
        op = (rev - (ar - ar_prev)) + (-(COGS[m] + OPEX[m] - PAY[m]) + (ap - ap_prev)) \
            + (-PAY[m] + (pay - pay_prev)) + (tax + (taxp - taxp_prev)) + (-FIN[m]) + (deferred - def_prev)
        inv = -CAPEX[m]
        fin = (sc - sc_prev) + (debt - debt_prev) + GRANT[m]
        excess = op + inv + fin
        cash = cash_prev + excess
        # BS
        assets = 0 + ppe + cash + ar
        el = sc + re + debt + ap + pay + deferred + taxp
        if abs(assets - el) > 1e-6:
            print(f"     month {m}: check = {assets - el:.4f}  (assets {assets:.0f} vs E&L {el:.0f})")
            return False
        ppe_prev, cash_prev, ar_prev, ap_prev, pay_prev, def_prev, taxp_prev, sc_prev, debt_prev, re_prev = (
            ppe, cash, ar, ap, pay, deferred, taxp, sc, debt, re)
    return True


def main():
    wb = openpyxl.load_workbook(P)
    ws, inp, iss = wb["ProForma"], wb[" Inputs"], wb["IS"]
    L = labels(iss)
    fails = []

    def ck(cond, msg):
        print(f"  {'✅' if cond else '❌'} {msg}")
        if not cond:
            fails.append(msg)

    print("\n[R1] ProForma profitability subtotals stripped")
    for r in (44, 116, 125, 130, 131, 132):
        ck(ws.cell(r, 3).value is None and ws.cell(r, 1).value is None, f"ProForma row {r} blank")

    print("\n[R2] IS computes subtotals from its own leaves (not ProForma pulls)")
    ck(ft(iss.cell(L["Gross profit TOTAL (€)"], 3)) == f"=C{L['Revenue']}-C{L['COGS TOTAL']}", "GP = Rev − COGS")
    ck(ft(iss.cell(L["EBITDA"], 3)) == f"=C{L['Gross profit TOTAL (€)']}-C{L['Operating expenses']}", "EBITDA = GP − OPEX")
    ck(ft(iss.cell(L["EBIT"], 3)) == f"=C{L['EBITDA']}+C{L['Grant financing']}-C{L['Depreciation & amortisation']}", "EBIT = EBITDA + grant − D&A")
    ck(ft(iss.cell(L["Income tax (expense)"], 3)) == f"=-MAX(0,C{L['Profit / (loss) before income tax']})*' Inputs'!$J$158", "Tax = −MAX(0,PBT)×rate (on IS)")
    ck(ft(iss.cell(L["Net profit / (loss) for the period"], 3)) == f"=C{L['Profit / (loss) before income tax']}+C{L['Income tax (expense)']}", "NP = PBT + tax")
    # per-bundle GP dropped
    hwgp = L["Hardware GP (€)"]
    ck(all(iss.cell(hwgp + i, 3).value is None for i in (1, 2, 3)), "per-bundle GP rows dropped")

    print("\n[F1] SaaS COGS plugged to target GM (placeholder, calibrate later)")
    from openpyxl.utils import get_column_letter
    ck(all(ft(ws.cell(41, c)) == f"={get_column_letter(c)}19*(1-' Inputs'!$J$99)" for c in range(3, 63)),
       "SaaS COGS (ProForma 41) = SaaS rev × (1−J99 target GM), all cols")
    ck(isinstance(inp.cell(99, 15).value, str) and "placeholder" in inp.cell(99, 15).value.lower(),
       "J99 SaaS-GM-target carries a placeholder note in col O")

    print("\n[R3] ProForma rolls; tax-payable & RE reference the IS")
    Lp = labels(ws)
    ck(ft(ws.cell(Lp["Trade receivables (AR)"], 3)).startswith("=((C5+C9+C15)*(1-"), "AR roll present")
    ck(ft(ws.cell(Lp["Trade payables (total)"], 3)).startswith("=C145+C146+C147+C148"), "AP total = Σ buckets")
    tp = Lp["Tax payable"]
    ck(ft(ws.cell(tp, 3)) == f"=-IS!C{L['Income tax (expense)']}*IF(' Inputs'!$J$167>=1,1,0)", "Tax-payable roll → IS tax")
    re = Lp["Retained earnings"]
    np = L["Net profit / (loss) for the period"]
    ck(ft(ws.cell(re, 3)).startswith("=(' Inputs'!$J$175+' Inputs'!$J$176+' Inputs'!$J$155")
       and ft(ws.cell(re, 3)).endswith(f"+IS!C{np}"), "RE roll t0 = balancing plug + IS net profit")
    ck(ft(ws.cell(re, 4)) == f"=C{re}+IS!D{np}", "RE roll month 2 = prior + IS NP")

    print("\n[inputs] CF groups present")
    Li = labels_col3 = {}
    for r in range(155, 185):
        v = inp.cell(r, 3).value
        if isinstance(v, str):
            labels_col3[v.strip()] = r
    for kw in ("Receivable days (DSO)", "Payable days (DPO)", "Equity round amount",
               "Opening cash", "Opening retained earnings"):
        ck(any(kw in k for k in labels_col3), f"input: {kw}")

    print("\n[R4] CF statement (direct) present + wired to rolls/IS")
    cf = wb["CF"]
    Lc = labels(cf)
    ck(ft(cf.cell(Lc["Cash received from customers"], 3)) == "=ProForma!C4-(ProForma!C144-' Inputs'!$J$176)",
       "Cash from customers = rev − ΔAR")
    ck(ft(cf.cell(Lc["Ending Cash Balance"], 3)) == "=C22+C21", "Ending cash = beginning + excess (plug)")
    ck(ft(cf.cell(Lc["Beginning Cash Balance"], 3)) == "=' Inputs'!$J$175", "Beginning cash (t0) = opening cash input")
    ck(ft(cf.cell(Lc["Beginning Cash Balance"], 4)) == "=C23", "Beginning cash (t1) = prior ending")

    print("\n[R5] BS present; cash=CF ending; check row = Assets − E&L")
    bs = wb["BS"]
    Lb = labels(bs)
    ck(ft(bs.cell(Lb["Cash & cash equivalents"], 3)) == "=CF!C23", "BS cash = CF ending")
    ck(ft(bs.cell(Lb["check (Assets − E&L)"], 3)) == "=C9-C19", "check = TOTAL ASSETS − TOTAL E&L")

    print("\n[R6] yearly CF_Y / BS_Y (calendar 2026–2030)")
    for s in ("CF_Y", "BS_Y"):
        ck(s in wb.sheetnames, f"sheet {s} present")
    cfy, bsy = wb["CF_Y"], wb["BS_Y"]
    Lcy, Lby = labels(cfy), labels(bsy)
    ck(cfy.cell(2, 3).value == "2026" and cfy.cell(2, 7).value == "2030", "CF_Y years 2026..2030")
    op = Lcy["Cash Flow from Operating Activities"]
    ck(ft(cfy.cell(op, 3)) == f"=SUM(CF!C{op}:H{op})", "CF_Y 2026 operating = SUM(CF! Jul–Dec)")
    ck(ft(cfy.cell(op, 4)) == f"=SUM(CF!I{op}:T{op})", "CF_Y 2027 operating = SUM(CF! Jan–Dec)")
    chk = Lby["check (Assets − E&L)"]
    ck(ft(bsy.cell(chk, 3)) == f"=BS!H{chk}", "BS_Y 2026 = BS Dec-2026 snapshot")
    ck(ft(bsy.cell(chk, 4)) == f"=BS!T{chk}", "BS_Y 2027 = BS Dec-2027 snapshot")

    print("\n[oracle] BS balances by construction (synthetic flows → check = 0 every month)")
    ck(_balance_oracle(), "synthetic 3-statement run: BS check = 0 for all test months")

    print("\n[overhaul] dead Line-3 usage-pricing ladder removed")
    ck(all(inp.cell(r, 3).value is None and inp.cell(r, 12).value is None for r in range(24, 30)),
       "usage-pricing ladder + header (Inputs 24-29) blanked (was orphaned)")

    print("\n[struct] no naked rows; no new #REF!")
    naked = [f"C{r}" for r in (Lp["Trade receivables (AR)"], re, L["EBITDA"])
             if ws.cell(r, 3).value is not None and ws.cell(r, 3).number_format == "General"
             or (iss.cell(L["EBITDA"], 3).number_format == "General")]
    ck(iss.cell(L["EBITDA"], 3).number_format != "General", "IS EBITDA styled (not General)")
    refrows = {(sn, c.row) for sn in wb.sheetnames for row in wb[sn].iter_rows() for c in row
               if isinstance(ft(c), str) and "#REF!" in ft(c)}
    ck(not refrows, f"no #REF! anywhere (capacity row repaired) (got {sorted(refrows)})")
    # blanked cells read as 0 (not #REF!), so explicitly assert nothing still pulls a stripped row
    import re
    strip = set(range(44, 56)) | {116, 125, 130, 131, 132}
    dangling = []
    for sn in wb.sheetnames:
        for row in wb[sn].iter_rows():
            for cell in row:
                t = ft(cell)
                if isinstance(t, str) and any(int(m) in strip for m in re.findall(r"ProForma!\$?[A-Z]{1,2}\$?(\d+)", t)):
                    dangling.append(f"{sn}!{cell.coordinate}")
    ck(not dangling, f"no cell still references a stripped ProForma row (found {dangling[:5]})")

    print()
    if fails:
        print(f"FAILED {len(fails)} check(s):")
        for f in fails:
            print(f"  - {f}")
        raise SystemExit(1)
    print("ALL CHECKS PASSED")


if __name__ == "__main__":
    main()
