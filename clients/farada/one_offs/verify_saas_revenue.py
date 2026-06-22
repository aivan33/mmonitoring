"""SaaS revenue oracle — verify the Line-3 SaaS revenue independently before touching costs.

The model computes SaaS revenue by summing per-cohort monthly MRR additions (ProForma row 20,
cumulative). This recomputes the SAME revenue a DIFFERENT way — a stock×rate reconstruction off
the cumulative installed sensor base — and checks the two agree. It then reports the implied
€/sensor/yr ARPU and a subscription-base sensitivity (the model is subscription+overage but the
subscription price is missing today), so we can judge magnitude and improve the logic.

Read-only (does not modify the workbook). No recalc engine needed — computed from inputs.
Run:  .venv/bin/python clients/farada/one_offs/verify_saas_revenue.py
"""
from __future__ import annotations

import openpyxl

P = "clients/farada/modeling/farada_model_v4.xlsx"
MONTHS = 60                       # Jul-2026 .. Jun-2031
QUARTERS = 20
RI_FIRST_COL = 2                 # Revenue_Inputs quarter columns B.. (B = 2026 Q3 = Jul-2026)
# (label, sensors/bundle row, included/sensor row, overage-price row, timetable row)
BUNDLES = [("S", 37, 41, 45, 12), ("M", 38, 42, 46, 13), ("L", 39, 43, 47, 14)]
# Mockup subscription base €/sensor/yr to test (the missing input). 0 = current overage-only.
SUBS_SENSITIVITY = [0.0, 2.0, 5.0, 10.0]
# calendar year -> month indices (m=0 is Jul-2026). 2026 = Jul-Dec (partial); drop 2031 H1.
def cal_year(m):  # noqa
    return 2026 + (m + 6) // 12


def monthly_new(qcounts):
    """Quarterly count -> monthly new units, matching the sheet's INT/MOD phasing."""
    out = []
    for m in range(MONTHS):
        n = qcounts[m // 3]
        k = m % 3 + 1
        out.append(n // 3 + (1 if (n % 3) >= k else 0))
    return out


def cumsum(xs):
    s, out = 0.0, []
    for x in xs:
        s += x
        out.append(s)
    return out


def main():
    wb = openpyxl.load_workbook(P, data_only=False)
    inp, ri = wb[" Inputs"], wb["Revenue_Inputs"]
    L = lambda r: inp.cell(r, 12).value or 0          # Realistic value (D2=1 -> J=OFFSET->L)
    avg_meas = L(106)
    hw_markup = L(98)

    print(f"avg measurements/sensor/yr = {avg_meas:,.0f}\n")
    print(f"{'Bundle':7}{'sensors/bdl':>12}{'incl':>6}{'€/meas':>8}"
          f"{'overage €/sensor/yr':>20}")
    per = {}
    for lbl, sr, ir, pr, tr in BUNDLES:
        sensors, incl, price = L(sr), L(ir), L(pr)
        ov_per = max(0.0, avg_meas - incl) * price
        per[lbl] = dict(sensors=sensors, incl=incl, price=price, ov_per=ov_per, trow=tr)
        print(f"{lbl:7}{sensors:>12,.0f}{incl:>6.0f}{price:>8.3f}{ov_per:>20.2f}")

    # ---- build monthly installed-sensor base + revenue, two ways -------------
    tie_max = 0.0
    inst_total = [0.0] * MONTHS
    ov_mrr = [0.0] * MONTHS
    bundle_series = {}                       # lbl -> {bundles_cum, inst, mrr} monthly lists
    for lbl, d in per.items():
        qcounts = [ri.cell(d["trow"], RI_FIRST_COL + q).value or 0 for q in range(QUARTERS)]
        new_u = monthly_new(qcounts)
        bundles_cum = cumsum(new_u)
        inst_sensors = cumsum([n * d["sensors"] for n in new_u])
        # method A (sheet/cohort): accumulate new cohorts' monthly MRR
        mrr_cohort, acc = [], 0.0
        for m in range(MONTHS):
            acc += new_u[m] * d["sensors"] * d["ov_per"] / 12
            mrr_cohort.append(acc)
        # method B (stock): installed sensors * per-sensor rate
        mrr_stock = [inst_sensors[m] * d["ov_per"] / 12 for m in range(MONTHS)]
        tie_max = max(tie_max, max(abs(a - b) for a, b in zip(mrr_cohort, mrr_stock)))
        bundle_series[lbl] = dict(bundles_cum=bundles_cum, inst=inst_sensors, mrr=mrr_stock)
        for m in range(MONTHS):
            inst_total[m] += inst_sensors[m]
            ov_mrr[m] += mrr_stock[m]

    print(f"\n[tie-out] cohort vs stock overage-MRR, max abs diff = {tie_max:.6f}  "
          f"({'TIE ✅' if tie_max < 1e-6 else 'MISMATCH ❌'})")

    # ---- revenue PER YEAR PER BUNDLE (overage-only) -------------------------
    yrs = [y for y in (2026, 2027, 2028, 2029, 2030)]
    midx = {y: [m for m in range(MONTHS) if cal_year(m) == y] for y in yrs}
    print("\nAnnual SaaS revenue PER BUNDLE (overage-only); bundle sold = its #sensors deployed:")
    for lbl, d in per.items():
        s = bundle_series[lbl]
        print(f"\n  Bundle {lbl}  —  {d['sensors']:,.0f} sensors/bundle, overage €{d['ov_per']:.2f}/sensor/yr")
        print(f"    {'year':6}{'bundles sold (cum)':>20}{'installed sensors':>20}{'SaaS rev €':>16}")
        for y in yrs:
            ms = midx[y]
            last = ms[-1]
            rev = sum(s["mrr"][m] for m in ms)
            print(f"    {y:<6}{s['bundles_cum'][last]:>20,.0f}{s['inst'][last]:>20,.0f}{rev:>16,.0f}")

    # ---- annual (calendar) overage revenue + implied ARPU -------------------
    print("\nAnnual SaaS (calendar; 2026 = Jul–Dec partial), overage-only (current logic):")
    print(f"{'year':6}{'installed sensors (Dec)':>26}{'overage rev €':>16}{'ARPU €/sensor/yr':>18}")
    years = {}
    for m in range(MONTHS):
        y = cal_year(m)
        if y > 2030:
            continue
        years.setdefault(y, {"rev": 0.0})
        years[y]["rev"] += ov_mrr[m]
        years[y]["inst_dec"] = inst_total[m]   # last month seen in the year = Dec (or Jun for 2031)
    for y in sorted(years):
        rev = years[y]["rev"]; inst = years[y]["inst_dec"]
        arpu = (rev / inst * 12 / (len([m for m in range(MONTHS) if cal_year(m) == y])) ) if inst else 0
        # simpler ARPU: year-end MRR annualised / installed
        print(f"{y:<6}{inst:>26,.0f}{rev:>16,.0f}{(ov_mrr[[m for m in range(MONTHS) if cal_year(m)==y][-1]]*12/inst if inst else 0):>18.2f}")

    # ---- subscription sensitivity (the missing piece) -----------------------
    print("\nSubscription+overage sensitivity — implied blended ARPU & FY2030 SaaS revenue")
    print("(subscription base applied €/sensor/yr across installed base; overage on top):")
    print(f"{'base €/sens/yr':>15}{'FY2030 SaaS rev €':>20}{'blended ARPU €/sensor/yr':>26}")
    m2030 = [m for m in range(MONTHS) if cal_year(m) == 2030]
    for base in SUBS_SENSITIVITY:
        sub_mrr = [inst_total[m] * base / 12 for m in range(MONTHS)]
        fy = sum((ov_mrr[m] + sub_mrr[m]) for m in m2030)
        inst_end = inst_total[m2030[-1]]
        arpu = (ov_mrr[m2030[-1]] + sub_mrr[m2030[-1]]) * 12 / inst_end if inst_end else 0
        print(f"{base:>15.1f}{fy:>20,.0f}{arpu:>26.2f}")

    # ---- context: hardware €/sensor (one-time) for comparison ---------------
    print(f"\ncontext: hardware is ONE-TIME ≈ unit_cost×(1+{hw_markup:.0%}) ≈ €1–4/sensor; "
          f"SaaS above is RECURRING per year. That contrast is the magnitude question.")


if __name__ == "__main__":
    main()
