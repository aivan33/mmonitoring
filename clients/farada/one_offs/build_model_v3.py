"""Build farada_model_v3 from farada_model_5y: split Line 3 (Hardware-enabled
SaaS #3) into Hardware>Bundle S/M/L and SaaS>Bundle S/M/L, and re-wire the
orphaned pricing inputs (J72 hardware markup, J59-61 included measurements).

P&L only. COGS Line 3 stays an aggregate row (decision 1). SaaS stays always-on
(decision 3). J73 (SaaS GM target) is left as a check, not a driver (flagged).

No recalc engine is available (formulas lib lacks OFFSET), so correctness is
safe-by-construction + checked: (a) per-bundle landed sensors must re-aggregate
to the workbook's Excel-cached Sensors-L3 row, (b) Lines 1&2 formulas must be
byte-identical after translation, (c) subtotals = sum of bundles by construction.

Run from repo root:  .venv/bin/python clients/farada/one_offs/build_model_v3.py
"""
from __future__ import annotations

import re
from copy import copy

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.formula import ArrayFormula

SRC = "clients/farada/modeling/farada_model_5y.xlsx"
DST = "clients/farada/modeling/farada_model_v3.xlsx"
FIRST, LAST = 3, 62  # ProForma month columns C..BJ (60 months, Jul-2026..Jun-2031)

# Bundle: (label, Revenue_Inputs row, sensors/bundle J, included-meas J, overage J)
BUNDLES = [("S", 12, 55, 59, 63), ("M", 13, 56, 60, 64), ("L", 14, 57, 61, 65)]

# ---- OPEX (copied from the reference model's `is` sheet, rows 36-64) ----
REF_MODEL = "clients/farada/modeling/model_farada_ 230326.xlsx"
# node = (label, reference is-row, children | None).  None children => leaf.
OPEX_TREE = ("Operating expenses", 36, [
    ("S&M", 37, [
        ("Total Payroll", 38, None),
        ("Events/Exhibitions", 39, None),
        ("Travel (Hotels, Food, flights)", 40, None),
        ("Digital Marketing", 41, None),
        ("Outsourced Marketing", 42, None),
        ("Content Marketing", 43, None),
        ("Sales Commissions", 44, None),
        ("Other marketing expenses", 45, None),
    ]),
    ("G&A", 46, [
        ("Total Payroll", 47, None),
        ("Office Expenses", 48, None),
        ("Travel and Representative", 49, None),
        ("Software and Tools", 50, None),
        ("Team Developments", 51, None),
        ("External Professional Services", 52, [
            ("Legal", 53, None),
            ("Accounting", 54, None),
            ("Other / Consulting Services", 55, None),
        ]),
        ("Miscellaneous expenses", 56, None),
    ]),
    ("R&D", 57, [
        ("Total Payroll (60% expensed)", 58, [
            ("Germany", 59, None),
            ("Serbia", 60, None),
        ]),
        ("Software and Tools", 61, None),
        ("R&D Sensors", 62, None),
        ("R&D Rent", 63, None),
        ("Other R&D expenses", 64, None),
    ]),
])


def flatten_opex(node, depth=0, out=None, bucket=None):
    """Pre-order flatten -> list of dicts; `bucket` = the depth-1 ancestor (S&M/G&A/R&D)."""
    if out is None:
        out = []
    label, is_row, kids = node
    b = label if depth == 1 else bucket
    rec = {"label": label, "is_row": is_row, "depth": depth, "leaf": kids is None,
           "bucket": b, "child_idx": []}
    out.append(rec)
    me = len(out) - 1
    if kids:
        for k in kids:
            out[me]["child_idx"].append(len(out))
            flatten_opex(k, depth + 1, out, b)
    return out


# Payroll leaves (is-rows) come from HR -> left blank in Inputs.
PAYROLL_IS_ROWS = {38, 47, 59, 60}

# ---------------------------------------------------------------- row map
# Insert 6 rows after original row 16 (revenue block) and 6 after original 43
# (GP block).  Original 14/15/16 and 41/42/43 are rewritten in place.
def newrow(r: int) -> int:
    if r <= 16:
        return r
    if r <= 43:
        return r + 6
    return r + 12


# References to the SaaS subtotals must point at their new semantic location:
# old 16 (SaaS revenue) -> 19, old 43 (SaaS GP) -> 52.
OVR = {16: 19, 43: 52}
def ref_row(r: int) -> int:
    return OVR.get(r, newrow(r))


_CELL = re.compile(r"(\$?)([A-Z]{1,3})(\$?)(\d+)")
_REF = re.compile(
    r"(?P<pre>(?:'[^']*'|[A-Za-z_][A-Za-z0-9_.]*)!)?"
    r"(?P<c1>\$?[A-Z]{1,3}\$?\d+)"
    r"(?P<rng>:(?P<c2>\$?[A-Z]{1,3}\$?\d+))?"
)


def _remap(tok: str) -> str:
    d1, col, d2, row = _CELL.match(tok).groups()
    return f"{d1}{col}{d2}{ref_row(int(row))}"


def translate(formula: str) -> str:
    """Remap same-sheet (ProForma-local) row refs; leave cross-sheet refs alone."""
    def repl(m: re.Match) -> str:
        if m.group("pre"):
            return m.group(0)  # cross-sheet ref/range -> untouched
        out = _remap(m.group("c1"))
        if m.group("rng"):
            out += ":" + _remap(m.group("c2"))
        return out
    return _REF.sub(repl, formula)


# ---------------------------------------------------------------- helpers
def landing_term(ri_letter: str, ri_row: int, k: int) -> str:
    """Bundles landing in a given month = quarterly count phased into 3 months."""
    cell = f"Revenue_Inputs!{ri_letter}${ri_row}"
    return f"(INT({cell}/3)+IF(MOD({cell},3)>={k},1,0))"


def col_axis(c: int):
    """(column letter, prev letter, Revenue_Inputs quarter column, month-in-qtr k)."""
    m = c - FIRST
    return get_column_letter(c), get_column_letter(c - 1), get_column_letter(2 + m // 3), m % 3 + 1


def build():
    wb = openpyxl.load_workbook(SRC)
    ws = wb["ProForma"]

    # Snapshot label-column + a detail row's style to harvest from.
    style_rev = ws.cell(15, 1)._style          # Line-3 sub-line label style
    style_gp = ws.cell(42, 1)._style           # Hardware-GP label style
    data_style_rev = {c: ws.cell(15, c)._style for c in range(FIRST, LAST + 1)}
    data_style_gp = {c: ws.cell(42, c)._style for c in range(FIRST, LAST + 1)}

    # 1) physically open space (styles of shifted cells preserved; formulas go stale)
    ws.insert_rows(17, 6)
    ws.insert_rows(49, 6)

    # 2) translate every carried formula's row refs old->new (skip array formulas:
    #    rows 6-12 are cross-sheet only and never shift) and skip rewritten blocks.
    rewritten = set(range(14, 23)) | set(range(47, 56))
    refs_16_43 = []
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if isinstance(v, ArrayFormula):
                continue
            if isinstance(v, str) and v.startswith("="):
                if cell.row in rewritten:
                    continue
                if re.search(r"(?<![A-Za-z'!])\$?[A-Z]{1,3}\$?(16|43)\b", v):
                    refs_16_43.append((cell.coordinate, v))
                cell.value = translate(v)

    # 3) rewrite Line-3 REVENUE block (rows 14-22)
    ws.cell(14, 1, "Hardware-enabled SaaS #3")
    ws.cell(15, 1, "    Hardware (device, cost + markup)")
    ws.cell(19, 1, "    SaaS (overage, recurring)")
    for i, (lbl, _, _, _, _) in enumerate(BUNDLES):
        ws.cell(16 + i, 1, f"        Bundle {lbl}")
        ws.cell(20 + i, 1, f"        Bundle {lbl}")
    for r in (14, 15, 19, 16, 17, 18, 20, 21, 22):
        ws.cell(r, 1)._style = copy(style_rev)

    for c in range(FIRST, LAST + 1):
        L, prev, ri, k = col_axis(c)
        ws.cell(14, c, f"={L}15+{L}19")
        ws.cell(15, c, f"={L}16+{L}17+{L}18")
        ws.cell(19, c, f"={L}20+{L}21+{L}22")
        for i, (lbl, rib, spb, incb, ovb) in enumerate(BUNDLES):
            land = landing_term(ri, rib, k)
            # Hardware = sensors * device_cost * (1 + markup J72).
            # device_cost = the 6-point component build = Σ driver cells (rows 69-73:
            # chip+packaging+sensor-test+final-test+ASIC), replacing the removed all-in
            # inputs J69/J70.  See align_cost_of_sales().
            unit_cost = f"({L}69+{L}70+{L}71+{L}72+{L}73)"
            hw = (f"={land}*' Inputs'!$J${spb}"
                  f"*({unit_cost}*(1+' Inputs'!$J$72))")
            ws.cell(16 + i, c, hw)
            # SaaS recurring = cumulative + new bundles * billable-overage measurements
            rev = (f"{land}*' Inputs'!$J${spb}"
                   f"*MAX(0,' Inputs'!$J$46-' Inputs'!$J${incb})/12*' Inputs'!$J${ovb}")
            ws.cell(20 + i, c, f"={rev}" if c == FIRST else f"={prev}{20 + i}+{rev}")
            for r in (16 + i, 20 + i):
                ws.cell(r, c)._style = copy(data_style_rev[c])

    # 4) rewrite Line-3 GP block (rows 47-55); COGS Line 3 stays aggregate at row 35
    ws.cell(47, 1, "  Gross profit Line 3 (€)")
    ws.cell(48, 1, "      Hardware GP (€)")
    ws.cell(52, 1, "      SaaS GP (€)")
    for i, (lbl, _, _, _, _) in enumerate(BUNDLES):
        ws.cell(49 + i, 1, f"          Bundle {lbl}")
        ws.cell(53 + i, 1, f"          Bundle {lbl}")
    for r in (47, 48, 52, 49, 50, 51, 53, 54, 55):
        ws.cell(r, 1)._style = copy(style_gp)

    for c in range(FIRST, LAST + 1):
        L, prev, ri, k = col_axis(c)
        ws.cell(47, c, f"={L}14-{L}35")          # rev #3 - aggregate COGS #3
        ws.cell(48, c, f"={L}49+{L}50+{L}51")
        ws.cell(52, c, f"={L}53+{L}54+{L}55")
        for i, (lbl, rib, spb, incb, ovb) in enumerate(BUNDLES):
            # Hardware GP = HW rev * markup/(1+markup)  (closed form, COGS = cost)
            ws.cell(49 + i, c,
                    f"={L}{16 + i}*' Inputs'!$J$72/(1+' Inputs'!$J$72)")
            # SaaS GP = cumulative(rev - usage cost); usage cost = meas * cloud J42
            land = landing_term(ri, rib, k)
            drev = (f"{land}*' Inputs'!$J${spb}"
                    f"*MAX(0,' Inputs'!$J$46-' Inputs'!$J${incb})/12*' Inputs'!$J${ovb}")
            dcost = f"{land}*' Inputs'!$J${spb}*' Inputs'!$J$46/12*' Inputs'!$J$42"
            delta = f"{drev}-({dcost})"
            ws.cell(53 + i, c, f"={delta}" if c == FIRST else f"={prev}{53 + i}+{delta}")
            for r in (49 + i, 53 + i):
                ws.cell(r, c)._style = copy(data_style_gp[c])

    # ---------------------------------------------------------------- OPEX
    opex = flatten_opex(OPEX_TREE)
    ref_fmt = openpyxl.load_workbook(REF_MODEL)["is"]                  # labels
    ref_val = openpyxl.load_workbook(REF_MODEL, data_only=True)["is"]  # cached values
    # single flat amount per category = reference's first-month run-rate (col C).
    amt = {rec["is_row"]: (ref_val.cell(rec["is_row"], FIRST).value or 0) for rec in opex}

    # ---- Inputs: one row per category, in the existing-input style ----------
    #   C=label  D=unit  G/H=start/end date  J=OFFSET active  L=Realistic amount
    inp = wb[" Inputs"]
    S_SEC_A, S_SEC_C = inp.cell(7, 1)._style, inp.cell(7, 3)._style      # section hdr
    S_SUB_B, S_SUB_C = inp.cell(15, 2)._style, inp.cell(15, 3)._style    # sub-number
    S_LAB, S_UNIT = inp.cell(8, 3)._style, inp.cell(8, 4)._style
    S_G = inp.cell(8, 7)._style                  # start-date style (mmm-yy, centered)
    S_H = S_G                                    # end-date: same date format (H8 was blank/General)
    S_J, S_L = inp.cell(8, 10)._style, inp.cell(8, 12)._style
    start_dt, end_dt = ws.cell(2, FIRST).value, ws.cell(2, LAST).value

    r = 87
    inp.cell(r, 1, "VIII.")._style = copy(S_SEC_A)
    inp.cell(r, 3, "OPERATING EXPENSES (single monthly run-rate, EUR; "
                   "step-changes = add rows later)")._style = copy(S_SEC_C)
    r += 1
    for bnum, blabel in [("8.1", "S&M"), ("8.2", "G&A"), ("8.3", "R&D")]:
        inp.cell(r, 2, bnum)._style = copy(S_SUB_B)
        inp.cell(r, 3, blabel)._style = copy(S_SUB_C)
        r += 1
        for rec in opex:
            if not (rec["leaf"] and rec["bucket"] == blabel):
                continue
            R = rec["is_row"]
            pay = R in PAYROLL_IS_ROWS
            label = ref_fmt.cell(R, 1).value + (" (from HR)" if pay else "")
            inp.cell(r, 3, label)._style = copy(S_LAB)
            inp.cell(r, 4, "EUR/mo")._style = copy(S_UNIT)
            inp.cell(r, 7, start_dt)._style = copy(S_G)
            inp.cell(r, 8, end_dt)._style = copy(S_H)
            inp.cell(r, 10, f"=OFFSET(K{r},0,$D$2)")._style = copy(S_J)
            inp.cell(r, 12, None if pay else amt[R])._style = copy(S_L)
            rec["inp_row"] = r
            r += 1

    # ---- ProForma: append OPEX, COLORS aligned to the existing ProForma ----
    # template rows: 43 section header (blue), 44 headline total (blue #,##0),
    # 56 headline % (blue 0.0%), 14 bold subtotal (no fill), 16 plain leaf.
    TPL = {"total": 44, "sub": 14, "leaf": 16}
    PR0 = ws.max_row + 3
    ws.cell(PR0, 1, "OPERATING EXPENSES")
    for c in range(1, LAST + 1):                                       # blue header band
        ws.cell(PR0, c)._style = copy(ws.cell(43, c)._style)
    base = PR0 + 1
    for i, rec in enumerate(opex):
        rec["pf_row"] = base + i
    for rec in opex:
        r = rec["pf_row"]
        role = "total" if rec["depth"] == 0 else ("leaf" if rec["leaf"] else "sub")
        trow = TPL[role]
        lc = ws.cell(r, 1, "  " * rec["depth"] + ref_fmt.cell(rec["is_row"], 1).value)
        lc._style = copy(ws.cell(trow, 1)._style)
        for c in range(FIRST, LAST + 1):
            L = get_column_letter(c)
            if rec["leaf"]:
                ir = rec["inp_row"]                # flat amount, gated by start/end date
                f = (f"=IF(AND({L}$2>=' Inputs'!$G${ir},{L}$2<=' Inputs'!$H${ir}),"
                     f"' Inputs'!$J${ir},0)")
            else:
                f = "=" + "+".join(f"{L}{opex[ci]['pf_row']}" for ci in rec["child_idx"])
            ws.cell(r, c, f)._style = copy(ws.cell(trow, c)._style)

    # EBITDA (blue headline) = Gross profit TOTAL (row 44) - Operating expenses total
    opex_total = base
    er = base + len(opex) + 1
    ws.cell(er, 1, "EBITDA")._style = copy(ws.cell(44, 1)._style)
    ws.cell(er + 1, 1, "EBITDA margin")._style = copy(ws.cell(56, 1)._style)
    for c in range(FIRST, LAST + 1):
        L = get_column_letter(c)
        ws.cell(er, c, f"={L}44-{L}{opex_total}")._style = copy(ws.cell(44, c)._style)
        ws.cell(er + 1, c, f"=IF({L}4=0,0,{L}{er}/{L}4)")._style = copy(ws.cell(56, c)._style)

    align_cost_of_sales(ws, inp)

    wb.save(DST)
    return refs_16_43, opex


# ---------------------------------------------------------------- cost of sales
# 6-POINT volume cost curve, ported from the live unit-economics file
# (`farada_unit_economics.xlsx`, sheet "Unit Economics v2").  Replaces the model's
# 2-stage (now/at-scale) cost + all-in plug.  Format is kept: COST OF SALES inputs +
# the per-component ProForma cost-driver rows; only values/wiring change.
#
# Volume points (annual sensors) and €/sensor per component.  Chip is derived from
# the file's wafer build (wafer-cost / sensors-per-wafer / yield); the rest are the
# file's hand-entered values.  Testing is split into Sensor + Final (the extra
# component the user flagged).  Last point (4M/yr) is where the own-ASIC cost-down
# lands -> "at-scale" now triggers at 4M, not 1M.
CURVE_THR = [1, 4000, 10000, 100000, 1000000, 4000000]
_SPW = 4000
_WAFER = [4000, 4000, 3735, 3068, 2401, 2000]
_YIELD = [0.70, 0.70, 0.73, 0.82, 0.90, 0.95]
CURVE = {
    "Chip":           [_WAFER[i] / (_SPW * _YIELD[i]) for i in range(6)],
    "Packaging":      [0.30, 0.30, 0.30, 0.10, 0.10, 0.10],
    "Sensor testing": [0.50, 0.50, 0.25, 0.10, 0.03, 0.01],
    "Final testing":  [0.20, 0.20, 0.187, 0.153, 0.12, 0.10],
    "ASIC / readout": [1.50, 1.50, 1.50, 1.50, 1.50, 0.32],
}
# component -> ProForma driver row (v3, after the +12 shift) and its label.
#   69 Chip, 70 Packaging, 71 Sensor test, 72 Final test, 73 ASIC (unified, all lines).
DRIVER = {"Chip": 69, "Packaging": 70, "Sensor testing": 71,
          "Final testing": 72, "ASIC / readout": 73}
DRIVER_LABEL = {69: "  Chip EUR/sensor", 70: "  Packaging €/sensor",
                71: "  Sensor testing €/sensor", 72: "  Final testing €/sensor",
                73: "  ASIC / readout €/sensor (all lines)"}
# COGS rewires (v3 rows): (cogs_row, sensors_row).  Testing = sensor+final drivers;
# ASIC = the single unified driver (row 73) for every line.
COGS_TEST = [(28, 64), (33, 65), (39, 66)]   # L1, L2, L3-HW
COGS_ASIC = [(29, 64), (34, 65), (40, 66)]   # L1, L2, L3-HW
GM_TIER_ROWS = range(85, 94)                  # GROSS MARGIN BY TIER (was 73-81)
# (band threshold, ladder Inputs row) for the GM-by-tier ASP pick, after the rung add.
ASP_LADDER = [(100, 17), (1000, 18), (10000, 19), (100000, 20),
              (100001, 21), (1000000, 22), (4000000, 23)]


def _pick(vol_cell: str, rows6: list[int]) -> str:
    """Nested-IF 6-point pick: highest threshold outermost, default = point 0."""
    expr = f"' Inputs'!$J${rows6[0]}"
    for i in range(1, 6):
        expr = f"IF({vol_cell}>=' Inputs'!$F${rows6[i]},' Inputs'!$J${rows6[i]},{expr})"
    return expr


def align_cost_of_sales(ws, inp):
    # 1) ASP ladder: add the €10 @ 1M rung (was €5) + a new €5 @ 4M rung at blank row 23.
    inp.cell(22, 3, "Price @ 1,000,000 pc")
    inp.cell(22, 12, 10.0)
    for col in (3, 4, 6, 10, 12):
        inp.cell(23, col)._style = copy(inp.cell(22, col)._style)
    inp.cell(23, 3, "Price @ 4,000,000 pc")
    inp.cell(23, 4, "EUR/sensor")
    inp.cell(23, 6, 4000000)
    inp.cell(23, 10, "=OFFSET(K23,0,$D$2)")
    inp.cell(23, 12, 5.0)

    # 2) extend the revenue array-formula ranges to include row 23.
    for r in (6, 7, 8, 10, 11, 12):
        for c in range(FIRST, LAST + 1):
            cell = ws.cell(r, c)
            v = cell.value
            t = v.text if isinstance(v, ArrayFormula) else v
            if not isinstance(t, str):
                continue
            t2 = t.replace("$J$16:$J$22", "$J$16:$J$23").replace("$F$16:$F$22", "$F$16:$F$23")
            cell.value = ArrayFormula(cell.coordinate, t2) if isinstance(v, ArrayFormula) else t2

    # 3) append the 6-point cost curve to Inputs (append-only => no row shift).
    S_SEC = inp.cell(31, 3)._style          # COST OF SALES section header
    S_SUB = inp.cell(32, 3)._style          # sub-header (component name)
    S_C, S_D = inp.cell(16, 3)._style, inp.cell(16, 4)._style   # ladder value-row styles
    S_F, S_J, S_L = inp.cell(16, 6)._style, inp.cell(16, 10)._style, inp.cell(16, 12)._style
    r = inp.max_row + 2
    inp.cell(r, 3, "COST OF SALES — 6-POINT VOLUME CURVE (€/sensor, from unit economics)")._style = copy(S_SEC)
    r += 1
    comp_rows: dict[str, list[int]] = {}
    for name, vals in CURVE.items():
        inp.cell(r, 3, name)._style = copy(S_SUB)
        r += 1
        rows6 = []
        for i, thr in enumerate(CURVE_THR):
            inp.cell(r, 3, f"@ {thr:,} /yr")._style = copy(S_C)
            inp.cell(r, 4, "EUR/sensor")._style = copy(S_D)
            inp.cell(r, 6, thr)._style = copy(S_F)
            inp.cell(r, 10, f"=OFFSET(K{r},0,$D$2)")._style = copy(S_J)
            lc = inp.cell(r, 12, round(vals[i], 4))
            lc._style = copy(S_L)
            lc.number_format = "€#,##0.00"
            rows6.append(r)
            r += 1
        comp_rows[name] = rows6

    # 4) blank the superseded inputs (the "remove"): old 2-pt chip/pkg/test/ASIC 33-40
    #    + all-in block 66-70.  Keep the rows (no shift); leave pointer notes.
    for br in (34, 35, 36, 37, 38, 39, 40, 67, 68, 69, 70):
        for col in (2, 3, 4, 6, 10, 11, 12):
            inp.cell(br, col).value = None
    inp.cell(33, 3, "→ see 6-POINT VOLUME CURVE below (sensor cost moved 2026-06-22)")
    for col in (4, 6, 10, 11, 12):
        inp.cell(33, col).value = None
    inp.cell(66, 3, "ALL-IN SENSOR COST removed 2026-06-22 — superseded by the 6-point curve")

    # 5) rewire the ProForma cost-driver rows to the 6-point curve, keyed on the
    #    column's total run-rate (row 67), exactly like the old 2-stage drivers.
    for name, drow in DRIVER.items():
        ws.cell(drow, 1, DRIVER_LABEL[drow])
        rows6 = comp_rows[name]
        for c in range(FIRST, LAST + 1):
            L = get_column_letter(c)
            ws.cell(drow, c, "=" + _pick(f"{L}67", rows6))

    # 6) rewire COGS: Testing = sensor+final drivers; ASIC = unified driver (row 73).
    for c in range(FIRST, LAST + 1):
        L = get_column_letter(c)
        for cog, sens in COGS_TEST:
            ws.cell(cog, c, f"={L}{sens}*({L}71+{L}72)")
        for cog, sens in COGS_ASIC:
            ws.cell(cog, c, f"={L}{sens}*{L}73")

    # 7) GROSS MARGIN BY TIER: ASP picks include the 4M rung; unit cost = 6-pt build
    #    on the band volume (col C) — preserves the block's "if this tier were the
    #    whole business" framing (stages on band vol, not aggregate run-rate).
    for r in GM_TIER_ROWS:
        cv = f"C{r}"
        asp = "' Inputs'!$J$16"
        for thr, jr in ASP_LADDER:
            asp = f"IF({cv}>={thr},' Inputs'!$J${jr},{asp})"
        ws.cell(r, 4, "=" + asp)
        cost = "+".join(f"({_pick(cv, rows6)})" for rows6 in comp_rows.values())
        ws.cell(r, 5, "=" + cost)


if __name__ == "__main__":
    refs, opex = build()
    print(f"Saved {DST}")
    for rec in opex:
        ir = rec.get("inp_row", "  -")
        kind = "leaf" if rec["leaf"] else "Σ"
        print(f"  PF{rec['pf_row']:>3} IN{ir:>3} [{kind:^4}] "
              f"{'  ' * rec['depth']}{rec['label']}")
