"""Reflow Farada ` Inputs` into the house skeleton (I. FUNDING … V. OTHER) + a FUNDING skeleton.

Permutation approach: existing input rows are MOVED intact (value + style preserved); new section
headers / sub-numbers / blank spacers / empty FUNDING-skeleton inputs are inserted around them.
Only cross-sheet `' Inputs'!$X$NN` references (X∈J/F/G/H/K) are rewritten via the old→new row map.

Safety gate (`_gate`): every input row that is *referenced* anywhere must be MOVEd (not dropped),
and each moved row must keep its label at the new position — so the ref graph is isomorphic and the
computed values are unchanged. Run after build_model_v5; rewrites farada_model_v6.xlsx in place
(.prereflow backup). No recalc engine → correctness = the equivalence gate + the balance oracle.

Run:  .venv/bin/python clients/farada/one_offs/reflow_inputs.py
"""
from __future__ import annotations

import re
import shutil
from copy import copy

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.formula import ArrayFormula

P = "clients/farada/modeling/farada_model_v6.xlsx"
SHEET = " Inputs"
NCOLS = 16  # A..P

# ── target layout ───────────────────────────────────────────────────────────
# items: ('keep', row) move an existing row | ('hdr','I.','TITLE') section band |
#        ('sub','1.1','Title') sub-group | ('blank',) spacer | ('new','label','unit') empty skeleton input
H = lambda a, t: ("hdr", a, t)
S = lambda n, t: ("sub", n, t)
B = ("blank",)
def K(*rows): return [("keep", r) for r in rows]
def NEW(*labels_units): return [("new", lbl, u) for lbl, u in labels_units]

LAYOUT = [
    # scenario selector MUST stay at its original rows — every OFFSET reads the fixed cell $D$2,
    # so row 2 (the selector) is pinned. Leading blank keeps row 1 empty as in the source.
    B, ("keep", 2), ("keep", 3), ("keep", 4), ("keep", 5), B,

    H("I.", "FUNDING ASSUMPTIONS"), B,
    S("1.1", "Equity injection"), *K(170, 171),
    *NEW(("Tranche 1 amount", "EUR"), ("Tranche 1 date", "date"),
         ("Tranche 2 amount", "EUR"), ("Tranche 2 date", "date")), B,
    S("1.2", "Debt / credit facility"), *K(172, 173, 174),
    *NEW(("Repayment period", "months"), ("PMT", "EUR/mo")), B,
    S("1.3", "Convertible loans"),
    *NEW(("Amount", "EUR"), ("Annual interest rate", "%"), ("Conversion date", "date")), B,
    S("1.4", "Grants"), *K(156), B,

    H("II.", "REVENUE ASSUMPTIONS"), B,
    S("2.1", "Hardware step-price ladder (volume tiers, all 3 lines)"), *K(16, 17, 18, 19, 20, 21, 22, 23), B,
    S("2.2", "Avg volume per tier (sensors / client)"), *K(53, 54, 55, 56, 57, 58), B,
    S("2.3", "Line 3 — sensors per bundle"), *K(37, 38, 39), B,
    S("2.4", "Line 3 — included measurements / sensor"), *K(41, 42, 43), B,
    S("2.5", "Line 3 — overage price (EUR / measurement)"), *K(45, 46, 47), B,
    S("2.6", "Pricing parameters"), *K(33, 98, 99, 106), B,

    H("III.", "PRODUCTION"), B,
    S("3.1", "Capacity ceiling (sensors / yr)"), *K(8, 9, 10, 11, 12), B,
    S("3.2", "Cost of sales — wafer cost (€/wafer, staged)"), *K(62, 63, 64, 65, 66, 67), B,
    S("3.3", "Cost of sales — packaging (€/sensor)"), *K(69, 70, 71, 72, 73, 74), B,
    S("3.4", "Cost of sales — sensor testing (€/sensor)"), *K(76, 77, 78, 79, 80, 81), B,
    S("3.5", "Cost of sales — final testing (€/sensor)"), *K(83, 84, 85, 86, 87, 88), B,
    S("3.6", "Cost of sales — ASIC / readout (€/sensor)"), *K(90, 91, 92, 93, 94, 95), B,
    S("3.7", "Cost of sales — usage & parameters"), *K(102, 107, 108), B,

    H("IV.", "OPERATING EXPENSES"), B,
    S("4.1", "S&M"), *K(112, 113, 114, 115, 116, 117, 118, 119), B,
    S("4.2", "G&A"), *K(121, 122, 123, 124, 125, 126, 127, 128, 129), B,
    S("4.3", "R&D"), *K(131, 132, 133, 134, 135, 136, 137), B,

    H("V.", "OTHER ASSUMPTIONS"), B,
    S("5.1", "Below EBITDA — D&A, finance & tax"), *K(153, 154, 155, 157, 158), B,
    S("5.2", "Working capital (payment terms)"), *K(162, 163, 164, 165, 166, 167), B,
    S("5.3", "Opening balances (Jul-2026)"), *K(175, 176, 177, 178, 179, 180, 181, 182),
]


def _ft(c):
    v = c.value
    return v.text if isinstance(v, ArrayFormula) else v


def _referenced_rows(wb):
    """Inputs rows referenced by any $J/$F/$G/$H/$K ref anywhere."""
    pat = re.compile(r"' Inputs'!\$[JFGHK]\$(\d+)")
    rows = set()
    for sn in wb.sheetnames:
        for row in wb[sn].iter_rows():
            for c in row:
                t = _ft(c)
                if isinstance(t, str):
                    rows.update(int(m) for m in pat.findall(t))
    return rows


def reflow(wb):
    inp = wb[SHEET]
    # snapshot existing rows (value + style) for cols A..P
    snap = {r: [(inp.cell(r, c).value, inp.cell(r, c)._style) for c in range(1, NCOLS + 1)]
            for r in range(1, inp.max_row + 1)}
    label_at = {r: inp.cell(r, 3).value for r in snap}          # old label by row (col C)
    band_style = inp.cell(7, 3)._style                          # section header style (grey band)
    sub_style = inp.cell(15, 2)._style                          # sub-number style
    lbl_style = inp.cell(16, 3)._style                          # input label style
    unit_style = inp.cell(16, 4)._style
    val_style = inp.cell(16, 12)._style                         # cream scenario value
    act_style = inp.cell(16, 10)._style                         # cyan active OFFSET

    # assign new rows
    old2new, plan = {}, []
    nr = 0
    for item in LAYOUT:
        nr += 1
        plan.append((nr, item))
        if item[0] == "keep":
            old2new[item[1]] = nr
    last = nr

    # RESET the whole working region to clean blank (value AND style) — the source sheet carries a
    # large tail of pre-formatted empty rows; clearing only values would leave stray formatting.
    maxr, maxc = inp.max_row, max(NCOLS, inp.max_column)
    for r in range(1, maxr + 1):
        for c in range(1, maxc + 1):
            cell = inp.cell(r, c)
            cell.value = None
            cell.style = "Normal"
    for nr, item in plan:
        if item[0] == "keep":
            for c, (val, st) in enumerate(snap[item[1]], start=1):
                cell = inp.cell(nr, c)
                cell.value = val
                cell._style = st
            # repoint the active OFFSET to the NEW row (J{nr}=OFFSET(K{nr},0,$D$2))
            if isinstance(_ft(inp.cell(nr, 10)), str) and "OFFSET" in _ft(inp.cell(nr, 10)):
                inp.cell(nr, 10, f"=OFFSET(K{nr},0,$D$2)")
        elif item[0] == "hdr":
            inp.cell(nr, 1, item[1])._style = band_style
            inp.cell(nr, 3, item[2])._style = band_style
        elif item[0] == "sub":
            inp.cell(nr, 2, item[1])._style = sub_style
            inp.cell(nr, 3, item[2])._style = sub_style
        elif item[0] == "new":                                   # empty skeleton input
            inp.cell(nr, 3, item[1])._style = lbl_style
            inp.cell(nr, 4, item[2])._style = unit_style
            inp.cell(nr, 10, f"=OFFSET(K{nr},0,$D$2)")._style = act_style
            inp.cell(nr, 12)._style = val_style                  # empty cream value cell
    # drop the now-blank trailing rows (no refs point below the skeleton)
    if maxr > last:
        inp.delete_rows(last + 1, maxr - last)
    return old2new, label_at, last


def unify_input_formats(wb):
    """Set each input value cell's (L/M/N) number format from its UNIT — the reflow's new skeleton
    inputs inherited one numeric format regardless of type (dates/%-as-plain-numbers). Unifies the
    whole sheet to: date → mmm-yyyy · % → 0.0% · EUR/sensor|wafer → €#,##0.00 · other EUR → €#,##0 ·
    counts/days/months/years → #,##0."""
    inp = wb[SHEET]

    def fmt(unit, label):
        u, lab = str(unit or "").strip().lower(), str(label or "").lower()
        if "date" in u or "date" in lab:
            return "mmm-yyyy"
        if "%" in u:
            return "0.0%"
        if "eur/sensor" in u or "eur/wafer" in u:
            return "€#,##0.00"
        if u.startswith("eur"):
            return "€#,##0"
        if u.startswith("#") or any(k in u for k in ("sensor", "meas", "pcs", "day", "month", "year")):
            return "#,##0"
        return "#,##0.00"

    n = 0
    for r in range(1, inp.max_row + 1):
        j = inp.cell(r, 10).value
        if isinstance(j, str) and "OFFSET" in j:                # a real input row
            f = fmt(inp.cell(r, 4).value, inp.cell(r, 3).value)
            for c in (12, 13, 14):                              # L/M/N scenario value cells
                cell = inp.cell(r, c)
                cell._style = copy(cell._style)                 # break shared-style aliasing
                cell.number_format = f
            n += 1
    print(f"  fmt: unified {n} input value rows by unit")


def remap_refs(wb, old2new):
    pat = re.compile(r"(' Inputs'!\$[JFGHK]\$)(\d+)")
    def repl(m):
        return f"{m.group(1)}{old2new.get(int(m.group(2)), int(m.group(2)))}"
    n = 0
    for sn in wb.sheetnames:
        for row in wb[sn].iter_rows():
            for c in row:
                t = _ft(c)
                if isinstance(t, str) and "' Inputs'!" in t:
                    nt = pat.sub(repl, t)
                    if nt != t:
                        c.value = nt
                        n += 1
    return n


def _gate(wb, old2new, label_at, referenced):
    inp = wb[SHEET]
    missing = sorted(referenced - set(old2new))
    assert not missing, f"referenced Inputs rows not relocated (refs would break): {missing}"
    bad = [(old, new) for old, new in old2new.items()
           if inp.cell(new, 3).value != label_at[old]]
    assert not bad, f"label moved/lost at: {[(o, n, label_at[o]) for o, n in bad[:8]]}"
    # the scenario selector ($D$2) is read by every OFFSET — it MUST stay numeric at row 2
    assert isinstance(inp.cell(2, 4).value, (int, float)), \
        f"scenario selector $D$2 not preserved (got {inp.cell(2, 4).value!r}) — OFFSETs would break"
    print(f"  gate ✓ {len(old2new)} rows relocated; all {len(referenced)} referenced preserved; selector $D$2 intact")


def main():
    shutil.copyfile(P, P.replace(".xlsx", ".prereflow.xlsx"))
    wb = openpyxl.load_workbook(P)
    referenced = _referenced_rows(wb)
    old2new, label_at, last = reflow(wb)
    _gate(wb, old2new, label_at, referenced)
    n = remap_refs(wb, old2new)
    wb.save(P)
    print(f"  reflow ✓ Inputs → I–V skeleton ({last} rows); remapped {n} formula refs")


if __name__ == "__main__":
    main()
