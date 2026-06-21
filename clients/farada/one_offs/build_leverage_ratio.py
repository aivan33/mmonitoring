"""Build a formatted 'Grant Leverage' workbook from the funding-programmes image.

Transcribes the grant/co-investment table (clients/farada/one_offs/image.png) and
adds a Leverage Ratio = Total Private Investment / Grant Funding, where:
  - Grant Funding         = the "Faraday Grant" column
  - Total Private Investm. = the "Co-investment" column
Totals and the leverage ratio are LIVE formulas so the numbers stay self-consistent
if a row is edited. ~approximate budgets (FastFOx) are entered as their point value.

Run:  .venv/bin/python clients/farada/one_offs/build_leverage_ratio.py
Out:  clients/farada/one_offs/farada_grant_leverage.xlsx
"""
from __future__ import annotations

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = "clients/farada/one_offs/farada_grant_leverage.xlsx"

# (programme, period, total_budget, grant, co_investment, coverage)
ROWS = [
    ("EIC Accelerator", "2023–2025", 3_507_500, 2_455_250, 1_052_250, 0.70),
    ("FastFOx (ILB Brandenburg + ERDF)", "2025–2028", 3_460_000, 2_420_000, 1_040_000, 0.70),
    ("GRW (equipment subsidy)", "2025–2028", 1_980_000, 396_000, 1_584_000, 0.20),
    ("EIC ACCESS+", "2026", 120_000, 60_000, 60_000, 0.50),
    ("InnoMatch (Cascade)", "Jun 2026–May 2027", 80_000, 60_000, 20_000, 0.75),
]
HEADERS = ["Grant / Programme", "Period", "Total Budget", "Grant",
           "Co-investment", "Coverage"]

# palette (mirrors the source image's navy header)
NAVY = "1F3864"
LIGHT = "D9E1F2"
GREY = "F2F2F2"
WHITE = "FFFFFF"
EUR = '#,##0;[Red]-#,##0'
EUR0 = '"€"#,##0'
PCT = "0%"

thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def build() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Grant Leverage"
    ws.sheet_view.showGridLines = False

    # --- Title -------------------------------------------------------------
    ws["A1"] = "Faraday — Grant Funding & Leverage"
    ws["A1"].font = Font(bold=True, size=14, color=NAVY)
    ws["A2"] = "Loans & grants received since founding (2023)"
    ws["A2"].font = Font(italic=True, size=10, color="808080")

    # --- Header row --------------------------------------------------------
    hdr = 4
    for c, name in enumerate(HEADERS, start=1):
        cell = ws.cell(row=hdr, column=c, value=name)
        cell.font = Font(bold=True, color=WHITE, size=11)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    # --- Data rows ---------------------------------------------------------
    first = hdr + 1
    for i, (prog, period, budget, grant, co, cov) in enumerate(ROWS):
        r = first + i
        shade = WHITE if i % 2 == 0 else GREY
        vals = [prog, period, budget, grant, co, cov]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.fill = PatternFill("solid", fgColor=shade)
            cell.border = border
            cell.alignment = Alignment(vertical="center",
                                       horizontal="left" if c <= 2 else "right",
                                       wrap_text=(c == 1))
            if c in (3, 4, 5):
                cell.number_format = EUR0
            elif c == 6:
                cell.number_format = PCT
                cell.alignment = Alignment(horizontal="center", vertical="center")
    last = first + len(ROWS) - 1

    # --- Total row (live SUM formulas) ------------------------------------
    tr = last + 1
    ws.cell(row=tr, column=1, value="Total")
    for c in (3, 4, 5):
        col = get_column_letter(c)
        ws.cell(row=tr, column=c, value=f"=SUM({col}{first}:{col}{last})")
        ws.cell(row=tr, column=c).number_format = EUR0
    # blended coverage = total grant / total budget
    ws.cell(row=tr, column=6, value=f"=D{tr}/C{tr}")
    ws.cell(row=tr, column=6).number_format = PCT
    for c in range(1, 7):
        cell = ws.cell(row=tr, column=c)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.border = border
        if cell.alignment.horizontal is None:
            cell.alignment = Alignment(horizontal="right" if c >= 3 else "left",
                                       vertical="center")
        else:
            cell.alignment = Alignment(horizontal=cell.alignment.horizontal,
                                       vertical="center")

    # --- Leverage ratio block ---------------------------------------------
    lr = tr + 2
    def kv(row, label, value, fmt=None, bold=False, big=False):
        lab = ws.cell(row=row, column=1, value=label)
        lab.font = Font(bold=bold, size=12 if big else 11, color=NAVY if big else "000000")
        lab.fill = PatternFill("solid", fgColor=LIGHT)
        lab.border = border
        val = ws.cell(row=row, column=2, value=value)
        val.font = Font(bold=bold, size=12 if big else 11, color=NAVY if big else "000000")
        val.fill = PatternFill("solid", fgColor=LIGHT)
        val.border = border
        if fmt:
            val.number_format = fmt
        val.alignment = Alignment(horizontal="right")
        return val

    ws.cell(row=lr - 1, column=1, value="Leverage Ratio").font = Font(bold=True, size=12, color=NAVY)
    kv(lr, "Grant Funding (Σ Grant)", f"=D{tr}", EUR0)
    kv(lr + 1, "Total Private Investment (Σ Co-investment)", f"=E{tr}", EUR0)
    kv(lr + 2, "Leverage Ratio  =  Private ÷ Grant", f"=E{tr}/D{tr}", "0.00", bold=True, big=True)
    note = ws.cell(row=lr + 4, column=1,
                   value="Interpretation: for every €1.00 of grant funding, Faraday "
                         "brought in this much private co-investment.")
    note.font = Font(italic=True, size=10, color="808080")

    # --- Column widths -----------------------------------------------------
    widths = {"A": 40, "B": 18, "C": 16, "D": 14, "E": 16, "F": 11}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.row_dimensions[hdr].height = 30

    wb.save(OUT)
    print(f"wrote {OUT}")


if __name__ == "__main__":
    build()
