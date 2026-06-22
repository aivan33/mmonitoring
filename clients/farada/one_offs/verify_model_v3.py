"""Verify farada_model_v3.xlsx after the Cost-of-Sales 6-point-curve alignment.

No recalc engine is available (formulas lib lacks OFFSET; LibreOffice absent), so this
is safe-by-construction + a Python oracle: it reimplements the 6-point cost curve and
the ASP ladder, then asserts the workbook's formulas reference the intended cells and
that the curve reproduces the unit-economics gross-margin table.  New formula cells have
no Excel cache, so it also prints the expected first-non-zero-month figures to eyeball.

Run from repo root:  .venv/bin/python clients/farada/one_offs/verify_model_v3.py
"""
from __future__ import annotations

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.formula import ArrayFormula

from build_model_v3 import CURVE, CURVE_THR  # type: ignore

DST = "clients/farada/modeling/farada_model_v3.xlsx"
FIRST, LAST = 3, 62

# Hand-derived oracle (unit economics v2): unit cost + GM at the 6 points.
ASP_THR = [1, 100, 1000, 10000, 100000, 100001, 1000000, 4000000]
ASP_VAL = [125, 75, 49, 29, 19, 15, 10, 5]


def asp(vol: float) -> float:
    out = ASP_VAL[0]
    for t, p in zip(ASP_THR, ASP_VAL):
        if vol >= t:
            out = p
    return out


def pick(vol: float, vals: list[float]) -> float:
    out = vals[0]
    for i in range(1, 6):
        if vol >= CURVE_THR[i]:
            out = vals[i]
    return out


def unit_cost(vol: float) -> float:
    return sum(pick(vol, vals) for vals in CURVE.values())


def main() -> None:
    wb = openpyxl.load_workbook(DST)
    ws, inp = wb["ProForma"], wb[" Inputs"]
    fails: list[str] = []

    def ck(cond: bool, msg: str) -> None:
        (print(f"  ✅ {msg}") if cond else (fails.append(msg) or print(f"  ❌ {msg}")))

    def ftext(cell) -> str:
        v = cell.value
        return (v.text if isinstance(v, ArrayFormula) else v) or ""

    print("\n[1] ASP ladder rung")
    ck(inp["F22"].value == 1000000 and inp["L22"].value == 10, "rung @1M repriced €5→€10 (F22=1M, L22=10)")
    ck(inp["F23"].value == 4000000 and inp["L23"].value == 5, "new rung @4M added (F23=4M, L23=5)")
    ck(inp["J23"].value == "=OFFSET(K23,0,$D$2)", "J23 uses the scenario OFFSET pattern")

    print("\n[2] revenue array formulas extended to row 23")
    for r in (6, 7, 8, 10, 11, 12):
        t = ftext(ws.cell(r, FIRST))
        ck("$J$16:$J$23" in t and "$F$16:$F$23" in t and "$J$16:$J$22" not in t,
           f"row {r} INDEX/SUMPRODUCT range -> :23")

    print("\n[3] 6-point cost block appended (values match the economics file)")
    # locate the block by its component sub-headers
    block: dict[str, list[int]] = {}
    cur = None
    for r in range(49, inp.max_row + 1):
        c = inp.cell(r, 3).value
        if c in CURVE:
            cur = c
            block[c] = []
        elif cur and isinstance(inp.cell(r, 6).value, (int, float)) and inp.cell(r, 12).value is not None:
            block[cur].append(r)
    for name, vals in CURVE.items():
        rows6 = block.get(name, [])
        ok = len(rows6) == 6
        if ok:
             for i, rr in enumerate(rows6):
                ok &= inp.cell(rr, 6).value == CURVE_THR[i]
                ok &= abs(inp.cell(rr, 12).value - round(vals[i], 4)) < 1e-9
        ck(ok, f"{name}: 6 rows, thresholds + €/sensor match")

    print("\n[4] superseded inputs blanked + nothing references them")
    for br in (34, 35, 36, 37, 38, 39, 40, 67, 68, 69, 70):
        ck(inp.cell(br, 12).value is None and inp.cell(br, 10).value is None,
           f"Inputs row {br} cleared")
    dead = {f"$J${n}" for n in (33, 34, 35, 36, 37, 38, 39, 40, 67, 68, 69, 70)}
    hits = []
    for row in ws.iter_rows():
        for cell in row:
            t = ftext(cell)
            if isinstance(t, str):
                for d in dead:
                    if d in t:
                        hits.append((cell.coordinate, d))
    ck(not hits, f"no ProForma refs to removed inputs J33-40/J67-70 (found {hits[:4]})")

    print("\n[5] cost drivers rewired to the 6-point curve (run-rate-keyed)")
    for name, drow in {"Chip": 69, "Packaging": 70, "Sensor testing": 71,
                       "Final testing": 72, "ASIC / readout": 73}.items():
        rows6 = block.get(name, [])
        t = ftext(ws.cell(drow, FIRST))
        ok = t.startswith("=IF(C67>=") and all(f"$J${rr}" in t for rr in rows6) and "$J$47" not in t
        ck(ok, f"driver row {drow} ({name}) = 6-pt IF on run-rate, no old scale-thr J47")

    print("\n[6] COGS rewired (Testing=sensor+final, ASIC=unified row 73)")
    ck(ftext(ws.cell(28, FIRST)) == "=C64*(C71+C72)", "L1 Testing = sensorsL1*(sensor+final)")
    ck(ftext(ws.cell(33, FIRST)) == "=C65*(C71+C72)", "L2 Testing = sensorsL2*(sensor+final)")
    ck(ftext(ws.cell(39, FIRST)) == "=C66*(C71+C72)", "L3 Testing = sensorsL3*(sensor+final)")
    for cog, sens in ((29, 64), (34, 65), (40, 66)):
        ck(ftext(ws.cell(cog, FIRST)) == f"=C{sens}*C73", f"ASIC row {cog} = sensors*unified-ASIC(73)")

    print("\n[7] L3 hardware revenue uses the component unit cost (not removed all-in)")
    t = ftext(ws.cell(16, FIRST))
    ck("(C69+C70+C71+C72+C73)" in t and "$J$69" not in t and "$J$70" not in t,
       "L3 HW device cost = Σ driver cells")

    print("\n[8] structural: no #REF!/empty in authored rows; no NEW #REF! introduced")
    # rows this builder authors/rewrites (cost of sales + L3 + drivers + GM-by-tier).
    touched = ({16, 17, 18, 28, 29, 33, 34, 39, 40, 69, 70, 71, 72, 73}
               | set(range(14, 23)) | set(range(47, 56)) | set(range(85, 94)))
    bad = []
    for r in touched:
        for c in range(FIRST, LAST + 1):
            t = ftext(ws.cell(r, c))
            if isinstance(t, str) and ("#REF!" in t or t == "="):
                bad.append(ws.cell(r, c).coordinate)
    ck(not bad, f"authored rows free of #REF!/empty (found {bad[:5]})")
    # any #REF! anywhere must be the PRE-EXISTING capacity-row junk (5y row 66 -> v3 78).
    ref_rows = {cell.row for row in ws.iter_rows() for cell in row
                if isinstance(ftext(cell), str) and "#REF!" in ftext(cell)}
    ck(ref_rows <= {78}, f"no new #REF! introduced; pre-existing only in capacity row 78 (got {sorted(ref_rows)})")

    print("\n[9] oracle — unit cost & gross margin curve (must match the economics file)")
    EXPECT = {1: 3.929, 10000: 3.516, 100000: 2.788, 1000000: 2.417, 4000000: 1.056}
    print(f"     {'vol':>10} {'unit€':>7} {'ASP€':>5} {'GM%':>6}")
    for v in (1, 4000, 10000, 100000, 1000000, 4000000):
        uc = unit_cost(v)
        a = asp(v)
        gm = (a - uc) / a * 100
        flag = ""
        if v in EXPECT:
            ck(abs(uc - EXPECT[v]) < 0.01, f"unit cost @{v:,} = €{uc:.3f}")
            flag = "  (checked)"
        print(f"     {v:>10,} {uc:>7.3f} {a:>5} {gm:>6.1f}{flag}")

    print("\n[10] eyeball — Realistic-scenario run-rate (5y cached) → expected cost/GM")
    cached = openpyxl.load_workbook("clients/farada/modeling/farada_model_5y.xlsx", data_only=True)["ProForma"]
    for col in range(FIRST, LAST + 1):
        rr = cached.cell(55, col).value  # 5y total run-rate row
        if isinstance(rr, (int, float)) and rr > 0:
            uc = unit_cost(rr)
            dt = cached.cell(2, col).value
            mon = dt.strftime("%b-%Y") if hasattr(dt, "strftime") else dt
            print(f"     first run-rate month {mon}: run-rate≈{rr:,.0f}/yr → unit cost €{uc:.3f}")
            break

    print()
    if fails:
        print(f"FAILED {len(fails)} check(s):")
        for f in fails:
            print(f"  - {f}")
        raise SystemExit(1)
    print("ALL CHECKS PASSED")


if __name__ == "__main__":
    main()
