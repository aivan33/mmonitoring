"""Verify farada_model_v6.5.xlsx — CF re-architecture R1–R3.

R1 ProForma stripped of profitability subtotals; R2 IS computes GP/EBITDA/EBIT/PBT/tax/NP from
its own leaves; R3 ProForma WC + financing rolls (tax-payable & RE reference the IS). No recalc
engine → structural + logic checks.

Run:  .venv/bin/python clients/farada/one_offs/verify_model_v6_5.py
"""
from __future__ import annotations

import openpyxl
from openpyxl.worksheet.formula import ArrayFormula

P = "clients/farada/modeling/farada_model_v6.5.xlsx"


def ft(cell):
    v = cell.value
    return (v.text if isinstance(v, ArrayFormula) else v)


def labels(ws):
    out = {}
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if isinstance(v, str) and v.strip():
            out.setdefault(v.strip(), r)
    return out


def bundle_rows(inp, subheader_prefix, n=3):
    """The n input rows (col-J OFFSET) immediately under an Inputs sub-header (col-C startswith)."""
    hdr = next((r for r in range(1, inp.max_row + 1)
                if isinstance(inp.cell(r, 3).value, str)
                and inp.cell(r, 3).value.strip().startswith(subheader_prefix)), None)
    rows = []
    r = (hdr or 0) + 1
    while hdr and len(rows) < n and r <= inp.max_row:
        if isinstance(inp.cell(r, 10).value, str) and "OFFSET" in inp.cell(r, 10).value:
            rows.append(r)
        r += 1
    return rows


def _balance_oracle():
    """Replicate the CF/BS formula logic on synthetic monthly leaves and assert the BS balances
    (Assets = Equity + Liabilities) every month — independent of the sheet's computed P&L."""
    rate, DSO, DPO, PAYD, PREPAY, SAASANN, LIFE_M = 0.25, 30, 30, 30, 0.5, 0.0, 60
    OB_CASH, OB_AR, OB_AP, OB_PAY, OB_DEF, OB_DEBT, OB_SC, OPEN_PPE = 2e6, 0, 0, 0, 0, 0, 5e6, 0
    H = [100000, 120000, 0, 200000]        # hardware+components revenue
    S = [10000, 12000, 15000, 18000]       # SaaS revenue
    COGS = [40000, 48000, 5000, 60000]
    OPEX = [50000, 50000, 50000, 50000]
    PAY = [30000, 30000, 30000, 30000]     # payroll (⊂ OPEX)
    FIN = [1000, 1000, 1000, 1000]
    GRANT = [0, 0, 50000, 0]
    CAPEX = [5000, 5000, 5000, 5000]
    plug_re = OB_CASH + OB_AR + OPEN_PPE - OB_AP - OB_PAY - OB_DEF - OB_DEBT - OB_SC

    ppe_prev, cash_prev = OPEN_PPE, OB_CASH
    ar_prev, ap_prev, pay_prev, def_prev, taxp_prev, sc_prev, debt_prev, re_prev = (
        OB_AR, OB_AP, OB_PAY, OB_DEF, 0.0, OB_SC, OB_DEBT, plug_re)
    for m in range(4):
        rev = H[m] + S[m]
        dep = min((ppe_prev + CAPEX[m]) / LIFE_M, ppe_prev + CAPEX[m])   # P&L D&A == schedule dep
        ppe = ppe_prev + CAPEX[m] - dep
        ebitda = rev - COGS[m] - OPEX[m]
        ebit = ebitda + GRANT[m] - dep
        pbt = ebit - FIN[m]                          # finance income 0
        tax = -max(0.0, pbt) * rate
        np = pbt + tax
        # rolls (closing balances)
        ar = (H[m] * (1 - PREPAY) + S[m] * (1 - SAASANN)) / 30 * DSO
        ap = (COGS[m] + (OPEX[m] - PAY[m])) / 30 * DPO
        pay = PAY[m] / 30 * PAYD
        deferred = S[m] * SAASANN * 6
        taxp = 0.0                                   # lag 0
        sc, debt = OB_SC, OB_DEBT                    # no tranches in test window
        re = re_prev + np
        # CF (direct)
        op = (rev - (ar - ar_prev)) + (-(COGS[m] + OPEX[m] - PAY[m]) + (ap - ap_prev)) \
            + (-PAY[m] + (pay - pay_prev)) + (tax + (taxp - taxp_prev)) + (-FIN[m]) + (deferred - def_prev)
        inv = -CAPEX[m]
        fin = (sc - sc_prev) + (debt - debt_prev) + GRANT[m]
        excess = op + inv + fin
        cash = cash_prev + excess
        # BS
        assets = 0 + ppe + cash + ar
        el = sc + re + debt + ap + pay + deferred + taxp
        if abs(assets - el) > 1e-6:
            print(f"     month {m}: check = {assets - el:.4f}  (assets {assets:.0f} vs E&L {el:.0f})")
            return False
        ppe_prev, cash_prev, ar_prev, ap_prev, pay_prev, def_prev, taxp_prev, sc_prev, debt_prev, re_prev = (
            ppe, cash, ar, ap, pay, deferred, taxp, sc, debt, re)
    return True


def main():
    wb = openpyxl.load_workbook(P)
    ws, inp, iss = wb["ProForma"], wb[" Inputs"], wb["IS"]
    L = labels(iss)
    fails = []

    def IJ(label):  # Inputs row by col-C label (startswith) — robust to the I–V reflow
        for r in range(1, inp.max_row + 1):
            v = inp.cell(r, 3).value
            if isinstance(v, str) and v.strip().startswith(label):
                return r
        raise KeyError(label)

    def JREF(label):
        return f"' Inputs'!$J${IJ(label)}"

    Lp = labels(ws)                      # ProForma label → row (robust to the ProForma reflow)
    PC = lambda label: f"C{Lp[label]}"   # 'C<row>' for an internal ProForma ref

    def ck(cond, msg):
        print(f"  {'✅' if cond else '❌'} {msg}")
        if not cond:
            fails.append(msg)

    print("\n[R1] ProForma carries NO profitability subtotals/margins (they live on the IS)")
    pf_labels = {str(ws.cell(r, 1).value).strip() for r in range(1, ws.max_row + 1)
                 if isinstance(ws.cell(r, 1).value, str)}
    for banned in ("GROSS PROFIT", "EBITDA", "EBIT", "Net profit"):
        ck(not any(banned in lbl for lbl in pf_labels), f"no '{banned}' row on the ProForma")

    print("\n[R2] IS computes subtotals from its own leaves (not ProForma pulls)")
    ck(ft(iss.cell(L["Gross profit TOTAL (€)"], 3)) == f"=C{L['Revenue']}-C{L['COGS TOTAL']}", "GP = Rev − COGS")
    ck(ft(iss.cell(L["EBITDA"], 3)) == f"=C{L['Gross profit TOTAL (€)']}-C{L['Operating expenses']}", "EBITDA = GP − OPEX")
    ck(ft(iss.cell(L["EBIT"], 3)) == f"=C{L['EBITDA']}+C{L['Grant financing']}-C{L['Depreciation & amortisation']}", "EBIT = EBITDA + grant − D&A")
    ck(ft(iss.cell(L["Income tax (expense)"], 3)) == f"=-MAX(0,C{L['Profit / (loss) before income tax']})*{JREF('Corporate tax rate')}", "Tax = −MAX(0,PBT)×rate (on IS)")
    ck(ft(iss.cell(L["Net profit / (loss) for the period"], 3)) == f"=C{L['Profit / (loss) before income tax']}+C{L['Income tax (expense)']}", "NP = PBT + tax")
    # per-bundle GP dropped
    hwgp = L["Hardware GP (€)"]
    ck(all(iss.cell(hwgp + i, 3).value is None for i in (1, 2, 3)), "per-bundle GP rows dropped")

    print("\n[F1] SaaS COGS plugged to target GM (placeholder, calibrate later)")
    from openpyxl.utils import get_column_letter
    saas_gm = IJ("SaaS gross margin target")
    saas_cogs, saas_rev = Lp["Usage (cloud / compute)"], Lp["SaaS (overage, recurring)"]
    ck(all(ft(ws.cell(saas_cogs, c)) == f"={get_column_letter(c)}{saas_rev}*(1-' Inputs'!$J${saas_gm})"
           for c in range(3, 63)), "SaaS COGS = SaaS rev × (1−SaaS-GM-target), all cols")
    ck(isinstance(inp.cell(saas_gm, 15).value, str) and "placeholder" in inp.cell(saas_gm, 15).value.lower(),
       "SaaS-GM-target carries a placeholder note in col O")

    print("\n[F2] yield explicit — chip derived in ProForma from wafer ÷ spw ÷ yield")
    ck(any(isinstance(inp.cell(r, 3).value, str) and "wafer cost" in inp.cell(r, 3).value.lower()
           for r in range(1, inp.max_row + 1)), "Inputs has a Wafer cost (€/wafer) block")
    ck("Yield (staged by run-rate)" in Lp, "ProForma has a Yield calc row (staged)")
    spw, yld = Lp["Sensors per wafer"], Lp["Yield (staged by run-rate)"]
    ck(ft(ws.cell(Lp["Chip EUR/sensor"], 3)).endswith(f"/(C{spw}*C{yld})"),
       "Chip €/sensor = wafer cascade ÷ (spw × yield)")

    print("\n[A] ProForma carries the skill-outline lower sections")
    pf_secs = {str(ws.cell(r, 1).value).strip() for r in range(1, ws.max_row + 1)
               if isinstance(ws.cell(r, 1).value, str)}
    for sec in ("BALANCE SHEET", "WC DRIVERS & RATIOS", "CASH FLOW", "TAXATION", "FUNDING"):
        ck(any(sec in s for s in pf_secs), f"ProForma section: {sec}")

    print("\n[D2] drivers ordered: Sensors → Run-rate → CoS/sensor → Measurements")
    ck(Lp["Total run-rate (sensors/yr)"] < Lp["Chip EUR/sensor"] < Lp["Measurements Line 3 (monthly)"],
       "run-rate < CoS/sensor (chip) < measurements")
    ck(Lp["Sensors per wafer"] < Lp["Chip EUR/sensor"] and Lp["Yield (staged by run-rate)"] < Lp["Chip EUR/sensor"],
       "sensors-per-wafer + yield sit above chip (its inputs)")

    print("\n[D1] run-rate = LTM trailing-12 (per-month, not a frozen constant)")
    rr = Lp["Total run-rate (sensors/yr)"]
    s1, s3 = Lp["Sensors Line 1 (monthly)"], Lp["Sensors Line 3 (monthly)"]
    ck(ft(ws.cell(rr, 3)) == f"=SUM(C{s1}:C{s3})", "run-rate month-1 = current-month sensors only")
    ck(ft(ws.cell(rr, 15)).startswith("=SUM(") and ft(ws.cell(rr, 15)) != ft(ws.cell(rr, 3)),
       "run-rate varies by month (LTM window grows; not the frozen =SUM(C5:N7))")

    print("\n[E] ProForma sum/subtotal lines are bold (readability)")
    ck(ws.cell(Lp["Total run-rate (sensors/yr)"], 1).font.bold, "'Total run-rate' (a SUM) is bold")
    ck(ws.cell(Lp["Revenue"], 1).font.bold, "'Revenue' (a subtotal) is bold")
    ck(not ws.cell(Lp["Sensors Line 1 (monthly)"], 1).font.bold, "a leaf driver is NOT bold")

    print("\n[F-meas] Line-3 measurements split into Included + Overage children")
    mr = Lp["Measurements Line 3 (monthly)"]
    kids = [str(ws.cell(mr + i, 1).value or "") for i in (1, 2)]
    ck(any("Included" in k for k in kids), "measurements has an Included (subscription) child")
    ck(any("Overage" in k for k in kids), "measurements has an Overage (beyond subscription) child")
    ck(ft(ws.cell(mr, 3)) == f"=C{mr + 1}+C{mr + 2}", "measurements total = Included + Overage")

    print("\n[D4] measurements children accumulate off their OWN prior column (installed base), not the total")
    def accum_own(child):  # month-2+ cols must self-accumulate: '={prevcol}{child}+...' (cumulative installed base)
        return all(isinstance(ft(ws.cell(child, c)), str)
                   and ft(ws.cell(child, c)).startswith(f"={get_column_letter(c - 1)}{child}+")
                   for c in range(4, 63))
    ck(accum_own(mr + 1), "Included child accumulates off its own prior column (not the total row)")
    ck(accum_own(mr + 2), "Overage child accumulates off its own prior column (not the total row)")

    print("\n[R3] ProForma rolls; tax-payable & RE reference the IS")
    ck(ft(ws.cell(Lp["Trade receivables (AR)"], 3)).startswith(
        f"=(({PC('Components #1 - Low Volume')}+{PC('Components #2 - High Volume')}+"
        f"{PC('Hardware (device, cost + markup)')})*(1-"), "AR roll present")
    ck(ft(ws.cell(Lp["Trade payables (total)"], 3)) ==
       f"={PC('Trade payables — COGS')}+{PC('Trade payables — S&M')}+"
       f"{PC('Trade payables — G&A')}+{PC('Trade payables — R&D')}", "AP total = Σ buckets")
    tp = Lp["Tax payable"]
    ck(ft(ws.cell(tp, 3)) == f"=-IS!C{L['Income tax (expense)']}*IF({JREF('Tax payment lag')}>=1,1,0)", "Tax-payable roll → IS tax")
    re = Lp["Retained earnings"]
    np = L["Net profit / (loss) for the period"]
    ck(ft(ws.cell(re, 3)).startswith(f"=({JREF('Opening cash')}+{JREF('Opening AR')}+{JREF('Opening PP&E (NBV)')}")
       and ft(ws.cell(re, 3)).endswith(f"+IS!C{np}"), "RE roll t0 = balancing plug + IS net profit")
    ck(ft(ws.cell(re, 4)) == f"=C{re}+IS!D{np}", "RE roll month 2 = prior + IS NP")

    print("\n[inputs] sections I–V + funding/WC/opening groups present (post-reflow)")
    secs = {inp.cell(r, 3).value.strip() for r in range(1, inp.max_row + 1)
            if isinstance(inp.cell(r, 1).value, str) and "." in str(inp.cell(r, 1).value)
            and isinstance(inp.cell(r, 3).value, str)}
    for sec in ("FUNDING ASSUMPTIONS", "REVENUE ASSUMPTIONS", "PRODUCTION",
                "OPERATING EXPENSES", "OTHER ASSUMPTIONS"):
        ck(any(sec in s for s in secs), f"section: {sec}")
    all_lbls = [inp.cell(r, 3).value for r in range(1, inp.max_row + 1) if isinstance(inp.cell(r, 3).value, str)]
    for kw in ("Receivable days (DSO)", "Payable days (DPO)", "Equity round amount",
               "Opening cash", "Opening retained earnings"):
        ck(any(k.strip().startswith(kw) for k in all_lbls), f"input: {kw}")

    print("\n[D5a] plan tier discounts + plan-heavy included + WC cash-timing settings")
    disc = bundle_rows(inp, "Line 3 — plan tier discount")
    ck(len(disc) == 3, "3 plan tier discount inputs present")
    ck([inp.cell(r, 12).value for r in disc] == [0.10, 0.15, 0.20], "discounts = 10/15/20% (S/M/L)")
    ck(bool(disc) and IJ("Line 3 — overage price") < disc[0], "discounts sorted after overage price (not appended)")
    incl = bundle_rows(inp, "Line 3 — included measurements")
    ck(bool(incl) and all((inp.cell(r, 12).value or 0) >= 800 for r in incl),
       "included re-set plan-heavy (≥800 of avg 1200)")
    ck(abs(inp.cell(IJ("Hardware prepayment"), 12).value or 0) < 1e-9, "PREPAY = 0 (30d net, no prepay)")
    ck(abs((inp.cell(IJ("SaaS billed annually"), 12).value or 0) - 1.0) < 1e-9,
       "SAAS_ANN = 100% (subscription billed annually upfront)")

    print("\n[fmt] input value cells formatted per unit (no dates/%-as-plain-numbers)")
    dfmt = inp.cell(IJ("Tranche 1 date"), 12).number_format.lower()
    ck("yy" in dfmt or "mmm" in dfmt, f"date input 'Tranche 1 date' has a date format (got {dfmt!r})")
    ck("%" in inp.cell(IJ("Annual interest rate"), 12).number_format,
       "rate input 'Annual interest rate' has a % format")
    ck("€" in inp.cell(IJ("Tranche 1 amount"), 12).number_format,
       "EUR input 'Tranche 1 amount' has a € format")

    print("\n[R4] CF statement (direct) present + wired to rolls/IS")
    cf = wb["CF"]
    Lc = labels(cf)
    ck(ft(cf.cell(Lc["Cash received from customers"], 3)) ==
       f"=ProForma!{PC('Revenue')}-(ProForma!{PC('Trade receivables (AR)')}-{JREF('Opening AR')})",
       "Cash from customers = rev − ΔAR")
    eb, bg, ex = Lc["Ending Cash Balance"], Lc["Beginning Cash Balance"], Lc["Excess Cash for the Period"]
    ck(ft(cf.cell(eb, 3)) == f"=C{bg}+C{ex}", "Ending cash = beginning + excess")
    for line in ("Recovery/(repayment) of VAT", "Distribution of dividends", "Net Cash Burn"):
        ck(any(isinstance(cf.cell(r, 1).value, str) and line in cf.cell(r, 1).value
               for r in range(1, cf.max_row + 1)), f"CF has '{line}'")
    ck(ft(cf.cell(Lc["Beginning Cash Balance"], 3)) == f"={JREF('Opening cash')}", "Beginning cash (t0) = opening cash input")
    ck(ft(cf.cell(Lc["Beginning Cash Balance"], 4)) == f"=C{eb}", "Beginning cash (t1) = prior ending")

    print("\n[R5] BS present; cash=CF ending; check row = Assets − E&L")
    bs = wb["BS"]
    Lb = labels(bs)
    ck(ft(bs.cell(Lb["Cash & cash equivalents"], 3)) == f"=CF!C{Lc['Ending Cash Balance']}", "BS cash = CF ending")
    ck(ft(bs.cell(Lb["check (Assets − E&L)"], 3)) == f"=C{Lb['TOTAL ASSETS']}-C{Lb['TOTAL EQUITY & LIABILITIES']}",
       "check = TOTAL ASSETS − TOTAL E&L")
    # reference structure present (blank-but-defined)
    for line in ("Inventory", "Business equipment", "Grants", "Other payables", "Current ratio"):
        ck(any(isinstance(bs.cell(r, 1).value, str) and line in bs.cell(r, 1).value
               for r in range(1, bs.max_row + 1)), f"BS has '{line}'")

    print("\n[R6] yearly CF_Y / BS_Y (calendar 2026–2030)")
    for s in ("CF_Y", "BS_Y"):
        ck(s in wb.sheetnames, f"sheet {s} present")
    cfy, bsy = wb["CF_Y"], wb["BS_Y"]
    Lcy, Lby = labels(cfy), labels(bsy)
    ck(cfy.cell(2, 3).value == "2026" and cfy.cell(2, 7).value == "2030", "CF_Y years 2026..2030")
    op = Lcy["Cash Flow from Operating Activities"]
    ck(ft(cfy.cell(op, 3)) == f"=SUM(CF!C{op}:H{op})", "CF_Y 2026 operating = SUM(CF! Jul–Dec)")
    ck(ft(cfy.cell(op, 4)) == f"=SUM(CF!I{op}:T{op})", "CF_Y 2027 operating = SUM(CF! Jan–Dec)")
    chk = Lby["check (Assets − E&L)"]
    ck(ft(bsy.cell(chk, 3)) == f"=BS!H{chk}", "BS_Y 2026 = BS Dec-2026 snapshot")
    ck(ft(bsy.cell(chk, 4)) == f"=BS!T{chk}", "BS_Y 2027 = BS Dec-2027 snapshot")

    print("\n[oracle] BS balances by construction (synthetic flows → check = 0 every month)")
    ck(_balance_oracle(), "synthetic 3-statement run: BS check = 0 for all test months")

    print("\n[overhaul] dead Line-3 usage-pricing ladder gone (dropped in the reflow)")
    ck(not any(isinstance(inp.cell(r, 3).value, str) and "meas." in inp.cell(r, 3).value
               for r in range(1, inp.max_row + 1)), "no 'Price @ N meas.' usage-ladder labels remain")

    print("\n[struct] no naked rows; no new #REF!")
    naked = [f"C{r}" for r in (Lp["Trade receivables (AR)"], re, L["EBITDA"])
             if ws.cell(r, 3).value is not None and ws.cell(r, 3).number_format == "General"
             or (iss.cell(L["EBITDA"], 3).number_format == "General")]
    ck(iss.cell(L["EBITDA"], 3).number_format != "General", "IS EBITDA styled (not General)")
    refrows = {(sn, c.row) for sn in wb.sheetnames for row in wb[sn].iter_rows() for c in row
               if isinstance(ft(c), str) and "#REF!" in ft(c)}
    ck(not refrows, f"no #REF! anywhere (capacity row repaired) (got {sorted(refrows)})")
    # (the old 'no ref to a stripped ProForma row' check is obsolete — the reflow DROPS those rows
    #  and reflow_proforma's gate already asserts no formula references a dropped row.)

    print()
    if fails:
        print(f"FAILED {len(fails)} check(s):")
        for f in fails:
            print(f"  - {f}")
        raise SystemExit(1)
    print("ALL CHECKS PASSED")


if __name__ == "__main__":
    main()
