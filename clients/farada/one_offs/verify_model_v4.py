"""Verify farada_model_v4.xlsx — ProForma (calc engine) split from a real Income Statement.

v4 adds a clean contiguous monthly `IS` (pulls values from ProForma, computes margins) and a
calendar-year `IS_Y` (2026 Jul–Dec partial … 2030), and moves the % margins OUT of the ProForma.
No recalc engine → structural + oracle checks.

Run from repo root:  .venv/bin/python clients/farada/one_offs/verify_model_v4.py
"""
from __future__ import annotations

import openpyxl
from openpyxl.worksheet.formula import ArrayFormula

P = "clients/farada/modeling/farada_model_v4.xlsx"


def ft(cell):
    v = cell.value
    return (v.text if isinstance(v, ArrayFormula) else v)


def rows_by_label(ws):
    out = {}
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if isinstance(v, str) and v.strip():
            out.setdefault(v.strip(), r)
    return out


def main() -> None:
    wb = openpyxl.load_workbook(P)
    fails: list[str] = []

    def ck(cond, msg):
        print(f"  {'✅' if cond else '❌'} {msg}")
        if not cond:
            fails.append(msg)

    print("\n[v4] sheets")
    for s in (" Inputs", "Revenue_Inputs", "ProForma", "HR", "IS", "IS_Y"):
        ck(s in wb.sheetnames, f"sheet {s!r} present")

    ws, isheet, y = wb["ProForma"], wb["IS"], wb["IS_Y"]
    L = rows_by_label(isheet)

    print("\n[v4] IS is a clean statement that PULLS from ProForma + computes margins")
    ck(ft(isheet.cell(L["Revenue"], 3)) == "=ProForma!C4", "IS Revenue = =ProForma!C4 (pull)")
    ck(ft(isheet.cell(L["EBITDA"], 3)) == "=ProForma!C116", "IS EBITDA = =ProForma!C116 (pull)")
    gm = L["Gross margin TOTAL"]
    ck(ft(isheet.cell(gm, 3)) == f"=IF(C{L['Revenue']}=0,0,C{L['Gross profit TOTAL (€)']}/C{L['Revenue']})",
       "IS Gross margin computed from IS's own GP/Revenue")
    em = L["EBITDA margin"]
    ck(ft(isheet.cell(em, 3)) == f"=IF(C{L['Revenue']}=0,0,C{L['EBITDA']}/C{L['Revenue']})",
       "IS EBITDA margin computed on IS")

    print("\n[v4] IS includes EBITDA + ALL below-EBITDA (no cut after GM)")
    for lbl in ("EBITDA", "EBIT", "Net profit / (loss) for the period", "Profit margin",
                "Depreciation & amortisation"):
        ck(lbl in L, f"IS has '{lbl}'")
    # contiguous: gap between last GM row and OPERATING EXPENSES must be small (was 23)
    gap = L["OPERATING EXPENSES"] - L["SaaS GM"]
    ck(gap <= 3, f"IS is contiguous (GM→OPEX gap = {gap} rows, not ~23)")

    print("\n[v4] IS_Y = calendar years (2026 Jul–Dec … 2030), full statement, margins recomputed")
    Ly = rows_by_label(y)
    ck(y["C2"].value == "2026" and y["G2"].value == "2030", "year header 2026..2030")
    ck("2031" not in [y.cell(2, c).value for c in range(3, 9)], "2031 H1 dropped from yearly")
    rev_y = Ly["Revenue"]
    ck(ft(y.cell(rev_y, 3)) == f"=SUM(IS!C{rev_y}:H{rev_y})", "2026 = SUM(IS! Jul–Dec) (6 mo partial)")
    ck(ft(y.cell(rev_y, 4)) == f"=SUM(IS!I{rev_y}:T{rev_y})", "2027 = SUM(IS! Jan–Dec)")
    for lbl in ("EBITDA", "EBIT", "Net profit / (loss) for the period"):
        ck(lbl in Ly, f"IS_Y has '{lbl}' (below-EBITDA present)")
    gmy = Ly["Gross margin TOTAL"]
    ck(ft(y.cell(gmy, 3)) == f"=IF(C{Ly['Revenue']}=0,0,C{Ly['Gross profit TOTAL (€)']}/C{Ly['Revenue']})",
       "IS_Y margin recomputed (not summed)")

    print("\n[v4] margins/KPIs moved OUT of the ProForma")
    for r in (56, 117, 133):
        ck(ws.cell(r, 3).value is None and ws.cell(r, 1).value is None, f"ProForma margin row {r} blanked")

    print("\n[v4] no naked rows; no #REF!")
    naked = [f"{co}{r}" for sh in (isheet, y) for r in (rows_by_label(sh).get("Revenue"),
             rows_by_label(sh).get("EBITDA")) if r for co in "CD"
             if sh[f"{co}{r}"].value is not None and sh[f"{co}{r}"].number_format == "General"]
    ck(not naked, f"IS/IS_Y data cells styled (found {naked[:5]})")
    refrows = {(sn, c.row) for sn in wb.sheetnames for row in wb[sn].iter_rows() for c in row
               if isinstance(ft(c), str) and "#REF!" in ft(c)}
    ck(refrows <= {("ProForma", 78)}, f"no new #REF! (got {sorted(refrows)})")

    print()
    if fails:
        print(f"FAILED {len(fails)} check(s):")
        for f in fails:
            print(f"  - {f}")
        raise SystemExit(1)
    print("ALL CHECKS PASSED")


if __name__ == "__main__":
    main()
