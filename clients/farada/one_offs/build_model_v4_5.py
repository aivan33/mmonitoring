"""Build farada_model_v4.5 from v4 — cash-flow groundwork + ProForma/IS re-architecture.

Per user review: (1) the ProForma must be a pure calc engine — every profitability SUBTOTAL and
MARGIN (Gross profit, EBITDA, EBIT, PBT, tax, Net profit, all margins) lives only on the Income
Statement, which now COMPUTES them; (2) build the cash flow WITH the working-capital engine
(direct method, Almacena-format), not orphaned rolls.

This file does R1–R3 (re-architecture + rolls). R4–R6 (CF/BS statements) follow.
  R1  strip ProForma profitability subtotals (rows 44-55, 116, 125, 130, 131, 132).
  R2  IS computes GP / EBITDA / EBIT / PBT / tax / NP from its own leaves (was =ProForma! pulls);
      per-bundle GP detail dropped (per-bundle COGS isn't split).
  R3  ProForma WC + financing rolls; the tax-payable & retained-earnings rolls reference IS.

Reads v4 (preserved), writes the DRAFT farada_model_v4.5.xlsx (final build → v5). Idempotent.
Run:  .venv/bin/python clients/farada/one_offs/build_model_v4_5.py
"""
from __future__ import annotations

from copy import copy
from datetime import datetime

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.formula import ArrayFormula

SRC = "clients/farada/modeling/farada_model_v4.xlsx"
DST = "clients/farada/modeling/farada_model_v4.5.xlsx"
FIRST, LAST = 3, 62

I = dict(WC_HDR=161, DSO=162, PREPAY=163, SAAS_ANN=164, DPO=165, PAYDAYS=166, TAXLAG=167,
         FIN_HDR=169, EQ_AMT=170, EQ_DATE=171, DEBT_AMT=172, DEBT_DATE=173, DEBT_INT=174,
         OB_CASH=175, OB_AR=176, OB_AP=177, OB_PAYROLL=178, OB_DEFERRED=179, OB_DEBT=180,
         OB_SC=181, OB_RE=182)
R = dict(HDR=143, AR=144, AP_COGS=145, AP_SM=146, AP_GA=147, AP_RD=148, AP_TOT=149,
         PAYROLL=150, DEFERRED=151, TAXPAY=152, SC=153, DEBT=154, RE=155)
# ProForma profitability subtotals to strip (margins 56-61/117/121/122/126/133 already blank in v4).
STRIP = list(range(44, 56)) + [116, 125, 130, 131, 132]


def _ft(c):
    v = c.value
    return v.text if isinstance(v, ArrayFormula) else v


def labels(ws):
    out = {}
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if isinstance(v, str) and v.strip():
            out.setdefault(v.strip(), r)
    return out


# ---------------------------------------------------------------- R1
def strip_proforma_subtotals(wb):
    ws = wb["ProForma"]
    for r in STRIP:
        for c in range(1, LAST + 1):
            ws.cell(r, c).value = None
    print(f"  R1: blanked ProForma subtotals {STRIP[:3]}…{STRIP[-3:]} (GP/EBITDA/EBIT/PBT/tax/NP)")


# ---------------------------------------------------------------- R2
def is_compute_subtotals(wb):
    iss, isy = wb["IS"], wb["IS_Y"]
    L = labels(iss)
    # subtotal label -> (lambda col -> formula referencing IS's own rows)
    rev, cogs = L["Revenue"], L["COGS TOTAL"]
    c1, c2, l3 = L["Components #1 - Low Volume"], L["Components #2 - High Volume"], L["Hardware-enabled SaaS #3"]
    cl1, cl2, cl3 = L["COGS Line #1"], L["COGS Line #2"], L["COGS Line #3"]
    hwr, hwc = L["Hardware (device, cost + markup)"], L["Hardware (device)"]
    sar, usg = L["SaaS (overage, recurring)"], L["Usage (cloud / compute)"]
    gp, opex = L["Gross profit TOTAL (€)"], L["Operating expenses"]
    eb, grant, da = L["EBITDA"], L["Grant financing"], L["Depreciation & amortisation"]
    ebit, finn = L["EBIT"], L["Finance (costs), net"]
    pbt, tax, npr = (L["Profit / (loss) before income tax"], L["Income tax (expense)"],
                     L["Net profit / (loss) for the period"])
    COMPUTE = {
        gp:   lambda x: f"={x}{rev}-{x}{cogs}",
        L["Gross profit Line 1 (€)"]: lambda x: f"={x}{c1}-{x}{cl1}",
        L["Gross profit Line 2 (€)"]: lambda x: f"={x}{c2}-{x}{cl2}",
        L["Gross profit Line 3 (€)"]: lambda x: f"={x}{l3}-{x}{cl3}",
        L["Hardware GP (€)"]: lambda x: f"={x}{hwr}-{x}{hwc}",
        L["SaaS GP (€)"]: lambda x: f"={x}{sar}-{x}{usg}",
        eb:   lambda x: f"={x}{gp}-{x}{opex}",
        ebit: lambda x: f"={x}{eb}+{x}{grant}-{x}{da}",
        pbt:  lambda x: f"={x}{ebit}+{x}{finn}",
        tax:  lambda x: f"=-MAX(0,{x}{pbt})*' Inputs'!$J$158",
        npr:  lambda x: f"={x}{pbt}+{x}{tax}",
    }
    for r, fn in COMPUTE.items():
        for c in range(FIRST, LAST + 1):
            iss.cell(r, c, fn(get_column_letter(c)))
    # drop per-bundle GP detail (per-bundle COGS isn't split) — blank in IS and IS_Y
    drop = [L["Hardware GP (€)"] + i for i in (1, 2, 3)] + [L["SaaS GP (€)"] + i for i in (1, 2, 3)]
    for sh in (iss, isy):
        for r in drop:
            for c in range(1, LAST + 1):
                sh.cell(r, c).value = None
    print(f"  R2: IS computes {len(COMPUTE)} subtotals from its own leaves; dropped bundle-GP rows {drop}")
    return L


# ---------------------------------------------------------------- R3 (rolls)
def add_cf_inputs(wb):
    inp = wb[" Inputs"]
    s_hdr = inp.cell(152, 3)._style
    s_lbl, s_unit = inp.cell(154, 3)._style, inp.cell(154, 4)._style
    s_j, s_l = inp.cell(154, 10)._style, inp.cell(154, 12)._style

    def hdr(r, text):
        inp.cell(r, 3, text)._style = copy(s_hdr)

    def put(r, label, unit, value, numfmt):
        inp.cell(r, 3, label)._style = copy(s_lbl)
        inp.cell(r, 4, unit)._style = copy(s_unit)
        inp.cell(r, 10, f"=OFFSET(K{r},0,$D$2)")._style = copy(s_j)
        cc = inp.cell(r, 12, value)
        cc._style = copy(s_l)
        cc.number_format = numfmt

    hdr(I["WC_HDR"], "WORKING CAPITAL (payment terms)  (mockup ← confirm)")
    put(I["DSO"], "Receivable days (DSO)", "days", 30, "#,##0")
    put(I["PREPAY"], "Hardware prepayment % (large orders)", "%", 0.5, "0.0%")
    put(I["SAAS_ANN"], "SaaS billed annually in advance", "%", 0.0, "0.0%")
    put(I["DPO"], "Payable days (DPO)", "days", 30, "#,##0")
    put(I["PAYDAYS"], "Payroll payable days", "days", 30, "#,##0")
    put(I["TAXLAG"], "Tax payment lag", "months", 0, "#,##0")
    hdr(I["FIN_HDR"], "FINANCING & OPENING BALANCES (Jul-2026)  (mockup ← confirm)")
    put(I["EQ_AMT"], "Equity round amount", "EUR", 5_000_000, "€#,##0")
    put(I["EQ_DATE"], "Equity round date", "date", datetime(2027, 7, 1), "mmm-yyyy")
    put(I["DEBT_AMT"], "Debt draw amount", "EUR", 0, "€#,##0")
    put(I["DEBT_DATE"], "Debt draw date", "date", datetime(2026, 7, 1), "mmm-yyyy")
    put(I["DEBT_INT"], "Debt annual interest", "%", 0.0, "0.0%")
    put(I["OB_CASH"], "Opening cash", "EUR", 2_000_000, "€#,##0")
    put(I["OB_AR"], "Opening AR", "EUR", 0, "€#,##0")
    put(I["OB_AP"], "Opening AP", "EUR", 0, "€#,##0")
    put(I["OB_PAYROLL"], "Opening payroll payable", "EUR", 0, "€#,##0")
    put(I["OB_DEFERRED"], "Opening deferred revenue", "EUR", 0, "€#,##0")
    put(I["OB_DEBT"], "Opening debt", "EUR", 0, "€#,##0")
    put(I["OB_SC"], "Opening share capital", "EUR", 5_000_000, "€#,##0")
    put(I["OB_RE"], "Opening retained earnings", "EUR", -3_000_000, "€#,##0")


def add_rolls(wb, L_is):
    ws = wb["ProForma"]
    leaf = {c: ws.cell(89, c)._style for c in range(1, LAST + 1)}
    band = {c: ws.cell(85, c)._style for c in range(1, LAST + 1)}
    jp = lambda n: f"' Inputs'!$J${I[n]}"
    np_r, tax_r = L_is["Net profit / (loss) for the period"], L_is["Income tax (expense)"]

    LABEL = {R["HDR"]: "WORKING CAPITAL & FINANCING ROLLS (balances)",
             R["AR"]: "  Trade receivables (AR)", R["AP_COGS"]: "  Trade payables — COGS",
             R["AP_SM"]: "  Trade payables — S&M", R["AP_GA"]: "  Trade payables — G&A",
             R["AP_RD"]: "  Trade payables — R&D", R["AP_TOT"]: "  Trade payables (total)",
             R["PAYROLL"]: "  Payroll payable", R["DEFERRED"]: "  Deferred revenue (SaaS annual)",
             R["TAXPAY"]: "  Tax payable", R["SC"]: "  Share capital", R["DEBT"]: "  Debt",
             R["RE"]: "  Retained earnings"}

    def fml(rk, x, prev, first):
        if rk == "AR":
            return f"=(({x}5+{x}9+{x}15)*(1-{jp('PREPAY')})+{x}19*(1-{jp('SAAS_ANN')}))/30*{jp('DSO')}"
        if rk == "AP_COGS": return f"={x}24/30*{jp('DPO')}"
        if rk == "AP_SM":   return f"=({x}87-{x}88)/30*{jp('DPO')}"
        if rk == "AP_GA":   return f"=({x}96-{x}97)/30*{jp('DPO')}"
        if rk == "AP_RD":   return f"=({x}107-{x}108)/30*{jp('DPO')}"
        if rk == "AP_TOT":  return f"={x}{R['AP_COGS']}+{x}{R['AP_SM']}+{x}{R['AP_GA']}+{x}{R['AP_RD']}"
        if rk == "PAYROLL": return f"=({x}88+{x}97+{x}108)/30*{jp('PAYDAYS')}"
        if rk == "DEFERRED": return f"={x}19*{jp('SAAS_ANN')}*6"
        if rk == "TAXPAY":  return f"=-IS!{x}{tax_r}*IF({jp('TAXLAG')}>=1,1,0)"
        if rk == "SC":
            inj = f"IF({x}2={jp('EQ_DATE')},{jp('EQ_AMT')},0)"
            return f"={jp('OB_SC')}+{inj}" if first else f"={prev}{R['SC']}+{inj}"
        if rk == "DEBT":
            dr = f"IF({x}2={jp('DEBT_DATE')},{jp('DEBT_AMT')},0)"
            return f"={jp('OB_DEBT')}+{dr}" if first else f"={prev}{R['DEBT']}+{dr}"
        if rk == "RE":
            return f"={jp('OB_RE')}+IS!{x}{np_r}" if first else f"={prev}{R['RE']}+IS!{x}{np_r}"
        return None

    ws.cell(R["HDR"], 1, LABEL[R["HDR"]])._style = copy(band[1])
    for c in range(2, LAST + 1):
        ws.cell(R["HDR"], c)._style = copy(band[c])
    for rk in ("AR", "AP_COGS", "AP_SM", "AP_GA", "AP_RD", "AP_TOT", "PAYROLL",
               "DEFERRED", "TAXPAY", "SC", "DEBT", "RE"):
        r = R[rk]
        ws.cell(r, 1, LABEL[r])._style = copy(leaf[1])
        for c in range(FIRST, LAST + 1):
            x, prev = get_column_letter(c), get_column_letter(c - 1)
            ws.cell(r, c, fml(rk, x, prev, c == FIRST))._style = copy(leaf[c])
    print(f"  R3: rolls {R['HDR']}-{R['RE']}; tax-payable & RE reference IS (tax r{tax_r}, NP r{np_r})")


def build():
    wb = openpyxl.load_workbook(SRC)
    add_cf_inputs(wb)
    strip_proforma_subtotals(wb)
    L_is = is_compute_subtotals(wb)
    add_rolls(wb, L_is)
    wb.save(DST)
    print(f"Saved {DST}")


if __name__ == "__main__":
    build()
