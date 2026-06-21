"""Add a 'Unit Economics v2' sheet with dynamic (descaling) unit costs.

Operates additively on the user's current farada_unit_economics.xlsx (preserving
their hand-edited formatting) — copies the 'Unit Economics' sheet, then rewrites
only the Chip and Testing columns to descale log-linearly between min/max over a
volume ramp. Packaging (lot @100k) and ASIC (own-ASIC @1M) stay stepped.

Idempotent: drops an existing 'Unit Economics v2' before re-adding.
Run:  .venv/bin/python clients/farada/one_offs/build_unit_economics_v2.py
"""
from __future__ import annotations

from copy import copy

import openpyxl
from openpyxl.styles import Font

PATH = "clients/farada/modeling/farada_unit_economics.xlsx"
V1, V2 = "Unit Economics", "Unit Economics v2"
ROWS = range(17, 26)   # table data rows

# log-linear descale between (max at v_lo) and (min at v_hi), clamped flat outside.
def descale(max_expr, min_expr, vol, v_lo="$B$14", v_hi="$B$13"):
    frac = f"MIN(1,MAX(0,(LN({vol})-LN({v_lo}))/(LN({v_hi})-LN({v_lo}))))"
    return f"=({max_expr})+(({min_expr})-({max_expr}))*{frac}"


def build():
    wb = openpyxl.load_workbook(PATH)
    if V2 in wb.sheetnames:
        del wb[V2]
    src = wb[V1]
    ws = wb.copy_worksheet(src)
    ws.title = V2
    ws.sheet_view.showGridLines = False
    # copy_worksheet misses column widths / merges on some versions — port them over.
    for k, dim in src.column_dimensions.items():
        ws.column_dimensions[k].width = dim.width
    for k, dim in src.row_dimensions.items():
        ws.row_dimensions[k].height = dim.height
    for mc in list(src.merged_cells.ranges):
        if str(mc) not in {str(x) for x in ws.merged_cells.ranges}:
            ws.merge_cells(str(mc))

    # 1) new input: Descale from (units) at row 14 (blank in v1), styled like an input
    ws["A14"].value = "Descale from (units)"
    ws["A14"]._style = copy(src["A3"]._style)
    ws["B14"].value = 3000
    ws["B14"]._style = copy(src["B3"]._style)
    ws["B14"].number_format = "#,##0"
    ws["C14"].value = "ramp end = Scale threshold (B13)"
    ws["C14"].font = Font(italic=True, size=8, color="FF666666")

    # 2) rewrite Chip (col C) and Testing (col E) to descale; D/F stay stepped (as copied)
    chip_max, chip_min = "$B$3/$B$2/$B$5", "$B$4/$B$2/$B$5"   # wafer std/scale ÷ spw ÷ yield
    test_max, test_min = "$B$8", "$B$9"
    for r in ROWS:
        ws[f"C{r}"].value = descale(chip_max, chip_min, f"A{r}")
        ws[f"E{r}"].value = descale(test_max, test_min, f"A{r}")

    # 3) label the section + mark which columns are dynamic
    ws["A15"].value = "UNIT ECONOMICS v2  —  Chip & Testing descale; Packaging & ASIC stepped"
    ws["C16"].value = "Chip €/u\n(descales)"
    ws["E16"].value = "Testing €/u\n(descales)"

    wb.save(PATH)
    print(f"Saved {V2} into {PATH}")


if __name__ == "__main__":
    build()
