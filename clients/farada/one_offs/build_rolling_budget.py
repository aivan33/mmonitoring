"""
Build the Farada rolling-budget v2.1.1 workbook (xlsx).

Output: clients/farada/reference/rolling_budget_v3.xlsx

v2.1.1 architecture (per `clients/farada/rolling_budget_spec.md` §9):
  * Three sheets: Inputs, Actuals, Pro Forma.
  * Past columns on Pro Forma pull from the Actuals sheet via direct
    cell reference (`=Actuals!<col><row>`). No IF wrappers.
  * Future columns on Pro Forma hold forecast formulas. User adjusts
    via the `=formula + X` delta convention (never overtype).
  * Actuals are loaded from `clients/farada/data/farada.db`
    (`financials` table, scenario='actual') at build time. The
    actuals-through month is `MAX(period_date)` in that table.
  * Only LEAF BS / IS rows pull from actuals; aggregates remain
    formulas (they sum the pulled leaves automatically).

v2.1.1 changes vs v2.1:
  * Compact R class layout (no scattered empty rows).
  * Section order: BS → IS → CF → Indirect Tie.
  * Tie-outs emitted AFTER each section (Balance Check after BS,
    NI/RE tie after IS, CF tie after Indirect, error count at end).
  * Chart of accounts aligned to Farada DB taxonomi (Insurance,
    Direct/Indirect Amortization, Finance income split, CAPEX combined,
    Other CA broken out into 3 sub-lines, etc.).
  * Inventory formula bug fixed (purchases = inventory out).
  * Cash chain (BS_CASH ↔ OPEN_CASH) reconnected in forecast.
  * Grant income clamped to MAX(0, deferred balance) so it can't
    drive deferred-grants negative.
  * Lighter visual styling (no navy banner, gray/border accents).

Run:  uv run python clients/farada/one_offs/build_rolling_budget.py
"""

from __future__ import annotations
import re
import sqlite3
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.formula.translate import Translator
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule


# ──────────────────────────────────────────────────────────────────────────────
# Layout constants
# ──────────────────────────────────────────────────────────────────────────────

PERIOD_COLS = ["E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P"]
MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
PERIOD_DATES = [date(2026, m, 1) for m in range(1, 13)]
# Populated at build time from `_load_actuals_from_db`. Drives PERIOD_FLAGS
# and the past-vs-future formula branch in `fill_right`.
PERIOD_FLAGS: list[str] = ["Forecast"] * 12  # overwritten by main()

OPEN_COL = "D"
FY_COL = "Q"
LABEL_COL = "B"
CODE_COL = "C"

# Path to the SQLite DB holding loaded taxonomi actuals.
DB_PATH = Path(__file__).resolve().parents[1] / "data" / "farada.db"

VERTICALS = ["Food Logistics", "Industrial IoT",
             "Consumer Electronics", "Medical Devices"]
PRODUCTS = ["Faraday-Ox Sensors", "Eval-Kits",
            "Integrated Sensors", "Services & Subscriptions"]


# ──────────────────────────────────────────────────────────────────────────────
# Row layout — single source of truth.
#
# v2.1.1 layout: BS first, then IS, then CF (direct), then Indirect Tie.
# Tie-outs sit immediately after each section so they're audited in-place
# rather than at the top of the sheet. Hidden helper rows (depreciation
# drivers, AR/AP/Inventory/VAT helpers, forecast-only splits, indirect-CF
# reconciliation) are interleaved with their visible parents and collapse
# via outlineLevel=1.
#
# Sub-section organisers (NCA_HEADER, CA_HEADER, EQUITY_HEADER, LIAB_HEADER,
# COS_HEADER, RD_HEADER, SM_HEADER, GA_HEADER, DA_HEADER) are visual labels
# only — no formula, no fill.
# ──────────────────────────────────────────────────────────────────────────────

class R:
    # Headers
    DATES = 2
    FLAG = 3
    # row 4 = blank spacer

    # ── BALANCE SHEET ─────────────────────────────────────────────
    BS_BANNER = 5

    # Non-current assets (DB display_order 0-2)
    NCA_HEADER = 6
    RD_NBV = 7              # DB: Non-tangible / R&D
    PPE_NBV = 8             # DB: Tangible / PP&E
    EQUIP_NBV = 9           # DB: Tangible / Business Equipment
    PPE_CHARGE = 10         # hidden depreciation driver
    EQUIP_CHARGE = 11       # hidden
    RD_CHARGE = 12          # hidden
    NCA_TOTAL = 13

    # Current assets (DB display_order 3-8)
    CA_HEADER = 14
    BS_PREPAID = 15         # DB ord=3
    BS_LOANS_NEG = 16       # DB ord=4
    BS_OTHER_RECV = 17      # DB ord=5
    BS_INVENTORY = 18       # DB ord=6
    PURCHASES_DRV = 19      # hidden helper
    INV_OUT_DRV = 20        # hidden helper
    BS_CASH = 21            # DB ord=7
    BS_AR = 22              # DB ord=8
    AR_HLP_SALES_ADD = 23   # hidden
    AR_HLP_COLL_SALES = 24  # hidden
    AR_HLP_COLL_DEP = 25    # hidden
    BS_VAT_RECV = 26        # hidden (forecast-only; zero in actuals)
    CA_TOTAL = 27

    TOTAL_ASSETS = 28

    # row 29 = blank spacer

    # Equity (DB display_order 9-10)
    EQUITY_HEADER = 30
    SHARE_CAPITAL = 31
    RETAINED_EARNINGS = 32
    EQUITY_TOTAL = 33

    # Liabilities (DB display_order 11-14)
    LIAB_HEADER = 34
    BS_LOAN = 35            # DB ord=11
    BS_PAYROLL_PAY = 36     # DB ord=12
    BS_OTHER_CL = 37        # DB ord=13
    BS_AP = 38              # DB ord=14
    AP_HLP_PURCH_ADD = 39   # hidden
    AP_HLP_PAYMENTS = 40    # hidden
    BS_VAT_PAY = 41         # hidden forecast-only
    VAT_HLP_OUTPUT = 42     # hidden
    VAT_HLP_INPUT = 43      # hidden
    VAT_HLP_SETTLED = 44    # hidden
    BS_TAX_PAY = 45         # hidden forecast-only
    BS_DEFERRED_GRANTS = 46 # hidden forecast-only
    BS_UNEARNED_REV = 47    # hidden forecast-only
    LIAB_TOTAL = 48

    LE_TOTAL = 49
    # row 50 = blank spacer

    BAL_CHECK = 51          # tie-out: TOTAL_ASSETS = LE_TOTAL

    # row 52 = blank spacer

    # ── INCOME STATEMENT ──────────────────────────────────────────
    IS_BANNER = 53
    REV_TOTAL = 54
    REV_CE_EVALKITS = 55    # DB ord=0
    REV_OTHER = 56          # DB ord=1 (incl. NRE)

    # Hidden forecast detail: 4 verticals × 4 products
    FL_FARADAOX = 57
    FL_EVALKITS = 58
    FL_INTEGRATED = 59
    FL_SERVICES = 60
    IIOT_FARADAOX = 61
    IIOT_EVALKITS = 62
    IIOT_INTEGRATED = 63
    IIOT_SERVICES = 64
    CE_FARADAOX = 65        # CE_EVALKITS is REV_CE_EVALKITS at row 55 (not duplicated)
    CE_INTEGRATED = 66
    CE_SERVICES = 67
    MD_FARADAOX = 68
    MD_EVALKITS = 69
    MD_INTEGRATED = 70
    MD_SERVICES = 71
    REV_NRE = 72            # part of REV_OTHER conceptually

    # Cost of Sales (DB display_order 2-6)
    COS_HEADER = 74
    COGS_MATERIALS = 75     # DB ord=2
    DIRECT_RENT = 76        # DB ord=3
    DIRECT_LOG = 77         # DB ord=4
    DIRECT_AMORT_COS = 78   # DB ord=5
    COS_OTHER = 79          # DB ord=6
    COS_TOTAL = 80

    GROSS_PROFIT = 81
    GROSS_MARGIN = 82

    # R&D (DB display_order 7-10)
    RD_HEADER = 84
    RD_PAYROLL_DE = 85
    RD_PAYROLL_RS = 86
    RD_SOFTWARE = 87
    RD_RENT = 88
    RD_OTHER = 89           # forecast-only (R&D Sensors + Other); not in DB
    RD_SUB = 90

    # S&M (DB display_order 11-17)
    SM_HEADER = 92
    SM_PAYROLL = 93
    SM_EVENTS = 94
    SM_TRAVEL = 95
    SM_DIGITAL = 96
    SM_OUTSOURCED = 97
    SM_CONTENT = 98
    SM_COMMISSIONS = 99
    SM_OTHER = 100          # forecast-only
    SM_SUB = 101

    # G&A (DB display_order 18-27, with Insurance ord=23)
    GA_HEADER = 103
    GA_PAYROLL = 104
    GA_OFFICE = 105
    GA_TRAVEL = 106
    GA_SOFTWARE = 107
    GA_TEAMDEV = 108
    GA_INSURANCE = 109      # NEW — DB ord=23
    GA_LEGAL = 110
    GA_ACCT = 111
    GA_CONSULT = 112        # NEW mapping (was missing)
    GA_MISC = 113
    GA_SUB = 114

    OPEX_TOTAL = 116
    EBITDA = 117
    EBITDA_MARGIN = 118

    # Other IS lines from DB (display_order 28-30)
    DIRECT_AMORT_ADJ = 120  # NEW — DB ord=28 (the negative "(adjustment)" line)
    FINANCE_INCOME = 121    # NEW — DB ord=29
    FINANCE_COSTS = 122     # DB ord=30

    # D&A (DB display_order 31-32 + legacy forecast splits)
    DA_HEADER = 124
    DA_DIRECT = 125         # NEW — DB ord=31
    DA_INDIRECT = 126       # NEW — DB ord=32
    DA_PPE = 127            # hidden legacy split (feeds DA_DIRECT in forecast)
    DA_EQUIP = 128          # hidden
    DA_RD = 129             # hidden (feeds DA_INDIRECT in forecast)
    DA_TOTAL = 130

    EBIT = 132
    GRANT_INCOME = 133      # forecast-only; clamped to deferred balance
    PRETAX = 134
    TAX_EXP = 135
    NI = 136

    # row 137 = blank spacer
    NI_RE_TIE = 138         # tie-out: NI vs ΔRE

    # ── CASH FLOW (Direct) ────────────────────────────────────────
    CF_BANNER = 140

    CASH_IN_TOTAL = 141     # DB ord=0
    CASH_IN_SALES = 142     # hidden forecast helper
    CASH_IN_DEPOSIT = 143   # hidden forecast helper

    CASH_OUT_SUPP = 145     # DB ord=1
    CASH_OUT_PAYROLL = 146  # DB ord=2
    CASH_OUT_TAX = 147      # DB ord=3 (Recovery/repayment of taxes)
    CASH_OUT_OTHER_OP = 148 # DB ord=4
    CASH_OUT_OPEX = 149     # hidden forecast helper
    CASH_OUT_DIRECT = 150   # hidden
    CASH_OUT_VAT = 151      # hidden

    OP_CF = 153

    CAPEX_TOTAL_DB = 155    # DB ord=5 (combined CAPEX)
    CAPEX_RD = 156          # DB ord=6
    CAPEX_OTHER_INV = 157   # DB ord=7
    CAPEX_PPE = 158         # hidden forecast split (feeds CAPEX_TOTAL_DB)
    CAPEX_EQUIP = 159       # hidden

    INV_CF = 161

    EQUITY_RAISE = 163      # DB ord=8
    GRANT_CASH = 164        # DB ord=9
    LOAN_FAC_DB = 165       # DB ord=10 (combined Loan Facility Financing)
    OTHER_FIN = 166         # DB ord=11
    LOAN_DRAW = 167         # hidden forecast helper
    LOAN_PRINCIPAL = 168    # hidden
    INTEREST_PAID = 169     # hidden
    DIVIDENDS = 170         # hidden

    FIN_CF = 172
    NET_CHG_CASH = 173
    OPEN_CASH = 174         # DB: Beginning Cash Balance
    CLOSE_CASH = 175

    # row 176 = blank spacer
    CF_TIE = 177            # tie-out: direct vs indirect

    # ── INDIRECT CF reconciliation (collapsed) ────────────────────
    TIE_BANNER = 179
    TIE_NI = 180
    TIE_DA = 181
    TIE_GRANT_NONCASH = 182
    TIE_INT_OFFSET = 183
    TIE_NONCASH_SUB = 184
    TIE_DAR = 185
    TIE_DINV = 186
    TIE_DAP = 187
    TIE_DPAYPAY = 188
    TIE_DTAXPAY = 189
    TIE_DVATPAY = 190
    TIE_DVATR = 191
    TIE_WC_SUB = 192
    TIE_OP_CF = 193
    TIE_CAPEX_PPE = 194
    TIE_CAPEX_EQ = 195
    TIE_CAPEX_RD = 196
    TIE_INV_CF = 197
    TIE_EQUITY = 198
    TIE_LOAN_DRAW = 199
    TIE_LOAN_PRINC = 200
    TIE_INT_PAID = 201
    TIE_GRANT_CASH = 202
    TIE_DIV = 203
    TIE_FIN_CF = 204
    TIE_NET_CHG = 205

    # row 206 = blank spacer
    ERR_COUNT = 207         # final error count tie-out

    # Hidden depreciable-base track (v3): rolls BUDGETED capex only, and mirrors
    # actual NBV in actual months (it's in ACTUALS_DB_MAP). The depreciation
    # charge reads this, not the live NBV line — so a manual +X appended to an
    # NBV cell is NOT depreciated (per the v3 "no useful life on manual entries").
    PPE_DEP_BASE = 208
    EQUIP_DEP_BASE = 209
    RD_DEP_BASE = 210
    # Hidden budgeted loan base (v3): interest is charged on this, not the live
    # loan line — so a manual +X loan draw doesn't accrue modelled interest.
    LOAN_INT_BASE = 211

    # v3 completeness: catch-all ΔWC for the BS leaves without their own indirect
    # line (Prepaid, Loans-neg, Other-recv, Other-CL, Unearned). Keeps the BS
    # balanced if the user manually edits any of those. Plus a check-vs-actual-
    # cash row (the client model's "CHECK VS ACTUAL CASH").
    CF_DWC_OTHER = 212
    CF_CASH_CHECK = 213

    # Aliases / convenience: last row used for styling/iter ranges.
    LAST_ROW = 213


# ──────────────────────────────────────────────────────────────────────────────
# Actuals from farada.db — leaf-row mapping
# ──────────────────────────────────────────────────────────────────────────────
#
# Each entry maps a Pro Forma row to one or more `(statement, grp_substr,
# subgroup_substr, data_substr)` keys against the `financials` table. Substring
# matching (LIKE 'X%') tolerates the truncations in the DB. Multiple keys per
# row sum their values.
#
# Only LEAF rows appear here. Aggregates (subtotals, EBITDA, NI, Total Assets)
# remain formulas in all columns — they sum the actuals-pulled leaves
# automatically and stay consistent.

ACTUALS_DB_MAP: dict[int, list[tuple[str, str, str, str]]] = {
    # ── BS leaf rows ──────────────────────────────────────────
    R.RD_NBV:            [("BS", "Non-tangible fixed asset", "Research and Develop", "Non-current")],
    R.PPE_NBV:           [("BS", "Tangible fixed assets", "PP&E", "Non-current")],
    R.EQUIP_NBV:         [("BS", "Tangible fixed assets", "Business Equipment", "Non-current")],
    # Depreciable-base track mirrors actual NBV in actual months (same DB keys),
    # then rolls budgeted capex in forecast. Keeps the charge neutral at the
    # actual→forecast boundary while excluding manual NBV top-ups.
    R.PPE_DEP_BASE:      [("BS", "Tangible fixed assets", "PP&E", "Non-current")],
    R.EQUIP_DEP_BASE:    [("BS", "Tangible fixed assets", "Business Equipment", "Non-current")],
    R.RD_DEP_BASE:       [("BS", "Non-tangible fixed asset", "Research and Develop", "Non-current")],
    R.LOAN_INT_BASE:     [("BS", "Loan facility", "Loan facility", "Non-Current Liab")],
    R.BS_PREPAID:        [("BS", "Prepaid expenses", "Prepaid expenses", "Current assets")],
    R.BS_LOANS_NEG:      [("BS", "Loans (other negot", "Loans (other negot", "Current assets")],
    R.BS_OTHER_RECV:     [("BS", "Other receivables", "Other receivables", "Current assets")],
    R.BS_INVENTORY:      [("BS", "Inventory", "Inventory", "Current assets")],
    R.BS_CASH:           [("BS", "Cash", "Cash", "Cash")],
    R.BS_AR:             [("BS", "Trade receivables", "Trade receivables", "Trade receivables")],
    R.SHARE_CAPITAL:     [("BS", "Share capital", "Share capital", "Equity")],
    R.RETAINED_EARNINGS: [("BS", "Retained earnings", "Retained earnings", "Equity")],
    R.BS_LOAN:           [("BS", "Loan facility", "Loan facility", "Non-Current Liab")],
    R.BS_PAYROLL_PAY:    [("BS", "Payables to person", "Payables to person", "Current Liab")],
    R.BS_OTHER_CL:       [("BS", "Other payables", "Other payables", "Current Liab")],
    R.BS_AP:             [("BS", "Trade payables", "Trade payables", "Trade payables")],

    # ── IS leaf rows ──────────────────────────────────────────
    R.REV_CE_EVALKITS:   [("IS", "Consumer Electronics", "Eval-Kits", "Sales")],
    R.REV_OTHER:         [("IS", "Other", "Other", "Sales")],

    R.COGS_MATERIALS:    [("IS", "COGS", "COGS", "Cost of Sales")],
    R.DIRECT_RENT:       [("IS", "Rent", "Rent", "Cost of Sales")],
    R.DIRECT_LOG:        [("IS", "Logistics", "Logistics", "Cost of Sales")],
    R.DIRECT_AMORT_COS:  [("IS", "Direct Amortization", "Direct Amortization", "Cost of Sales")],
    R.COS_OTHER:         [("IS", "Other", "Other", "Cost of Sales")],

    R.RD_PAYROLL_DE:     [("IS", "Total Payroll", "Germany", "R&D")],
    R.RD_PAYROLL_RS:     [("IS", "Total Payroll", "Serbia", "R&D")],
    R.RD_SOFTWARE:       [("IS", "Software and Tools", "Software and Tools", "R&D")],
    R.RD_RENT:           [("IS", "R&D Rent", "R&D Rent", "R&D")],

    R.SM_PAYROLL:        [("IS", "Total Payroll", "Total Payroll", "S&M")],
    R.SM_EVENTS:         [("IS", "Events/Exhibitions", "Events/Exhibitions", "S&M")],
    R.SM_TRAVEL:         [("IS", "Travel", "Travel", "S&M")],
    R.SM_DIGITAL:        [("IS", "Digital Marketing", "Digital Marketing", "S&M")],
    R.SM_OUTSOURCED:     [("IS", "Outsourced Marketi", "Outsourced Marketi", "S&M")],
    R.SM_CONTENT:        [("IS", "Content Marketing", "Content Marketing", "S&M")],
    R.SM_COMMISSIONS:    [("IS", "Sales Commissions", "Sales Commissions", "S&M")],

    R.GA_PAYROLL:        [("IS", "Total Payroll", "Total Payroll", "G&A")],
    R.GA_OFFICE:         [("IS", "Office Expenses", "Office Expenses", "G&A")],
    R.GA_TRAVEL:         [("IS", "Travel", "Travel and Repres", "G&A")],
    R.GA_SOFTWARE:       [("IS", "Software and Tools", "Software and Tools", "G&A")],
    R.GA_TEAMDEV:        [("IS", "Team Developments", "Team Developments", "G&A")],
    R.GA_INSURANCE:      [("IS", "Insurance", "Insurance", "G&A")],
    R.GA_LEGAL:          [("IS", "External Professio", "Legal", "G&A")],
    R.GA_ACCT:           [("IS", "External Professio", "Accounting", "G&A")],
    R.GA_CONSULT:        [("IS", "External Professio", "Other/ Consulting", "G&A")],
    R.GA_MISC:           [("IS", "Miscellaneous expe", "Miscellaneous expe", "G&A")],

    R.DIRECT_AMORT_ADJ:  [("IS", "Direct Amortization (adj", "Direct Amortization (adj", "Direct Amortization (adjustmen")],
    R.FINANCE_INCOME:    [("IS", "Finance income", "Finance income", "Finance income")],
    R.FINANCE_COSTS:     [("IS", "Finance costs", "Finance costs", "Finance costs")],

    R.DA_DIRECT:         [("IS", "Direct Amortization", "Direct Amortization", "Depreciation and amortization")],
    R.DA_INDIRECT:       [("IS", "Indirect Amortization", "Indirect Amortization", "Depreciation and amortization")],

    # ── CF leaf rows ──────────────────────────────────────────
    R.CASH_IN_TOTAL:     [("CF", "Cash received from custo", "Cash received from custo", "Cash Flow from Operat")],
    R.CASH_OUT_SUPP:     [("CF", "Cash paid to suppl", "Cash paid to suppl", "Cash Flow from Operat")],
    R.CASH_OUT_PAYROLL:  [("CF", "Payment for person", "Payment for person", "Cash Flow from Operat")],
    R.CASH_OUT_TAX:      [("CF", "Recovery/(repayment) of", "Recovery/(repayment) of", "Cash Flow from Operat")],
    R.CASH_OUT_OTHER_OP: [("CF", "Other", "Other", "Cash Flow from Operat")],

    R.CAPEX_TOTAL_DB:    [("CF", "CAPEX", "CAPEX", "Cash Flow from Invest")],
    R.CAPEX_RD:          [("CF", "Research and Devel", "Research and Devel", "Cash Flow from Invest")],
    R.CAPEX_OTHER_INV:   [("CF", "Other", "Other", "Cash Flow from Invest")],

    R.EQUITY_RAISE:      [("CF", "Capital Increase", "Capital Increase", "Cash Flow from Financ")],
    R.GRANT_CASH:        [("CF", "Grants", "Grants", "Cash Flow from Financ")],
    R.LOAN_FAC_DB:       [("CF", "Loan Facility", "Loan Facility", "Cash Flow from Financ")],
    R.OTHER_FIN:         [("CF", "Other payments", "Other payments", "Cash Flow from Financ")],
}

# Populated by `_load_actuals_from_db`. Two views:
#   _ACTUALS_VALUES:  {row: {period_date: value}}   — for build_actuals
#   _ACTUALS_PRESENT: {row: set(period_date)}       — for fill_right's branch
_ACTUALS_VALUES: dict[int, dict[date, float]] = {}
_ACTUALS_PRESENT: dict[int, set[date]] = {}


def _load_actuals_from_db(db_path: Path) -> tuple[dict, dict, list[date]]:
    """Read `financials` table; return (values, present, all_actual_months)."""
    values: dict[int, dict[date, float]] = {}
    present: dict[int, set[date]] = {}
    actual_months: set[date] = set()

    with sqlite3.connect(db_path) as conn:
        conn.row_factory = sqlite3.Row
        # Consolidated entity only. April also carries farada_de / farada_rs
        # entity rows (loaded for the L2 reconciliation check); without this
        # filter every April line would be summed across entities (~2x inflated).
        cur = conn.execute(
            "SELECT DISTINCT period_date FROM financials "
            "WHERE scenario='actual' AND entity='farada' ORDER BY period_date"
        )
        for row in cur:
            actual_months.add(date.fromisoformat(row[0]))

        for pf_row, key_list in ACTUALS_DB_MAP.items():
            per_period: dict[date, float] = {}
            for stmt, grp_sub, sg_sub, data_sub in key_list:
                cur = conn.execute(
                    "SELECT period_date, value FROM financials "
                    "WHERE scenario='actual' "
                    "  AND entity='farada' "
                    "  AND statement=? "
                    "  AND grp LIKE ? "
                    "  AND subgroup LIKE ? "
                    "  AND data LIKE ?",
                    (stmt, f"{grp_sub}%", f"{sg_sub}%", f"{data_sub}%"),
                )
                for r in cur:
                    if r["value"] is None:
                        continue
                    d = date.fromisoformat(r["period_date"])
                    per_period[d] = per_period.get(d, 0.0) + float(r["value"])
            if per_period:
                values[pf_row] = per_period
                present[pf_row] = set(per_period.keys())

    return values, present, sorted(actual_months)


# ──────────────────────────────────────────────────────────────────────────────
# Styling helpers — v2.1.1 lighter palette
# ──────────────────────────────────────────────────────────────────────────────

FONT_NAME = "Century Gothic"
BASE_FONT = Font(name=FONT_NAME, size=10)
BOLD = Font(name=FONT_NAME, size=10, bold=True)
BOLD_ITALIC = Font(name=FONT_NAME, size=10, bold=True, italic=True)
INPUT_FONT = Font(name=FONT_NAME, size=10, color="0000FF")
CHECK_FONT = Font(name=FONT_NAME, size=10, color="C00000", bold=True)
HELPER_FONT = Font(name=FONT_NAME, size=10, color="595959", italic=True)
ITALIC = Font(name=FONT_NAME, size=10, italic=True)

# v2.1.1: lighter palette — no navy banner, rely on borders for separation.
BANNER_FONT = Font(name=FONT_NAME, size=11, bold=True, color="002060")
BANNER_FILL = PatternFill("solid", fgColor="F4F4F4")          # light gray
SECTION_HEADER_FONT = BOLD_ITALIC                              # for NCA/CA/etc.
KEY_TOTAL_FILL = PatternFill(fill_type=None)                   # no fill; borders only
SUBTOTAL_FILL = PatternFill(fill_type=None)                    # no fill
DATA_FILL = PatternFill(fill_type=None)
HEADER_DATE_FILL = PatternFill("solid", fgColor="EAEAEA")      # subtle date strip

THIN = Side(style="thin", color="000000")
DOUBLE = Side(style="double", color="000000")
BORDER_TOTAL = Border(top=THIN, bottom=THIN)
BORDER_GRAND_TOTAL = Border(top=THIN, bottom=DOUBLE)
BORDER_SUBTOTAL = Border(top=THIN)                             # subtotal: top thin border only
BORDER_MARGIN = Border(bottom=DOUBLE)

# Number formats
FMT_AMOUNT = '_(* #,##0_);_(* (#,##0);_(* "-"??_);_(@_)'       # accounting w/ parens
FMT_PCT = '0.0%;(0.0%);"-"'
FMT_DATE = "mmm-yy"


def fill_right(ws, row: int, formula_origin_col: str, formula: str) -> None:
    """Write formula in origin column then fill across all 12 period columns.

    v2.1 behaviour: for rows in `_ACTUALS_PRESENT`, any column whose period
    appears in that row's actuals set is overwritten with `=Actuals!<col><row>`.
    The forecast formula remains in unmapped columns. Aggregate rows (not in
    `_ACTUALS_PRESENT`) fall back to pure formula-fill across all columns.
    """
    origin_cell = f"{formula_origin_col}{row}"
    actuals_periods = _ACTUALS_PRESENT.get(row, set())

    for i, col in enumerate(PERIOD_COLS):
        target = f"{col}{row}"
        if PERIOD_DATES[i] in actuals_periods:
            ws[target] = f"=Actuals!{col}{row}"
        elif col == formula_origin_col:
            ws[target] = formula
        else:
            ws[target] = Translator(formula, origin=origin_cell).translate_formula(target)


def fy_sum(ws, row: int) -> None:
    ws[f"{FY_COL}{row}"] = f"=SUM({PERIOD_COLS[0]}{row}:{PERIOD_COLS[-1]}{row})"


def fy_last(ws, row: int) -> None:
    ws[f"{FY_COL}{row}"] = f"={PERIOD_COLS[-1]}{row}"


def label(ws, row: int, text: str, code: str = "", bold: bool = False,
          banner: bool = False, helper: bool = False,
          section_header: bool = False) -> None:
    cell = ws[f"{LABEL_COL}{row}"]
    cell.value = text
    if banner:
        cell.font = BANNER_FONT
        cell.fill = BANNER_FILL
    elif section_header:
        cell.font = SECTION_HEADER_FONT
    elif bold:
        cell.font = BOLD
    elif helper:
        cell.font = HELPER_FONT
    if code:
        ws[f"{CODE_COL}{row}"] = code


# ──────────────────────────────────────────────────────────────────────────────
# Inputs sheet
# ──────────────────────────────────────────────────────────────────────────────

class I:
    # Funding
    LOAN_OPEN = 5
    LOAN_RATE = 6
    LOAN_PMT = 7
    EQUITY_TR1_DATE = 10
    EQUITY_TR1_AMT = 11
    PROFIT_RECOG_R = 14
    PROFIT_RECV_R = 15
    EIC_RECV_R = 16
    EIC_RECOG_R = 17

    # Revenue
    QTY_BASE = 20    # 16 monthly rows
    PRICE_BASE = 40  # 16 single values
    NRE_QTY_R = 60
    NRE_PRICE = 61
    NRE_COST_PCT = 62

    # Cost
    UNIT_COST_BASE = 65   # 16 single values
    MAINT_PCT = 85
    PROD_RENT = 86
    LOGISTICS = 87
    OTHER_DIRECT = 88
    DIRECT_HC = 89
    DIRECT_COST_PER_HEAD = 90

    SM_PAYROLL_TOTAL = 92
    EVENTS = 93
    SM_TRAVEL = 94
    DIGITAL_MKT = 95
    OUTSOURCED_MKT = 96
    CONTENT_MKT = 97
    SALES_COMMISSION_PCT = 98
    OTHER_MKT_PCT = 99

    GA_PAYROLL_TOTAL = 101
    OFFICE_EXP = 102
    GA_TRAVEL = 103
    GA_SOFTWARE = 104
    TEAM_DEV = 105
    GA_INSURANCE = 106       # NEW
    LEGAL = 107
    ACCT = 108
    OTHER_CONSULT_R = 109
    MISC_PCT = 110

    RD_PAYROLL_DE = 112
    RD_PAYROLL_RS = 113
    RD_SOFTWARE = 114
    RD_RENT = 115
    OTHER_RD = 116

    DSO_DAYS = 119
    DPO_DAYS = 120
    PAYROLL_ACCRUAL_PCT = 121
    VAT_RATE = 122
    DEPOSIT_PCT = 123
    TAX_RATE = 124
    TAX_LAG_MONTHS = 125

    PPE_OPEN = 128
    PPE_CAPEX_R = 129
    PPE_LIFE = 130
    EQUIP_OPEN = 131
    EQUIP_CAPEX_R = 132
    EQUIP_LIFE = 133
    RD_OPEN = 134
    RD_CAPEX_R = 135
    RD_LIFE = 136

    OPEN_CASH = 139
    OPEN_AR = 140
    OPEN_INV = 141
    OPEN_VAT_R = 142
    OPEN_AP = 143
    OPEN_PAYROLL_PAY = 144
    OPEN_VAT_PAY = 145
    OPEN_TAX_PAY = 146
    OPEN_DEFERRED_GRANTS = 147
    OPEN_UNEARNED_REV = 148
    OPEN_SHARE_CAP = 149
    OPEN_RE = 150
    # v2.1.1: OPEN_OTHER_CA split into three (matches DB sub-lines)
    OPEN_PREPAID = 151
    OPEN_LOANS_NEG = 152
    OPEN_OTHER_RECV = 153
    OPEN_OTHER_CL = 154
    DIRECT_AMORT_R = 155

    # v2.1.1: forecast values for new IS lines (default 0 — user edits)
    DA_DIRECT_R = 158         # monthly DA-Direct (default 0; legacy split feeds it)
    DA_INDIRECT_R = 159       # monthly DA-Indirect (default 0)
    FINANCE_INCOME_R = 160    # monthly Finance income (default 0)
    DIRECT_AMORT_ADJ_R = 161  # monthly Direct Amort adjustment (default 0)
    COS_OTHER_R = 162         # monthly COS Other (default 0)


# Sold-quantity schedule for FaradaOx sensors — mirrors source Pro Forma r34.
FARADAOX_TOTAL_2026 = 3_000
FARADAOX_QUARTER_PCT = [0.0, 0.25, 0.25, 0.50]
FARADAOX_VERT_PCT = {"Food Logistics": 0.50, "Industrial IoT": 0.30,
                     "Consumer Electronics": 0.10, "Medical Devices": 0.10}


def faradaox_monthly_qty(vertical: str) -> list:
    monthly = []
    for q in range(4):
        q_total = FARADAOX_TOTAL_2026 * FARADAOX_QUARTER_PCT[q] * FARADAOX_VERT_PCT[vertical]
        per_month = q_total / 3
        monthly.extend([per_month] * 3)
    return monthly


EVAL_KIT_QTY = {
    "Food Logistics":       [0, 0, 1, 1, 1, 0, 1, 0, 1, 0, 0, 0],
    "Industrial IoT":       [0, 0, 0, 0, 0, 1, 0, 1, 0, 1, 0, 0],
    "Consumer Electronics": [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 1],
    "Medical Devices":      [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
}


def build_inputs(wb: Workbook) -> None:
    ws = wb.create_sheet("Inputs")
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 14
    for col in PERIOD_COLS:
        ws.column_dimensions[col].width = 12

    for i, col in enumerate(PERIOD_COLS):
        ws[f"{col}3"] = f"{MONTHS[i]}-26"
        ws[f"{col}3"].font = Font(name=FONT_NAME, size=10, bold=True)
        ws[f"{col}3"].alignment = Alignment(horizontal="center")
        ws[f"{col}3"].fill = HEADER_DATE_FILL

    def putval(row: int, lbl: str, value, section: str = ""):
        if section:
            ws[f"B{row}"] = section
            ws[f"B{row}"].font = BOLD
        ws[f"C{row}"] = lbl
        ws[f"C{row}"].font = Font(name=FONT_NAME, size=10)
        ws[f"D{row}"] = value
        ws[f"D{row}"].font = INPUT_FONT
        if isinstance(value, (int, float)):
            ws[f"D{row}"].number_format = FMT_AMOUNT if abs(value) > 1 else FMT_PCT

    def putrow(row: int, lbl: str, values: list):
        ws[f"C{row}"] = lbl
        ws[f"C{row}"].font = Font(name=FONT_NAME, size=10)
        for col, v in zip(PERIOD_COLS, values):
            ws[f"{col}{row}"] = v
            ws[f"{col}{row}"].font = INPUT_FONT
            ws[f"{col}{row}"].number_format = FMT_AMOUNT

    # I. FUNDING
    ws["B1"] = "I. FUNDING"
    ws["B1"].font = BOLD
    putval(I.LOAN_OPEN, "Loan opening balance (EUR) [REF: 0 in 2026]", 0)
    putval(I.LOAN_RATE, "Loan annual interest rate [REF: 6% assumption]", 0.06)
    putval(I.LOAN_PMT, "Loan monthly principal payment (EUR) [REF: 0]", 0)
    putval(I.EQUITY_TR1_DATE, "Equity tranche 1 date [FILL]", date(2026, 6, 1))
    putval(I.EQUITY_TR1_AMT, "Equity tranche 1 amount (EUR) [FILL]", 0)
    putrow(I.PROFIT_RECOG_R, "ProFIT monthly recognition (EUR) [CALIB source IS r68 - EIC amort]",
           [0, 0, 0, 77_477, 65_417, 65_417, 65_417, 65_417, 63_742, 75_802, 75_802, 75_802])
    putrow(I.PROFIT_RECV_R, "ProFIT cash receipts (EUR) [CALIB source CF r17]",
           [0, 0, 0, 95_560, 0, 0, 250_500, 0, 0, 259_210, 0, 0])
    putrow(I.EIC_RECV_R, "EIC cash receipts (EUR) [REF Inputs!J54]",
           [0, 434_000, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0])
    putrow(I.EIC_RECOG_R, "EIC monthly recognition (EUR) [DERIVED 434K/24mo]",
           [434_000 / 24] * 12)

    # II. REVENUE
    ws["B19"] = "II. REVENUE"
    ws["B19"].font = BOLD
    ws["C19"] = "Sold quantities by vertical × product (monthly)"
    r = I.QTY_BASE
    qty_schedule = {}
    for v in VERTICALS:
        qty_schedule[(v, "Faraday-Ox Sensors")] = faradaox_monthly_qty(v)
        qty_schedule[(v, "Eval-Kits")] = EVAL_KIT_QTY[v]
        qty_schedule[(v, "Integrated Sensors")] = [0] * 12
        qty_schedule[(v, "Services & Subscriptions")] = [0] * 12

    for v in VERTICALS:
        for p in PRODUCTS:
            putrow(r, f"  {v} – {p}", qty_schedule[(v, p)])
            r += 1

    ws[f"C{I.PRICE_BASE - 1}"] = "Unit prices (EUR) [REF Inputs!J60:J78]"
    ws[f"C{I.PRICE_BASE - 1}"].font = BOLD
    r = I.PRICE_BASE
    prices = {
        ("Food Logistics", "Faraday-Ox Sensors"): 20,
        ("Industrial IoT", "Faraday-Ox Sensors"): 20,
        ("Consumer Electronics", "Faraday-Ox Sensors"): 20,
        ("Medical Devices", "Faraday-Ox Sensors"): 20,
        ("Food Logistics", "Eval-Kits"): 1500,
        ("Industrial IoT", "Eval-Kits"): 1500,
        ("Consumer Electronics", "Eval-Kits"): 1500,
        ("Medical Devices", "Eval-Kits"): 1500,
        ("Food Logistics", "Integrated Sensors"): 7.5,
        ("Industrial IoT", "Integrated Sensors"): 7.5,
        ("Consumer Electronics", "Integrated Sensors"): 7.5,
        ("Medical Devices", "Integrated Sensors"): 7.5,
        ("Food Logistics", "Services & Subscriptions"): 0,
        ("Industrial IoT", "Services & Subscriptions"): 0,
        ("Consumer Electronics", "Services & Subscriptions"): 0,
        ("Medical Devices", "Services & Subscriptions"): 0,
    }
    for v in VERTICALS:
        for p in PRODUCTS:
            putval(r, f"  {v} – {p}", prices[(v, p)])
            r += 1

    putrow(I.NRE_QTY_R, "NRE deals per month [CALIB source: 1/mo Apr-Dec]",
           [0, 0, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1])
    putval(I.NRE_PRICE, "NRE price per deal (EUR) [REF Inputs!J78]", 30_000)
    putval(I.NRE_COST_PCT, "NRE cost % of price [REF Inputs!J192]", 0.20)

    # III. COSTS
    ws["B64"] = "III. COSTS"
    ws["B64"].font = BOLD
    ws["C64"] = "Unit costs (EUR per unit) [REF Inputs!J173-J181]"
    ws["C64"].font = BOLD
    r = I.UNIT_COST_BASE
    unit_costs = {
        ("Food Logistics", "Faraday-Ox Sensors"): 5.596,
        ("Industrial IoT", "Faraday-Ox Sensors"): 5.596,
        ("Consumer Electronics", "Faraday-Ox Sensors"): 5.596,
        ("Medical Devices", "Faraday-Ox Sensors"): 5.596,
        ("Food Logistics", "Eval-Kits"): 500,
        ("Industrial IoT", "Eval-Kits"): 500,
        ("Consumer Electronics", "Eval-Kits"): 500,
        ("Medical Devices", "Eval-Kits"): 500,
        ("Food Logistics", "Integrated Sensors"): 0,
        ("Industrial IoT", "Integrated Sensors"): 0,
        ("Consumer Electronics", "Integrated Sensors"): 0,
        ("Medical Devices", "Integrated Sensors"): 0,
        ("Food Logistics", "Services & Subscriptions"): 0,
        ("Industrial IoT", "Services & Subscriptions"): 0,
        ("Consumer Electronics", "Services & Subscriptions"): 0,
        ("Medical Devices", "Services & Subscriptions"): 0,
    }
    for v in VERTICALS:
        for p in PRODUCTS:
            putval(r, f"  {v} – {p}", unit_costs[(v, p)])
            r += 1

    putval(I.MAINT_PCT, "Maintenance % of revenue [REF Inputs!J219]", 0.005)
    putval(I.PROD_RENT, "Production rent EUR/month [REF Inputs!J221]", 5_000)
    putval(I.LOGISTICS, "Logistics EUR/month [REF Inputs!J223]", 2_000)
    putval(I.OTHER_DIRECT, "Other direct EUR/month [FILL]", 0)
    putrow(I.DIRECT_HC, "Direct production headcount [REF: 0 in 2026]", [0] * 12)
    putval(I.DIRECT_COST_PER_HEAD, "Direct cost per head (EUR/month) [FILL]", 0)

    sm_payroll = [17_967.32] * 4 + [22_133.99] * 8
    putrow(I.SM_PAYROLL_TOTAL, "S&M Payroll EUR/month [REF HR sheet]", sm_payroll)

    events = [5_000, 5_000, 10_000, 0, 10_000, 5_000, 0, 0, 10_000, 0, 5_000, 0]
    putrow(I.EVENTS, "Events/exhibitions EUR/month [REF Inputs!J230-J236]", events)
    putrow(I.SM_TRAVEL, "S&M Travel EUR/month [REF Inputs!J240]", [3_000] * 12)
    putval(I.DIGITAL_MKT, "Digital marketing EUR/month [REF Inputs!J247]", 2_500)
    putval(I.OUTSOURCED_MKT, "Outsourced marketing EUR/month [REF Inputs!J254]", 1_000)
    putval(I.CONTENT_MKT, "Content marketing EUR/month [REF Inputs!J261]", 2_500)
    putval(I.SALES_COMMISSION_PCT, "Sales commission % of revenue [REF Inputs!J267]", 0.20)
    putval(I.OTHER_MKT_PCT, "Other marketing % of S&M [REF Inputs!J269]", 0.05)

    putrow(I.GA_PAYROLL_TOTAL, "G&A Payroll EUR/month [REF HR sheet]",
           [19_924.53] * 12)
    putval(I.OFFICE_EXP, "Office expenses EUR/month [REF Inputs!J272]", 1_500)
    putval(I.GA_TRAVEL, "G&A Travel EUR/month [REF Inputs!J273]", 3_000)
    putval(I.GA_SOFTWARE, "G&A Software EUR/month [REF Inputs!J274]", 1_250)
    putval(I.TEAM_DEV, "Team development EUR/month [REF Inputs!J275]", 4_166.67)
    # v2.1.1: Insurance — DB shows Jan=3,743, Feb=7,305 (avg ~5,500; use 3,500 baseline
    # since Feb may include a one-off prepay). User can edit.
    putval(I.GA_INSURANCE, "Insurance EUR/month [DB-cal ~3,500/mo baseline]", 3_500)
    putval(I.LEGAL, "Legal EUR/month [REF Inputs!J277]", 10_000)
    putval(I.ACCT, "Accountancy EUR/month [REF Inputs!J278]", 1_500)
    other_consult = [1_100] * 3 + [910] * 3 + [1_820] * 6
    putrow(I.OTHER_CONSULT_R, "Other consulting EUR/month [REF Inputs!J279-J281]",
           other_consult)
    putval(I.MISC_PCT, "Misc % of G&A [REF Inputs!J282]", 0.05)

    full_rd_de = 58_128.19
    full_rd_rs = 33_832.53
    capitalize_pct = 0.40
    rd_de_expensed = full_rd_de * (1 - capitalize_pct)
    rd_rs_expensed = full_rd_rs * (1 - capitalize_pct)
    putrow(I.RD_PAYROLL_DE,
           f"R&D Payroll Germany 60% expensed [CALIB: full={full_rd_de:,.0f} × 60%]",
           [rd_de_expensed] * 12)
    putrow(I.RD_PAYROLL_RS,
           f"R&D Payroll Serbia 60% expensed [CALIB: full={full_rd_rs:,.0f} × 60%]",
           [rd_rs_expensed] * 12)
    putval(I.RD_SOFTWARE, "R&D Software EUR/month [REF Inputs!J286]", 1_250)
    putval(I.RD_RENT, "R&D Rent Serbia EUR/month [REF Inputs!J290]", 3_000)
    other_rd_monthly = (
        [0, 0, 0]
        + [4_397] * 6
        + [2_998] * 3
    )
    putrow(I.OTHER_RD, "R&D Sensors + Other R&D EUR/month [CALIB source IS r63+r65]",
           other_rd_monthly)

    # IV. WORKING CAPITAL
    ws[f"B{I.DSO_DAYS - 1}"] = "IV. WORKING CAPITAL"
    ws[f"B{I.DSO_DAYS - 1}"].font = BOLD
    putval(I.DSO_DAYS, "DSO days [REF Inputs!J323]", 30)
    putval(I.DPO_DAYS, "DPO days [REF Inputs!J325]", 40)
    putval(I.PAYROLL_ACCRUAL_PCT, "Payroll paid next month % [FILL: 0=in-month]", 0)
    putval(I.VAT_RATE, "VAT rate [FILL: raw cell empty; using 0]", 0.0)
    putval(I.DEPOSIT_PCT, "Service revenue deposit % [REF Inputs §346: 10%]", 0.10)
    putval(I.TAX_RATE, "Corporate tax rate [REF Inputs!J329 = 15.83%]", 0.1583)
    putval(I.TAX_LAG_MONTHS, "Tax payment lag months [FILL: annual = 12]", 12)

    # V. CAPEX
    ws[f"B{I.PPE_OPEN - 1}"] = "V. CAPEX & DEPRECIATION"
    ws[f"B{I.PPE_OPEN - 1}"].font = BOLD
    putval(I.PPE_OPEN, "PPE opening NBV [REF Inputs!J312]", 317_583)
    putrow(I.PPE_CAPEX_R, "PPE monthly capex [CALIB source CF r12]",
           [0, 0, 5_000, 0, 1_500, 5_000, 0, 0, 5_000, 0, 0, 5_000])
    putval(I.PPE_LIFE, "PPE useful life (months) [REF Inputs!J315 = 10y]", 120)
    putval(I.EQUIP_OPEN, "Equipment opening NBV [REF Inputs!J318]", 461_417)
    putrow(I.EQUIP_CAPEX_R, "Equipment monthly capex [REF Inputs!J319 × headcount]",
           [0, 0, 0] + [3_000] * 9)
    putval(I.EQUIP_LIFE, "Equipment useful life (months) [REF Inputs!J320 = 5y]", 60)
    putval(I.RD_OPEN, "R&D capitalised opening NBV [REF Inputs!J301]", 21_748)
    putrow(I.RD_CAPEX_R, "R&D monthly capex [CALIB source CF r13: 36.7K/mo Apr-Dec]",
           [0, 0, 0] + [36_784] * 9)
    putval(I.RD_LIFE, "R&D useful life (months) [REF Inputs!J303 = 4y]", 48)

    # VI. OPENING BALANCES
    ws[f"B{I.OPEN_CASH - 1}"] = "VI. OPENING BALANCES (Dec-25 close)"
    ws[f"B{I.OPEN_CASH - 1}"].font = BOLD
    open_cash = 1_531_776 + 31_310
    open_ar = 144_676 + 10_344
    open_inv = 19_661
    open_vat_r = 0
    open_ppe = 317_583
    open_equip = 434_143
    open_rd = 9_648 + 12_100
    open_ap = 329_508 + 65_802
    open_payroll_pay = 17_837
    open_vat_pay = 0
    open_tax_pay = 0
    open_loan = 0
    open_deferred = 0
    open_unearned = 0
    open_share_cap = 5_425_012 + 190
    # v2.1.1: Other CA broken out into three (matches DB sub-lines).
    # Jan-26 actuals from DB: Prepaid=21, Loans_neg=200, Other_recv=19,169.
    open_prepaid = 21
    open_loans_neg = 200
    open_other_recv = 19_169
    open_other_cl = 305_924

    open_assets = (open_cash + open_ar + open_inv + open_vat_r
                   + open_prepaid + open_loans_neg + open_other_recv
                   + open_ppe + open_equip + open_rd)
    open_lne_ex_re = (open_ap + open_payroll_pay + open_vat_pay + open_tax_pay
                      + open_loan + open_deferred + open_unearned + open_other_cl
                      + open_share_cap)
    open_re = open_assets - open_lne_ex_re

    putval(I.OPEN_CASH, "Cash", open_cash)
    putval(I.OPEN_AR, "Accounts receivable [REF C17+C52]", open_ar)
    putval(I.OPEN_INV, "Inventory [REF C54]", open_inv)
    putval(I.OPEN_VAT_R, "VAT receivable [FILL]", open_vat_r)
    putval(I.OPEN_AP, "Accounts payable [REF C27+C62]", open_ap)
    putval(I.OPEN_PAYROLL_PAY, "Payroll payable [REF C29]", open_payroll_pay)
    putval(I.OPEN_VAT_PAY, "VAT payable [FILL]", open_vat_pay)
    putval(I.OPEN_TAX_PAY, "Tax payable [FILL]", open_tax_pay)
    putval(I.OPEN_DEFERRED_GRANTS, "Deferred grant income [FILL]", open_deferred)
    putval(I.OPEN_UNEARNED_REV, "Unearned revenue [FILL]", open_unearned)
    putval(I.OPEN_SHARE_CAP, "Share capital", open_share_cap)
    putval(I.OPEN_RE, "Retained earnings", open_re)
    putval(I.OPEN_PREPAID, "Prepaid expenses [DB Jan-26]", open_prepaid)
    putval(I.OPEN_LOANS_NEG, "Loans (other negotiable) [DB Jan-26]", open_loans_neg)
    putval(I.OPEN_OTHER_RECV, "Other receivables [DB Jan-26]", open_other_recv)
    putval(I.OPEN_OTHER_CL, "Other current liabilities", open_other_cl)
    # v2.1.1: DIRECT_AMORT_R is the CoS Direct Amortization line. The new
    # DA_DIRECT row below EBITDA already captures depreciation via PPE_CHARGE
    # + EQUIP_CHARGE — leaving this schedule nonzero would double-count D&A
    # against EBIT. In v2.1.1 forecasts emit 0; actual months get the DB value
    # via fill_right on DIRECT_AMORT_COS. The (adjustment) line at row 120
    # reverses the CoS impact in DB so EBIT is correct.
    putrow(I.DIRECT_AMORT_R, "Direct Amortization (CoS) monthly [forecast 0; actuals = DB]",
           [0] * 12)

    # v2.1.1: forecast inputs for new IS lines (default 0; user edits to model)
    ws[f"B{I.DA_DIRECT_R - 1}"] = "VII. NEW IS-LINE FORECASTS (v2.1.1)"
    ws[f"B{I.DA_DIRECT_R - 1}"].font = BOLD
    putrow(I.DA_DIRECT_R, "DA Direct monthly [forecast — default 0]", [0] * 12)
    putrow(I.DA_INDIRECT_R, "DA Indirect monthly [forecast — default 0]", [0] * 12)
    putrow(I.FINANCE_INCOME_R, "Finance income monthly [forecast — default 0]", [0] * 12)
    putrow(I.DIRECT_AMORT_ADJ_R, "Direct Amort adjustment monthly [forecast — default 0]", [0] * 12)
    putrow(I.COS_OTHER_R, "COS Other monthly [forecast — default 0]", [0] * 12)

    clean_inputs_labels(ws)
    return ws


def clean_inputs_labels(ws) -> None:
    """Strip (parens), [tags], and verbose unit suffixes from Inputs labels."""
    paren_re = re.compile(r'\s*\([^)]*\)')
    bracket_re = re.compile(r'\s*\[[^\]]*\]')
    suffix_patterns = [
        re.compile(r'\s+EUR/month\s*$', re.IGNORECASE),
        re.compile(r'\s+EUR/year\s*$', re.IGNORECASE),
        re.compile(r'\s+EUR per unit\s*$', re.IGNORECASE),
        re.compile(r'\s+EUR per deal\s*$', re.IGNORECASE),
        re.compile(r'\s+EUR p\.m\.\s*$', re.IGNORECASE),
        re.compile(r'\s+EUR\s*$'),
        re.compile(r'\s+%\s+of\s+revenue\s*$', re.IGNORECASE),
        re.compile(r'\s+%\s+of\s+G&A\s*$', re.IGNORECASE),
        re.compile(r'\s+%\s+of\s+S&M\s*$', re.IGNORECASE),
    ]
    for row in ws.iter_rows(min_col=3, max_col=3):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                cleaned = bracket_re.sub('', cell.value)
                cleaned = paren_re.sub('', cleaned)
                for pat in suffix_patterns:
                    cleaned = pat.sub('', cleaned)
                cell.value = cleaned.rstrip()


# ──────────────────────────────────────────────────────────────────────────────
# Pro Forma sheet
# ──────────────────────────────────────────────────────────────────────────────

def inp_abs(row: int, col: str = "D") -> str:
    return f"Inputs!${col}${row}"


def inp_period(row: int, col: str) -> str:
    return f"Inputs!{col}${row}"


def _prev_col(col: str) -> str:
    """Return the column to the left of `col` (PERIOD_COLS or OPEN_COL)."""
    if col == PERIOD_COLS[0]:
        return OPEN_COL
    idx = PERIOD_COLS.index(col)
    return PERIOD_COLS[idx - 1]


def build_pro_forma(wb: Workbook) -> None:
    ws = wb.create_sheet("Pro Forma")
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions[LABEL_COL].width = 44
    ws.column_dimensions[CODE_COL].width = 5
    ws.column_dimensions[OPEN_COL].width = 13
    for col in PERIOD_COLS:
        ws.column_dimensions[col].width = 12
    ws.column_dimensions[FY_COL].width = 14

    # ── Header ──
    ws[f"{LABEL_COL}1"] = "Farada Rolling Budget — FY2026 (EUR)"
    ws[f"{LABEL_COL}1"].font = Font(name=FONT_NAME, size=12, bold=True, color="002060")
    ws[f"{PERIOD_COLS[0]}{R.DATES}"] = date(2026, 1, 1)
    ws[f"{PERIOD_COLS[0]}{R.DATES}"].number_format = FMT_DATE
    ws[f"{PERIOD_COLS[0]}{R.DATES}"].font = BOLD
    for i in range(1, 12):
        prev = PERIOD_COLS[i - 1]
        curr = PERIOD_COLS[i]
        ws[f"{curr}{R.DATES}"] = f"=EDATE({prev}{R.DATES},1)"
        ws[f"{curr}{R.DATES}"].number_format = FMT_DATE
        ws[f"{curr}{R.DATES}"].font = BOLD
    ws[f"{FY_COL}{R.DATES}"] = "FY2026"
    ws[f"{FY_COL}{R.DATES}"].font = BOLD
    for col, flag in zip(PERIOD_COLS, PERIOD_FLAGS):
        ws[f"{col}{R.FLAG}"] = flag
        ws[f"{col}{R.FLAG}"].alignment = Alignment(horizontal="center")
        if flag == "Actual":
            ws[f"{col}{R.FLAG}"].font = Font(name=FONT_NAME, size=9, bold=True, color="C00000")
        else:
            ws[f"{col}{R.FLAG}"].font = Font(name=FONT_NAME, size=9, bold=True, color="595959")

    # ─────────────────────────────────────────────────────────────
    # SECTION 1 — BALANCE SHEET (built first; tie-out at end)
    # ─────────────────────────────────────────────────────────────
    label(ws, R.BS_BANNER, "BALANCE SHEET", banner=True)

    # ── Non-current assets ──
    label(ws, R.NCA_HEADER, "Non-current assets", section_header=True)

    label(ws, R.RD_NBV, "  R&D capitalised (NBV)", "BS")
    ws[f"D{R.RD_NBV}"] = f"={inp_abs(I.RD_OPEN)}"
    fill_right(ws, R.RD_NBV, "E",
               f"=D{R.RD_NBV}+E{R.CAPEX_RD}-E{R.RD_CHARGE}")
    fy_last(ws, R.RD_NBV)

    label(ws, R.PPE_NBV, "  PP&E (NBV)", "BS")
    ws[f"D{R.PPE_NBV}"] = f"={inp_abs(I.PPE_OPEN)}"
    fill_right(ws, R.PPE_NBV, "E",
               f"=D{R.PPE_NBV}+E{R.CAPEX_PPE}-E{R.PPE_CHARGE}")
    fy_last(ws, R.PPE_NBV)

    label(ws, R.EQUIP_NBV, "  Business equipment (NBV)", "BS")
    ws[f"D{R.EQUIP_NBV}"] = f"={inp_abs(I.EQUIP_OPEN)}"
    fill_right(ws, R.EQUIP_NBV, "E",
               f"=D{R.EQUIP_NBV}+E{R.CAPEX_EQUIP}-E{R.EQUIP_CHARGE}")
    fy_last(ws, R.EQUIP_NBV)

    # v3: depreciation is charged on the budgeted DEP_BASE (not the live NBV), so
    # a manual +X on an NBV cell isn't depreciated. DEP_BASE ≡ NBV when unedited.
    label(ws, R.PPE_CHARGE, "    PP&E depreciation", "Drv", helper=True)
    fill_right(ws, R.PPE_CHARGE, "E",
               f"=(D{R.PPE_DEP_BASE}+E{R.CAPEX_PPE})/{inp_abs(I.PPE_LIFE)}")
    fy_sum(ws, R.PPE_CHARGE)
    label(ws, R.EQUIP_CHARGE, "    Equipment depreciation", "Drv", helper=True)
    fill_right(ws, R.EQUIP_CHARGE, "E",
               f"=(D{R.EQUIP_DEP_BASE}+E{R.CAPEX_EQUIP})/{inp_abs(I.EQUIP_LIFE)}")
    fy_sum(ws, R.EQUIP_CHARGE)
    label(ws, R.RD_CHARGE, "    R&D amortization", "Drv", helper=True)
    fill_right(ws, R.RD_CHARGE, "E",
               f"=(D{R.RD_DEP_BASE}+E{R.CAPEX_RD})/{inp_abs(I.RD_LIFE)}")
    fy_sum(ws, R.RD_CHARGE)

    # Hidden depreciable-base rolls (rows 208-210). Budgeted capex only — the
    # NBV roll above adds capex+manual, this one adds only the capex driver, so
    # the charge derived from it ignores manual top-ups. Mirrors actual NBV in
    # actual months (ACTUALS_DB_MAP), keeping the boundary charge neutral.
    label(ws, R.PPE_DEP_BASE, "  PP&E depreciable base (budget)", "Drv", helper=True)
    ws[f"D{R.PPE_DEP_BASE}"] = f"={inp_abs(I.PPE_OPEN)}"
    fill_right(ws, R.PPE_DEP_BASE, "E",
               f"=D{R.PPE_DEP_BASE}+E{R.CAPEX_PPE}-E{R.PPE_CHARGE}")
    fy_last(ws, R.PPE_DEP_BASE)
    label(ws, R.EQUIP_DEP_BASE, "  Equipment depreciable base (budget)", "Drv", helper=True)
    ws[f"D{R.EQUIP_DEP_BASE}"] = f"={inp_abs(I.EQUIP_OPEN)}"
    fill_right(ws, R.EQUIP_DEP_BASE, "E",
               f"=D{R.EQUIP_DEP_BASE}+E{R.CAPEX_EQUIP}-E{R.EQUIP_CHARGE}")
    fy_last(ws, R.EQUIP_DEP_BASE)
    label(ws, R.RD_DEP_BASE, "  R&D depreciable base (budget)", "Drv", helper=True)
    ws[f"D{R.RD_DEP_BASE}"] = f"={inp_abs(I.RD_OPEN)}"
    fill_right(ws, R.RD_DEP_BASE, "E",
               f"=D{R.RD_DEP_BASE}+E{R.CAPEX_RD}-E{R.RD_CHARGE}")
    fy_last(ws, R.RD_DEP_BASE)

    # Budgeted loan base for interest (rolls draw + principal drivers only).
    label(ws, R.LOAN_INT_BASE, "  Loan interest base (budget)", "Drv", helper=True)
    ws[f"D{R.LOAN_INT_BASE}"] = f"={inp_abs(I.LOAN_OPEN)}"
    fill_right(ws, R.LOAN_INT_BASE, "E",
               f"=D{R.LOAN_INT_BASE}+E{R.LOAN_DRAW}+E{R.LOAN_PRINCIPAL}")
    fy_last(ws, R.LOAN_INT_BASE)

    label(ws, R.NCA_TOTAL, "Total non-current assets", "BS", bold=True)
    ws[f"D{R.NCA_TOTAL}"] = f"=D{R.RD_NBV}+D{R.PPE_NBV}+D{R.EQUIP_NBV}"
    fill_right(ws, R.NCA_TOTAL, "E",
               f"=E{R.RD_NBV}+E{R.PPE_NBV}+E{R.EQUIP_NBV}")
    fy_last(ws, R.NCA_TOTAL)

    # ── Current assets ──
    label(ws, R.CA_HEADER, "Current assets", section_header=True)

    # Flat-carry-forward forecast formulas; in actual months `fill_right`
    # overwrites with `=Actuals!<col><row>` (these rows are in ACTUALS_DB_MAP).
    label(ws, R.BS_PREPAID, "  Prepaid expenses", "BS")
    ws[f"D{R.BS_PREPAID}"] = f"={inp_abs(I.OPEN_PREPAID)}"
    fill_right(ws, R.BS_PREPAID, "E", f"=D{R.BS_PREPAID}")
    fy_last(ws, R.BS_PREPAID)

    label(ws, R.BS_LOANS_NEG, "  Loans (other negotiable)", "BS")
    ws[f"D{R.BS_LOANS_NEG}"] = f"={inp_abs(I.OPEN_LOANS_NEG)}"
    fill_right(ws, R.BS_LOANS_NEG, "E", f"=D{R.BS_LOANS_NEG}")
    fy_last(ws, R.BS_LOANS_NEG)

    label(ws, R.BS_OTHER_RECV, "  Other receivables", "BS")
    ws[f"D{R.BS_OTHER_RECV}"] = f"={inp_abs(I.OPEN_OTHER_RECV)}"
    fill_right(ws, R.BS_OTHER_RECV, "E", f"=D{R.BS_OTHER_RECV}")
    fy_last(ws, R.BS_OTHER_RECV)

    # Inventory + hidden helpers
    label(ws, R.BS_INVENTORY, "  Inventory", "BS")
    ws[f"D{R.BS_INVENTORY}"] = f"={inp_abs(I.OPEN_INV)}"
    fill_right(ws, R.BS_INVENTORY, "E",
               f"=D{R.BS_INVENTORY}+E{R.PURCHASES_DRV}-E{R.INV_OUT_DRV}")
    fy_last(ws, R.BS_INVENTORY)
    label(ws, R.PURCHASES_DRV, "    Purchases", "Hlp", helper=True)
    fill_right(ws, R.PURCHASES_DRV, "E",
               f"=SUMPRODUCT(Inputs!E${I.QTY_BASE}:E${I.QTY_BASE + 15},"
               f"Inputs!$D${I.UNIT_COST_BASE}:$D${I.UNIT_COST_BASE + 15})")
    fy_sum(ws, R.PURCHASES_DRV)
    # FIX (a): inventory-out must equal purchases (no NRE term, which is a service cost).
    label(ws, R.INV_OUT_DRV, "    Inventory out", "Hlp", helper=True)
    fill_right(ws, R.INV_OUT_DRV, "E",
               f"=SUMPRODUCT(Inputs!E${I.QTY_BASE}:E${I.QTY_BASE + 15},"
               f"Inputs!$D${I.UNIT_COST_BASE}:$D${I.UNIT_COST_BASE + 15})")
    fy_sum(ws, R.INV_OUT_DRV)

    # Cash
    label(ws, R.BS_CASH, "  Cash & cash equivalents", "BS")
    ws[f"D{R.BS_CASH}"] = f"={inp_abs(I.OPEN_CASH)}"
    fill_right(ws, R.BS_CASH, "E", f"=E{R.CLOSE_CASH}")
    fy_last(ws, R.BS_CASH)

    # Trade receivables + hidden helpers
    label(ws, R.BS_AR, "  Trade receivables", "BS")
    ws[f"D{R.BS_AR}"] = f"={inp_abs(I.OPEN_AR)}"
    fill_right(ws, R.BS_AR, "E",
               f"=D{R.BS_AR}+E{R.AR_HLP_SALES_ADD}"
               f"+E{R.AR_HLP_COLL_SALES}+E{R.AR_HLP_COLL_DEP}")
    fy_last(ws, R.BS_AR)
    label(ws, R.AR_HLP_SALES_ADD, "    + Gross sales", "Hlp", helper=True)
    fill_right(ws, R.AR_HLP_SALES_ADD, "E",
               f"=E{R.REV_TOTAL}*(1+{inp_abs(I.VAT_RATE)})")
    fy_sum(ws, R.AR_HLP_SALES_ADD)
    label(ws, R.AR_HLP_COLL_SALES, "    − Sales collections", "Hlp", helper=True)
    fill_right(ws, R.AR_HLP_COLL_SALES, "E", f"=-E{R.CASH_IN_SALES}")
    fy_sum(ws, R.AR_HLP_COLL_SALES)
    label(ws, R.AR_HLP_COLL_DEP, "    − Deposit collections", "Hlp", helper=True)
    fill_right(ws, R.AR_HLP_COLL_DEP, "E", f"=-E{R.CASH_IN_DEPOSIT}")
    fy_sum(ws, R.AR_HLP_COLL_DEP)

    # VAT receivable (hidden, forecast-only flat carry)
    label(ws, R.BS_VAT_RECV, "  VAT receivable", "BS", helper=True)
    ws[f"D{R.BS_VAT_RECV}"] = f"={inp_abs(I.OPEN_VAT_R)}"
    fill_right(ws, R.BS_VAT_RECV, "E", f"=D{R.BS_VAT_RECV}")
    fy_last(ws, R.BS_VAT_RECV)

    # Current assets total
    label(ws, R.CA_TOTAL, "Total current assets", "BS", bold=True)
    ws[f"D{R.CA_TOTAL}"] = (
        f"=D{R.BS_PREPAID}+D{R.BS_LOANS_NEG}+D{R.BS_OTHER_RECV}"
        f"+D{R.BS_INVENTORY}+D{R.BS_CASH}+D{R.BS_AR}+D{R.BS_VAT_RECV}"
    )
    fill_right(ws, R.CA_TOTAL, "E",
               f"=E{R.BS_PREPAID}+E{R.BS_LOANS_NEG}+E{R.BS_OTHER_RECV}"
               f"+E{R.BS_INVENTORY}+E{R.BS_CASH}+E{R.BS_AR}+E{R.BS_VAT_RECV}")
    fy_last(ws, R.CA_TOTAL)

    # Total assets
    label(ws, R.TOTAL_ASSETS, "TOTAL ASSETS", "BS", bold=True)
    ws[f"D{R.TOTAL_ASSETS}"] = f"=D{R.NCA_TOTAL}+D{R.CA_TOTAL}"
    fill_right(ws, R.TOTAL_ASSETS, "E", f"=E{R.NCA_TOTAL}+E{R.CA_TOTAL}")
    fy_last(ws, R.TOTAL_ASSETS)

    # ── Equity ──
    label(ws, R.EQUITY_HEADER, "Equity", section_header=True)

    label(ws, R.SHARE_CAPITAL, "  Share capital", "BS")
    ws[f"D{R.SHARE_CAPITAL}"] = f"={inp_abs(I.OPEN_SHARE_CAP)}"
    fill_right(ws, R.SHARE_CAPITAL, "E",
               f"=D{R.SHARE_CAPITAL}+E{R.EQUITY_RAISE}")
    fy_last(ws, R.SHARE_CAPITAL)

    label(ws, R.RETAINED_EARNINGS, "  Retained earnings", "BS")
    ws[f"D{R.RETAINED_EARNINGS}"] = f"={inp_abs(I.OPEN_RE)}"
    fill_right(ws, R.RETAINED_EARNINGS, "E",
               f"=D{R.RETAINED_EARNINGS}+E{R.NI}-E{R.DIVIDENDS}")
    fy_last(ws, R.RETAINED_EARNINGS)

    label(ws, R.EQUITY_TOTAL, "Total equity", "BS", bold=True)
    ws[f"D{R.EQUITY_TOTAL}"] = f"=D{R.SHARE_CAPITAL}+D{R.RETAINED_EARNINGS}"
    fill_right(ws, R.EQUITY_TOTAL, "E",
               f"=E{R.SHARE_CAPITAL}+E{R.RETAINED_EARNINGS}")
    fy_last(ws, R.EQUITY_TOTAL)

    # ── Liabilities ──
    label(ws, R.LIAB_HEADER, "Liabilities", section_header=True)

    label(ws, R.BS_LOAN, "  Loan facility financing", "BS")
    ws[f"D{R.BS_LOAN}"] = f"={inp_abs(I.LOAN_OPEN)}"
    fill_right(ws, R.BS_LOAN, "E",
               f"=D{R.BS_LOAN}+E{R.LOAN_DRAW}+E{R.LOAN_PRINCIPAL}")
    fy_last(ws, R.BS_LOAN)

    label(ws, R.BS_PAYROLL_PAY, "  Payables to personnel", "BS")
    ws[f"D{R.BS_PAYROLL_PAY}"] = f"={inp_abs(I.OPEN_PAYROLL_PAY)}"
    fill_right(ws, R.BS_PAYROLL_PAY, "E",
               f"=D{R.BS_PAYROLL_PAY}"
               f"+(E{R.SM_PAYROLL}+E{R.GA_PAYROLL}"
               f"+E{R.RD_PAYROLL_DE}+E{R.RD_PAYROLL_RS})"
               f"-E{R.CASH_OUT_PAYROLL}")
    fy_last(ws, R.BS_PAYROLL_PAY)

    label(ws, R.BS_OTHER_CL, "  Other payables", "BS")
    ws[f"D{R.BS_OTHER_CL}"] = f"={inp_abs(I.OPEN_OTHER_CL)}"
    fill_right(ws, R.BS_OTHER_CL, "E", f"=D{R.BS_OTHER_CL}")
    fy_last(ws, R.BS_OTHER_CL)

    # Trade payables + hidden helpers
    label(ws, R.BS_AP, "  Trade payables", "BS")
    ws[f"D{R.BS_AP}"] = f"={inp_abs(I.OPEN_AP)}"
    fill_right(ws, R.BS_AP, "E",
               f"=D{R.BS_AP}+E{R.AP_HLP_PURCH_ADD}+E{R.AP_HLP_PAYMENTS}")
    fy_last(ws, R.BS_AP)
    label(ws, R.AP_HLP_PURCH_ADD, "    + Gross purchases", "Hlp", helper=True)
    fill_right(ws, R.AP_HLP_PURCH_ADD, "E",
               f"=E{R.PURCHASES_DRV}*(1+{inp_abs(I.VAT_RATE)})")
    fy_sum(ws, R.AP_HLP_PURCH_ADD)
    label(ws, R.AP_HLP_PAYMENTS, "    − Supplier payments", "Hlp", helper=True)
    fill_right(ws, R.AP_HLP_PAYMENTS, "E", f"=-E{R.CASH_OUT_SUPP}")
    fy_sum(ws, R.AP_HLP_PAYMENTS)

    # VAT payable (hidden, forecast-only)
    label(ws, R.BS_VAT_PAY, "  VAT payable", "BS", helper=True)
    ws[f"D{R.BS_VAT_PAY}"] = f"={inp_abs(I.OPEN_VAT_PAY)}"
    fill_right(ws, R.BS_VAT_PAY, "E",
               f"=D{R.BS_VAT_PAY}+E{R.VAT_HLP_OUTPUT}"
               f"-E{R.VAT_HLP_INPUT}-E{R.VAT_HLP_SETTLED}")
    fy_last(ws, R.BS_VAT_PAY)
    label(ws, R.VAT_HLP_OUTPUT, "    + Output VAT", "Hlp", helper=True)
    fill_right(ws, R.VAT_HLP_OUTPUT, "E",
               f"=E{R.REV_TOTAL}*{inp_abs(I.VAT_RATE)}")
    fy_sum(ws, R.VAT_HLP_OUTPUT)
    label(ws, R.VAT_HLP_INPUT, "    − Input VAT", "Hlp", helper=True)
    fill_right(ws, R.VAT_HLP_INPUT, "E",
               f"=E{R.PURCHASES_DRV}*{inp_abs(I.VAT_RATE)}")
    fy_sum(ws, R.VAT_HLP_INPUT)
    label(ws, R.VAT_HLP_SETTLED, "    − Settled", "Hlp", helper=True)
    fill_right(ws, R.VAT_HLP_SETTLED, "E",
               f"=MAX(0,E{R.VAT_HLP_OUTPUT}-E{R.VAT_HLP_INPUT})")
    fy_sum(ws, R.VAT_HLP_SETTLED)

    # Tax payable (hidden, forecast-only)
    label(ws, R.BS_TAX_PAY, "  Tax payable", "BS", helper=True)
    ws[f"D{R.BS_TAX_PAY}"] = f"={inp_abs(I.OPEN_TAX_PAY)}"
    fill_right(ws, R.BS_TAX_PAY, "E",
               f"=D{R.BS_TAX_PAY}+E{R.TAX_EXP}-E{R.CASH_OUT_TAX}")
    fy_last(ws, R.BS_TAX_PAY)

    # Deferred grants (hidden, forecast-only) — uses GRANT_INCOME which is
    # itself clamped to MAX(0, prior_deferred + new_cash) so this can't go negative.
    label(ws, R.BS_DEFERRED_GRANTS, "  Deferred grants", "BS", helper=True)
    ws[f"D{R.BS_DEFERRED_GRANTS}"] = f"={inp_abs(I.OPEN_DEFERRED_GRANTS)}"
    fill_right(ws, R.BS_DEFERRED_GRANTS, "E",
               f"=D{R.BS_DEFERRED_GRANTS}+E{R.GRANT_CASH}-E{R.GRANT_INCOME}")
    fy_last(ws, R.BS_DEFERRED_GRANTS)

    # Unearned revenue (hidden, forecast-only)
    label(ws, R.BS_UNEARNED_REV, "  Unearned revenue", "BS", helper=True)
    ws[f"D{R.BS_UNEARNED_REV}"] = f"={inp_abs(I.OPEN_UNEARNED_REV)}"
    fill_right(ws, R.BS_UNEARNED_REV, "E", f"=D{R.BS_UNEARNED_REV}")
    fy_last(ws, R.BS_UNEARNED_REV)

    label(ws, R.LIAB_TOTAL, "Total liabilities", "BS", bold=True)
    ws[f"D{R.LIAB_TOTAL}"] = (
        f"=D{R.BS_LOAN}+D{R.BS_PAYROLL_PAY}+D{R.BS_OTHER_CL}+D{R.BS_AP}"
        f"+D{R.BS_VAT_PAY}+D{R.BS_TAX_PAY}+D{R.BS_DEFERRED_GRANTS}+D{R.BS_UNEARNED_REV}"
    )
    fill_right(ws, R.LIAB_TOTAL, "E",
               f"=E{R.BS_LOAN}+E{R.BS_PAYROLL_PAY}+E{R.BS_OTHER_CL}+E{R.BS_AP}"
               f"+E{R.BS_VAT_PAY}+E{R.BS_TAX_PAY}+E{R.BS_DEFERRED_GRANTS}+E{R.BS_UNEARNED_REV}")
    fy_last(ws, R.LIAB_TOTAL)

    label(ws, R.LE_TOTAL, "TOTAL EQUITY AND LIABILITIES", "BS", bold=True)
    ws[f"D{R.LE_TOTAL}"] = f"=D{R.LIAB_TOTAL}+D{R.EQUITY_TOTAL}"
    fill_right(ws, R.LE_TOTAL, "E", f"=E{R.LIAB_TOTAL}+E{R.EQUITY_TOTAL}")
    fy_last(ws, R.LE_TOTAL)

    # Balance Check (tie-out at end of BS)
    label(ws, R.BAL_CHECK, "Balance check", "Chk", bold=True)
    fill_right(ws, R.BAL_CHECK, "E", f"=E{R.TOTAL_ASSETS}-E{R.LE_TOTAL}")
    ws[f"{FY_COL}{R.BAL_CHECK}"] = (
        f"=MAX(ABS({PERIOD_COLS[0]}{R.BAL_CHECK}:{PERIOD_COLS[-1]}{R.BAL_CHECK}))"
    )
    for col in PERIOD_COLS + [FY_COL]:
        ws[f"{col}{R.BAL_CHECK}"].font = CHECK_FONT

    # ─────────────────────────────────────────────────────────────
    # SECTION 2 — INCOME STATEMENT (tie-out at end)
    # ─────────────────────────────────────────────────────────────
    label(ws, R.IS_BANNER, "INCOME STATEMENT", banner=True)

    # Revenue: REV_TOTAL = REV_CE_EVALKITS + REV_OTHER (DB structure).
    # Forecast: REV_CE_EVALKITS = CE_EVALKITS hidden product line; REV_OTHER =
    # sum of every other forecast product line + NRE.
    label(ws, R.REV_TOTAL, "Sales", "IS", bold=True)
    fill_right(ws, R.REV_TOTAL, "E",
               f"=E{R.REV_CE_EVALKITS}+E{R.REV_OTHER}")
    fy_sum(ws, R.REV_TOTAL)

    label(ws, R.REV_CE_EVALKITS, "  Consumer Electronics — Eval-Kits", "IS")
    # Forecast formula: CE eval-kit qty × price (offset 9 in QTY_BASE & PRICE_BASE)
    fill_right(ws, R.REV_CE_EVALKITS, "E",
               f"={inp_period(I.QTY_BASE + 9, 'E')}*{inp_abs(I.PRICE_BASE + 9)}")
    fy_sum(ws, R.REV_CE_EVALKITS)

    label(ws, R.REV_OTHER, "  Other (sum of remaining product lines + NRE)", "IS")
    # All product detail lines below + NRE − the one CE-EvalKit row that's
    # already counted as REV_CE_EVALKITS. Simpler: sum the 15 detail rows
    # (excluding CE_EVALKITS which is REV_CE_EVALKITS itself, not duplicated).
    detail_rows = [R.FL_FARADAOX, R.FL_EVALKITS, R.FL_INTEGRATED, R.FL_SERVICES,
                   R.IIOT_FARADAOX, R.IIOT_EVALKITS, R.IIOT_INTEGRATED, R.IIOT_SERVICES,
                   R.CE_FARADAOX, R.CE_INTEGRATED, R.CE_SERVICES,
                   R.MD_FARADAOX, R.MD_EVALKITS, R.MD_INTEGRATED, R.MD_SERVICES]
    fill_right(ws, R.REV_OTHER, "E",
               "=" + "+".join(f"E{r}" for r in detail_rows) + f"+E{R.REV_NRE}")
    fy_sum(ws, R.REV_OTHER)

    # Hidden product detail (forecast-only; rows still produce qty × price formulas)
    # qty_offset maps: FL=0-3, IIoT=4-7, CE=8-11, MD=12-15
    product_detail = [
        (R.FL_FARADAOX,    0, "Food Logistics – Faraday-Ox"),
        (R.FL_EVALKITS,    1, "Food Logistics – Eval-Kits"),
        (R.FL_INTEGRATED,  2, "Food Logistics – Integrated"),
        (R.FL_SERVICES,    3, "Food Logistics – Services"),
        (R.IIOT_FARADAOX,  4, "Industrial IoT – Faraday-Ox"),
        (R.IIOT_EVALKITS,  5, "Industrial IoT – Eval-Kits"),
        (R.IIOT_INTEGRATED,6, "Industrial IoT – Integrated"),
        (R.IIOT_SERVICES,  7, "Industrial IoT – Services"),
        (R.CE_FARADAOX,    8, "Consumer Electronics – Faraday-Ox"),
        # NOTE: CE_EVALKITS (offset 9) IS REV_CE_EVALKITS row above; not duplicated.
        (R.CE_INTEGRATED, 10, "Consumer Electronics – Integrated"),
        (R.CE_SERVICES,   11, "Consumer Electronics – Services"),
        (R.MD_FARADAOX,   12, "Medical Devices – Faraday-Ox"),
        (R.MD_EVALKITS,   13, "Medical Devices – Eval-Kits"),
        (R.MD_INTEGRATED, 14, "Medical Devices – Integrated"),
        (R.MD_SERVICES,   15, "Medical Devices – Services"),
    ]
    for pf_row, qty_offset, prod_label in product_detail:
        label(ws, pf_row, f"    {prod_label}", "Hlp", helper=True)
        fill_right(ws, pf_row, "E",
                   f"={inp_period(I.QTY_BASE + qty_offset, 'E')}"
                   f"*{inp_abs(I.PRICE_BASE + qty_offset)}")
        fy_sum(ws, pf_row)

    label(ws, R.REV_NRE, "    Other (NREs)", "Hlp", helper=True)
    fill_right(ws, R.REV_NRE, "E",
               f"={inp_period(I.NRE_QTY_R, 'E')}*{inp_abs(I.NRE_PRICE)}")
    fy_sum(ws, R.REV_NRE)

    # ── Cost of Sales (DB structure: COGS / Rent / Logistics / Direct Amort / Other) ──
    label(ws, R.COS_HEADER, "Cost of Sales", section_header=True)
    label(ws, R.COGS_MATERIALS, "  COGS materials", "IS")
    # COGS materials = products only. NRE service cost lives in COS_OTHER
    # below (and is paid cash-in-month via CASH_OUT_DIRECT) so it traces to
    # the BS via a Cash decrease rather than vanishing as an IS-only hit.
    fill_right(ws, R.COGS_MATERIALS, "E",
               f"=SUMPRODUCT(Inputs!E${I.QTY_BASE}:E${I.QTY_BASE + 15},"
               f"Inputs!$D${I.UNIT_COST_BASE}:$D${I.UNIT_COST_BASE + 15})")
    fy_sum(ws, R.COGS_MATERIALS)

    label(ws, R.DIRECT_RENT, "  Rent (production)", "IS")
    fill_right(ws, R.DIRECT_RENT, "E", f"={inp_abs(I.PROD_RENT)}")
    fy_sum(ws, R.DIRECT_RENT)

    label(ws, R.DIRECT_LOG, "  Logistics", "IS")
    fill_right(ws, R.DIRECT_LOG, "E", f"={inp_abs(I.LOGISTICS)}")
    fy_sum(ws, R.DIRECT_LOG)

    label(ws, R.DIRECT_AMORT_COS, "  Direct Amortization (in CoS)", "IS")
    fill_right(ws, R.DIRECT_AMORT_COS, "E", f"={inp_period(I.DIRECT_AMORT_R, 'E')}")
    fy_sum(ws, R.DIRECT_AMORT_COS)

    label(ws, R.COS_OTHER, "  Other (CoS)", "IS")
    # NRE service cost lives here (not in COGS materials). Paid cash-in-month
    # via CASH_OUT_DIRECT — keeps BS balanced via Cash decrease.
    fill_right(ws, R.COS_OTHER, "E",
               f"={inp_period(I.COS_OTHER_R, 'E')}"
               f"+{inp_period(I.NRE_QTY_R, 'E')}*{inp_abs(I.NRE_PRICE)}*{inp_abs(I.NRE_COST_PCT)}")
    fy_sum(ws, R.COS_OTHER)

    label(ws, R.COS_TOTAL, "Total Cost of Sales", "IS", bold=True)
    fill_right(ws, R.COS_TOTAL, "E",
               f"=SUM(E{R.COGS_MATERIALS}:E{R.COS_OTHER})")
    fy_sum(ws, R.COS_TOTAL)

    label(ws, R.GROSS_PROFIT, "Gross profit", "IS", bold=True)
    fill_right(ws, R.GROSS_PROFIT, "E", f"=E{R.REV_TOTAL}-E{R.COS_TOTAL}")
    fy_sum(ws, R.GROSS_PROFIT)
    label(ws, R.GROSS_MARGIN, "  Gross margin", "IS")
    fill_right(ws, R.GROSS_MARGIN, "E",
               f"=IFERROR(E{R.GROSS_PROFIT}/E{R.REV_TOTAL},0)")
    ws[f"{FY_COL}{R.GROSS_MARGIN}"] = (
        f"=IFERROR({FY_COL}{R.GROSS_PROFIT}/{FY_COL}{R.REV_TOTAL},0)"
    )
    for col in PERIOD_COLS + [FY_COL]:
        ws[f"{col}{R.GROSS_MARGIN}"].number_format = "0.0%"

    # ── R&D ──
    label(ws, R.RD_HEADER, "R&D", section_header=True)
    rd_items = [
        (R.RD_PAYROLL_DE, "  Total Payroll – Germany",
         f"={inp_period(I.RD_PAYROLL_DE, 'E')}"),
        (R.RD_PAYROLL_RS, "  Total Payroll – Serbia",
         f"={inp_period(I.RD_PAYROLL_RS, 'E')}"),
        (R.RD_SOFTWARE, "  Software and Tools", f"={inp_abs(I.RD_SOFTWARE)}"),
        (R.RD_RENT, "  R&D Rent", f"={inp_abs(I.RD_RENT)}"),
        (R.RD_OTHER, "  R&D Sensors + Other R&D", f"={inp_period(I.OTHER_RD, 'E')}"),
    ]
    for row, lbl, formula in rd_items:
        label(ws, row, lbl, "IS")
        fill_right(ws, row, "E", formula)
        fy_sum(ws, row)
    label(ws, R.RD_SUB, "R&D subtotal", "IS", bold=True)
    fill_right(ws, R.RD_SUB, "E", f"=SUM(E{R.RD_PAYROLL_DE}:E{R.RD_OTHER})")
    fy_sum(ws, R.RD_SUB)

    # ── S&M ──
    label(ws, R.SM_HEADER, "S&M", section_header=True)
    sm_items = [
        (R.SM_PAYROLL, "  Total Payroll", f"={inp_period(I.SM_PAYROLL_TOTAL, 'E')}"),
        (R.SM_EVENTS, "  Events/Exhibitions", f"={inp_period(I.EVENTS, 'E')}"),
        (R.SM_TRAVEL, "  Travel", f"={inp_period(I.SM_TRAVEL, 'E')}"),
        (R.SM_DIGITAL, "  Digital Marketing", f"={inp_abs(I.DIGITAL_MKT)}"),
        (R.SM_OUTSOURCED, "  Outsourced Marketing", f"={inp_abs(I.OUTSOURCED_MKT)}"),
        (R.SM_CONTENT, "  Content Marketing", f"={inp_abs(I.CONTENT_MKT)}"),
        (R.SM_COMMISSIONS, "  Sales Commissions",
         f"=E{R.REV_TOTAL}*{inp_abs(I.SALES_COMMISSION_PCT)}"),
        (R.SM_OTHER, "  Other marketing expenses",
         f"=(E{R.SM_EVENTS}+E{R.SM_TRAVEL}+E{R.SM_DIGITAL}"
         f"+E{R.SM_OUTSOURCED}+E{R.SM_CONTENT}+E{R.SM_COMMISSIONS})"
         f"*{inp_abs(I.OTHER_MKT_PCT)}"),
    ]
    for row, lbl, formula in sm_items:
        label(ws, row, lbl, "IS")
        fill_right(ws, row, "E", formula)
        fy_sum(ws, row)
    label(ws, R.SM_SUB, "S&M subtotal", "IS", bold=True)
    fill_right(ws, R.SM_SUB, "E", f"=SUM(E{R.SM_PAYROLL}:E{R.SM_OTHER})")
    fy_sum(ws, R.SM_SUB)

    # ── G&A (with Insurance) ──
    label(ws, R.GA_HEADER, "G&A", section_header=True)
    ga_items = [
        (R.GA_PAYROLL, "  Total Payroll", f"={inp_period(I.GA_PAYROLL_TOTAL, 'E')}"),
        (R.GA_OFFICE, "  Office Expenses", f"={inp_abs(I.OFFICE_EXP)}"),
        (R.GA_TRAVEL, "  Travel and Representative", f"={inp_abs(I.GA_TRAVEL)}"),
        (R.GA_SOFTWARE, "  Software and Tools", f"={inp_abs(I.GA_SOFTWARE)}"),
        (R.GA_TEAMDEV, "  Team Developments", f"={inp_abs(I.TEAM_DEV)}"),
        (R.GA_INSURANCE, "  Insurance", f"={inp_abs(I.GA_INSURANCE)}"),
        (R.GA_LEGAL, "  Legal", f"={inp_abs(I.LEGAL)}"),
        (R.GA_ACCT, "  Accounting", f"={inp_abs(I.ACCT)}"),
        (R.GA_CONSULT, "  Other/Consulting Services", f"={inp_period(I.OTHER_CONSULT_R, 'E')}"),
        (R.GA_MISC, "  Miscellaneous expenses",
         f"=(E{R.GA_OFFICE}+E{R.GA_LEGAL}+E{R.GA_ACCT}+E{R.GA_CONSULT})"
         f"*{inp_abs(I.MISC_PCT)}"),
    ]
    for row, lbl, formula in ga_items:
        label(ws, row, lbl, "IS")
        fill_right(ws, row, "E", formula)
        fy_sum(ws, row)
    label(ws, R.GA_SUB, "G&A subtotal", "IS", bold=True)
    fill_right(ws, R.GA_SUB, "E", f"=SUM(E{R.GA_PAYROLL}:E{R.GA_MISC})")
    fy_sum(ws, R.GA_SUB)

    label(ws, R.OPEX_TOTAL, "Operating expenses", "IS", bold=True)
    fill_right(ws, R.OPEX_TOTAL, "E",
               f"=E{R.RD_SUB}+E{R.SM_SUB}+E{R.GA_SUB}")
    fy_sum(ws, R.OPEX_TOTAL)

    label(ws, R.EBITDA, "EBITDA", "IS", bold=True)
    fill_right(ws, R.EBITDA, "E", f"=E{R.GROSS_PROFIT}-E{R.OPEX_TOTAL}")
    fy_sum(ws, R.EBITDA)
    label(ws, R.EBITDA_MARGIN, "  EBITDA margin", "IS")
    fill_right(ws, R.EBITDA_MARGIN, "E",
               f"=IFERROR(E{R.EBITDA}/E{R.REV_TOTAL},0)")
    ws[f"{FY_COL}{R.EBITDA_MARGIN}"] = (
        f"=IFERROR({FY_COL}{R.EBITDA}/{FY_COL}{R.REV_TOTAL},0)"
    )
    for col in PERIOD_COLS + [FY_COL]:
        ws[f"{col}{R.EBITDA_MARGIN}"].number_format = "0.0%"

    # ── Direct Amort adjustment + Finance income / costs (DB ord 28-30) ──
    label(ws, R.DIRECT_AMORT_ADJ, "Direct Amort. adjustment", "IS")
    fill_right(ws, R.DIRECT_AMORT_ADJ, "E", f"={inp_period(I.DIRECT_AMORT_ADJ_R, 'E')}")
    fy_sum(ws, R.DIRECT_AMORT_ADJ)

    label(ws, R.FINANCE_INCOME, "Finance income", "IS")
    fill_right(ws, R.FINANCE_INCOME, "E", f"={inp_period(I.FINANCE_INCOME_R, 'E')}")
    fy_sum(ws, R.FINANCE_INCOME)

    # v3: interest is charged on the budgeted LOAN_INT_BASE (not the live loan
    # line), so a manual +X loan draw doesn't accrue modelled interest.
    label(ws, R.FINANCE_COSTS, "Finance costs", "IS")
    fill_right(ws, R.FINANCE_COSTS, "E",
               f"=D{R.LOAN_INT_BASE}*{inp_abs(I.LOAN_RATE)}/12")
    fy_sum(ws, R.FINANCE_COSTS)

    # ── D&A (DB lines + hidden forecast splits) ──
    # Forecast policy: DA_DIRECT = PPE_CHARGE + EQUIP_CHARGE (tangible),
    # DA_INDIRECT = RD_CHARGE (intangible/R&D amortisation). This matches
    # the DB convention where Direct = depreciation on tangible production
    # assets and Indirect = amortisation of capitalised R&D / overhead.
    label(ws, R.DA_HEADER, "Depreciation & amortization", section_header=True)
    label(ws, R.DA_DIRECT, "  Direct Amortization (D&A)", "IS")
    fill_right(ws, R.DA_DIRECT, "E", f"=E{R.PPE_CHARGE}+E{R.EQUIP_CHARGE}")
    fy_sum(ws, R.DA_DIRECT)
    label(ws, R.DA_INDIRECT, "  Indirect Amortization (D&A)", "IS")
    fill_right(ws, R.DA_INDIRECT, "E", f"=E{R.RD_CHARGE}")
    fy_sum(ws, R.DA_INDIRECT)
    # Hidden legacy splits (kept so downstream tie-out helpers continue to work).
    label(ws, R.DA_PPE, "    PP&E (legacy split)", "Hlp", helper=True)
    fill_right(ws, R.DA_PPE, "E", f"=E{R.PPE_CHARGE}")
    fy_sum(ws, R.DA_PPE)
    label(ws, R.DA_EQUIP, "    Equipment (legacy split)", "Hlp", helper=True)
    fill_right(ws, R.DA_EQUIP, "E", f"=E{R.EQUIP_CHARGE}")
    fy_sum(ws, R.DA_EQUIP)
    label(ws, R.DA_RD, "    R&D (legacy split)", "Hlp", helper=True)
    fill_right(ws, R.DA_RD, "E", f"=E{R.RD_CHARGE}")
    fy_sum(ws, R.DA_RD)
    label(ws, R.DA_TOTAL, "Total D&A", "IS", bold=True)
    fill_right(ws, R.DA_TOTAL, "E", f"=E{R.DA_DIRECT}+E{R.DA_INDIRECT}")
    fy_sum(ws, R.DA_TOTAL)

    label(ws, R.EBIT, "EBIT", "IS", bold=True)
    fill_right(ws, R.EBIT, "E", f"=E{R.EBITDA}-E{R.DA_TOTAL}")
    fy_sum(ws, R.EBIT)

    # FIX (c): clamp grant income recognition so it can't drive deferred balance negative.
    # GRANT_INCOME = MIN(PROFIT_RECOG + EIC_RECOG, MAX(0, prior_deferred + new_cash))
    label(ws, R.GRANT_INCOME, "Grant financing", "IS")
    for i, col in enumerate(PERIOD_COLS):
        prev = _prev_col(col)
        ws[f"{col}{R.GRANT_INCOME}"] = (
            f"=MIN({inp_period(I.PROFIT_RECOG_R, col)}+{inp_period(I.EIC_RECOG_R, col)},"
            f"MAX(0,{prev}{R.BS_DEFERRED_GRANTS}+{col}{R.GRANT_CASH}))"
        )
    fy_sum(ws, R.GRANT_INCOME)

    label(ws, R.PRETAX, "Profit / (loss) before income tax", "IS", bold=True)
    fill_right(ws, R.PRETAX, "E",
               f"=E{R.EBIT}+E{R.DIRECT_AMORT_ADJ}+E{R.FINANCE_INCOME}-E{R.FINANCE_COSTS}+E{R.GRANT_INCOME}")
    fy_sum(ws, R.PRETAX)

    label(ws, R.TAX_EXP, "Income tax (expense)", "IS")
    fill_right(ws, R.TAX_EXP, "E",
               f"=MAX(0,E{R.PRETAX})*{inp_abs(I.TAX_RATE)}")
    fy_sum(ws, R.TAX_EXP)

    label(ws, R.NI, "Profit / (loss) for the period", "IS", bold=True)
    fill_right(ws, R.NI, "E", f"=E{R.PRETAX}-E{R.TAX_EXP}")
    fy_sum(ws, R.NI)

    # NI vs ΔRE tie-out (at end of IS)
    label(ws, R.NI_RE_TIE, "NI vs ΔRE", "Chk", bold=True)
    fill_right(ws, R.NI_RE_TIE, "E",
               f"=E{R.NI}-(E{R.RETAINED_EARNINGS}-D{R.RETAINED_EARNINGS})")
    ws[f"{FY_COL}{R.NI_RE_TIE}"] = (
        f"=MAX(ABS({PERIOD_COLS[0]}{R.NI_RE_TIE}:{PERIOD_COLS[-1]}{R.NI_RE_TIE}))"
    )
    for col in PERIOD_COLS + [FY_COL]:
        ws[f"{col}{R.NI_RE_TIE}"].font = CHECK_FONT

    # ─────────────────────────────────────────────────────────────
    # SECTION 3 — CASH FLOW (Direct method)
    # ─────────────────────────────────────────────────────────────
    label(ws, R.CF_BANNER, "CASH FLOW", banner=True)

    # ── Cash inflows ──
    label(ws, R.CASH_IN_TOTAL, "Cash received from customers", "CF", bold=True)
    fill_right(ws, R.CASH_IN_TOTAL, "E",
               f"=E{R.CASH_IN_SALES}+E{R.CASH_IN_DEPOSIT}")
    fy_sum(ws, R.CASH_IN_TOTAL)

    label(ws, R.CASH_IN_SALES, "  Sales collections", "Hlp", helper=True)
    ws[f"E{R.CASH_IN_SALES}"] = f"=D{R.BS_AR}"
    for i in range(1, 12):
        prev = PERIOD_COLS[i - 1]
        curr = PERIOD_COLS[i]
        ws[f"{curr}{R.CASH_IN_SALES}"] = (
            f"={prev}{R.REV_TOTAL}*(1+{inp_abs(I.VAT_RATE)})"
        )
    fy_sum(ws, R.CASH_IN_SALES)

    label(ws, R.CASH_IN_DEPOSIT, "  Service deposits", "Hlp", helper=True)
    fill_right(ws, R.CASH_IN_DEPOSIT, "E",
               f"=(E{R.FL_SERVICES}+E{R.IIOT_SERVICES}+E{R.CE_SERVICES}+E{R.MD_SERVICES})"
               f"*{inp_abs(I.DEPOSIT_PCT)}")
    fy_sum(ws, R.CASH_IN_DEPOSIT)

    # ── Cash outflows (operating) ──
    label(ws, R.CASH_OUT_SUPP, "Cash paid to suppliers", "CF")
    ws[f"E{R.CASH_OUT_SUPP}"] = f"=D{R.BS_AP}"
    for i in range(1, 12):
        prev = PERIOD_COLS[i - 1]
        curr = PERIOD_COLS[i]
        ws[f"{curr}{R.CASH_OUT_SUPP}"] = (
            f"={prev}{R.PURCHASES_DRV}*(1+{inp_abs(I.VAT_RATE)})"
        )
    fy_sum(ws, R.CASH_OUT_SUPP)

    label(ws, R.CASH_OUT_PAYROLL, "Payment for personnel", "CF")
    fill_right(ws, R.CASH_OUT_PAYROLL, "E",
               f"=(E{R.SM_PAYROLL}+E{R.GA_PAYROLL}"
               f"+E{R.RD_PAYROLL_DE}+E{R.RD_PAYROLL_RS})"
               f"*(1-{inp_abs(I.PAYROLL_ACCRUAL_PCT)})")
    fy_sum(ws, R.CASH_OUT_PAYROLL)

    label(ws, R.CASH_OUT_TAX, "Recovery/(repayment) of taxes", "CF")
    # Stored as cash OUT (positive value = repayment / negative = recovery).
    # Forecast: tax payment ties to TAX_EXP * lag flag.
    for col in PERIOD_COLS:
        ws[f"{col}{R.CASH_OUT_TAX}"] = (
            f"=IF({inp_abs(I.TAX_LAG_MONTHS)}>=12,0,{col}{R.TAX_EXP})"
        )
    fy_sum(ws, R.CASH_OUT_TAX)

    # CASH_OUT_OTHER_OP catches all non-payroll, non-supplier, non-tax operating
    # cash items: direct production costs (rent + logistics) + non-payroll opex
    # (office, software, events, etc.) + VAT settlement. In actual months this
    # is overwritten by the DB "Other (Op)" value via `fill_right`.
    label(ws, R.CASH_OUT_OTHER_OP, "Other (Op)", "CF")
    fill_right(ws, R.CASH_OUT_OTHER_OP, "E",
               f"=E{R.CASH_OUT_DIRECT}+E{R.CASH_OUT_OPEX}+E{R.CASH_OUT_VAT}")
    fy_sum(ws, R.CASH_OUT_OTHER_OP)

    label(ws, R.CASH_OUT_OPEX, "  Operating expenses (helper)", "Hlp", helper=True)
    fill_right(ws, R.CASH_OUT_OPEX, "E",
               f"=(E{R.SM_SUB}+E{R.GA_SUB}+E{R.RD_SUB})"
               f"-(E{R.SM_PAYROLL}+E{R.GA_PAYROLL}+E{R.RD_PAYROLL_DE}+E{R.RD_PAYROLL_RS})")
    fy_sum(ws, R.CASH_OUT_OPEX)

    label(ws, R.CASH_OUT_DIRECT, "  Direct costs (helper)", "Hlp", helper=True)
    # Includes Rent, Logistics, and COS_OTHER (NRE service cost) — all paid
    # cash-in-month. Routing COS_OTHER through here closes the NRE-cost
    # balance leak (cost flows through IS NI → RE, matched by Cash decrease).
    fill_right(ws, R.CASH_OUT_DIRECT, "E",
               f"=E{R.DIRECT_RENT}+E{R.DIRECT_LOG}+E{R.COS_OTHER}")
    fy_sum(ws, R.CASH_OUT_DIRECT)

    label(ws, R.CASH_OUT_VAT, "  VAT settlement (helper)", "Hlp", helper=True)
    fill_right(ws, R.CASH_OUT_VAT, "E", f"=E{R.VAT_HLP_SETTLED}")
    fy_sum(ws, R.CASH_OUT_VAT)

    # Operating CF net (DB: Cash IN − Cash paid to suppliers − Payment for personnel
    # − Recovery/repayment of taxes − Other Op). Recovery/repayment is positive = repaid;
    # we subtract it as a cash outflow.
    # v3: the visible operating CF is the INDIRECT subtotal (NI + non-cash + ΔWC).
    # The direct cash-in/out rows above remain as the working-capital ENGINE that
    # drives the BS rolls (AR/AP/payroll/tax/VAT), but are no longer the headline.
    label(ws, R.OP_CF, "Cash Flow from Operating Activities", "CF", bold=True)
    fill_right(ws, R.OP_CF, "E", f"=E{R.TIE_OP_CF}")
    fy_sum(ws, R.OP_CF)

    # ── Investing ──
    label(ws, R.CAPEX_TOTAL_DB, "CAPEX (combined)", "CF")
    # Forecast: sum of hidden PPE + Equipment splits. Actuals: pulled from DB.
    fill_right(ws, R.CAPEX_TOTAL_DB, "E", f"=E{R.CAPEX_PPE}+E{R.CAPEX_EQUIP}")
    fy_sum(ws, R.CAPEX_TOTAL_DB)
    label(ws, R.CAPEX_RD, "R&D capitalized", "CF")
    fill_right(ws, R.CAPEX_RD, "E", f"={inp_period(I.RD_CAPEX_R, 'E')}")
    fy_sum(ws, R.CAPEX_RD)
    label(ws, R.CAPEX_OTHER_INV, "Other (Inv)", "CF")
    fill_right(ws, R.CAPEX_OTHER_INV, "E", "=0")
    fy_sum(ws, R.CAPEX_OTHER_INV)
    label(ws, R.CAPEX_PPE, "  CAPEX PP&E (hidden split)", "Hlp", helper=True)
    fill_right(ws, R.CAPEX_PPE, "E", f"={inp_period(I.PPE_CAPEX_R, 'E')}")
    fy_sum(ws, R.CAPEX_PPE)
    label(ws, R.CAPEX_EQUIP, "  CAPEX Equipment (hidden split)", "Hlp", helper=True)
    fill_right(ws, R.CAPEX_EQUIP, "E", f"={inp_period(I.EQUIP_CAPEX_R, 'E')}")
    fy_sum(ws, R.CAPEX_EQUIP)

    label(ws, R.INV_CF, "Cash Flow from Investing Activities", "CF", bold=True)
    fill_right(ws, R.INV_CF, "E", f"=E{R.TIE_INV_CF}")
    fy_sum(ws, R.INV_CF)

    # ── Financing ──
    label(ws, R.EQUITY_RAISE, "Capital Increase", "CF")
    for col in PERIOD_COLS:
        ws[f"{col}{R.EQUITY_RAISE}"] = (
            f"=IF(AND(YEAR({col}{R.DATES})=YEAR({inp_abs(I.EQUITY_TR1_DATE)}),"
            f"MONTH({col}{R.DATES})=MONTH({inp_abs(I.EQUITY_TR1_DATE)})),"
            f"{inp_abs(I.EQUITY_TR1_AMT)},0)"
        )
    fy_sum(ws, R.EQUITY_RAISE)

    label(ws, R.GRANT_CASH, "Grants received", "CF")
    fill_right(ws, R.GRANT_CASH, "E",
               f"={inp_period(I.PROFIT_RECV_R, 'E')}+{inp_period(I.EIC_RECV_R, 'E')}")
    fy_sum(ws, R.GRANT_CASH)

    label(ws, R.LOAN_FAC_DB, "Loan Facility Financing", "CF")
    # Forecast: combined draw + principal repayments.
    fill_right(ws, R.LOAN_FAC_DB, "E", f"=E{R.LOAN_DRAW}+E{R.LOAN_PRINCIPAL}")
    fy_sum(ws, R.LOAN_FAC_DB)

    label(ws, R.OTHER_FIN, "Other payments/proceeds (Fin)", "CF")
    fill_right(ws, R.OTHER_FIN, "E", "=0")
    fy_sum(ws, R.OTHER_FIN)

    label(ws, R.LOAN_DRAW, "  Loan drawdown (hidden)", "Hlp", helper=True)
    fill_right(ws, R.LOAN_DRAW, "E", "=0")
    fy_sum(ws, R.LOAN_DRAW)
    label(ws, R.LOAN_PRINCIPAL, "  Loan principal repaid (hidden)", "Hlp", helper=True)
    fill_right(ws, R.LOAN_PRINCIPAL, "E", f"=-{inp_abs(I.LOAN_PMT)}")
    fy_sum(ws, R.LOAN_PRINCIPAL)
    label(ws, R.INTEREST_PAID, "  Interest paid (hidden)", "Hlp", helper=True)
    fill_right(ws, R.INTEREST_PAID, "E", f"=-E{R.FINANCE_COSTS}")
    fy_sum(ws, R.INTEREST_PAID)
    label(ws, R.DIVIDENDS, "  Dividends paid (hidden)", "Hlp", helper=True)
    fill_right(ws, R.DIVIDENDS, "E", "=0")
    fy_sum(ws, R.DIVIDENDS)

    label(ws, R.FIN_CF, "Cash Flow from Financing Activities", "CF", bold=True)
    fill_right(ws, R.FIN_CF, "E", f"=E{R.TIE_FIN_CF}")
    fy_sum(ws, R.FIN_CF)

    # ── Net change & closing cash ──
    label(ws, R.NET_CHG_CASH, "Net Δ Cash for the Period", "CF", bold=True)
    fill_right(ws, R.NET_CHG_CASH, "E",
               f"=E{R.OP_CF}+E{R.INV_CF}+E{R.FIN_CF}")
    fy_sum(ws, R.NET_CHG_CASH)

    # FIX (b): OPEN_CASH = previous column's BS_CASH (which itself = CLOSE_CASH or
    # Actuals!BS_CASH). This wires the BS cash chain to CF cleanly in forecast AND
    # makes the first forecast month pick up the last actual BS cash.
    label(ws, R.OPEN_CASH, "Beginning Cash Balance", "CF")
    ws[f"E{R.OPEN_CASH}"] = f"=D{R.BS_CASH}"
    for i in range(1, 12):
        prev = PERIOD_COLS[i - 1]
        curr = PERIOD_COLS[i]
        ws[f"{curr}{R.OPEN_CASH}"] = f"={prev}{R.BS_CASH}"
    fy_last(ws, R.OPEN_CASH)

    label(ws, R.CLOSE_CASH, "Ending Cash Balance", "CF", bold=True)
    fill_right(ws, R.CLOSE_CASH, "E",
               f"=E{R.OPEN_CASH}+E{R.NET_CHG_CASH}")
    fy_last(ws, R.CLOSE_CASH)

    # v3: check the derived CF's ending cash against the BS cash. In forecast
    # months BS Cash = the plug (= this row's CLOSE_CASH) so the check is 0; in
    # actual months BS Cash is the pulled actual, so a non-zero value flags any
    # gap between the derived CF and reported actual cash (client model's r100).
    label(ws, R.CF_CASH_CHECK, "Check vs actual cash", "Chk", bold=True)
    fill_right(ws, R.CF_CASH_CHECK, "E", f"=E{R.CLOSE_CASH}-E{R.BS_CASH}")
    ws[f"{FY_COL}{R.CF_CASH_CHECK}"] = (
        f"=MAX(ABS({PERIOD_COLS[0]}{R.CF_CASH_CHECK}:{PERIOD_COLS[-1]}{R.CF_CASH_CHECK}))"
    )
    for col in PERIOD_COLS + [FY_COL]:
        ws[f"{col}{R.CF_CASH_CHECK}"].font = CHECK_FONT

    # CF tie-out (direct vs indirect) — emitted just before the indirect block
    label(ws, R.CF_TIE, "CF tie (direct − indirect)", "Chk", bold=True)
    fill_right(ws, R.CF_TIE, "E", f"=E{R.NET_CHG_CASH}-E{R.TIE_NET_CHG}")
    ws[f"{FY_COL}{R.CF_TIE}"] = (
        f"=MAX(ABS({PERIOD_COLS[0]}{R.CF_TIE}:{PERIOD_COLS[-1]}{R.CF_TIE}))"
    )
    for col in PERIOD_COLS + [FY_COL]:
        ws[f"{col}{R.CF_TIE}"].font = CHECK_FONT

    # ─────────────────────────────────────────────────────────────
    # SECTION 4 — INDIRECT CF reconciliation (collapsed by default)
    # ─────────────────────────────────────────────────────────────
    label(ws, R.TIE_BANNER, "CASH FLOW — Operating detail (indirect)", banner=True)

    # v3: start the indirect CF from Δ Retained earnings (+ dividends), not the
    # IS's computed NI. In forecast RE = prevRE + NI − div, so ΔRE + div ≡ NI
    # (neutral); in ACTUAL months RE pulls actuals, so this is the *actual* net
    # income — making the derived CF reconcile to actual cash (check row).
    label(ws, R.TIE_NI, "  Net income", "CF")
    fill_right(ws, R.TIE_NI, "E",
               f"=(E{R.RETAINED_EARNINGS}-D{R.RETAINED_EARNINGS})+E{R.DIVIDENDS}")
    fy_sum(ws, R.TIE_NI)
    label(ws, R.TIE_DA, "  + Depreciation & amortization", "CF")
    fill_right(ws, R.TIE_DA, "E", f"=E{R.DA_TOTAL}")
    fy_sum(ws, R.TIE_DA)
    label(ws, R.TIE_GRANT_NONCASH, "  − Grant income (non-cash)", "Chk")
    fill_right(ws, R.TIE_GRANT_NONCASH, "E", f"=-E{R.GRANT_INCOME}")
    fy_sum(ws, R.TIE_GRANT_NONCASH)
    label(ws, R.TIE_INT_OFFSET, "  + Finance costs (non-cash)", "CF")
    fill_right(ws, R.TIE_INT_OFFSET, "E", f"=E{R.FINANCE_COSTS}")
    fy_sum(ws, R.TIE_INT_OFFSET)
    label(ws, R.TIE_NONCASH_SUB, "  Subtotal: NI + non-cash", "Chk", bold=True)
    fill_right(ws, R.TIE_NONCASH_SUB, "E",
               f"=E{R.TIE_NI}+E{R.TIE_DA}+E{R.TIE_GRANT_NONCASH}+E{R.TIE_INT_OFFSET}")
    fy_sum(ws, R.TIE_NONCASH_SUB)

    label(ws, R.TIE_DAR, "  − ΔTrade receivables", "Chk")
    fill_right(ws, R.TIE_DAR, "E", f"=-(E{R.BS_AR}-D{R.BS_AR})")
    fy_sum(ws, R.TIE_DAR)
    label(ws, R.TIE_DINV, "  − ΔInventory", "Chk")
    fill_right(ws, R.TIE_DINV, "E", f"=-(E{R.BS_INVENTORY}-D{R.BS_INVENTORY})")
    fy_sum(ws, R.TIE_DINV)
    label(ws, R.TIE_DAP, "  + ΔTrade payables", "Chk")
    fill_right(ws, R.TIE_DAP, "E", f"=+(E{R.BS_AP}-D{R.BS_AP})")
    fy_sum(ws, R.TIE_DAP)
    label(ws, R.TIE_DPAYPAY, "  + ΔPayables to personnel", "Chk")
    fill_right(ws, R.TIE_DPAYPAY, "E",
               f"=+(E{R.BS_PAYROLL_PAY}-D{R.BS_PAYROLL_PAY})")
    fy_sum(ws, R.TIE_DPAYPAY)
    label(ws, R.TIE_DTAXPAY, "  + ΔTax payable", "Chk")
    fill_right(ws, R.TIE_DTAXPAY, "E", f"=+(E{R.BS_TAX_PAY}-D{R.BS_TAX_PAY})")
    fy_sum(ws, R.TIE_DTAXPAY)
    label(ws, R.TIE_DVATPAY, "  + ΔVAT payable", "Chk")
    fill_right(ws, R.TIE_DVATPAY, "E", f"=+(E{R.BS_VAT_PAY}-D{R.BS_VAT_PAY})")
    fy_sum(ws, R.TIE_DVATPAY)
    label(ws, R.TIE_DVATR, "  − ΔVAT receivable", "Chk")
    fill_right(ws, R.TIE_DVATR, "E", f"=-(E{R.BS_VAT_RECV}-D{R.BS_VAT_RECV})")
    fy_sum(ws, R.TIE_DVATR)
    # Catch-all ΔWC for BS leaves without their own line — assets negative,
    # liabilities positive. Flat (zero) in the budget; non-zero only if the user
    # manually edits one of these lines, in which case it keeps the BS balanced.
    label(ws, R.CF_DWC_OTHER, "  ± Δ Other working capital", "Chk")
    fill_right(ws, R.CF_DWC_OTHER, "E",
               f"=-(E{R.BS_PREPAID}-D{R.BS_PREPAID})"
               f"-(E{R.BS_LOANS_NEG}-D{R.BS_LOANS_NEG})"
               f"-(E{R.BS_OTHER_RECV}-D{R.BS_OTHER_RECV})"
               f"+(E{R.BS_OTHER_CL}-D{R.BS_OTHER_CL})"
               f"+(E{R.BS_UNEARNED_REV}-D{R.BS_UNEARNED_REV})")
    fy_sum(ws, R.CF_DWC_OTHER)
    label(ws, R.TIE_WC_SUB, "  Subtotal: working capital", "Chk", bold=True)
    fill_right(ws, R.TIE_WC_SUB,
               "E", f"=SUM(E{R.TIE_DAR}:E{R.TIE_DVATR})+E{R.CF_DWC_OTHER}")
    fy_sum(ws, R.TIE_WC_SUB)

    label(ws, R.TIE_OP_CF, "  Cash Flow from Operating Activities", "CF", bold=True)
    fill_right(ws, R.TIE_OP_CF, "E", f"=E{R.TIE_NONCASH_SUB}+E{R.TIE_WC_SUB}")
    fy_sum(ws, R.TIE_OP_CF)

    # v3: investing derives from the BS NBV delta (+ depreciation charge), so a
    # manual +X on a fixed-asset line flows straight through. Algebraically equals
    # −capex when unedited (NBV = prev + capex − charge).
    label(ws, R.TIE_CAPEX_PPE, "  − CAPEX PP&E", "Chk")
    fill_right(ws, R.TIE_CAPEX_PPE, "E",
               f"=-(E{R.PPE_NBV}-D{R.PPE_NBV}+E{R.PPE_CHARGE})")
    fy_sum(ws, R.TIE_CAPEX_PPE)
    label(ws, R.TIE_CAPEX_EQ, "  − CAPEX Equipment", "Chk")
    fill_right(ws, R.TIE_CAPEX_EQ, "E",
               f"=-(E{R.EQUIP_NBV}-D{R.EQUIP_NBV}+E{R.EQUIP_CHARGE})")
    fy_sum(ws, R.TIE_CAPEX_EQ)
    label(ws, R.TIE_CAPEX_RD, "  − R&D capitalized", "Chk")
    fill_right(ws, R.TIE_CAPEX_RD, "E",
               f"=-(E{R.RD_NBV}-D{R.RD_NBV}+E{R.RD_CHARGE})")
    fy_sum(ws, R.TIE_CAPEX_RD)
    label(ws, R.TIE_INV_CF, "  Investing CF, indirect", "Chk", bold=True)
    fill_right(ws, R.TIE_INV_CF, "E",
               f"=SUM(E{R.TIE_CAPEX_PPE}:E{R.TIE_CAPEX_RD})")
    fy_sum(ws, R.TIE_INV_CF)

    # v3: financing derives from the BS deltas (share capital, loan balance), so
    # manual +X edits to those BS lines flow through. ΔShareCapital ≡ EQUITY_RAISE,
    # ΔLoan ≡ draw + principal when unedited.
    label(ws, R.TIE_EQUITY, "  Δ Share capital", "Chk")
    fill_right(ws, R.TIE_EQUITY, "E",
               f"=+(E{R.SHARE_CAPITAL}-D{R.SHARE_CAPITAL})")
    fy_sum(ws, R.TIE_EQUITY)
    label(ws, R.TIE_LOAN_DRAW, "  Δ Loan facility", "Chk")
    fill_right(ws, R.TIE_LOAN_DRAW, "E", f"=+(E{R.BS_LOAN}-D{R.BS_LOAN})")
    fy_sum(ws, R.TIE_LOAN_DRAW)
    label(ws, R.TIE_LOAN_PRINC, "  (loan principal — folded into Δ Loan)", "Chk")
    fill_right(ws, R.TIE_LOAN_PRINC, "E", "=0")
    fy_sum(ws, R.TIE_LOAN_PRINC)
    label(ws, R.TIE_INT_PAID, "  Interest paid", "Chk")
    fill_right(ws, R.TIE_INT_PAID, "E", f"=E{R.INTEREST_PAID}")
    fy_sum(ws, R.TIE_INT_PAID)
    label(ws, R.TIE_GRANT_CASH, "  Grants received", "Chk")
    fill_right(ws, R.TIE_GRANT_CASH, "E", f"=E{R.GRANT_CASH}")
    fy_sum(ws, R.TIE_GRANT_CASH)
    label(ws, R.TIE_DIV, "  Dividends paid", "Chk")
    fill_right(ws, R.TIE_DIV, "E", f"=E{R.DIVIDENDS}")
    fy_sum(ws, R.TIE_DIV)
    label(ws, R.TIE_FIN_CF, "  Financing CF, indirect", "Chk", bold=True)
    fill_right(ws, R.TIE_FIN_CF, "E",
               f"=SUM(E{R.TIE_EQUITY}:E{R.TIE_DIV})")
    fy_sum(ws, R.TIE_FIN_CF)

    label(ws, R.TIE_NET_CHG, "Net Δ Cash, indirect", "Chk", bold=True)
    fill_right(ws, R.TIE_NET_CHG, "E",
               f"=E{R.TIE_OP_CF}+E{R.TIE_INV_CF}+E{R.TIE_FIN_CF}")
    fy_sum(ws, R.TIE_NET_CHG)

    # ── Final error count tie-out ──
    label(ws, R.ERR_COUNT, "Error cells", "Chk", bold=True)
    for col in PERIOD_COLS:
        ws[f"{col}{R.ERR_COUNT}"] = (
            f"=SUMPRODUCT(--ISERROR({col}{R.BS_BANNER}:{col}{R.TIE_NET_CHG}))"
        )
    fy_sum(ws, R.ERR_COUNT)
    for col in PERIOD_COLS + [FY_COL]:
        ws[f"{col}{R.ERR_COUNT}"].font = CHECK_FONT

    # ── Styling, outline groups, conditional formatting ──
    apply_source_styling(ws)
    setup_outline_groups(ws)

    # Conditional formatting: tint Actual columns subtle gray
    actual_fill = PatternFill("solid", fgColor="EFEFEF")
    for col in PERIOD_COLS:
        rng = f"{col}4:{col}{R.ERR_COUNT}"
        ws.conditional_formatting.add(
            rng,
            FormulaRule(formula=[f'${col}$3="Actual"'], fill=actual_fill, stopIfTrue=False)
        )

    ws.freeze_panes = "E4"


# ──────────────────────────────────────────────────────────────────────────────
# Source-style formatting (v2.1.1 — lighter palette)
# ──────────────────────────────────────────────────────────────────────────────

# Row classification — what gets which visual treatment
KEY_TOTAL_ROWS = {
    R.REV_TOTAL, R.GROSS_PROFIT, R.EBITDA, R.EBIT, R.PRETAX, R.NI,
    R.COS_TOTAL, R.OPEX_TOTAL, R.DA_TOTAL,
    R.CASH_IN_TOTAL,
    R.OP_CF, R.INV_CF, R.FIN_CF, R.NET_CHG_CASH, R.CLOSE_CASH,
    R.NCA_TOTAL, R.CA_TOTAL, R.TOTAL_ASSETS,
    R.LIAB_TOTAL, R.EQUITY_TOTAL, R.LE_TOTAL,
    R.TIE_NONCASH_SUB, R.TIE_WC_SUB, R.TIE_OP_CF, R.TIE_INV_CF,
    R.TIE_FIN_CF, R.TIE_NET_CHG,
}
SUBTOTAL_ROWS = {
    R.REV_CE_EVALKITS, R.REV_OTHER,
    R.SM_SUB, R.GA_SUB, R.RD_SUB,
    R.DA_DIRECT, R.DA_INDIRECT,
}
SECTION_HEADER_ROWS = {
    R.NCA_HEADER, R.CA_HEADER, R.EQUITY_HEADER, R.LIAB_HEADER,
    R.COS_HEADER, R.RD_HEADER, R.SM_HEADER, R.GA_HEADER, R.DA_HEADER,
}
BANNER_ROWS = {R.IS_BANNER, R.CF_BANNER, R.BS_BANNER, R.TIE_BANNER}
MARGIN_ROWS = {R.GROSS_MARGIN, R.EBITDA_MARGIN}
CHECK_ROWS = {R.BAL_CHECK, R.CF_TIE, R.ERR_COUNT, R.NI_RE_TIE}


def apply_source_styling(ws) -> None:
    """v2.1.1 styling — light banner fill, borders for separation, no heavy navy."""
    all_data_cols = [OPEN_COL] + PERIOD_COLS + [FY_COL]
    max_row = R.LAST_ROW

    # 1. Reset font to Century Gothic for everything, preserving bold/italic/color.
    for r in range(1, max_row + 1):
        for col in [LABEL_COL, CODE_COL] + all_data_cols:
            cell = ws[f"{col}{r}"]
            f = cell.font
            new_color = f.color if (f.color and f.color.value not in (None, "00000000")) else None
            cell.font = Font(name=FONT_NAME, size=10,
                             bold=f.bold or False,
                             italic=f.italic or False,
                             color=new_color)

    # 2. Number formats — accounting on data cells; date on row 2; pct on margins
    for r in range(R.BS_BANNER, max_row + 1):
        for col in all_data_cols:
            cell = ws[f"{col}{r}"]
            if cell.number_format == "General":
                cell.number_format = FMT_AMOUNT
    for col in PERIOD_COLS:
        ws[f"{col}{R.DATES}"].number_format = FMT_DATE
        ws[f"{col}{R.DATES}"].fill = HEADER_DATE_FILL
        ws[f"{col}{R.DATES}"].alignment = Alignment(horizontal="center")

    # 3. Row-by-row visual treatment
    for r in range(1, max_row + 1):
        if r in BANNER_ROWS:
            for col in [LABEL_COL, CODE_COL] + all_data_cols:
                cell = ws[f"{col}{r}"]
                cell.fill = BANNER_FILL
                cell.font = Font(name=FONT_NAME, size=11, bold=True, color="002060")
            ws.row_dimensions[r].height = 18
        elif r in SECTION_HEADER_ROWS:
            for col in [LABEL_COL, CODE_COL]:
                cell = ws[f"{col}{r}"]
                cell.font = SECTION_HEADER_FONT
        elif r in KEY_TOTAL_ROWS:
            for col in [LABEL_COL, CODE_COL] + all_data_cols:
                cell = ws[f"{col}{r}"]
                cell.border = BORDER_TOTAL
                cell.font = Font(name=FONT_NAME, size=10, bold=True,
                                 color=cell.font.color)
        elif r in SUBTOTAL_ROWS:
            for col in [LABEL_COL, CODE_COL] + all_data_cols:
                cell = ws[f"{col}{r}"]
                cell.border = BORDER_SUBTOTAL
                cell.font = Font(name=FONT_NAME, size=10, bold=True, italic=True,
                                 color=cell.font.color)
        elif r in MARGIN_ROWS:
            for col in all_data_cols:
                cell = ws[f"{col}{r}"]
                cell.number_format = FMT_PCT
                cell.font = ITALIC
                cell.border = BORDER_MARGIN
            ws[f"{LABEL_COL}{r}"].font = ITALIC
        elif r in CHECK_ROWS:
            for col in [LABEL_COL] + all_data_cols:
                ws[f"{col}{r}"].font = CHECK_FONT

    # 4. Grand totals get a double-bottom border
    for r in (R.NI, R.TOTAL_ASSETS, R.LE_TOTAL, R.CLOSE_CASH, R.TIE_NET_CHG):
        for col in [LABEL_COL, CODE_COL] + all_data_cols:
            cell = ws[f"{col}{r}"]
            cell.border = BORDER_GRAND_TOTAL

    # 5. Column widths
    ws.column_dimensions[LABEL_COL].width = 44
    ws.column_dimensions[CODE_COL].width = 5
    ws.column_dimensions[OPEN_COL].width = 13
    for col in PERIOD_COLS:
        ws.column_dimensions[col].width = 12
    ws.column_dimensions[FY_COL].width = 14

    # 6. Header rows 1-3 styling
    ws[f"{LABEL_COL}1"].font = Font(name=FONT_NAME, size=12, bold=True, color="002060")
    for col in PERIOD_COLS:
        ws[f"{col}{R.FLAG}"].alignment = Alignment(horizontal="center")
        flag = ws[f"{col}{R.FLAG}"].value
        if flag == "Actual":
            ws[f"{col}{R.FLAG}"].font = Font(name=FONT_NAME, size=9, bold=True, color="C00000")
        else:
            ws[f"{col}{R.FLAG}"].font = Font(name=FONT_NAME, size=9, bold=True, color="595959")


# ──────────────────────────────────────────────────────────────────────────────
# Row outline groups (collapsible sections)
# ──────────────────────────────────────────────────────────────────────────────

def setup_outline_groups(ws) -> None:
    """v2.1.1: hide depreciation drivers, AR/AP/VAT/INV helpers, forecast-only
    BS lines, forecast splits (CAPEX, loan helpers, payroll helpers), vertical×
    product detail, legacy D&A splits, and the entire indirect-CF reconciliation."""
    ws.sheet_properties.outlinePr.summaryBelow = False
    ws.sheet_properties.outlinePr.summaryRight = False

    # (start_row, end_row, hidden_by_default)
    detail_ranges = [
        # BS — depreciation drivers (hidden)
        (R.PPE_CHARGE, R.RD_CHARGE, True),
        # BS — inventory helpers
        (R.PURCHASES_DRV, R.INV_OUT_DRV, True),
        # BS — AR helpers
        (R.AR_HLP_SALES_ADD, R.AR_HLP_COLL_DEP, True),
        # BS — AP helpers
        (R.AP_HLP_PURCH_ADD, R.AP_HLP_PAYMENTS, True),
        # BS — VAT helpers
        (R.VAT_HLP_OUTPUT, R.VAT_HLP_SETTLED, True),
        # BS — forecast-only liabilities (VAT/Tax/Deferred grants/Unearned rev)
        (R.BS_VAT_RECV, R.BS_VAT_RECV, True),
        (R.BS_VAT_PAY, R.BS_VAT_PAY, True),
        (R.BS_TAX_PAY, R.BS_TAX_PAY, True),
        (R.BS_DEFERRED_GRANTS, R.BS_DEFERRED_GRANTS, True),
        (R.BS_UNEARNED_REV, R.BS_UNEARNED_REV, True),

        # IS — vertical×product detail (rows 57-72, collapsed)
        (R.FL_FARADAOX, R.REV_NRE, True),

        # IS — D&A legacy split helpers
        (R.DA_PPE, R.DA_RD, True),

        # CF — direct-method operating ENGINE (drives the BS working-capital
        # rolls; hidden, since v3 presents the indirect operating breakdown).
        (R.CASH_IN_TOTAL, R.CASH_OUT_VAT, True),
        # CF — CAPEX hidden splits
        (R.CAPEX_PPE, R.CAPEX_EQUIP, True),
        # CF — financing helpers (loan draw / principal / interest / dividends)
        (R.LOAN_DRAW, R.DIVIDENDS, True),
        # CF — the direct/indirect tie row is always 0 now (one method); hide it.
        (R.CF_TIE, R.CF_TIE, True),

        # v3: SHOW the indirect operating detail (NI + D&A + ΔWC) as the headline
        # breakdown; hide only the investing/financing duplicates + net (already
        # shown in the CF summary above).
        (R.TIE_CAPEX_PPE, R.TIE_NET_CHG, True),
        # Hidden depreciable-base + loan-interest-base track (v3)
        (R.PPE_DEP_BASE, R.LOAN_INT_BASE, True),
        # Catch-all ΔWC helper (hidden); cash-check stays visible.
        (R.CF_DWC_OTHER, R.CF_DWC_OTHER, True),
    ]
    for start, end, hidden in detail_ranges:
        for r in range(start, end + 1):
            ws.row_dimensions[r].outlineLevel = 1
            ws.row_dimensions[r].hidden = hidden


# ──────────────────────────────────────────────────────────────────────────────
# Actuals sheet
# ──────────────────────────────────────────────────────────────────────────────

def build_actuals(wb: Workbook) -> None:
    """Mirror Pro Forma row layout; only `ACTUALS_DB_MAP` rows carry values."""
    ws = wb.create_sheet("Actuals")
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions[LABEL_COL].width = 42
    ws.column_dimensions[CODE_COL].width = 5
    ws.column_dimensions[OPEN_COL].width = 13
    for col in PERIOD_COLS:
        ws.column_dimensions[col].width = 12
    ws.column_dimensions[FY_COL].width = 14

    ws[f"{LABEL_COL}1"] = "Farada Rolling Budget — FY2026 Actuals (EUR)"
    ws[f"{LABEL_COL}1"].font = Font(name=FONT_NAME, size=12, bold=True, color="002060")
    ws[f"{PERIOD_COLS[0]}{R.DATES}"] = date(2026, 1, 1)
    ws[f"{PERIOD_COLS[0]}{R.DATES}"].number_format = FMT_DATE
    ws[f"{PERIOD_COLS[0]}{R.DATES}"].font = BOLD
    for i in range(1, 12):
        prev = PERIOD_COLS[i - 1]
        curr = PERIOD_COLS[i]
        ws[f"{curr}{R.DATES}"] = f"=EDATE({prev}{R.DATES},1)"
        ws[f"{curr}{R.DATES}"].number_format = FMT_DATE
        ws[f"{curr}{R.DATES}"].font = BOLD
        ws[f"{curr}{R.DATES}"].fill = HEADER_DATE_FILL
        ws[f"{curr}{R.DATES}"].alignment = Alignment(horizontal="center")
    ws[f"{PERIOD_COLS[0]}{R.DATES}"].fill = HEADER_DATE_FILL
    ws[f"{PERIOD_COLS[0]}{R.DATES}"].alignment = Alignment(horizontal="center")
    ws[f"{FY_COL}{R.DATES}"] = "FY2026"
    ws[f"{FY_COL}{R.DATES}"].font = BOLD

    for col, flag in zip(PERIOD_COLS, PERIOD_FLAGS):
        cell = ws[f"{col}{R.FLAG}"]
        cell.value = flag
        cell.alignment = Alignment(horizontal="center")
        if flag == "Actual":
            cell.font = Font(name=FONT_NAME, size=9, bold=True, color="C00000")
        else:
            cell.font = Font(name=FONT_NAME, size=9, bold=True, color="595959")

    actual_font = Font(name=FONT_NAME, size=10, color="00008B")
    for pf_row, period_to_value in _ACTUALS_VALUES.items():
        for period, val in period_to_value.items():
            if period not in PERIOD_DATES:
                continue
            col = PERIOD_COLS[PERIOD_DATES.index(period)]
            cell = ws[f"{col}{pf_row}"]
            cell.value = val
            cell.font = actual_font
            cell.number_format = FMT_AMOUNT

    actual_col_fill = PatternFill("solid", fgColor="F4F4F4")
    for i, flag in enumerate(PERIOD_FLAGS):
        if flag != "Actual":
            continue
        col = PERIOD_COLS[i]
        ws[f"{col}{R.DATES}"].fill = HEADER_DATE_FILL
        for pf_row in _ACTUALS_VALUES:
            cell = ws[f"{col}{pf_row}"]
            if cell.value is not None:
                cell.fill = actual_col_fill

    ws.freeze_panes = "E4"


# ──────────────────────────────────────────────────────────────────────────────

def main() -> None:
    if not DB_PATH.exists():
        raise FileNotFoundError(
            f"Expected farada.db at {DB_PATH}. Run `scripts/build_db.py farada` first."
        )
    values, present, all_actual_months = _load_actuals_from_db(DB_PATH)
    _ACTUALS_VALUES.update(values)
    _ACTUALS_PRESENT.update(present)

    # Force-zero hidden BS lines that aren't in the DB chart of accounts as
    # standalone lines (deferred grants, unearned revenue, VAT recv/pay, tax pay).
    # Without this, Pro Forma's forecast formula would compute a nonzero value
    # in actuals months and double-count against `BS_OTHER_CL`.
    ZERO_IN_ACTUALS = (
        R.BS_DEFERRED_GRANTS,
        R.BS_UNEARNED_REV,
        R.BS_VAT_RECV,
        R.BS_VAT_PAY,
        R.BS_TAX_PAY,
    )
    for pf_row in ZERO_IN_ACTUALS:
        _ACTUALS_VALUES.setdefault(pf_row, {})
        _ACTUALS_PRESENT.setdefault(pf_row, set())
        for month in all_actual_months:
            _ACTUALS_VALUES[pf_row].setdefault(month, 0.0)
            _ACTUALS_PRESENT[pf_row].add(month)

    actual_month_set = set(all_actual_months)
    for i, period in enumerate(PERIOD_DATES):
        PERIOD_FLAGS[i] = "Actual" if period in actual_month_set else "Forecast"

    n_actual = sum(1 for f in PERIOD_FLAGS if f == "Actual")
    n_rows_with_actuals = len(_ACTUALS_PRESENT)
    print(f"actuals: {n_actual} months ({all_actual_months}), "
          f"{n_rows_with_actuals} mapped Pro Forma rows")

    wb = Workbook()
    wb.remove(wb.active)
    build_actuals(wb)
    build_inputs(wb)
    build_pro_forma(wb)

    desired_order = ["Pro Forma", "Inputs", "Actuals"]
    wb._sheets = [wb[name] for name in desired_order]

    out = Path(__file__).resolve().parents[1] / "reference" / "rolling_budget_v3.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"wrote {out}")

    # Self-check: reload and scan for unresolved error cells.
    reloaded = openpyxl.load_workbook(out, data_only=False)
    error_strings = ("#REF!", "#NAME?", "#VALUE!", "#DIV/0!", "#N/A", "#NULL!")
    error_count = 0
    for sheet in reloaded.sheetnames:
        ws = reloaded[sheet]
        for row in ws.iter_rows(values_only=False):
            for cell in row:
                v = cell.value
                if isinstance(v, str) and any(err in v for err in error_strings):
                    if v.startswith("=") and any(fn in v.upper()
                                                  for fn in ("ISERROR", "IFERROR")):
                        continue
                    error_count += 1
                    if error_count <= 5:
                        print(f"  ERROR in {sheet}!{cell.coordinate}: {v}")
    if error_count > 5:
        print(f"  ... and {error_count - 5} more error cells")
    if error_count > 0:
        raise SystemExit(f"build failed: {error_count} error cells in output")

    pf = reloaded["Pro Forma"]
    inp = reloaded["Inputs"]
    act = reloaded["Actuals"]
    print(f"Pro Forma: {pf.max_row} rows × {pf.max_column} cols")
    print(f"Inputs:    {inp.max_row} rows × {inp.max_column} cols")
    print(f"Actuals:   {act.max_row} rows × {act.max_column} cols")
    print("self-check passed: no error cells")


if __name__ == "__main__":
    main()
