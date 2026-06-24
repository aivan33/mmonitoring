"""Build farada_model_v7 (FINAL) from v4 — cash-flow + ProForma/IS re-architecture + completion.

Per user review: (1) the ProForma must be a pure calc engine — every profitability SUBTOTAL and
MARGIN (Gross profit, EBITDA, EBIT, PBT, tax, Net profit, all margins) lives only on the Income
Statement, which now COMPUTES them; (2) build the cash flow WITH the working-capital engine
(direct method, Almacena-format), not orphaned rolls.

This file does R1–R3 (re-architecture + rolls). R4–R6 (CF/BS statements) follow.
  R1  strip ProForma profitability subtotals (rows 44-55, 116, 125, 130, 131, 132).
  R2  IS computes GP / EBITDA / EBIT / PBT / tax / NP from its own leaves (was =ProForma! pulls);
      per-bundle GP detail dropped (per-bundle COGS isn't split).
  R3  ProForma WC + financing rolls; the tax-payable & retained-earnings rolls reference IS.

Reads v4 (preserved), writes the FINAL farada_model_v7.xlsx. Idempotent.
Run:  .venv/bin/python clients/farada/one_offs/build_model_v7.py
"""
from __future__ import annotations

from copy import copy
from datetime import datetime

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.formula import ArrayFormula

SRC = "clients/farada/modeling/farada_model_v4.xlsx"
DST = "clients/farada/modeling/farada_model_v7.xlsx"
FIRST, LAST = 3, 62

I = dict(WC_HDR=161, DSO=162, PREPAY=163, SAAS_ANN=164, DPO=165, PAYDAYS=166, TAXLAG=167,
         FIN_HDR=169, EQ_AMT=170, EQ_DATE=171, DEBT_AMT=172, DEBT_DATE=173, DEBT_INT=174,
         OB_CASH=175, OB_AR=176, OB_AP=177, OB_PAYROLL=178, OB_DEFERRED=179, OB_DEBT=180,
         OB_SC=181, OB_RE=182)
R = dict(HDR=143, AR=144, AP_COGS=145, AP_SM=146, AP_GA=147, AP_RD=148, AP_TOT=149,
         PAYROLL=150, DEFERRED=151, TAXPAY=152, SC=153, DEBT=154, RE=155)
# ProForma profitability subtotals to strip (margins 56-61/117/121/122/126/133 already blank in v4).
STRIP = list(range(44, 56)) + [116, 125, 130, 131, 132]


def _ft(c):
    v = c.value
    return v.text if isinstance(v, ArrayFormula) else v


def labels(ws):
    out = {}
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if isinstance(v, str) and v.strip():
            out.setdefault(v.strip(), r)
    return out


# ---------------------------------------------------------------- R1
def strip_proforma_subtotals(wb):
    ws = wb["ProForma"]
    for r in STRIP:
        for c in range(1, LAST + 1):
            ws.cell(r, c).value = None
    print(f"  R1: blanked ProForma subtotals {STRIP[:3]}…{STRIP[-3:]} (GP/EBITDA/EBIT/PBT/tax/NP)")


# ---------------------------------------------------------------- R2
def is_compute_subtotals(wb):
    iss, isy = wb["IS"], wb["IS_Y"]
    L = labels(iss)
    # subtotal label -> (lambda col -> formula referencing IS's own rows)
    rev, cogs = L["Revenue"], L["COGS TOTAL"]
    c1, c2, l3 = L["Components #1 - Low Volume"], L["Components #2 - High Volume"], L["Hardware-enabled SaaS #3"]
    cl1, cl2, cl3 = L["COGS Line #1"], L["COGS Line #2"], L["COGS Line #3"]
    hwr, hwc = L["Hardware (device, cost + markup)"], L["Hardware (device)"]
    sar, usg = L["SaaS (overage, recurring)"], L["Usage (cloud / compute)"]
    gp, opex = L["Gross profit TOTAL (€)"], L["Operating expenses"]
    eb, grant, da = L["EBITDA"], L["Grant financing"], L["Depreciation & amortisation"]
    ebit, finn = L["EBIT"], L["Finance (costs), net"]
    pbt, tax, npr = (L["Profit / (loss) before income tax"], L["Income tax (expense)"],
                     L["Net profit / (loss) for the period"])
    COMPUTE = {
        gp:   lambda x: f"={x}{rev}-{x}{cogs}",
        L["Gross profit Line 1 (€)"]: lambda x: f"={x}{c1}-{x}{cl1}",
        L["Gross profit Line 2 (€)"]: lambda x: f"={x}{c2}-{x}{cl2}",
        L["Gross profit Line 3 (€)"]: lambda x: f"={x}{l3}-{x}{cl3}",
        L["Hardware GP (€)"]: lambda x: f"={x}{hwr}-{x}{hwc}",
        L["SaaS GP (€)"]: lambda x: f"={x}{sar}-{x}{usg}",
        eb:   lambda x: f"={x}{gp}-{x}{opex}",
        ebit: lambda x: f"={x}{eb}+{x}{grant}-{x}{da}",
        pbt:  lambda x: f"={x}{ebit}+{x}{finn}",
        tax:  lambda x: f"=-MAX(0,{x}{pbt})*' Inputs'!$J$158",
        npr:  lambda x: f"={x}{pbt}+{x}{tax}",
    }
    for r, fn in COMPUTE.items():
        for c in range(FIRST, LAST + 1):
            iss.cell(r, c, fn(get_column_letter(c)))
    # drop per-bundle GP detail (per-bundle COGS isn't split) — blank in IS and IS_Y
    drop = [L["Hardware GP (€)"] + i for i in (1, 2, 3)] + [L["SaaS GP (€)"] + i for i in (1, 2, 3)]
    for sh in (iss, isy):
        for r in drop:
            for c in range(1, LAST + 1):
                sh.cell(r, c).value = None
    print(f"  R2: IS computes {len(COMPUTE)} subtotals from its own leaves; dropped bundle-GP rows {drop}")
    return L


# ---------------------------------------------------------------- R3 (rolls)
def add_cf_inputs(wb):
    inp = wb[" Inputs"]
    s_hdr = inp.cell(152, 3)._style
    s_lbl, s_unit = inp.cell(154, 3)._style, inp.cell(154, 4)._style
    s_j, s_l = inp.cell(154, 10)._style, inp.cell(154, 12)._style

    def hdr(r, text):
        inp.cell(r, 3, text)._style = copy(s_hdr)

    def put(r, label, unit, value, numfmt):
        inp.cell(r, 3, label)._style = copy(s_lbl)
        inp.cell(r, 4, unit)._style = copy(s_unit)
        inp.cell(r, 10, f"=OFFSET(K{r},0,$D$2)")._style = copy(s_j)
        cc = inp.cell(r, 12, value)
        cc._style = copy(s_l)
        cc.number_format = numfmt

    hdr(I["WC_HDR"], "WORKING CAPITAL (payment terms)  (mockup ← confirm)")
    put(I["DSO"], "Receivable days (DSO)", "days", 30, "#,##0")
    put(I["PREPAY"], "Hardware prepayment % (large orders)", "%", 0.0, "0.0%")
    put(I["SAAS_ANN"], "SaaS billed annually in advance", "%", 1.0, "0.0%")
    put(I["DPO"], "Payable days (DPO)", "days", 30, "#,##0")
    put(I["PAYDAYS"], "Payroll payable days", "days", 30, "#,##0")
    put(I["TAXLAG"], "Tax payment lag", "months", 0, "#,##0")
    hdr(I["FIN_HDR"], "FINANCING & OPENING BALANCES (Jul-2026)  (mockup ← confirm)")
    put(I["EQ_AMT"], "Equity round amount", "EUR", 5_000_000, "€#,##0")
    put(I["EQ_DATE"], "Equity round date", "date", datetime(2027, 7, 1), "mmm-yyyy")
    put(I["DEBT_AMT"], "Debt draw amount", "EUR", 0, "€#,##0")
    put(I["DEBT_DATE"], "Debt draw date", "date", datetime(2026, 7, 1), "mmm-yyyy")
    put(I["DEBT_INT"], "Debt annual interest", "%", 0.0, "0.0%")
    put(I["OB_CASH"], "Opening cash", "EUR", 2_000_000, "€#,##0")
    put(I["OB_AR"], "Opening AR", "EUR", 0, "€#,##0")
    put(I["OB_AP"], "Opening AP", "EUR", 0, "€#,##0")
    put(I["OB_PAYROLL"], "Opening payroll payable", "EUR", 0, "€#,##0")
    put(I["OB_DEFERRED"], "Opening deferred revenue", "EUR", 0, "€#,##0")
    put(I["OB_DEBT"], "Opening debt", "EUR", 0, "€#,##0")
    put(I["OB_SC"], "Opening share capital", "EUR", 5_000_000, "€#,##0")
    # Opening retained earnings is NOT a free input — it's the balancing plug so the opening
    # balance sheet ties (assets = equity + liabilities). Computed in the RE roll.
    inp.cell(I["OB_RE"], 3, "Opening retained earnings (= balancing plug)")._style = copy(s_lbl)


def _bundle_rows(inp, subheader_prefix, n=3):
    """The n input rows (col-J OFFSET) immediately under an Inputs sub-header (col-C startswith).
    Label-based → robust to the I–V reflow renumbering."""
    hdr = next((r for r in range(1, inp.max_row + 1)
                if isinstance(inp.cell(r, 3).value, str)
                and inp.cell(r, 3).value.strip().startswith(subheader_prefix)), None)
    rows, r = [], (hdr or 0) + 1
    while hdr and len(rows) < n and r <= inp.max_row:
        if isinstance(inp.cell(r, 10).value, str) and "OFFSET" in inp.cell(r, 10).value:
            rows.append(r)
        r += 1
    return rows


def set_d5_inputs(wb):
    """D5a — populate the new plan tier-discount skeletons (S/M/L = 10/15/20% off list) and re-set the
    included-measurements quota plan-heavy (~80% of avg=1200 → 960). Both flagged placeholders to
    calibrate. Runs POST-reflow, by sub-header context (rows renumbered by the I–V reflow). Values go
    in col L (Realistic; the OFFSET reads it at D2=1)."""
    inp = wb[" Inputs"]
    for r, v in zip(_bundle_rows(inp, "Line 3 — plan tier discount"), (0.10, 0.15, 0.20)):
        inp.cell(r, 12, v)
    note = next((r for r in range(1, inp.max_row + 1) if isinstance(inp.cell(r, 3).value, str)
                 and inp.cell(r, 3).value.strip().startswith("Line 3 — plan tier discount")), None)
    if note:
        inp.cell(note, 15, "← PLACEHOLDER: plan rate = list × (1−discount); calibrate later")
    incl = _bundle_rows(inp, "Line 3 — included measurements")
    for r in incl:
        inp.cell(r, 12, 960)
    if incl:
        inp.cell(incl[0] - 1, 15, "← PLACEHOLDER: plan-heavy (~80% of avg 1200); calibrate later")
    # cloud cost calibrated up off the €0.0005 placeholder → implied SaaS GM ~90–94% (worst bundle 90.5%)
    cloud = next(r for r in range(1, inp.max_row + 1) if isinstance(inp.cell(r, 3).value, str)
                 and inp.cell(r, 3).value.strip().startswith("Cloud / compute per measurement"))
    inp.cell(cloud, 12, 0.0016)
    # OD1 — overage ramp delay: clients don't overuse credits until `delay` months after onboarding
    delay = next(r for r in range(1, inp.max_row + 1) if isinstance(inp.cell(r, 3).value, str)
                 and inp.cell(r, 3).value.strip().startswith("Overage ramp delay"))
    inp.cell(delay, 12, 3)
    inp.cell(delay, 15, "← PLACEHOLDER: months before a cohort starts using overage; calibrate later")
    print(f"  D5a: tier discounts + plan-heavy included + cloud_cost 0.0016 + overage delay 3mo set")


def seed_yield_inputs(wb):
    """V1 — seed the new staged-yield + sensors-per-wafer Inputs with today's curve so chip €/sensor is
    unchanged: spw=4000; yield rungs (run-rate threshold in col F, yield in col L) 1→0.70, 10k→0.73,
    100k→0.82, 1M→0.90, 4M→0.95. Runs after the Inputs reflow (rung rows exist), by label."""
    inp = wb[" Inputs"]
    def row(prefix):
        return next(r for r in range(1, inp.max_row + 1) if isinstance(inp.cell(r, 3).value, str)
                    and inp.cell(r, 3).value.strip().startswith(prefix))
    inp.cell(row("Sensors per wafer"), 12, 4000)
    for lbl, thr, y in [("Yield @ 1 /yr", 1, 0.70), ("Yield @ 10,000 /yr", 10000, 0.73),
                        ("Yield @ 100,000 /yr", 100000, 0.82), ("Yield @ 1,000,000 /yr", 1000000, 0.90),
                        ("Yield @ 4,000,000 /yr", 4000000, 0.95)]:
        r = row(lbl)
        inp.cell(r, 6, thr)        # F = run-rate threshold (like the cost-of-sales curves)
        inp.cell(r, 12, y)         # L = Realistic yield
    print("  V1: seeded sensors-per-wafer (4000) + staged yield (0.70..0.95)")


def add_rolls(wb, L_is):
    ws = wb["ProForma"]
    leaf = {c: ws.cell(89, c)._style for c in range(1, LAST + 1)}
    band = {c: ws.cell(85, c)._style for c in range(1, LAST + 1)}
    jp = lambda n: f"' Inputs'!$J${I[n]}"
    np_r, tax_r = L_is["Net profit / (loss) for the period"], L_is["Income tax (expense)"]

    LABEL = {R["HDR"]: "WORKING CAPITAL & FINANCING ROLLS (balances)",
             R["AR"]: "  Trade receivables (AR)", R["AP_COGS"]: "  Trade payables — COGS",
             R["AP_SM"]: "  Trade payables — S&M", R["AP_GA"]: "  Trade payables — G&A",
             R["AP_RD"]: "  Trade payables — R&D", R["AP_TOT"]: "  Trade payables (total)",
             R["PAYROLL"]: "  Payroll payable", R["DEFERRED"]: "  Deferred revenue (SaaS annual)",
             R["TAXPAY"]: "  Tax payable", R["SC"]: "  Share capital", R["DEBT"]: "  Debt",
             R["RE"]: "  Retained earnings"}

    def fml(rk, x, prev, first):
        if rk == "AR":
            return f"=(({x}5+{x}9+{x}15)*(1-{jp('PREPAY')})+{x}19*(1-{jp('SAAS_ANN')}))/30*{jp('DSO')}"
        if rk == "AP_COGS": return f"={x}24/30*{jp('DPO')}"
        if rk == "AP_SM":   return f"=({x}87-{x}88)/30*{jp('DPO')}"
        if rk == "AP_GA":   return f"=({x}96-{x}97)/30*{jp('DPO')}"
        if rk == "AP_RD":   return f"=({x}107-{x}108)/30*{jp('DPO')}"
        if rk == "AP_TOT":  return f"={x}{R['AP_COGS']}+{x}{R['AP_SM']}+{x}{R['AP_GA']}+{x}{R['AP_RD']}"
        if rk == "PAYROLL": return f"=({x}88+{x}97+{x}108)/30*{jp('PAYDAYS')}"
        if rk == "DEFERRED": return f"={x}19*{jp('SAAS_ANN')}*6"
        if rk == "TAXPAY":  return f"=-IS!{x}{tax_r}*IF({jp('TAXLAG')}>=1,1,0)"
        if rk == "SC":
            inj = f"IF({x}2={jp('EQ_DATE')},{jp('EQ_AMT')},0)"
            return f"={jp('OB_SC')}+{inj}" if first else f"={prev}{R['SC']}+{inj}"
        if rk == "DEBT":
            dr = f"IF({x}2={jp('DEBT_DATE')},{jp('DEBT_AMT')},0)"
            return f"={jp('OB_DEBT')}+{dr}" if first else f"={prev}{R['DEBT']}+{dr}"
        if rk == "RE":
            # opening RE = balancing plug = opening assets − opening other-L&E (so BS ties at t0)
            plug = (f"({jp('OB_CASH')}+{jp('OB_AR')}+' Inputs'!$J$155"
                    f"-{jp('OB_AP')}-{jp('OB_PAYROLL')}-{jp('OB_DEFERRED')}-{jp('OB_DEBT')}-{jp('OB_SC')})")
            return f"={plug}+IS!{x}{np_r}" if first else f"={prev}{R['RE']}+IS!{x}{np_r}"
        return None

    ws.cell(R["HDR"], 1, LABEL[R["HDR"]])._style = copy(band[1])
    for c in range(2, LAST + 1):
        ws.cell(R["HDR"], c)._style = copy(band[c])
    for rk in ("AR", "AP_COGS", "AP_SM", "AP_GA", "AP_RD", "AP_TOT", "PAYROLL",
               "DEFERRED", "TAXPAY", "SC", "DEBT", "RE"):
        r = R[rk]
        ws.cell(r, 1, LABEL[r])._style = copy(leaf[1])
        for c in range(FIRST, LAST + 1):
            x, prev = get_column_letter(c), get_column_letter(c - 1)
            ws.cell(r, c, fml(rk, x, prev, c == FIRST))._style = copy(leaf[c])
    print(f"  R3: rolls {R['HDR']}-{R['RE']}; tax-payable & RE reference IS (tax r{tax_r}, NP r{np_r})")


# ---------------------------------------------------------------- R4 (CF) / R5 (BS)
def _new_sheet(wb, name, after):
    if name in wb.sheetnames:
        del wb[name]
    s = wb.create_sheet(name, index=wb.sheetnames.index(after) + 1)
    s.sheet_view.showGridLines = False
    s.freeze_panes = "C3"
    return s


def build_cf(wb, L_is):
    iss = wb["IS"]
    tax_r = L_is["Income tax (expense)"]
    s = _new_sheet(wb, "CF", "IS_Y")
    s.column_dimensions["A"].width = iss.column_dimensions["A"].width or 42
    leaf = {c: iss.cell(5, c)._style for c in range(1, LAST + 1)}
    tot = {c: iss.cell(4, c)._style for c in range(1, LAST + 1)}
    s.cell(1, 1, "Cash Flow Statement — Monthly (Direct method, EUR)")._style = copy(iss.cell(1, 1)._style)
    s.cell(2, 1, "Currency: EUR")._style = copy(iss.cell(2, 1)._style)
    for c in range(FIRST, LAST + 1):
        s.cell(2, c, f"=IS!{get_column_letter(c)}2")._style = copy(iss.cell(2, c)._style)
    PF = "ProForma!"
    LBL = {4: "Cash received from customers", 5: "Cash paid to suppliers",
           6: "Payment for personnel and social security", 7: "Corporate and other taxes, net",
           8: "Bank charges paid", 9: "Movement in deferred revenue",
           10: "Cash Flow from Operating Activities", 12: "CAPEX", 13: "R&D (capitalised)",
           14: "Cash Flow from Investing Activities", 16: "Capital Increase",
           17: "Loan facility financing", 18: "Grants", 19: "Cash Flow from Financing Activities",
           21: "Excess Cash for the Period", 22: "Beginning Cash Balance",
           23: "Ending Cash Balance", 24: "% Change in cash"}
    TOT = {10, 14, 19, 21, 23}
    for r, lab in LBL.items():
        s.cell(r, 1, lab)._style = copy((tot if r in TOT else leaf)[1])

    def delta(rollrow, ob, x, p, first):
        base = (f"' Inputs'!$J${ob}" if ob else "0")
        return f"({PF}{x}{rollrow}-{base})" if first else f"({PF}{x}{rollrow}-{PF}{p}{rollrow})"

    for c in range(FIRST, LAST + 1):
        x, p, first = get_column_letter(c), get_column_letter(c - 1), c == FIRST
        D = lambda rr, ob: delta(rr, ob, x, p, first)
        f = {
            4: f"={PF}{x}4-{D(R['AR'], I['OB_AR'])}",
            5: f"=-({PF}{x}24+{PF}{x}86-({PF}{x}88+{PF}{x}97+{PF}{x}108))+{D(R['AP_TOT'], I['OB_AP'])}",
            6: f"=-({PF}{x}88+{PF}{x}97+{PF}{x}108)+{D(R['PAYROLL'], I['OB_PAYROLL'])}",
            7: f"=IS!{x}{tax_r}+{D(R['TAXPAY'], None)}",
            8: f"=-{PF}{x}128",
            9: f"={D(R['DEFERRED'], I['OB_DEFERRED'])}",
            10: f"=SUM({x}4:{x}9)",
            12: f"=-{PF}{x}136", 13: "0", 14: f"={x}12+{x}13",
            16: f"={D(R['SC'], I['OB_SC'])}", 17: f"={D(R['DEBT'], I['OB_DEBT'])}",
            18: f"={PF}{x}120", 19: f"={x}16+{x}17+{x}18",
            21: f"={x}10+{x}14+{x}19",
            22: (f"=' Inputs'!$J$175" if first else f"={p}23"),
            23: f"={x}22+{x}21",
            24: (f"=IF(' Inputs'!$J$175=0,0,{x}23/' Inputs'!$J$175-1)" if first else f"=IF({p}23=0,0,{x}23/{p}23-1)"),
        }
        for r, formula in f.items():
            cell = s.cell(r, c, formula)
            cell._style = copy((tot if r in TOT else leaf)[c])
            if r == 24:
                cell.number_format = "0.0%"
    print("  R4: CF (monthly, direct) — operating/investing/financing; cash = plug")


def build_bs(wb):
    iss = wb["IS"]
    s = _new_sheet(wb, "BS", "CF")
    s.column_dimensions["A"].width = iss.column_dimensions["A"].width or 42
    leaf = {c: iss.cell(5, c)._style for c in range(1, LAST + 1)}
    tot = {c: iss.cell(4, c)._style for c in range(1, LAST + 1)}
    band = {c: iss.cell(63, c)._style for c in range(1, LAST + 1)}
    s.cell(1, 1, "Balance Sheet — Monthly (EUR)")._style = copy(iss.cell(1, 1)._style)
    s.cell(2, 1, "Currency: EUR")._style = copy(iss.cell(2, 1)._style)
    for c in range(FIRST, LAST + 1):
        s.cell(2, c, f"=IS!{get_column_letter(c)}2")._style = copy(iss.cell(2, c)._style)
    PF = "ProForma!"
    LBL = {4: "ASSETS", 5: "  Intangible fixed assets (R&D)", 6: "  Tangible fixed assets (PP&E)",
           7: "  Cash & cash equivalents", 8: "  Trade receivable", 9: "TOTAL ASSETS",
           11: "EQUITY & LIABILITIES", 12: "  Share capital", 13: "  Retained earnings",
           14: "  Loan facility financing", 15: "  Trade payables",
           16: "  Personnel & social security payables", 17: "  Deferred revenue",
           18: "  Tax payables", 19: "TOTAL EQUITY & LIABILITIES", 20: "check (Assets − E&L)"}
    BANDS, TOT = {4, 11}, {9, 19}
    for r, lab in LBL.items():
        st = band if r in BANDS else (tot if r in TOT else leaf)
        s.cell(r, 1, lab)._style = copy(st[1])
    for c in range(FIRST, LAST + 1):
        x = get_column_letter(c)
        f = {5: "0", 6: f"={PF}{x}139", 7: f"=CF!{x}23", 8: f"={PF}{x}144",
             9: f"={x}5+{x}6+{x}7+{x}8",
             12: f"={PF}{x}153", 13: f"={PF}{x}155", 14: f"={PF}{x}154", 15: f"={PF}{x}149",
             16: f"={PF}{x}150", 17: f"={PF}{x}151", 18: f"={PF}{x}152",
             19: f"=SUM({x}12:{x}18)", 20: f"={x}9-{x}19"}
        for r, formula in f.items():
            st = band if r in BANDS else (tot if r in TOT else leaf)
            s.cell(r, c, formula)._style = copy(st[c])
        for r in (4, 11):
            for cc in range(2, LAST + 1):
                s.cell(r, cc)._style = copy(band[cc])
    print("  R5: BS (monthly) — cash = CF ending; check row = Assets − E&L (target 0)")


# ---------------------------------------------------------------- R6 (yearly)
CAL_YEARS = [("2026", 3, 8), ("2027", 9, 20), ("2028", 21, 32), ("2029", 33, 44), ("2030", 45, 56)]


def build_yearly(wb):
    cf, bs = wb["CF"], wb["BS"]
    ycols = [get_column_letter(3 + k) for k in range(5)]
    # --- CF_Y: SUM 12 monthly flows; cash balances = year-open / year-end roll ---
    y = _new_sheet(wb, "CF_Y", "BS")
    y.column_dimensions["A"].width = cf.column_dimensions["A"].width or 42
    y.cell(1, 1, "Cash Flow — Yearly (calendar; 2026 = Jul–Dec)")._style = copy(cf.cell(1, 1)._style)
    for k, (lbl, _, _) in enumerate(CAL_YEARS):
        c = y.cell(2, 3 + k, lbl)
        c._style = copy(cf.cell(2, 3)._style)
        c.number_format = "General"
        y.column_dimensions[ycols[k]].width = 14
    FLOW = {4, 5, 6, 7, 8, 9, 10, 12, 13, 14, 16, 17, 18, 19, 21}
    for r in range(4, 25):
        lab = cf.cell(r, 1).value
        if lab is None:
            continue
        y.cell(r, 1, lab)._style = copy(cf.cell(r, 1)._style)
        for k, (_, c0, c1) in enumerate(CAL_YEARS):
            yc, f0, f1 = ycols[k], get_column_letter(c0), get_column_letter(c1)
            cell = y.cell(r, 3 + k)
            cell._style = copy(cf.cell(r, 3)._style)
            if r in FLOW:
                cell.value = f"=SUM(CF!{f0}{r}:{f1}{r})"
            elif r == 22:
                cell.value = f"=CF!{f0}22"
            elif r == 23:
                cell.value = f"={yc}22+{yc}21"
            elif r == 24:
                cell.value = f"=IF({yc}22=0,0,{yc}23/{yc}22-1)"
                cell.number_format = "0.0%"
    # --- BS_Y: December period-end snapshot ---
    yb = _new_sheet(wb, "BS_Y", "CF_Y")
    yb.column_dimensions["A"].width = bs.column_dimensions["A"].width or 42
    yb.cell(1, 1, "Balance Sheet — Yearly (calendar year-end; 2026 = Dec)")._style = copy(bs.cell(1, 1)._style)
    for k, (lbl, _, _) in enumerate(CAL_YEARS):
        c = yb.cell(2, 3 + k, lbl)
        c._style = copy(bs.cell(2, 3)._style)
        c.number_format = "General"
        yb.column_dimensions[ycols[k]].width = 14
    BANDS = {4, 11}
    for r in range(4, 21):
        lab = bs.cell(r, 1).value
        if lab is None:
            continue
        yb.cell(r, 1, lab)._style = copy(bs.cell(r, 1)._style)
        for k, (_, _, c1) in enumerate(CAL_YEARS):
            decL = get_column_letter(c1)
            cell = yb.cell(r, 3 + k)
            cell._style = copy(bs.cell(r, 3)._style)
            if r not in BANDS:
                cell.value = f"=BS!{decL}{r}"
    print("  R6: CF_Y / BS_Y (calendar 2026-2030; CF = SUM flows, BS = Dec snapshot)")


# ---------------------------------------------------------------- overhaul
def fix_capacity_ref(wb):
    """Repair the pre-existing capacity #REF! — cols BE-BJ carried a broken outer IF branch
    (a deleted capacity tier). Rewrite the whole row uniformly to the clean 5-tier cascade
    (Inputs G/J 8-12 = the 'Capacity from Jul-20xx' rows)."""
    ws = wb["ProForma"]
    R = 78
    for c in range(FIRST, LAST + 1):
        x = get_column_letter(c)
        ws.cell(R, c, f"=IF({x}$2>=' Inputs'!$G$12,' Inputs'!$J$12,"
                      f"IF({x}$2>=' Inputs'!$G$11,' Inputs'!$J$11,"
                      f"IF({x}$2>=' Inputs'!$G$10,' Inputs'!$J$10,"
                      f"IF({x}$2>=' Inputs'!$G$9,' Inputs'!$J$9,' Inputs'!$J$8))))")
    print("  overhaul: capacity row 78 repaired (no more #REF!)")


def clean_dead_inputs(wb):
    """Blank the orphaned Line-3 usage-pricing ladder (Inputs 24-29 — header + 5 rungs):
    superseded by the per-bundle overage prices (J45-47); 0 references anywhere. Blank in place
    (no row-shift); the rows get removed wholesale in the later Inputs re-sequence."""
    inp = wb[" Inputs"]
    for r in range(24, 30):
        for col in (2, 3, 4, 6, 10, 11, 12, 15):
            inp.cell(r, col).value = None
    print("  overhaul: blanked orphaned usage-pricing ladder (Inputs 24-29)")


def calibrate_saas_placeholder(wb):
    """F1 (completion) — the SaaS overage-only line gave a ~98% GM (the one clearly-wrong line).
    PLACEHOLDER: plug SaaS COGS (ProForma 41 'Usage cloud/compute') to the existing 'SaaS gross
    margin target' input (J99=80%): COGS = SaaS revenue (row 19) × (1−target). Wires J99 (was
    orphaned), flagged for real calibration (subscription + realistic cloud cost) later. The
    measurement-driven J102 input stays in place but unreferenced — noted as superseded."""
    pf, inp = wb["ProForma"], wb[" Inputs"]
    for c in range(FIRST, LAST + 1):
        x = get_column_letter(c)
        pf.cell(41, c, f"={x}19*(1-' Inputs'!$J$99)")
    inp.cell(99, 15, "← PLACEHOLDER: drives SaaS COGS (= SaaS rev × (1−this)); "
                     "calibrate real subscription + cloud cost later")
    inp.cell(102, 15, "← superseded by the SaaS-GM-target placeholder (J99); kept for calibration")
    print("  completion F1: SaaS COGS plugged to target GM (J99=80%), flagged")


def expose_yield(wb):
    """F2 (completion) — make yield explicit. Chip €/sensor = wafer ÷ sensors-per-wafer ÷ yield.
    Reuse the Inputs chip rows (62-67) as the WAFER COST assumption (€/wafer, staged on run-rate);
    add sensors-per-wafer + yield as ProForma calc rows (the engine, empty rows 82-83) and derive
    chip there (per user: the yield calc belongs in the ProForma). No row-shift; chip values
    unchanged by construction (4000/(4000·0.70)=1.43 … 2000/(4000·0.95)=0.53)."""
    inp, pf = wb[" Inputs"], wb["ProForma"]
    WAFER = [4000, 4000, 3735, 3068, 2401, 2000]
    inp.cell(61, 3, "Wafer cost (€/wafer)")
    inp.cell(61, 15, "← chip €/sensor derived in ProForma = wafer ÷ sensors-per-wafer ÷ yield")
    for i, r in enumerate(range(62, 68)):
        inp.cell(r, 12, WAFER[i])                       # L = Realistic wafer cost
    pf.cell(82, 1, "  Sensors per wafer")
    pf.cell(83, 1, "  Yield (staged by run-rate)")
    for c in range(FIRST, LAST + 1):
        x = get_column_letter(c)
        pf.cell(82, c, 4000)
        pf.cell(83, c, f"=IF({x}67>=4000000,0.95,IF({x}67>=1000000,0.9,IF({x}67>=100000,0.82,"
                       f"IF({x}67>=10000,0.73,0.7))))")
        wafer = (f"IF({x}67>=' Inputs'!$F$67,' Inputs'!$J$67,IF({x}67>=' Inputs'!$F$66,' Inputs'!$J$66,"
                 f"IF({x}67>=' Inputs'!$F$65,' Inputs'!$J$65,IF({x}67>=' Inputs'!$F$64,' Inputs'!$J$64,"
                 f"IF({x}67>=' Inputs'!$F$63,' Inputs'!$J$63,' Inputs'!$J$62)))))")
        pf.cell(69, c, f"=({wafer})/({x}82*{x}83)")
    for c in range(1, LAST + 1):                         # style the new rows like the chip driver
        pf.cell(82, c)._style = copy(pf.cell(69, c)._style)
        pf.cell(83, c)._style = copy(pf.cell(69, c)._style)
    print("  completion F2: yield exposed — chip = wafer÷spw÷yield (ProForma calc rows)")


def build():
    wb = openpyxl.load_workbook(SRC)
    fix_capacity_ref(wb)
    clean_dead_inputs(wb)
    calibrate_saas_placeholder(wb)
    expose_yield(wb)
    add_cf_inputs(wb)
    strip_proforma_subtotals(wb)
    L_is = is_compute_subtotals(wb)
    add_rolls(wb, L_is)
    build_cf(wb, L_is)
    build_bs(wb)
    build_yearly(wb)
    # re-sequence Inputs into the house skeleton (I. FUNDING … V. OTHER) + remap refs (equivalence-gated)
    import reflow_inputs as rf
    referenced = rf._referenced_rows(wb)
    old2new, label_at, _ = rf.reflow(wb)
    rf._gate(wb, old2new, label_at, referenced)
    rf.remap_refs(wb, old2new)
    set_d5_inputs(wb)            # D5a — populate tier-discount skeletons + plan-heavy included
    seed_yield_inputs(wb)        # V1 — seed staged-yield + sensors-per-wafer Inputs (today's curve)
    # re-sequence the ProForma engine into the skill-outline order (drivers first) + context-aware remap
    import reflow_proforma as rfp
    rfp.reflow(wb)
    rfp.add_installed_base(wb)         # V2: per-bundle installed-base (cumulative sensors) driver rows
    rfp.add_measurement_children(wb)   # V4: measurements off IB (Included≠Overage, clean total)
    rfp.add_subscription_lines(wb)     # D5b: Subscription (recurring) revenue + billings memo
    rfp.rewire_saas_off_ib(wb)         # V3: subscription/overage/billings off the installed base
    rfp.cloud_cogs_measurement_driven(wb)  # D5d: cloud COGS = measurements × cloud_cost (retire GM plug)
    rfp.rework_wc_rolls(wb)            # D5e: deferred=running(billings−sub); AR excludes subscription
    rfp.fix_run_rate(wb)               # D1: LTM trailing-12 run-rate (was a frozen constant)
    rfp.add_proforma_sections(wb)      # skill-outline lower sections (BS rolls + WC/CF/Tax/Funding)
    rfp.style_subtotals(wb)            # bold the ProForma sum/subtotal lines (readability)
    rfp.add_overage_delay(wb)          # OD: ramp-delay overage rev + measurements (OFFSET, guarded)
    rfp.wire_yield_inputs(wb)          # V1: ProForma yield/spw → staged Inputs (no hardcoded curve)
    import restructure_statements as rs
    rs.restructure_bs(wb)              # BS → reference structure (sub-groups + blank lines + ratios)
    rs.restructure_cf(wb)             # CF → reference lines (VAT/Other/Dividends/burn) + CF_Y mirror
    rfp.build_cupffee_cf(wb)          # RB1: CASH FLOW = Cupffee by-component engine; CF statement pull+sum
    rfp.populate_wc_ratios(wb)        # CB3: WC drivers & ratios in ProForma; BS pulls them
    rfp.populate_tax_funding(wb)      # CB4: TAXATION + FUNDING sections (thin refs)
    rf.unify_input_formats(wb)   # LAST — final word on Inputs value-cell number formats
    wb.save(DST)
    print(f"Saved {DST}")


if __name__ == "__main__":
    build()
