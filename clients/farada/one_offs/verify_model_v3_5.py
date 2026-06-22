"""Verify farada_model_v3.5.xlsx — Phase 1 (HR payroll wiring + salary indexation).

v3.5 is the hand-built source of truth (HR sheet + grouped inputs); we PATCH it in place
(patch_model_v3_5.py). No recalc engine, so this is safe-by-construction + structural checks.

Run from repo root:  .venv/bin/python clients/farada/one_offs/verify_model_v3_5.py
"""
from __future__ import annotations

import openpyxl
from openpyxl.worksheet.formula import ArrayFormula

P = "clients/farada/modeling/farada_model_v3.5.xlsx"
SAL_IDX_ROW = 137          # salary indexation input (end of OPEX group)


def ft(cell):
    v = cell.value
    return (v.text if isinstance(v, ArrayFormula) else v) or "" if v is not None else None


def main() -> None:
    wb = openpyxl.load_workbook(P)
    ws, inp, hr = wb["ProForma"], wb[" Inputs"], wb["HR"]
    fails: list[str] = []

    def ck(cond, msg):
        print(f"  {'✅' if cond else '❌'} {msg}")
        if not cond:
            fails.append(msg)

    print("\n[T3] HR payroll wired into ProForma OPEX (HR-total-row convention)")
    expect = {"C88": "=HR!O16", "C97": "=HR!O48", "C109": "=HR!O67", "C110": "=HR!O66"}
    for coord, f in expect.items():
        ck(ft(ws[coord]) == f, f"{coord} = {f}")
    ck(ft(ws["D88"]) == "=HR!P16", "fill-right aligns (D88=HR!P16, calendar offset C↔O)")
    # the OPEX payroll subtotals must roll into Operating expenses -> EBITDA
    ck(ft(ws["C108"]) == "=C109+C110", "R&D Total Payroll = Germany+Serbia")

    print("\n[T2] Salary indexation input present, in the OPEX group, wired to HR")
    lbl = ft(inp.cell(SAL_IDX_ROW, 3))
    ck(isinstance(lbl, str) and "indexation" in lbl.lower(), f"Inputs r{SAL_IDX_ROW} label = salary indexation ({lbl!r})")
    ck(ft(inp.cell(SAL_IDX_ROW, 10)) == f"=OFFSET(K{SAL_IDX_ROW},0,$D$2)", "uses the scenario OFFSET pattern")
    ck(isinstance(inp.cell(SAL_IDX_ROW, 12).value, (int, float)), "has a Realistic value in L")
    n_old = sum(1 for row in hr.iter_rows() for c in row if isinstance(ft(c), str) and "$J$250" in ft(c))
    n_new = sum(1 for row in hr.iter_rows() for c in row if isinstance(ft(c), str) and f"$J${SAL_IDX_ROW}" in ft(c))
    ck(n_old == 0, f"HR no longer references the empty J250 (found {n_old})")
    ck(n_new > 2000, f"HR salary-escalation cells now reference J{SAL_IDX_ROW} ({n_new} cells)")

    print("\n[struct] no NEW #REF! (pre-existing capacity row 78 only)")
    ref_rows = {cell.row for row in ws.iter_rows() for cell in row
                if isinstance(ft(cell), str) and "#REF!" in ft(cell)}
    ck(ref_rows <= {78}, f"#REF! rows ⊆ {{78}} (got {sorted(ref_rows)})")

    print("\n[Phase2] P&L completed to Net Profit (mirrors reference is)")
    chain = {
        "C125": "=C116+C120-C123",                       # EBIT = EBITDA + grant - D&A
        "C129": "=C127-C128",                            # finance net
        "C130": "=C125+C129",                            # PBT
        "C131": "=-MAX(0,C130)*' Inputs'!$J$158",        # tax on positive PBT
        "C132": "=C130+C131",                            # net profit
    }
    for coord, f in chain.items():
        ck(ft(ws[coord]) == f, f"{coord} = {f}")
    ck(ft(ws["C133"]) == "=IF(C4=0,0,C132/C4)", "C133 profit margin = NP/Sales")

    print("\n[Phase2] D&A is BUILT from a capex schedule (not a direct input)")
    ck(ft(ws["C123"]) == "=C124", "D&A total = depreciation sub-line")
    ck(ft(ws["C124"]) == "=C138", "Depreciation (P&L) ← schedule row 138")
    ck(ft(ws["C136"]) == "=IF(AND(C2>=' Inputs'!$G$153,C2<=' Inputs'!$H$153),' Inputs'!$J$153,0)",
       "Capex row gated by date from the capex input")
    ck(ft(ws["C137"]) == "=' Inputs'!$J$155" and ft(ws["D137"]) == "=C139",
       "Opening PP&E = opening input then prior closing (rollforward)")
    ck(ft(ws["C138"]) == "=MIN((C137+C136)/(' Inputs'!$J$154*12),C137+C136)",
       "Depreciation = straight-line on (opening+capex) over life×12 months")
    ck(ft(ws["C139"]) == "=C137+C136-C138", "Closing PP&E = opening + capex − depreciation")
    # depreciation oracle: mockup capex €/mo, life yrs -> month-by-month NBV roll
    capex = inp.cell(153, 12).value or 0
    life_m = (inp.cell(154, 12).value or 0) * 12
    open_nbv = inp.cell(155, 12).value or 0
    nbv, dep3 = open_nbv, []
    for _ in range(3):
        d = min((nbv + capex) / life_m, nbv + capex) if life_m else 0
        dep3.append(round(d, 2))
        nbv = nbv + capex - d
    print(f"     oracle: capex €{capex:,.0f}/mo, life {life_m/12:.0f}y → dep first 3 mo = {dep3}")
    ck(life_m > 0 and all(d >= 0 for d in dep3), "depreciation schedule yields non-negative dep")

    print("\n[Phase2] no naked rows — added rows are styled (number_format set, not General)")
    naked = []
    for r in list(range(120, 134)) + list(range(136, 140)):
        for coord in (f"C{r}", f"D{r}"):
            if ws[coord].value is not None and ws[coord].number_format == "General":
                naked.append(coord)
    ck(not naked, f"no General-format data cells in the new P&L/schedule rows (found {naked[:5]})")

    print("\n[Phase2] below-EBITDA / D&A inputs present (grouped after OPEX)")
    for row, kw in ((153, "capex"), (154, "life"), (155, "opening"), (156, "grant"),
                    (157, "finance"), (158, "tax")):
        lbl = ft(inp.cell(row, 3))
        ck(isinstance(lbl, str) and kw in lbl.lower(), f"Inputs r{row} = {kw} input ({lbl!r})")

    print("\n[Phase3] Yearly P&L sheet IS_Y (SUM 12 months / recompute margins)")
    ck("IS_Y" in wb.sheetnames, "IS_Y sheet exists")
    if "IS_Y" in wb.sheetnames:
        y = wb["IS_Y"]
        ck(ft(y["C4"]) == "=SUM(ProForma!C4:N4)", "FY1 Revenue = SUM(ProForma!C4:N4)")
        ck(ft(y["D4"]) == "=SUM(ProForma!O4:Z4)", "FY2 Revenue = SUM(ProForma!O4:Z4)")
        ck(ft(y["G132"]) == "=SUM(ProForma!AY132:BJ132)", "FY5 Net profit = SUM(ProForma!AY132:BJ132)")
        ck(ft(y["C116"]) == "=SUM(ProForma!C116:N116)", "FY1 EBITDA = SUM monthly")
        ck(ft(y["C56"]) == "=IF(C4=0,0,C44/C4)", "GM% recomputed from IS_Y's own lines")
        ck(ft(y["C133"]) == "=IF(C4=0,0,C132/C4)", "Profit margin recomputed (not summed)")
        ck(ft(y["C121"]) == "=IF(C4=0,0,(C116+C120)/C4)", "EBITDA margin incl. grant recomputed")
        ck(ft(y["A43"]) == "GROSS PROFIT" and ft(y["A85"]) == "OPERATING EXPENSES", "section headers mirrored")
        naked_y = [f"{co}{r}" for r in (4, 56, 116, 132) for co in "CDEFG"
                   if y[f"{co}{r}"].value is not None and y[f"{co}{r}"].number_format == "General"]
        ck(not naked_y, f"IS_Y has no naked (General-format) data cells (found {naked_y[:5]})")

    print("\n[flag] HR subtotal labels (informational, not changed — format is a given):")
    print(f"     HR r39 {hr.cell(39,1).value!r} sums R&D engineers; HR r48 {hr.cell(48,1).value!r} sums G&A people")
    print("     → labels look swapped, but ProForma pulls the correct CONTENT. Flagged for the user.")

    print()
    if fails:
        print(f"FAILED {len(fails)} check(s):")
        for f in fails:
            print(f"  - {f}")
        raise SystemExit(1)
    print("ALL CHECKS PASSED")


if __name__ == "__main__":
    main()
