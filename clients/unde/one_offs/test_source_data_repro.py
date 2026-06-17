"""Checkpoint A: the Categorization->Source-Data transform must reproduce the
April schedule's appended rows (2524-2555) from (April Cat - March Cat).

Run: python clients/unde/one_offs/test_source_data_repro.py
"""
import datetime as dt
import sys
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))
from mrr_source_data import new_invoices, to_source_row

C = Path(__file__).resolve().parents[1] / "raw"
APR_CAT = C / "04/Undelucram Categorization - April 2026.xlsx"
MAR_CAT = C / "03/Undelucram Categorization March 2026.xlsx"
APR_SCHED = C / "04/MRR_Schedule_Undelucram April 2026.xlsx"
FX_RON, FX_USD = 5.0735, 1.0  # backed out of the April file (neutralizes FX)


def actual_appended(first=2524):
    wb = load_workbook(APR_SCHED, data_only=True)
    ws = wb["1.1 Source Data"]
    rows = []
    for r in range(first, ws.max_row + 1):
        if ws.cell(r, 2).value is None:
            continue
        rows.append({
            "currency": ws.cell(r, 1).value, "amount": ws.cell(r, 2).value,
            "client": ws.cell(r, 3).value, "valoare": ws.cell(r, 4).value,
            "start": ws.cell(r, 5).value, "period": ws.cell(r, 7).value,
            "country": ws.cell(r, 8).value, "mrr": ws.cell(r, 9).value,
            "monthly": ws.cell(r, 10).value, "produs": ws.cell(r, 11).value,
        })
    wb.close()
    return rows


def norm(row):
    def n(x):
        return round(x, 2) if isinstance(x, float) else (
            x.date() if isinstance(x, dt.datetime) else x)
    return (row["currency"], n(row["amount"]), row["client"], n(row["valoare"]),
            n(row["start"]), row["period"], row["country"], row["mrr"],
            n(row["monthly"]), row["produs"])


def main():
    # Selection (candidate rule) + transform must reproduce April's appended
    # rows. PASS = every actual appended row is produced by a candidate. Extra
    # candidates beyond the actual set are the manual-exclusion review items
    # (e.g. ASTREA renewal) — reported, not a failure.
    from mrr_source_data import candidates
    cand = [to_source_row(i, FX_RON, FX_USD)
            for i in candidates(APR_CAT, 2026, 4)]
    actual = actual_appended()

    def fkey(r):
        def n(x):
            return round(x, 2) if isinstance(x, float) else x
        return (r["currency"], n(r["amount"]), r["client"], n(r["valoare"]),
                r["period"], r["country"], r["mrr"], n(r["monthly"]), r["produs"])

    from collections import Counter
    cb, ca = Counter(fkey(r) for r in cand), Counter(fkey(r) for r in actual)
    missing = list((ca - cb).elements())   # actual rows no candidate produced
    extra = list((cb - ca).elements())     # candidates the colleague excluded
    print(f"candidates={len(cand)}  actual appended={len(actual)}  "
          f"reproduced={sum((cb & ca).values())}/{len(actual)}")
    for r in missing:
        print("  MISSING (actual not reproduced):", r)
    for r in extra:
        print("  REVIEW (candidate the colleague excluded):", r)
    print("\nRESULT:", "PASS" if not missing else "FAIL")


if __name__ == "__main__":
    main()
