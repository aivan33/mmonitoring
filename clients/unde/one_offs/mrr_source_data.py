"""Categorization -> MRR Schedule '1.1 Source Data' rows (workstream B, stage 1a).

The Categorization 'Income Invoices' sheet is a cumulative DB. Each month the
NEW invoices (this month's file minus last month's) are appended to Source Data,
EUR-converted and keyed for the downstream MRR calc.

Column indices in 'Income Invoices' (0-based, header on row 7):
  0 Date | 2 Billing start | 3 Billing end | 4 Currency | 5 Amount
  8 Commercial name | 9 Country | 10 Product | 11 Contract length
  12 Subscription type | 14 MRR flag | 15 Start date | 16 End date
"""
from __future__ import annotations

import datetime as dt
from pathlib import Path

from openpyxl import load_workbook

# Categorization product label -> Source Data 'Produs' label.
PRODUS = {
    "Employer branding": "Employer Branding",
    "Jobbing": "Corporate Jobbing",
    "Advertising": "Advertising",
    "LinkedIn Learning": "LinkedIn Learning",
    "Linkedin Learning": "LinkedIn Learning",
    "LinkedIn": "LinkedIn Learning",
    "Salary report": "Salary report",
    "Brand perception": "Brand perception",
}


# Country-code prefix (combined "RO - Employer branding" format) -> Country.
COUNTRY = {
    "RO": "Romania", "GR": "Greece", "BG": "Bulgaria", "HU": "Hungary",
    "MD": "Moldova", "CZ": "Czech Republic", "PL": "Poland", "MENA": "MENA",
}
# Header labels we locate by name (the column layout drifts between monthly
# Categorization versions, so we never hard-code indices).
_HEAD = {
    "date": "Date", "billstart": "Billing start date", "currency": "Currency",
    "amount": "Amount", "commercial": "Commercial name",
    "category": "Management Report Category", "length": "Contract length",
    "subtype": "Subscription type", "mrr": "MRR", "start": "Start date",
    "end": "End date",
}


def _resolve_country_product(cat, next_cell):
    """Category may be combined ('RO - Employer branding') or already split
    (country in the cell, product in the next, unlabelled column — April's
    newer layout)."""
    if isinstance(cat, str) and " - " in cat:
        code, prod = cat.split(" - ", 1)
        return COUNTRY.get(code.strip(), code.strip()), prod.strip()
    return cat, next_cell  # already-split layout


def load_invoices(path: str | Path) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Income Invoices"]
    rows = list(ws.iter_rows(min_row=7, values_only=True))
    header = rows[0]
    idx = {}
    for key, label in _HEAD.items():
        idx[key] = next((i for i, v in enumerate(header) if v == label), None)
    cat_i = idx["category"]
    out = []
    for r in rows[1:]:
        if (idx["date"] is None or r[idx["date"]] is None) and \
           (idx["amount"] is None or r[idx["amount"]] is None):
            continue
        cat = r[cat_i] if cat_i is not None else None
        nxt = r[cat_i + 1] if cat_i is not None and cat_i + 1 < len(r) else None
        country, product = _resolve_country_product(cat, nxt)
        g = lambda k: r[idx[k]] if idx[k] is not None else None
        out.append({
            "date": g("date"), "billstart": g("billstart"), "currency": g("currency"),
            "amount": g("amount"), "commercial": g("commercial"),
            "country": country, "product": product,
            "length": g("length"), "subtype": g("subtype"), "mrr": g("mrr"),
            "start": g("start"), "end": g("end"),
        })
    wb.close()
    return out


def new_invoices(curr_path, year: int, month: int) -> list[dict]:
    """Invoices dated in the reporting month (docx: 'all new invoices in the
    month'). Cross-file diffing is NOT viable — the Categorization column schema
    drifts between monthly versions (April added a Commercial-name column and
    split Country/Product), so selection is by invoice Date within one file."""
    out = []
    for inv in load_invoices(curr_path):
        d = inv["date"]
        if isinstance(d, dt.datetime) and d.year == year and d.month == month:
            out.append(inv)
    return out


def effective_date(inv: dict):
    """Invoice date, de-typo'd: an implausible year (e.g. 3036-04-24) is the
    colleague's typo for the current reporting year — fix the year (taken from
    the contract Start date), keeping month/day. Both PORSCHE 3036-04-24 rows
    thus resolve to 2026-04-24, matching the April append."""
    d = inv["date"]
    if isinstance(d, dt.datetime) and 2020 <= d.year <= 2030:
        return d, False
    s = inv["start"]
    yr = s.year if isinstance(s, dt.datetime) else 2026
    if isinstance(d, dt.datetime):
        try:
            return dt.datetime(yr, d.month, d.day), True
        except ValueError:
            pass
    if isinstance(s, dt.datetime):
        return s, True
    return d, True


def candidates(path: str | Path, year: int, month: int) -> list[dict]:
    """Proposed 'new this month' invoices for human review. Selection is a
    manual judgment in practice, so this is generous + flagged, not final.

    Each candidate carries: `_date_typo` (raw invoice date was implausible) and
    `_dup_of_count` (how many earlier invoices share client+product+amount —
    a possible renewal/duplicate to confirm)."""
    invs = load_invoices(path)
    from collections import Counter
    seen = Counter()
    out = []
    for inv in invs:
        ed, typo = effective_date(inv)
        sig = (inv["commercial"], inv["product"], inv["amount"])
        if isinstance(ed, dt.datetime) and ed.year == year and ed.month == month:
            inv = dict(inv)
            inv["_eff_date"] = ed
            inv["_date_typo"] = typo
            inv["_dup_of_count"] = seen[sig]
            out.append(inv)
        seen[sig] += 1
    return out


def to_source_row(inv: dict, fx_ron: float, fx_usd: float) -> dict:
    cur = inv["currency"]
    rate = {"LEI": fx_ron, "RON": fx_ron, "USD": fx_usd, "EUR": 1.0}.get(cur, 1.0)
    amount = inv["amount"] or 0.0
    valoare = amount / rate
    period = inv["length"]
    monthly = valoare / period if isinstance(period, (int, float)) and period else None
    return {
        "currency": cur, "amount": amount, "client": inv["commercial"],
        "valoare": valoare, "start": inv["start"], "period": period,
        "country": inv["country"], "mrr": inv["mrr"], "monthly": monthly,
        "produs": PRODUS.get(inv["product"], inv["product"]),
    }
