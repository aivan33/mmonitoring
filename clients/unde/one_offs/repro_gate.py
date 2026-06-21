"""Reproduction gate: rebuild a month already in the taxonomi from the MR and
diff cell-by-cell. The mapping is trusted only when every MR-sourced row
reproduces to < EUR 1. Derived KPI rows are reported separately (best-effort).

Usage: python clients/unde/one_offs/repro_gate.py 2026-03
"""
from __future__ import annotations

import sys
from pathlib import Path

import yaml
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[3]))
from core.report.mr import extract_month
from core.report.mr_to_taxonomi import populate_taxonomi

CLIENT = Path(__file__).resolve().parents[1]
SHEET_FOR = {"IS": "IS (Actual)", "CF": "CF Indirect (Actual)", "BS": "BS (Actual)"}


def col_for_month(month):  # Data,Group,Subgroup = 1..3, Jan = 4
    return 3 + month


def read_col(path, month):
    wb = load_workbook(path, data_only=True)
    vals = {}
    for code, sheet in SHEET_FOR.items():
        ws = wb[sheet]
        col = col_for_month(month)
        for r in range(2, ws.max_row + 1):
            key = (ws.cell(r, 1).value, ws.cell(r, 2).value, ws.cell(r, 3).value)
            if key[0] is None:
                continue
            vals[(code, key)] = ws.cell(r, col).value
    wb.close()
    return vals


def num(v):
    return float(v) if isinstance(v, (int, float)) else 0.0


def main():
    period = sys.argv[1] if len(sys.argv) > 1 else "2026-03"
    year, month = int(period[:4]), int(period[5:7])
    mr_path = CLIENT / "raw/04/Undelucram - Monthly reports 2026.xlsx"
    mapping = yaml.safe_load((CLIENT / "mapping.yaml").read_text())
    prev = CLIENT / "raw/taxonomi_act_2026-03.xlsx"

    original = read_col(prev, month)

    extracts = {s: extract_month(mr_path, mapping, year, month, s)
                for s in ("IS", "CF", "BS")}
    out = CLIENT / "raw" / f"_repro_{period}.xlsx"
    populate_taxonomi(prev, extracts, year, month, out, mapping=mapping)
    repro = read_col(out, month)

    # Which keys are MR-sourced (direct), which are derived KPIs / null.
    direct = set()
    for code, mkey in (("IS", "mapping_is"), ("CF", "mapping_cf"), ("BS", "mapping_bs")):
        for e in mapping[mkey]:
            if e["mr_row"] is not None:
                direct.add((code, (e["data"], e["grp"], e["subgroup"])))
    derived = set()
    for e in mapping.get("kpi_derivations", []):
        derived.add((e["statement_for_target"], tuple(e["target"])))

    direct_fail, kpi_report, other = [], [], []
    for k, orig in original.items():
        d = abs(num(repro.get(k)) - num(orig))
        if d <= 0.01:
            continue
        line = (k[0], k[1][0], k[1][1], k[1][2], num(orig), num(repro.get(k)), d)
        if k in direct:
            if d > 1.0:
                direct_fail.append(line)
        elif k in derived:
            kpi_report.append(line)
        else:
            other.append(line)

    def show(title, rows):
        print(f"\n=== {title} ({len(rows)}) ===")
        for c, dd, g, s, o, rp, dl in sorted(rows, key=lambda x: -x[6]):
            print(f" {c} {dd}|{g}|{s}: orig={o:,.1f} repro={rp:,.1f} d={dl:,.1f}")

    show("DIRECT-MAPPED FAILURES (>EUR1) -- must be empty to pass", direct_fail)
    show("DERIVED KPI deltas (best-effort)", kpi_report)
    show("OTHER unmapped/derived rows w/ deltas", other)

    # Subtotal reconciliation against MR's own subtotals.
    print("\n=== SUBTOTAL RECONCILIATION (repro vs MR subtotal) ===")
    mr = load_workbook(mr_path, data_only=True)
    isw = mr["Income Statement"]
    c = col_for_month(month) - 3 + 4  # MR IS Jan=col5
    def mrval(ws, row, jan_col):
        return num(ws.cell(row, jan_col + (month - 1)).value)
    is_sales = sum(num(v) for (code, key), v in repro.items()
                   if code == "IS" and key[0] == "Sales")
    print(f" IS Sales: repro Σ={is_sales:,.1f}  MR 'Sales ' r7={mrval(isw,7,5):,.1f}  MR MRR r5={mrval(isw,5,5):,.1f}")
    mr.close()

    print("\nRESULT:", "PASS" if not direct_fail else f"FAIL ({len(direct_fail)} direct rows off)")
    out.unlink(missing_ok=True)


if __name__ == "__main__":
    main()
