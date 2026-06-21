"""Build Undelucram's taxonomi-actual column for a period.

Wraps the shared report pipeline (core/report/mr.py + mr_to_taxonomi.py) and
adds the one Undelucram-specific seam: the MRR headline is sourced from the
MRR schedule's 'Reporting (1)' sheet, not the management report.

Usage: python clients/unde/one_offs/build_taxonomi.py 2026-04
"""
from __future__ import annotations

import datetime as dt
import sys
from pathlib import Path

import yaml
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[3]))
from core.report.mr import extract_month
from core.report.mr_to_taxonomi import populate_taxonomi

CLIENT = Path(__file__).resolve().parents[1]
MRR_KEY = ("MRR", "MRR", "MRR")


def mrr_from_schedule(path: Path, sheet: str, year: int, month: int) -> float | None:
    """Read the 'MRR' row at the column whose row-2 date header matches the
    target month, from the MRR schedule's reporting sheet."""
    wb = load_workbook(path, data_only=True)
    try:
        ws = wb[sheet]
        col = None
        for c in range(1, ws.max_column + 1):
            v = ws.cell(2, c).value
            if isinstance(v, dt.datetime) and v.year == year and v.month == month:
                col = c
                break
        if col is None:
            raise ValueError(f"{sheet}: no date header for {year}-{month:02d}")
        for r in range(1, ws.max_row + 1):
            lab = ws.cell(r, 1).value
            if isinstance(lab, str) and lab.strip() == "MRR":
                v = ws.cell(r, col).value
                return float(v) if isinstance(v, (int, float)) else None
        raise ValueError(f"{sheet}: no 'MRR' row found")
    finally:
        wb.close()


def main():
    period = sys.argv[1]
    year, month = int(period[:4]), int(period[5:7])
    cfg = yaml.safe_load((CLIENT / "config.yaml").read_text())
    rep = cfg["reporting"]
    mapping = yaml.safe_load((CLIENT / rep["mapping"]).read_text())
    mr_path = CLIENT / rep["mr_source"]

    extracts = {s: extract_month(mr_path, mapping, year, month, s)
                for s in ("IS", "CF", "BS")}

    # Inject MRR from the MRR schedule.
    mrr = mrr_from_schedule(CLIENT / rep["mrr_source"], rep["mrr_sheet"], year, month)
    extracts["IS"][MRR_KEY] = mrr
    print(f"  MRR from schedule ({year}-{month:02d}): {mrr:,.1f}")

    prev = CLIENT / "raw/taxonomi_act_2026-03.xlsx"
    out = CLIENT / "raw" / f"taxonomi_act_{period}.xlsx"
    populate_taxonomi(prev, extracts, year, month, out, mapping=mapping)
    print(f"  wrote {out.relative_to(CLIENT.parents[1])}")


if __name__ == "__main__":
    main()
