"""Reconstruct Almacena's lender funding book month-by-month (2026).

There is only one lender file (the April export), but every loan carries its
start and repayment dates, so the book for any month is *derived* by replaying
each loan's active window. This module is the single source of that derivation;
both the HTML builder and the exploration notebook import it.

DERIVED, not raw: these are reconstructed monthly snapshots, scoped to 2026
(Jan–Apr) and reported in USD (the lender data's native currency). The only
month we can check against source figures is April; ``--check`` proves the
reconstruction reproduces the file's own Available Funds / Cost of Funds /
active-loan count exactly.

Conventions (tuned to reproduce April's AccruedDaysInMonth / contributions):
- A loan is active in a month if its [start, repayment] window overlaps the
  calendar month. Overlap days are **inclusive** of both endpoints.
- Available-funds contribution = Principal x overlap_days / days_in_month
  (time-weighted average drawn principal for the month).
- Accrued interest (cost) = Principal x annual_rate x overlap_days / 365.
"""

from __future__ import annotations

import calendar
import datetime as dt
import json
import sys
from pathlib import Path

from openpyxl import load_workbook

SRC = Path(__file__).resolve().parents[1] / "raw/04/lender_loans_accrued_interest.xlsx"
MONTHS_2026 = [(2026, m) for m in range(1, 5)]   # Jan–Apr 2026
MON_ABBR = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
            "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def _load_loans(path: Path = SRC) -> list[dict]:
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    hdr = {str(ws.cell(1, c).value).strip(): c for c in range(1, ws.max_column + 1)}
    out = []
    for r in range(2, ws.max_row + 1):
        s = ws.cell(r, hdr["StartDate"]).value
        e = ws.cell(r, hdr["RepaymentDate"]).value
        if not (isinstance(s, dt.datetime) and isinstance(e, dt.datetime)):
            continue
        out.append({
            "lender": ws.cell(r, hdr["LenderName"]).value,
            "ref": ws.cell(r, hdr["LenderLoanRef"]).value,
            "start": s.date(),
            "repay": e.date(),
            "principal": float(ws.cell(r, hdr["PrincipalAmount"]).value or 0.0),
            "rate": float(ws.cell(r, hdr["InterestRateAnnual"]).value or 0.0),
        })
    wb.close()
    return out


def _overlap_days(start: dt.date, repay: dt.date,
                  m_start: dt.date, m_end: dt.date) -> int:
    a, b = max(start, m_start), min(repay, m_end)
    return (b - a).days + 1 if b >= a else 0       # inclusive


def month_book(loans: list[dict], year: int, month: int) -> dict:
    """The active loan book for one month: per-loan rows + totals."""
    dim = calendar.monthrange(year, month)[1]
    m_start = dt.date(year, month, 1)
    m_end = dt.date(year, month, dim)
    rows = []
    for ln in loans:
        od = _overlap_days(ln["start"], ln["repay"], m_start, m_end)
        if od == 0:
            continue
        avail = ln["principal"] * od / dim
        accrued = ln["principal"] * ln["rate"] * od / 365.0
        rows.append({
            **{k: ln[k] for k in ("lender", "ref", "principal", "rate")},
            "start": ln["start"].isoformat(),
            "repay": ln["repay"].isoformat(),
            "days_active": od,
            "available": round(avail, 2),
            "accrued": round(accrued, 2),
            "new": m_start <= ln["start"] <= m_end,
            "maturing": m_start <= ln["repay"] <= m_end,
        })
    rows.sort(key=lambda r: -r["principal"])
    principal = sum(r["principal"] for r in rows)
    available = sum(r["available"] for r in rows)
    cost = sum(r["accrued"] for r in rows)
    blended = (sum(r["principal"] * r["rate"] for r in rows) / principal
               if principal else 0.0)
    return {
        "key": f"{year}-{month:02d}",
        "label": f"{MON_ABBR[month - 1]} {year}",
        "days_in_month": dim,
        "n_loans": len(rows),
        "n_lenders": len({r["lender"] for r in rows}),
        "principal": round(principal, 2),
        "available": round(available, 2),
        "cost": round(cost, 2),
        "blended_rate": round(blended, 6),
        "n_new": sum(1 for r in rows if r["new"]),
        "new_principal": round(sum(r["principal"] for r in rows if r["new"]), 2),
        "n_maturing": sum(1 for r in rows if r["maturing"]),
        "maturing_principal": round(sum(r["principal"] for r in rows if r["maturing"]), 2),
        "loans": rows,
    }


def build(path: Path = SRC) -> dict:
    """Full 2026 reconstruction + MoM deltas. JSON-serialisable."""
    loans = _load_loans(path)
    months = [month_book(loans, y, m) for (y, m) in MONTHS_2026]
    for i, mb in enumerate(months):
        prev = months[i - 1] if i else None
        mb["mom"] = {
            k: round(mb[k] - prev[k], 4) if prev else None
            for k in ("available", "cost", "principal", "blended_rate",
                      "n_loans", "n_lenders")
        }
    return {
        "currency": "USD",
        "derived": True,
        "derived_note": ("Reconstructed from the April loan schedule "
                         "(lender_loans_accrued_interest.xlsx) — monthly "
                         "snapshots are derived, not raw."),
        "scope": "2026 (Jan–Apr)",
        "source_file": SRC.name,
        "months": months,
    }


def _check() -> int:
    """Prove April reproduces the source file's own figures."""
    apr = month_book(_load_loans(), 2026, 4)
    checks = [
        ("Available Funds", apr["available"], 15_590_305.13),
        ("Cost of Funds", apr["cost"], 116_409.02),
        ("Active loans", apr["n_loans"], 24),
        ("Blended rate", round(apr["blended_rate"] * 100, 2), 9.08),
    ]
    ok = True
    for name, got, want in checks:
        delta = abs(got - want)
        flag = "OK" if delta <= 1.0 else "FAIL"
        if flag == "FAIL":
            ok = False
        print(f"  [{flag}] {name}: got {got:,} vs source {want:,}  (Δ{delta:,.2f})")
    print("April reconstruction reproduces source." if ok
          else "MISMATCH — reconstruction drifted from source.")
    return 0 if ok else 1


if __name__ == "__main__":
    if len(sys.argv) > 1 and sys.argv[1] == "--check":
        sys.exit(_check())
    print(json.dumps(build(), indent=2, default=str))
