"""Build the funding-maturity KPI source for the Almacena deck.

Reads the lender-loan schedule (`raw/04/lender_loans_accrued_interest.xlsx`) and
buckets the principal of the *currently-active* funding book by repayment month,
emitting a `kpi_wide` xlsx with a single row `Maturing Principal` (USD; the DB
loader converts to EUR at the config rate). This feeds the Funding Maturity
chart that replaced the dropped Cash Drag chart in the April pack.

"Active" = loan started on/before the reporting month-end and not yet repaid
(RepaymentDate >= reporting month start). Maturities are shown Apr–Dec 2026;
any 2027 tail is printed for the analyst but not charted (near-term wall only).

    uv run python clients/almacena/one_offs/build_funding_maturity.py
    # -> raw/04/funding_maturity_apr.xlsx
"""

from __future__ import annotations

import datetime as dt
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "raw/04/lender_loans_accrued_interest.xlsx"
OUT = ROOT / "raw/04/funding_maturity_apr.xlsx"

AS_OF_START = dt.date(2026, 4, 1)
AS_OF_END = dt.date(2026, 4, 30)
CHART_MONTHS = [(2026, m) for m in range(4, 13)]   # Apr–Dec 2026
MON_ABBR = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
            "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def _col(ws, name: str) -> int:
    for c in range(1, ws.max_column + 1):
        if str(ws.cell(1, c).value).strip() == name:
            return c
    raise KeyError(name)


def main() -> int:
    wb = load_workbook(SRC, data_only=True)
    ws = wb[wb.sheetnames[0]]
    c_start = _col(ws, "StartDate")
    c_rep = _col(ws, "RepaymentDate")
    c_prin = _col(ws, "PrincipalAmount")

    by_month: dict[tuple[int, int], float] = defaultdict(float)
    tail_2027 = 0.0
    active = 0
    for r in range(2, ws.max_row + 1):
        start = ws.cell(r, c_start).value
        rep = ws.cell(r, c_rep).value
        prin = ws.cell(r, c_prin).value or 0.0
        if not (isinstance(start, dt.datetime) and isinstance(rep, dt.datetime)):
            continue
        # active in the reporting month: started by month-end, not yet repaid
        if start.date() > AS_OF_END or rep.date() < AS_OF_START:
            continue
        active += 1
        if rep.year == 2026:
            by_month[(rep.year, rep.month)] += prin
        elif rep.year >= 2027:
            tail_2027 += prin
    wb.close()

    # Write kpi_wide: row1 = Month + chart-month headers; row2 = the KPI.
    out = Workbook()
    o = out.active
    o.title = "Funding Maturity"
    o.cell(1, 1, "Month")
    for j, (y, m) in enumerate(CHART_MONTHS, start=2):
        o.cell(1, j, f"{MON_ABBR[m - 1]} {y % 100}")
    o.cell(2, 1, "Maturing Principal")
    for j, (y, m) in enumerate(CHART_MONTHS, start=2):
        o.cell(2, j, round(by_month.get((y, m), 0.0), 2))
    OUT.parent.mkdir(parents=True, exist_ok=True)
    out.save(OUT)
    out.close()

    print(f"Wrote {OUT.relative_to(ROOT.parent.parent)}  (active loans: {active})")
    print("Maturing principal by month (USD):")
    for (y, m) in CHART_MONTHS:
        v = by_month.get((y, m), 0.0)
        if v:
            print(f"  {MON_ABBR[m - 1]} {y}: {v:,.0f}  (EUR @1.087 ≈ {v / 1.087:,.0f})")
    if tail_2027:
        print(f"  2027 tail (not charted): {tail_2027:,.0f} USD")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
