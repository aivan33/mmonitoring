"""Monthly loans-DB update for the Almacena model.

Reads a month's lender loan ledger (`raw/<MM>/lender_loans_accrued_interest.xlsx`),
keeps the loans outstanding at month-end, and maps them to the model's
`Loans Database` column schema — a **paste-ready** snapshot for a clean-slate update
of the model's loan book. Lender **abbreviations are kept as-is** (no legal names).

Usage:
    python clients/almacena/one_offs/build_loans_db_update.py <YYYY-MM>

    e.g.  python clients/almacena/one_offs/build_loans_db_update.py 2026-04

Writes `clients/almacena/budget/loans_db_update_<YYYY-MM>.csv` (gitignored — loan
detail) with the 10 `Loans Database` columns in order, and prints a summary. To
update the model: open the workbook, clear the old `Loans Database` rows, and paste
the CSV under the header row.

Monthly: drop next month's ledger in `raw/<MM>/` and re-run with the new YYYY-MM.
"""

from __future__ import annotations

import argparse
import calendar
import csv
import datetime as dt
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[3]
CLIENT = ROOT / "clients/almacena"

# Output columns, in the model's `Loans Database` order (cols 1-10).
COLUMNS = [
    "Loan ID", "Lender Name", "Principal Amount (EUR)", "Start Date",
    "Tenor (M) Testing", "r (% p.a)", "Payment Frequency (M)",
    "Repayment date", "Repayment Amount", "Active (T/F)",
]

DEFAULT_FREQ = 3  # months — matches the model's existing Payment Frequency


def read_ledger(path: str | Path) -> list[dict]:
    """Read the lender loan ledger into records keyed by our internal names."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        header = [str(c) if c is not None else "" for c in
                  next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        idx = {name: i for i, name in enumerate(header)}
        recs = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[idx["LenderName"]] is None:
                continue
            recs.append({
                "lender": row[idx["LenderName"]],
                "start": row[idx["StartDate"]],
                "repayment": row[idx["RepaymentDate"]],
                "principal": row[idx["PrincipalAmount"]] or 0.0,
                "rate": row[idx["InterestRateAnnual"]],
                "total_interest": row[idx["TotalInterestLoan"]] or 0.0,
            })
        return recs
    finally:
        wb.close()


def _tenor_months(start, repayment) -> int | None:
    if not (isinstance(start, dt.datetime) and isinstance(repayment, dt.datetime)):
        return None
    return round((repayment - start).days / 30.4375)


def to_model_rows(
    records: list[dict], asof: dt.datetime, freq: int = DEFAULT_FREQ,
    roll_to: dt.datetime | None = None,
) -> list[dict]:
    """Map active (outstanding at ``asof``) ledger records to model-schema rows.

    A loan is outstanding at ``asof`` only if it has already been drawn
    (``start <= asof``) and not yet repaid (``repayment > asof``).

    Why both bounds matter: the ledger is the full schedule, not a month snapshot —
    it pre-books **roll-overs** as new loans that start the day after the maturing
    loan, with principal = old principal + capitalized interest. Those renewals are
    dated into the next month (or later) and carry zero accrued interest at this
    month-end; the ``start <= asof`` bound excludes them here and the same filter
    rolls them in automatically next month (no double-count).

    ``roll_to``: if set, treat the book as **evergreen** — push every loan's
    Repayment date out to ``roll_to`` (a date past the model's forecast horizon).
    The book then stays live and accruing across the whole forecast, so no principal
    repayment lands in-window and the model never books the phantom repay+redraw that
    a real mid-forecast maturity would create. Filtering still uses the *actual*
    dates; only the output repayment date / tenor / repayment amount are rolled.
    """
    active = [
        r for r in records
        if r["start"] and r["start"] <= asof and r["repayment"] and r["repayment"] > asof
    ]
    active.sort(key=lambda r: (str(r["lender"]).lower(), -r["principal"]))
    rows = []
    for i, r in enumerate(active, 1):
        repay = roll_to or r["repayment"]
        if roll_to:
            years = (roll_to - r["start"]).days / 365.0
            repay_amt = round(r["principal"] * (1 + (r["rate"] or 0) * years), 2)
        else:
            repay_amt = round(r["principal"] + r["total_interest"], 2)
        rows.append({
            "Loan ID": f"LN-{i:03d}",
            "Lender Name": r["lender"],
            "Principal Amount (EUR)": r["principal"],
            "Start Date": r["start"],
            "Tenor (M) Testing": _tenor_months(r["start"], repay),
            "r (% p.a)": r["rate"],
            "Payment Frequency (M)": freq,
            "Repayment date": repay,
            "Repayment Amount": repay_amt,
            "Active (T/F)": 1,
        })
    return rows


def _fmt(v):
    if isinstance(v, dt.datetime):
        return v.strftime("%Y-%m-%d")
    return v


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("period", help="month to build, YYYY-MM (e.g. 2026-04)")
    ap.add_argument("--roll-to", metavar="YYYY-MM-DD",
                    help="evergreen mode: push every repayment to this date (past the model "
                         "horizon, e.g. 2029-12-31 — the model runs to Dec-2028) so no "
                         "principal repayment lands in the forecast (no phantom roll CF)")
    args = ap.parse_args()
    year, month = (int(x) for x in args.period.split("-"))
    asof = dt.datetime(year, month, calendar.monthrange(year, month)[1])
    roll_to = dt.datetime(*(int(x) for x in args.roll_to.split("-"))) if args.roll_to else None

    ledger = CLIENT / "raw" / f"{month:02d}" / "lender_loans_accrued_interest.xlsx"
    if not ledger.exists():
        print(f"ledger not found: {ledger}", file=sys.stderr)
        return 1

    rows = to_model_rows(read_ledger(ledger), asof=asof, roll_to=roll_to)

    suffix = "_evergreen" if roll_to else ""
    out = CLIENT / "budget" / f"loans_db_update_{args.period}{suffix}.csv"
    with out.open("w", newline="") as f:
        w = csv.writer(f)
        w.writerow(COLUMNS)
        for r in rows:
            w.writerow([_fmt(r[c]) for c in COLUMNS])

    total = sum(r["Principal Amount (EUR)"] for r in rows)
    blended = sum(r["Principal Amount (EUR)"] * r["r (% p.a)"] for r in rows) / total
    print(f"{len(rows)} active loans @ {asof:%d-%b-%Y} | total {total:,.0f} EUR | blended {blended:.2%}")
    print(f"paste-ready CSV -> {out}")
    print(f"\n{'Loan ID':8}{'Lender':24}{'Principal':>13}{'Rate':>6}  {'Start':>10}{'Maturity':>11}")
    for r in rows:
        print(f"{r['Loan ID']:8}{str(r['Lender Name'])[:23]:24}{r['Principal Amount (EUR)']:>13,.0f}"
              f"{r['r (% p.a)']:>6.1%}  {_fmt(r['Start Date']):>10}{_fmt(r['Repayment date']):>11}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
