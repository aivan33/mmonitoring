"""Build Almacena canonical taxonomi workbooks from the raw monthly statements.

Two tracked entities (`consolidated`, `ap_foundation`) driven by
``taxonomi_mapping.yaml``. Each canonical row pulls one or more source terms
(summed, with per-term sign) from the raw IS/CF/BS sheets; the builder asserts
each source row's label still matches the mapping's expectation and WARNs on
drift. Months after the requested period are left blank (an April run yields
Jan–Apr populated, May–Dec empty), mirroring the prior taxonomi.

The prior canonical workbook (`template`) doubles as the **reproduction
reference**: ``--validate`` recomputes Jan–Mar and diffs every cell against it.
It must report "Q1 reproduction clean" before an emitted month is trusted.

    uv run python clients/almacena/one_offs/build_taxonomi.py --validate
    uv run python clients/almacena/one_offs/build_taxonomi.py 2026-04
    # -> raw/taxonomi_act_04.xlsx (consolidated), raw/ap_04_act.xlsx (ap_foundation)
"""

from __future__ import annotations

import sys
from pathlib import Path

import yaml
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]          # clients/almacena
RAW = ROOT / "raw"
MAPPING = Path(__file__).resolve().parent / "taxonomi_mapping.yaml"
TOL = 1.0                                            # euro tolerance for reproduction


def _norm(s) -> str:
    return " ".join(str(s or "").split()).strip().lower()


def _num(v) -> float:
    return float(v) if isinstance(v, (int, float)) else 0.0


class Source:
    """One raw workbook; lazily materialises the sheets the mapping touches."""

    def __init__(self, path: Path):
        self.path = path
        self._wb = load_workbook(path, data_only=True, read_only=True)
        self._rows: dict[str, list] = {}

    def rows(self, sheet: str) -> list:
        if sheet not in self._rows:
            self._rows[sheet] = list(self._wb[sheet].iter_rows(values_only=True))
        return self._rows[sheet]

    def close(self):
        self._wb.close()


def _resolve_entity(ent: dict, n_months: int):
    """Return {(sheet, data, grp, subgroup): [v_jan, ...]} and a list of warnings."""
    # open each distinct source file once
    files = {cfg["file"] for cfg in ent["sheets"].values()}
    srcs = {f: Source(RAW / f) for f in files}
    warnings: list[str] = []

    def cell(scfg, row, mi):
        rows = srcs[scfg["file"]].rows(scfg["sheet"])
        stride = scfg.get("stride", 1)
        voff = scfg.get("voff", 0)
        col = scfg["jan_col"] + voff + mi * stride          # 1-based
        if row - 1 >= len(rows) or col - 1 >= len(rows[row - 1]):
            return 0.0
        return _num(rows[row - 1][col - 1])

    def label_at(scfg, row):
        rows = srcs[scfg["file"]].rows(scfg["sheet"])
        if row - 1 >= len(rows):
            return ""
        cells = rows[row - 1]
        c = scfg["label_col"] - 1
        return cells[c] if c < len(cells) else ""

    out: dict = {}
    for sheet_name, entries in ent["statements"].items():
        for entry in entries:
            data, grp, sub = entry["k"]
            spec = entry["r"]
            if spec == "none":
                out[(sheet_name, data, grp, sub)] = [None] * n_months
                continue
            if spec == "const0":
                out[(sheet_name, data, grp, sub)] = [0.0] * n_months
                continue
            vec = [0.0] * n_months
            for alias, row, exp_label, sign in spec:
                scfg = ent["sheets"][alias]
                actual = label_at(scfg, row)
                if _norm(exp_label) not in _norm(actual) and _norm(actual) not in _norm(exp_label):
                    warnings.append(
                        f"{data}/{grp}/{sub}: {alias} r{row} expected "
                        f"{exp_label!r} but found {str(actual).strip()!r}")
                for mi in range(n_months):
                    vec[mi] += sign * cell(scfg, row, mi)
            out[(sheet_name, data, grp, sub)] = vec
    for s in srcs.values():
        s.close()
    return out, warnings


def _template_keys(path: Path):
    """Yield (sheet, row, (data, grp, subgroup)) for every data row in a template."""
    wb = load_workbook(path, data_only=True, read_only=True)
    for sheet in wb.sheetnames:
        for ri, row in enumerate(wb[sheet].iter_rows(values_only=True), 1):
            if ri == 1:
                continue
            key = tuple(str(row[c]).strip() if c < len(row) and row[c] is not None else ""
                        for c in (0, 1, 2))
            if key != ("", "", ""):
                yield sheet, ri, key
    wb.close()


def validate(mapping: dict) -> int:
    bad = 0
    for ename, ent in mapping["entities"].items():
        values, warnings = _resolve_entity(ent, 3)
        for w in warnings:
            print(f"  WARNING [{ename}] {w}")
        # diff Jan–Mar against the template's own cells
        wb = load_workbook(RAW / ent["template"], data_only=True, read_only=True)
        diffs = []
        for sheet in wb.sheetnames:
            for ri, row in enumerate(wb[sheet].iter_rows(values_only=True), 1):
                if ri == 1:
                    continue
                key = tuple(str(row[c]).strip() if c < len(row) and row[c] is not None else ""
                            for c in (0, 1, 2))
                if key == ("", "", ""):
                    continue
                vec = values.get((sheet, *key))
                if vec is None:
                    diffs.append(f"{sheet} {key}: no mapping for template row")
                    continue
                for mi in range(3):
                    want = row[3 + mi] if 3 + mi < len(row) else None
                    got = vec[mi]
                    wv = _num(want)
                    gv = _num(got) if got is not None else 0.0
                    if abs(wv - gv) > TOL:
                        diffs.append(f"{sheet} {'/'.join(key)} m{mi+1}: "
                                     f"want {wv:,.2f} got {gv:,.2f} (Δ{gv-wv:+,.2f})")
        wb.close()
        if diffs:
            bad += len(diffs)
            print(f"\n[{ename}] {len(diffs)} reproduction diff(s):")
            for d in diffs:
                print(f"  {d}")
        else:
            print(f"[{ename}] Q1 reproduction clean "
                  f"({sum(1 for _ in _template_keys(RAW / ent['template']))} rows).")
    return 1 if bad else 0


def emit(mapping: dict, year: int, month: int) -> int:
    rc = 0
    for ename, ent in mapping["entities"].items():
        values, warnings = _resolve_entity(ent, month)
        for w in warnings:
            print(f"  WARNING [{ename}] {w}")
        wb = load_workbook(RAW / ent["template"])      # keep formatting
        for sheet, ri, key in list(_template_keys(RAW / ent["template"])):
            vec = values.get((sheet, *key))
            ws = wb[sheet]
            for mi in range(12):                       # Jan..Dec -> cols 4..15
                col = 4 + mi
                if vec is None or mi >= month or vec[mi] is None:
                    ws.cell(ri, col).value = None
                else:
                    ws.cell(ri, col).value = round(float(vec[mi]), 2)
        out = RAW / ent["output"].format(mm=f"{month:02d}")
        wb.save(out)
        wb.close()
        print(f"[{ename}] wrote {out.relative_to(ROOT.parent.parent)} "
              f"(Jan–month {month:02d} {year}).")
    return rc


def main(argv: list[str]) -> int:
    mapping = yaml.safe_load(MAPPING.read_text())
    if len(argv) == 1 and argv[0] == "--validate":
        return validate(mapping)
    if len(argv) == 1 and "-" in argv[0]:
        year, month = (int(x) for x in argv[0].split("-"))
        return emit(mapping, year, month)
    print(__doc__)
    return 2


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
