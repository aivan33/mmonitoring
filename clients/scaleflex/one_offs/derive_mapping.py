"""Derive Scaleflex's MR -> taxonomi mapping from the 2025 "Rosetta Stone".

Each 2025 management-report workbook carries BOTH the raw ``Income Statement`` /
``Cash Flow statement`` sheets AND fully-populated ``Taxonomy IS`` / ``Taxonomy CF``
sheets whose cells are pure references (e.g. ``='Income Statement'!C8``). Reading
those formulas gives the exact, authoritative row->taxonomi-key mapping the prior
engineer used — no value-matching guesswork.

This script:
  1. Reads the Taxonomy-sheet formulas straight from the xlsx XML (openpyxl's
     formula loader chokes on the workbook's pivot cache, so we parse the parts
     directly).
  2. Resolves each reference to a (section, label) in the raw sheet — section
     anchoring is required because ``Payroll BG/FR/VN`` repeat across
     CoS / R&D / S&M / G&A.
  3. VALIDATES the discovered mapping by re-extracting all 12 months of 2025 and
     comparing to the canonical standalone export. Must be clean (one documented
     exception) before the mapping is trusted.
  4. Emits ``mapping.yaml`` translated to the 2026 workbook's labels (the IS was
     re-laid-out for 2026: FileRobot->DAM, Storage->Infrastructure Cloud, etc.)
     plus the sign-off decisions for the new 2026 revenue lines.

Run once at onboarding (or whenever the source layout changes):

    uv run python clients/scaleflex/one_offs/derive_mapping.py
"""

from __future__ import annotations

import re
import sys
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path

import yaml
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_2025 = ROOT / "raw/2025/12/2025 - Management report 2025 (4).xlsx"
STANDALONE_2025 = ROOT / "raw/2025_taxonomi/Actuals EUR December 2025.xlsx"
OUT_MAPPING = ROOT / "mapping.yaml"

MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

# Section header labels per raw sheet — used to disambiguate repeated labels.
IS_SECTIONS = {"Sales", "Cost of Sales", "R&D", "S&M", "G&A"}
CF_SECTIONS = {"Money in", "Money out", "CoS", "R&D", "S&M", "G&A"}

# Per statement: (taxonomy sheet, raw sheet, label column 0-indexed, sections).
STATEMENTS = [
    ("IS", "Taxonomy IS", "Income Statement", 1, IS_SECTIONS),
    ("CF", "Taxonomy CF", "Cash Flow statement", 0, CF_SECTIONS),
]

# 2025 raw label -> 2026 raw label. The 2026 IS was rebuilt; CF barely changed.
RENAMES = {
    "Income Statement": {
        "FileRobot": "DAM", "CloudImage": "DMO",
        "Storage": "Infrastructure Cloud", "CDN": "Infrastructure CDN", "DNS": "Misc",
        "Events": "Event marketing", "Partners commissions": "Partner Commissions",
        "Finance income": "Financing income", "Finance costs": "Financing costs",
        "Travel, Events & Retreats": "Travel & Events",
        "Other operating expense (utilities, equipment, hardware, insurance etc.)":
            "Other operating expense",
    },
    "Cash Flow statement": {
        "Other operat expenses": "Other operating expenses",
    },
}

# Sign-off decisions for 2026 lines that had NO source in 2025 (see
# rolling discussion / AskUserQuestion answers). Keyed by taxonomi tuple.
OVERRIDES_2026 = {
    # 2026 single "VXP" revenue line -> VXP Developers; VXP Digital Teams left empty.
    ("Sales", "VXP Developers", "VXP Developers"):
        {"kind": "sum", "sheet": "Income Statement", "section": "Sales", "labels": ["VXP"]},
    # Visual AI is now a populated line in 2026.
    ("Sales", "Visual AI", "Visual AI"):
        {"kind": "src", "sheet": "Income Statement", "section": "Sales", "label": "Visual AI"},
    # "Services" + "Professional Services" both fold into Additional Services.
    ("Sales", "Additional Services", "Additional Services"):
        {"kind": "sum", "sheet": "Income Statement", "section": "Sales",
         "labels": ["Services", "Professional Services"]},
    # 2026 CF dropped the "Change in cash" ratio row — no source.
    ("% Change in cash", "% Change in cash", "% Change in cash"): {"kind": "none"},
}

# Known stale cell in the 2025 standalone export: the R&D "Software & Tools - R&D"
# row diverges from the workbook formula (=R&D section total). The workbook formula
# is authoritative, so this row is excepted from the standalone-reproduction gate.
VALIDATION_EXCEPTIONS = {("R&D", "Software & Tools - R&D", "Software & Tools - R&D")}

_NS = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
       "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships"}


def _taxonomy_formulas(path: Path, sheet: str) -> dict[int, str | None]:
    """Return {row -> formula string in column D (the Jan column)} for a sheet,
    parsed from the xlsx XML to dodge the pivot-cache loader bug."""
    z = zipfile.ZipFile(path)
    name2rid = {s.get("name"): s.get(f"{{{_NS['r']}}}id")
                for s in ET.fromstring(z.read("xl/workbook.xml")).find("m:sheets", _NS)}
    rid2tgt = {x.get("Id"): x.get("Target")
               for x in ET.fromstring(z.read("xl/_rels/workbook.xml.rels"))}
    sx = ET.fromstring(z.read("xl/" + rid2tgt[name2rid[sheet]].lstrip("/")))
    z.close()
    out: dict[int, str | None] = {}
    for c in sx.iter(f"{{{_NS['m']}}}c"):
        m = re.match(r"([A-Z]+)(\d+)", c.get("r"))
        if m.group(1) != "D":
            continue
        f = c.find("m:f", _NS)
        out[int(m.group(2))] = f.text if f is not None else None
    return out


def _parse_ref(formula: str | None) -> int | None:
    """Row number of a single-cell cross-sheet reference, else None."""
    m = re.match(r"'?[^'!]+'?!\$?[A-Z]+\$?(\d+)$", formula or "")
    return int(m.group(1)) if m else None


def _raw_row_meta(path: Path, sheet: str, label_col: int,
                  sections: set[str]) -> dict[int, tuple[str | None, str]]:
    """{row -> (section, label)} for every labelled row of a raw sheet."""
    wb = load_workbook(path, data_only=True, read_only=True)
    rows = list(wb[sheet].iter_rows(values_only=True))
    wb.close()
    meta: dict[int, tuple[str | None, str]] = {}
    current: str | None = None
    for i, row in enumerate(rows, 1):
        cell = row[label_col] if len(row) > label_col else None
        if cell is None or str(cell).strip() == "":
            continue
        label = str(cell).strip()
        if label in sections:
            current = label
        meta[i] = (current, label)
    return meta


def discover_mapping() -> list[dict]:
    """Return the discovered mapping entries (with 2025 source labels)."""
    entries: list[dict] = []
    for stmt, tax_sheet, raw_sheet, label_col, sections in STATEMENTS:
        formulas = _taxonomy_formulas(WORKBOOK_2025, tax_sheet)
        meta = _raw_row_meta(WORKBOOK_2025, raw_sheet, label_col, sections)
        wb = load_workbook(WORKBOOK_2025, data_only=True, read_only=True)
        tax_rows = list(wb[tax_sheet].iter_rows(values_only=True))
        wb.close()
        for trow, row in enumerate(tax_rows, 1):
            if trow == 1 or row[0] is None or str(row[0]).strip() == "":
                continue
            key = tuple(str(row[i]).strip() for i in range(3))
            formula = formulas.get(trow)
            entry = {"statement": stmt, "key": key}
            if formula is None:
                entry["source"] = {"kind": "none"}
            elif _parse_ref(formula) is None:
                entry["source"] = {"kind": "derived", "formula": formula.replace("D", "")}
            else:
                section, label = meta[_parse_ref(formula)]
                entry["source"] = {"kind": "src", "sheet": raw_sheet,
                                   "section": section, "label": label}
            entries.append(entry)
    return entries


# --- value extraction (shared shape with build_taxonomi.py) ----------------

def _jan_col(path: Path, sheet: str, label_col: int) -> int:
    wb = load_workbook(path, data_only=True, read_only=True)
    rows = list(wb[sheet].iter_rows(values_only=True))
    wb.close()
    for row in rows[:8]:
        for ci, c in enumerate(row):
            if str(c).strip() == "Jan":
                return ci
    return label_col + 1


def _lut(path: Path, sheet: str, label_col: int, sections: set[str]) -> dict:
    lut: dict[tuple[str | None, str], int] = {}
    for row, (section, label) in _raw_row_meta(path, sheet, label_col, sections).items():
        lut[(section, label)] = row
        lut.setdefault((None, label), row)
    return lut


def extract(path: Path, entries: list[dict], n_months: int,
            use_renames: bool) -> dict[tuple, list]:
    """Resolve every entry to an n-month value vector against ``path``."""
    sheets = {"Income Statement": (1, IS_SECTIONS),
              "Cash Flow statement": (0, CF_SECTIONS)}
    jan = {s: _jan_col(path, s, lc) for s, (lc, _) in sheets.items()}
    luts = {s: _lut(path, s, lc, sec) for s, (lc, sec) in sheets.items()}
    wb = load_workbook(path, data_only=True, read_only=True)
    raw = {s: list(wb[s].iter_rows(values_only=True)) for s in sheets}
    wb.close()

    def cell(sheet: str, row: int, mi: int):
        rows = raw[sheet]
        v = rows[row - 1][jan[sheet] + mi] if row - 1 < len(rows) else None
        return float(v) if isinstance(v, (int, float)) else 0.0  # Excel: blank ref -> 0

    def find(sheet: str, section, label):
        if use_renames:
            label = RENAMES.get(sheet, {}).get(label, label)
        return luts[sheet].get((section, label)) or luts[sheet].get((None, label))

    out: dict[tuple, list] = {}
    unmatched: list[tuple] = []
    for entry in entries:
        key, src = entry["key"], entry["source"]
        kind = src["kind"]
        if kind == "none":
            out[key] = [None] * n_months
        elif kind == "derived":
            out[key] = None  # filled below
        elif kind == "src":
            row = find(src["sheet"], src["section"], src["label"])
            if row is None:
                unmatched.append((key, src["label"]))
                out[key] = [None] * n_months
            else:
                out[key] = [cell(src["sheet"], row, mi) for mi in range(n_months)]
        elif kind == "sum":
            vec = [0.0] * n_months
            for label in src["labels"]:
                row = find(src["sheet"], src["section"], label)
                if row is None:
                    unmatched.append((key, label))
                    continue
                for mi in range(n_months):
                    vec[mi] += cell(src["sheet"], row, mi)
            out[key] = vec
    # derived: MRR = DAM + DMO
    dam = out[("Sales", "DAM", "DAM")]
    dmo = out[("Sales", "DMO", "DMO")]
    out[("MRR", "MRR", "MRR")] = [(dam[i] or 0) + (dmo[i] or 0) for i in range(n_months)]
    return out, unmatched


def _standalone_values(path: Path) -> dict[tuple, list]:
    wb = load_workbook(path, data_only=True, read_only=True)
    got: dict[tuple, list] = {}
    for sheet in ("IS (Actual)", "CF (Actual)"):
        for row in wb[sheet].iter_rows(min_row=2, values_only=True):
            if row[0] is None or str(row[0]).strip() == "":
                continue
            key = tuple(str(row[i]).strip() for i in range(3))
            got[key] = [float(v) if isinstance(v, (int, float)) else None
                        for v in row[3:15]]
    wb.close()
    return got


def validate_2025(entries: list[dict]) -> int:
    """Re-extract 2025 and diff against the standalone export. Returns diff count."""
    got, _ = extract(WORKBOOK_2025, entries, 12, use_renames=False)
    expected = _standalone_values(STANDALONE_2025)
    diffs = 0
    for key, exp in expected.items():
        if key in VALIDATION_EXCEPTIONS or key not in got or got[key] is None:
            continue
        for mi in range(12):
            a, b = got[key][mi], exp[mi]
            if a is None or b is None:
                if a == 0.0 and b is None:  # blank source rendered as 0 — benign
                    continue
                if a is None and b is None:
                    continue
                print(f"  NULL {key} {MONTHS[mi]}: got={a} exp={b}")
                diffs += 1
            elif abs(a - b) > 0.5:
                print(f"  DIFF {key} {MONTHS[mi]}: got={a} exp={b}")
                diffs += 1
    return diffs


def to_2026_mapping(entries: list[dict]) -> dict:
    """Translate discovered entries to 2026 labels + sign-off overrides."""
    by_stmt: dict[str, list] = {"IS": [], "CF": []}
    for entry in entries:
        key = entry["key"]
        src = OVERRIDES_2026.get(key, dict(entry["source"]))
        if src.get("kind") == "src" and "label" in src:
            src["label"] = RENAMES.get(src["sheet"], {}).get(src["label"], src["label"])
        out = {"data": key[0], "grp": key[1], "subgroup": key[2], "source": src}
        by_stmt[entry["statement"]].append(out)
    return {"mapping_is": by_stmt["IS"], "mapping_cf": by_stmt["CF"]}


def main() -> int:
    entries = discover_mapping()
    print(f"Discovered {len(entries)} taxonomi rows from 2025 formulas.")
    diffs = validate_2025(entries)
    if diffs:
        print(f"VALIDATION FAILED: {diffs} unexpected diff(s) vs 2025 standalone.")
        return 1
    print("2025 reproduction clean (1 documented exception). Mapping trusted.")
    mapping = to_2026_mapping(entries)
    OUT_MAPPING.write_text(
        "# Scaleflex MR -> taxonomi mapping (IS + CF; no BS).\n"
        "# Generated by one_offs/derive_mapping.py from the 2025 Taxonomy-sheet\n"
        "# formulas, translated to 2026 workbook labels. Section-anchored because\n"
        "# Payroll BG/FR/VN repeat across CoS/R&D/S&M/G&A. Sign off before locking.\n\n"
        + yaml.safe_dump(mapping, sort_keys=False, allow_unicode=True, width=100)
    )
    print(f"Wrote {OUT_MAPPING.relative_to(ROOT.parent.parent)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
