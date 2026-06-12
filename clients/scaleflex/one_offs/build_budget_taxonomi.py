"""Build the Scaleflex 2026 BUDGET taxonomi from the Business Model workbook.

Companion to ``build_taxonomi.py`` (which builds the *actuals* taxonomi from the
monthly management report). This reads ``budget_mapping.yaml`` and the 2026 month
columns of the Business Model workbook's ``IS_Monthly`` / ``CF_Monthly`` sheets,
then writes the canonical taxonomi (IS + CF; no BS) into ``IS (Budget)`` /
``CF (Budget)`` sheets by load-modify-save on the prior-year standalone template
so formatting carries over. The keys are identical to the actuals taxonomi, so
the report can compare Actual vs Plan row-for-row.

Resolution is section-anchored on the BM top-level blocks (``Payroll BG/FR/VN``
repeat across blocks) and treats a blank source cell as 0. The whole year
(Jan-Dec) is populated -- a budget is a full-year plan, not month-gated.

    uv run python clients/scaleflex/one_offs/build_budget_taxonomi.py
    # -> raw/taxonomi_budget_2026.xlsx
"""

from __future__ import annotations

import datetime as dt
import sys
from pathlib import Path

import yaml
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
MAPPING = ROOT / "budget_mapping.yaml"
TEMPLATE = ROOT / "raw/2025_taxonomi/Actuals EUR December 2025.xlsx"
BM = ROOT / "raw/Business Model_Scaleflex_until_2028_Weekly_CF.xlsx"
YEAR = 2026
OUT = ROOT / "raw/taxonomi_budget_2026.xlsx"

# BM label column (0-indexed col B) + top-level section headers per sheet.
LABEL_COL = 1
IS_SECTIONS = {"Sales", "Cost of Sales", "R&D", "S&M", "G&A"}
CF_SECTIONS = {"Money in", "CoS", "R&D", "S&M", "G&A"}
# source sheet -> (section set, target taxonomi sheet, renamed-from)
SHEETS = {
    "IS_Monthly": (IS_SECTIONS, "IS (Budget)", "IS (Actual)"),
    "CF_Monthly": (CF_SECTIONS, "CF (Budget)", "CF (Actual)"),
}
# ratio rows stored as float rather than rounded
FLOAT_KEYS = {("% Change in cash", "% Change in cash", "% Change in cash")}


def _jan_col(rows):
    """Column index whose row-2 header is the Jan of YEAR (datetime)."""
    for row in rows[:6]:
        for ci, c in enumerate(row):
            if isinstance(c, dt.datetime) and c.year == YEAR and c.month == 1:
                return ci
    raise SystemExit(f"Could not find a {YEAR}-01 date header in the BM sheet.")


def _row_meta(rows, sections):
    """Map row-number -> (current top-level section, label)."""
    meta, current = {}, None
    for i, row in enumerate(rows, 1):
        cell = row[LABEL_COL] if len(row) > LABEL_COL else None
        if cell is None or str(cell).strip() == "":
            continue
        label = str(cell).strip()
        if label in sections:
            current = label
        meta[i] = (current, label)
    return meta


def extract(mapping: dict):
    wb = load_workbook(BM, data_only=True, read_only=True)
    raw = {s: list(wb[s].iter_rows(values_only=True)) for s in SHEETS}
    wb.close()
    jan = {s: _jan_col(raw[s]) for s in SHEETS}
    lut = {}
    for s, (sec, _, _) in SHEETS.items():
        d = {}
        for r, (section, label) in _row_meta(raw[s], sec).items():
            d[(section, label)] = r
            d.setdefault((None, label), r)
        lut[s] = d

    def cell(sheet, row, mi):
        rows = raw[sheet]
        v = rows[row - 1][jan[sheet] + mi] if row - 1 < len(rows) else None
        return float(v) if isinstance(v, (int, float)) else 0.0

    def find(sheet, section, label):
        return lut[sheet].get((section, label)) or lut[sheet].get((None, label))

    out, unmatched = {}, []
    for section_key in ("mapping_is", "mapping_cf"):
        for entry in mapping[section_key]:
            key = (entry["data"], entry["grp"], entry["subgroup"])
            src = entry["source"]
            kind = src["kind"]
            if kind in ("none", "derived"):
                out[key] = [None] * 12
            elif kind == "src":
                row = find(src["sheet"], src.get("section"), src["label"])
                if row is None:
                    unmatched.append((key, src["label"]))
                    out[key] = [None] * 12
                else:
                    vec = [cell(src["sheet"], row, mi) for mi in range(12)]
                    if src.get("transform") == "abs":
                        vec = [abs(v) for v in vec]
                    out[key] = vec
            elif kind == "sum":
                vec = [0.0] * 12
                for label in src["labels"]:
                    row = find(src["sheet"], src.get("section"), label)
                    if row is None:
                        unmatched.append((key, label))
                        continue
                    for mi in range(12):
                        vec[mi] += cell(src["sheet"], row, mi)
                out[key] = vec
    dam, dmo = out[("Sales", "DAM", "DAM")], out[("Sales", "DMO", "DMO")]
    out[("MRR", "MRR", "MRR")] = [(dam[i] or 0) + (dmo[i] or 0) for i in range(12)]
    return out, unmatched


def write(values: dict):
    wb = load_workbook(TEMPLATE)  # clean file -- no pivot cache
    # keep only IS/CF; rename the Actual sheets to Budget (Scaleflex scope = IS+CF).
    for drop in ("BS (Actual)", "CF Indirect (Actual)"):
        if drop in wb.sheetnames:
            del wb[drop]
    for _, (_, tax_sheet, src_name) in SHEETS.items():
        ws = wb[src_name]
        ws.title = tax_sheet
        for r in range(2, ws.max_row + 1):
            key = tuple(
                (str(ws.cell(r, c).value).strip() if ws.cell(r, c).value is not None else "")
                for c in (1, 2, 3)
            )
            if key == ("", "", ""):
                continue
            vec = values.get(key)
            for mi in range(12):  # Jan..Dec live in columns 4..15
                col = 4 + mi
                if vec is None or vec[mi] is None:
                    ws.cell(r, col).value = None
                elif key in FLOAT_KEYS:
                    ws.cell(r, col).value = float(vec[mi])
                else:
                    ws.cell(r, col).value = round(float(vec[mi]), 2)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    wb.close()


def main() -> int:
    mapping = yaml.safe_load(MAPPING.read_text())
    values, unmatched = extract(mapping)
    if unmatched:
        print(f"WARNING: {len(unmatched)} source label(s) not found:")
        for key, label in unmatched:
            print(f"  {'/'.join(key)} <- {label!r}")
    write(values)
    print(f"Wrote {OUT.relative_to(ROOT.parent.parent)} (Budget {YEAR}, Jan-Dec).")
    return 0


if __name__ == "__main__":
    sys.exit(main())
