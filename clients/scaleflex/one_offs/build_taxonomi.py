"""Build a Scaleflex taxonomi-actual xlsx from a management-report workbook.

Reads ``mapping.yaml`` (generated/signed-off via ``derive_mapping.py``) and the
month columns of the MR workbook's ``Income Statement`` / ``Cash Flow statement``
sheets, then writes the canonical taxonomi (IS + CF; no BS) by load-modify-save on
the prior-year standalone template so formatting carries over.

Resolution is section-anchored (``Payroll BG/FR/VN`` repeat across blocks) and
treats a blank source cell as 0 (mirrors the workbook's ``='IS'!C8`` formulas).
Months after the requested period are left blank, so an April run yields Jan–Apr
populated and May–Dec empty.

    uv run python clients/scaleflex/one_offs/build_taxonomi.py \
        raw/management_report_26 2026-04
    # -> raw/taxonomi_act_apr26.xlsx
"""

from __future__ import annotations

import sys
from pathlib import Path

import yaml
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
MAPPING = ROOT / "mapping.yaml"
TEMPLATE = ROOT / "raw/2025_taxonomi/Actuals EUR December 2025.xlsx"

MONTH_ABBR = ["jan", "feb", "mar", "apr", "may", "jun",
              "jul", "aug", "sep", "oct", "nov", "dec"]
IS_SECTIONS = {"Sales", "Cost of Sales", "R&D", "S&M", "G&A"}
CF_SECTIONS = {"Money in", "Money out", "CoS", "R&D", "S&M", "G&A"}
# raw sheet -> (label column 0-indexed, section set, target taxonomi sheet)
SHEETS = {
    "Income Statement": (1, IS_SECTIONS, "IS (Actual)"),
    "Cash Flow statement": (0, CF_SECTIONS, "CF (Actual)"),
}
# ratio rows stored as float rather than rounded
FLOAT_KEYS = {("% Change in cash", "% Change in cash", "% Change in cash")}


def _row_meta(rows, label_col, sections):
    meta, current = {}, None
    for i, row in enumerate(rows, 1):
        cell = row[label_col] if len(row) > label_col else None
        if cell is None or str(cell).strip() == "":
            continue
        label = str(cell).strip()
        if label in sections:
            current = label
        meta[i] = (current, label)
    return meta


def _jan_col(rows, label_col):
    for row in rows[:8]:
        for ci, c in enumerate(row):
            if str(c).strip() == "Jan":
                return ci
    return label_col + 1


def extract(mr_path: Path, mapping: dict, n_months: int):
    wb = load_workbook(mr_path, data_only=True, read_only=True)
    raw = {s: list(wb[s].iter_rows(values_only=True)) for s in SHEETS}
    wb.close()
    jan = {s: _jan_col(raw[s], lc) for s, (lc, _, _) in SHEETS.items()}
    lut = {}
    for s, (lc, sec, _) in SHEETS.items():
        d = {}
        for r, (section, label) in _row_meta(raw[s], lc, sec).items():
            d[(section, label)] = r
            d.setdefault((None, label), r)
        lut[s] = d

    def cell(sheet, row, mi):
        rows = raw[sheet]
        v = rows[row - 1][jan[sheet] + mi] if row - 1 < len(rows) else None
        return float(v) if isinstance(v, (int, float)) else 0.0

    def find(sheet, section, label):
        return lut[sheet].get((section, label)) or lut[sheet].get((None, label))

    out, unmatched = {}, []
    for section_key in ("mapping_is", "mapping_cf"):
        for entry in mapping[section_key]:
            key = (entry["data"], entry["grp"], entry["subgroup"])
            src = entry["source"]
            kind = src["kind"]
            if kind in ("none", "derived"):
                out[key] = [None] * n_months
            elif kind == "src":
                row = find(src["sheet"], src.get("section"), src["label"])
                if row is None:
                    unmatched.append((key, src["label"]))
                    out[key] = [None] * n_months
                else:
                    out[key] = [cell(src["sheet"], row, mi) for mi in range(n_months)]
            elif kind == "sum":
                vec = [0.0] * n_months
                for label in src["labels"]:
                    row = find(src["sheet"], src.get("section"), label)
                    if row is None:
                        unmatched.append((key, label))
                        continue
                    for mi in range(n_months):
                        vec[mi] += cell(src["sheet"], row, mi)
                out[key] = vec
    dam, dmo = out[("Sales", "DAM", "DAM")], out[("Sales", "DMO", "DMO")]
    out[("MRR", "MRR", "MRR")] = [(dam[i] or 0) + (dmo[i] or 0) for i in range(n_months)]

    # Statement reconciliation: a source line that no mapping entry consumes is
    # otherwise dropped silently. Derive the CF money-in / money-out totals from
    # the mapped leaves and compare to the MR's own total rows; any gap means a
    # line was skipped (or double-counted) and must be mapped or categorised,
    # not lost. Catches e.g. an unmapped "Trainings" outflow.
    cf_in = {"Cash sales", "Financing inflows", "Money in"}
    cf_out = {"CoS", "Money out", "Financing outflows"}

    def _derived(cats):
        v = [0.0] * n_months
        for k, vec in out.items():
            if k[0] in cats and vec:
                for mi in range(n_months):
                    v[mi] += vec[mi] or 0.0
        return v

    recon = []
    for label, cats in (("Total money in", cf_in), ("Total money out", cf_out)):
        trow = find("Cash Flow statement", None, label)
        if trow is None:
            continue
        src_tot = [cell("Cash Flow statement", trow, mi) for mi in range(n_months)]
        der = _derived(cats)
        gaps = [(MONTH_ABBR[mi], round(src_tot[mi] - der[mi], 2))
                for mi in range(n_months) if abs(src_tot[mi] - der[mi]) > 1.0]
        if gaps:
            recon.append((label, gaps, round(sum(src_tot) - sum(der), 2)))
    return out, unmatched, recon


def write(values: dict, n_months: int, out_path: Path):
    wb = load_workbook(TEMPLATE)  # clean file — no pivot cache
    for _, _, tax_sheet in SHEETS.values():
        ws = wb[tax_sheet]
        for r in range(2, ws.max_row + 1):
            key = tuple(
                (str(ws.cell(r, c).value).strip() if ws.cell(r, c).value is not None else "")
                for c in (1, 2, 3)
            )
            if key == ("", "", ""):
                continue
            vec = values.get(key)
            for mi in range(12):  # Jan..Dec live in columns 4..15
                col = 4 + mi
                if vec is None or mi >= n_months or vec[mi] is None:
                    ws.cell(r, col).value = None
                elif key in FLOAT_KEYS:
                    ws.cell(r, col).value = float(vec[mi])
                else:
                    ws.cell(r, col).value = round(float(vec[mi]), 2)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    wb.close()


def main(argv: list[str]) -> int:
    if len(argv) != 2:
        print(__doc__)
        return 2
    mr_path = (ROOT / argv[0]) if not Path(argv[0]).is_absolute() else Path(argv[0])
    year, month = (int(x) for x in argv[1].split("-"))
    mapping = yaml.safe_load(MAPPING.read_text())
    values, unmatched, recon = extract(mr_path, mapping, month)
    if unmatched:
        print(f"WARNING: {len(unmatched)} source label(s) not found:")
        for key, label in unmatched:
            print(f"  {'/'.join(key)} <- {label!r}")
    if recon:
        print(f"WARNING: {len(recon)} CF total(s) don't reconcile — a source "
              f"line is unmapped (categorise it in 'Other' or add a mapping):")
        for label, gaps, ytd_gap in recon:
            detail = ", ".join(f"{m} {g:+,.0f}" for m, g in gaps)
            print(f"  {label}: YTD gap {ytd_gap:+,.0f}  (months: {detail})")
    out_path = ROOT / f"raw/taxonomi_act_{MONTH_ABBR[month - 1]}{year % 100}.xlsx"
    write(values, month, out_path)
    print(f"Wrote {out_path.relative_to(ROOT.parent.parent)} "
          f"(months Jan–{MONTH_ABBR[month - 1].capitalize()} {year}).")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
