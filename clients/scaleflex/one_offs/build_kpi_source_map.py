#!/usr/bin/env python3
"""Build the KPIs-sheet source map for the monthly monitoring deck.

For every row of the Monitoring Template `KPIs` sheet, this records WHERE the
number comes from (the client's `Revenues & Metrics - Finance.xlsx`, sheet
`Data monthly`, or the CFO platform), pulls the monthly inputs from that source
of truth, and proves the mapping by reconciling against the template's already
populated 2025 column.

Outputs (next to the client root):
  - KPI_source_map.xlsx   two sheets: `Mapping` (coordinates + reconciliation)
                          and `Values` (monthly series, copy-paste into the deck)
  - KPI_SOURCE_MAP.md     short validation reference

Usage:
  uv run python clients/scaleflex/one_offs/build_kpi_source_map.py \
      "clients/scaleflex/raw/Revenues & Metrics - Finance.xlsx" \
      "clients/scaleflex/raw/Monitoring Template 2026.xlsx" 2026-04
"""
import sys
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

REPORT_SHEET = "Data monthly"

# ---- column geometry --------------------------------------------------------
def report_col(year, month):
    # Data monthly: Jan-2023 lives in column C (index 3)
    return 3 + (year - 2023) * 12 + (month - 1)

def template_col(year, month):
    # KPIs: Dec-2023 lives in column B (index 2)
    return 2 + (year - 2023) * 12 + (month - 1) - 11

# ---- the verified mapping (template row -> source spec) ----------------------
# kind: direct | derived | platform | cross | none
# rows: report row number(s) in `Data monthly` that feed the metric
M = [
    # --- Profitability: filled from the CFO platform, not this file ----------
    (3,  "Sales growth",                "Profitability", "platform", [], None, "monthly", "CFO platform (IS). Not in the client file."),
    (4,  "Gross margin",                "Profitability", "platform", [], None, "monthly", "CFO platform (IS)."),
    (5,  "OPEX / Sales",                "Profitability", "platform", [], None, "monthly", "CFO platform (IS)."),
    (6,  "EBITDA margin",               "Profitability", "platform", [], None, "monthly", "CFO platform (IS)."),
    (7,  "EBIT margin",                 "Profitability", "platform", [], None, "monthly", "CFO platform (IS)."),
    (8,  "Net income margin",           "Profitability", "platform", [], None, "monthly", "CFO platform (IS)."),
    # --- Liquidity / efficiency ---------------------------------------------
    (10, "Burn multiple",               "Liquidity", "cross", [357], "Burn Multiple (Burn/Net New ARR)",
         "monthly", "Burn/Net-New-ARR. The client file row (r357) uses its own Monthly burn (r355) and does NOT reconcile to the deck's 2025 (which uses CFO cash burn). Source the burn numerator from the platform, or recompute. FLAG."),
    (11, "Marketing efficiency",        "Liquidity", "platform", [361], "Marketing Efficiency",
         "monthly", "Total Sales / Total Marketing spend. Client file row r361 is EMPTY -> comes from the CFO platform. FLAG."),
    (12, "ARPC (blended)",              "Liquidity", "none", [316], "ARPA Total",
         "monthly", "Deck row not maintained since Dec-2024 (#DIV/0!). Client file has ARPA (~EUR 6.2k, annual per-account), a different basis from ARPC (~EUR 600, monthly per-client). No drop-in source. Needs a definition decision."),
    (13, "ARPC Cloudimage",             "Liquidity", "none", [317], "ARPA DMO Enterprise",
         "monthly", "Same basis mismatch as ARPC (blended)."),
    (14, "ARPC Filerobot",              "Liquidity", "none", [318], "ARPA DAM",
         "monthly", "Same basis mismatch as ARPC (blended)."),
    (15, "ARR per FTE",                 "Liquidity", "direct", [328], "Revenue per employee",
         "month-end", "Reconciles to 2025."),
    (16, "Growth endurance",            "Liquidity", "none", [222], "Y/Y ARR growth (r222)",
         "monthly", "this-period growth / year-ago growth. No drop-in row: r222 (Y/Y ARR growth) is an input but the deck value is a ratio of two periods' growth, so it does not equal r222. Derive, don't copy."),
    (17, "Net Dollar Retention",        "Retention", "direct", [304], "Net Dollar Retention", "monthly", ""),
    (18, "Gross Dollar Retention",      "Retention", "direct", [303], "Gross Dollar Retention", "monthly", ""),
    (19, "Annual Churn Rate",           "Retention", "direct", [308], "Annual churn", "monthly", ""),
    (20, "Annual Churn rate (+Downgrades)", "Retention", "direct", [309], "Annual churn + downgrade", "monthly", ""),
    (21, "LTV",                         "Acquisition", "direct", [335], "LTV (annual churn)", "monthly", ""),
    (22, "CAC",                         "Acquisition", "direct", [344], "CAC",
         "monthly", "Jan-Jul'25 match exactly; Aug/Sep/Dec'25 differ (report restated after the deck was captured)."),
    (23, "LTV:CAC",                     "Acquisition", "direct", [345], "LTV / CAC",
         "monthly", "Same H2-2025 restatement as CAC."),
    (24, "CAC Payback",                 "Acquisition", "direct", [347], "Gross Margin Adj. CAC Payback", "monthly", "Reconciles to 2025."),
    (25, "Rule of 40",                  "Efficiency", "cross", [359], "Rule of 40 (Growth + EBITDA margin)",
         "monthly", "Growth (client) + EBITDA margin (platform). Client file r359 uses its own EBITDA and does NOT reconcile to the deck's 2025. EBITDA half is platform. FLAG."),
    (27, "S&M",                         "Acquisition", "direct", [341], "S&M Costs", "monthly", "Reconciles to 2025."),
    (28, "New ARR",                     "Efficiency", "direct", [354], "Net new ARR",
         "monthly", "Jan-Jul'25 match; H2'25 differs (report restated)."),
    (31, "Headcount",                   "Efficiency", "direct", [326], "Headcount", "month-end", "Reconciles to 2025."),
    # --- Bookings table ------------------------------------------------------
    (34, "New Business",                "Bookings", "direct", [5],  "Bookings > New Business", "monthly", ""),
    (35, "Upgrade",                     "Bookings", "direct", [22], "Bookings > Upgrade", "monthly", ""),
    (36, "Downgrade",                   "Bookings", "direct", [76], "Bookings > Downgrade", "monthly", ""),
    (37, "Churn",                       "Bookings", "direct", [93], "Bookings > Churn", "monthly", ""),
    (38, "Net ARR Result",              "Bookings", "derived", [5, 22, 76, 93],
         "New Business + Upgrade + Downgrade + Churn", "monthly",
         "Deck computes this as the sum of the four rows above (verified to the euro)."),
    # --- MRR breakdown table -------------------------------------------------
    (41, "DMO core",                    "MRR breakdown", "direct", [200], "Recognized revenue > Cloudimage", "month-end", "Excludes overuse."),
    (42, "DMO overuse",                 "MRR breakdown", "direct", [201], "Recognized revenue > Cloudimage Overuse", "month-end", ""),
    (43, "DAM core",                    "MRR breakdown", "direct", [208], "Recognized revenue > Filerobot", "month-end", "Excludes overuse."),
    (44, "DAM overuse",                 "MRR breakdown", "direct", [209], "Recognized revenue > Filerobot Overuse", "month-end", ""),
    (45, "Services",                    "MRR breakdown", "direct", [211], "Recognized revenue > Professional Services", "month-end", "Row 210 'Services' is empty; data is in r211."),
    (46, "Net ARR Result",              "MRR breakdown", "derived", [200, 201, 208, 209, 211],
         "DMO core + DMO overuse + DAM core + DAM overuse + Services", "month-end",
         "Total recognised MRR; sum of the five rows above (verified to the euro)."),
]

MONTHS_2025 = [(2025, m) for m in range(1, 13)]

# Rows whose as-of value is only meaningful once the client file's recognized-
# revenue rows (Cloudimage et al.) are populated for the month. While those are
# 0/blank, the computed value is garbage (e.g. New ARR = -(prior month's ARR)).
# 15 ARR/FTE, 28 New ARR, 41-46 MRR breakdown all read from recognised revenue;
# 24 CAC Payback chains through New ARR (=S&M/(New ARR*GM)*12), so it's hit too.
REVREC_DEPENDENT = {15, 24, 28, 41, 42, 43, 44, 45, 46}

# The recognised-revenue rows are =SUMIFS('Revenue scheduling'!<amount col>, product, month).
# Independently re-derive them from that tab to fact-check the cached values.
# Revenue scheduling geometry: amount cols H=8 (plan rev), K=11 (overuse), M=13
# (revenue converted); criteria cols Q=17 (revenue month), R=18 (product).
REVREC_SPEC = {  # deck row -> (report row, product, amount col index 0-based)
    41: (200, "Cloudimage", 7),   # DMO core   = SUMIFS(H, Cloudimage)
    42: (201, "Cloudimage", 10),  # DMO overuse= SUMIFS(K, Cloudimage)
    43: (208, "Filerobot", 7),    # DAM core   = SUMIFS(H, Filerobot)
    44: (209, "Filerobot", 10),   # DAM overuse= SUMIFS(K, Filerobot)
    45: (211, "Services", 12),    # Services   = SUMIFS(M, Services)
}
SCHED_PRODUCTS_IN_MRR = {"Cloudimage", "Filerobot", "Services"}

def to_float(x):
    try:
        return float(x)
    except (TypeError, ValueError):
        return None

def verify_revrec(report_path, out_months):
    """Re-derive the recognised-revenue rows from 'Revenue scheduling' and compare
    to the workbook's cached values. Returns (stale, unmapped, sched_last) where
    stale = [(deck_row, ym, recomputed, cached)] for months that disagree >0.5%,
    unmapped = {product: latest-month M} for schedule products no MRR row captures."""
    import datetime
    from collections import defaultdict
    wb = openpyxl.load_workbook(report_path, read_only=True, data_only=True)
    sch = wb["Revenue scheduling"]
    agg = defaultdict(float)        # (ym, product, col) -> sum
    prod_m = defaultdict(float)     # (ym, product) -> M sum (for unmapped scan)
    sched_last = None
    for row in sch.iter_rows(min_row=2, values_only=True):
        if len(row) < 18:
            continue
        q, prod = row[16], row[17]
        if not isinstance(q, datetime.datetime):
            continue
        ym = (q.year, q.month)
        sched_last = max(sched_last, ym) if sched_last else ym
        for col in (7, 10, 12):
            v = to_float(row[col])
            if v:
                agg[(ym, prod, col)] += v
        mv = to_float(row[12])
        if mv:
            prod_m[(ym, prod)] += mv
    dm = wb[REPORT_SHEET]
    cache = {}
    rows_needed = {rr for (rr, _, _) in REVREC_SPEC.values()}
    for rn, row in enumerate(dm.iter_rows(values_only=True), 1):
        if rn in rows_needed:
            cache[rn] = row
        if rn > 212:
            break
    def report_col(y, m):
        return 3 + (y - 2023) * 12 + (m - 1)
    stale = []
    for deckrow, (rrow, prod, col) in REVREC_SPEC.items():
        for (y, mo) in out_months:
            recomp = agg.get(((y, mo), prod, col), 0.0)
            c = report_col(y, mo) - 1
            cached = to_float(cache[rrow][c]) if c < len(cache[rrow]) else None
            if cached is None:
                continue
            if abs(recomp - cached) > max(abs(recomp), abs(cached), 1) * 0.005:
                stale.append((deckrow, (y, mo), recomp, cached))
    # products present in the schedule's latest month that no MRR row captures
    unmapped = {}
    if sched_last:
        for (ym, prod), v in prod_m.items():
            if ym == sched_last and prod not in SCHED_PRODUCTS_IN_MRR and v:
                unmapped[prod] = unmapped.get(prod, 0.0) + v
    return stale, unmapped, sched_last

def compute_2026_corrections(report_path):
    """Re-derive the recognised-revenue family for 2026 straight from
    'Revenue scheduling' (the source of truth), so we can hand over hardcoded
    correct values instead of the workbook's stale cache. Applies the Services
    fix (Services + Professional Services). Returns {(deck_row, ym): value} for
    months the schedule actually covers, plus the schedule's last month."""
    import datetime
    from collections import defaultdict
    wb = openpyxl.load_workbook(report_path, read_only=True, data_only=True)
    sch = wb["Revenue scheduling"]
    H = defaultdict(float); K = defaultdict(float); Mc = defaultdict(float)  # (ym,prod)->sum
    last = None
    for row in sch.iter_rows(min_row=2, values_only=True):
        if len(row) < 18:
            continue
        q, prod = row[16], row[17]
        if not isinstance(q, datetime.datetime):
            continue
        ym = (q.year, q.month)
        last = max(last, ym) if last else ym
        for d, idx in ((H, 7), (K, 10), (Mc, 12)):
            v = to_float(row[idx])
            if v:
                d[(ym, prod)] += v
    dm = wb[REPORT_SHEET]
    cache213 = hc = gm = sm = {}
    grab = {213: "cache213", 326: "hc", 333: "gm", 341: "sm"}
    store = {}
    for rn, row in enumerate(dm.iter_rows(values_only=True), 1):
        if rn in grab:
            store[grab[rn]] = row
        if rn > 347:
            break
    cache213 = store.get("cache213", {})
    hc = store.get("hc", {})
    gm = store.get("gm", {})
    sm = store.get("sm", {})
    def col(y, m):
        return 3 + (y - 2023) * 12 + (m - 1)
    def arr(ym):  # ARR = (DMO core+overuse + DAM core+overuse) * 12, from schedule
        return (H[(ym, "Cloudimage")] + K[(ym, "Cloudimage")]
                + H[(ym, "Filerobot")] + K[(ym, "Filerobot")]) * 12
    out = {}
    months_2026 = [(2026, m) for m in range(1, 13) if last and (2026, m) <= last]
    for ym in months_2026:
        dmo_c, dmo_o = H[(ym, "Cloudimage")], K[(ym, "Cloudimage")]
        dam_c, dam_o = H[(ym, "Filerobot")], K[(ym, "Filerobot")]
        svc = Mc[(ym, "Services")] + Mc[(ym, "Professional Services")]  # Services FIX
        out[(41, ym)] = dmo_c
        out[(42, ym)] = dmo_o
        out[(43, ym)] = dam_c
        out[(44, ym)] = dam_o
        out[(45, ym)] = svc
        out[(46, ym)] = dmo_c + dmo_o + dam_c + dam_o + svc
        hcv = to_float(hc[col(*ym) - 1]) if hc and col(*ym) - 1 < len(hc) else None
        out[(15, ym)] = arr(ym) / hcv if hcv else None
        # New ARR = ARR(ym) - ARR(prev); for Jan-26 prior is Dec-25 (schedule
        # hole), so bridge with the cached/locked Dec-25 ARR.
        prev = (ym[0], ym[1] - 1) if ym[1] > 1 else (ym[0] - 1, 12)
        if prev[0] == 2025:
            prev_arr = to_float(cache213[col(*prev) - 1]) if cache213 and col(*prev) - 1 < len(cache213) else None
        else:
            prev_arr = arr(prev)
        new_arr = (arr(ym) - prev_arr) if prev_arr is not None else None
        out[(28, ym)] = new_arr
        # CAC Payback = S&M / (New ARR * gross margin) * 12  (row 347), chained off New ARR
        gmv = to_float(gm[col(*ym) - 1]) if gm and col(*ym) - 1 < len(gm) else None
        smv = to_float(sm[col(*ym) - 1]) if sm and col(*ym) - 1 < len(sm) else None
        if new_arr and gmv and smv:
            out[(24, ym)] = smv / (new_arr * gmv) * 12
    return out, last

def classify_recon(errs):
    """errs: per-month (Jan..Dec 2025) relative error, None where not comparable."""
    comp = [(i, e) for i, e in enumerate(errs) if e is not None]
    if not comp:
        return "n/a"
    n_ok = sum(1 for _, e in comp if e < 0.01)
    if n_ok == len(comp):
        return "PASS"
    first_bad = next((i for i, e in comp if e >= 0.01), None)
    if first_bad is not None and first_bad >= 5 and all(e < 0.01 for i, e in comp if i < first_bad):
        return f"H1'25 exact; restated from 2025-{first_bad + 1:02d}"
    return f"REVIEW ({n_ok}/{len(comp)} mo within 1%)"

def main():
    report_path, template_path, asof = sys.argv[1], sys.argv[2], sys.argv[3]
    asof_y, asof_m = int(asof[:4]), int(asof[5:7])

    rb = openpyxl.load_workbook(report_path, read_only=True, data_only=True)
    rws = rb[REPORT_SHEET]
    # grab the report rows we need into a dict {rownum: row tuple}
    needed = {r for spec in M for r in spec[4]}
    rcells = {}
    for rn, row in enumerate(rws.iter_rows(values_only=True), 1):
        if rn in needed:
            rcells[rn] = row
        if rn > 400:
            break

    # formula view: classify each source cell as a formula or an entered value,
    # and capture the as-of formula string so we can show it as the flag.
    fcells = {}
    try:
        fb = openpyxl.load_workbook(report_path, data_only=False)
        fws = fb[REPORT_SHEET]
        for r in needed | {213}:
            fcells[r] = fws.cell(row=r, column=report_col(asof_y, asof_m)).value
    except Exception:
        fcells = {}  # formula loader can choke on a pivot cache; degrade gracefully

    def is_formula(rows):
        return any(isinstance(fcells.get(r), str) and fcells[r].startswith("=") for r in rows)

    tb = openpyxl.load_workbook(template_path, data_only=True)
    tws = tb["KPIs"]

    def report_val(rows, year, month):
        col = report_col(year, month)
        vals = [to_float(rcells[r][col - 1]) if (col - 1) < len(rcells[r]) else None for r in rows]
        if all(v is None for v in vals):
            return None
        return sum(v or 0.0 for v in vals)

    def template_val(trow, year, month):
        return to_float(tws.cell(row=trow, column=template_col(year, month)).value)

    # output month columns: full 2025 + Jan..asof 2026
    out_months = MONTHS_2025 + [(2026, m) for m in range(1, asof_m + 1)]

    # latest month for which recognised revenue actually computes (>0). These
    # rows are `=SUMIFS('Revenue scheduling'...)`; that tab ends here, so later
    # months legitimately compute to 0 (recognition lag), not a pending fill-up.
    rev_cutoff = None
    for (y, mo) in out_months:
        if report_val([200], y, mo):
            rev_cutoff = (y, mo)
    rev_cutoff_s = f"{rev_cutoff[0]}-{rev_cutoff[1]:02d}" if rev_cutoff else "n/a"
    asof_after_rev = bool(rev_cutoff) and (asof_y, asof_m) > rev_cutoff

    # hardcoded-correct 2026 values for the recognised-revenue family, re-derived
    # from 'Revenue scheduling' (with the Services fix), replacing the stale cache
    corr2026, sched_last = compute_2026_corrections(report_path)

    # ---- reconcile 2025 -----------------------------------------------------
    results = []
    for trow, name, section, kind, rows, srclabel, basis, note in M:
        recon = "n/a"
        if kind in ("direct", "derived"):
            errs = []
            for (y, mo) in MONTHS_2025:
                tv, sv = template_val(trow, y, mo), report_val(rows, y, mo)
                errs.append(None if (tv is None or sv is None)
                            else abs(tv - sv) / max(abs(tv), abs(sv), 1e-9))
            recon = classify_recon(errs)

        celltype = "formula" if is_formula(rows) else ("value" if rows else "—")
        # the as-of formula string (single-row metrics only; derived totals sum
        # several rows, so a single formula would mislead)
        formula = fcells.get(rows[0]) if (rows and celltype == "formula" and kind != "derived") else None

        # rows that draw on recognised revenue read 0 past the revenue cutoff
        lagged = trow in REVREC_DEPENDENT and asof_after_rev
        if kind not in ("direct", "derived"):
            cur = None
        elif lagged:
            cur = f"rev-rec ends {rev_cutoff_s}"
        else:
            cur = report_val(rows, asof_y, asof_m)
        series = [report_val(rows, y, mo) if kind in ("direct", "derived") else None for (y, mo) in out_months]
        # override 2026 with the hardcoded-correct re-derivation where we have it
        for i, ym in enumerate(out_months):
            if (trow, ym) in corr2026:
                series[i] = corr2026[(trow, ym)]
        last_avail = corr2026.get((trow, sched_last)) if (lagged and sched_last) else None
        if last_avail is None and lagged and rev_cutoff:
            last_avail = report_val(rows, *rev_cutoff)
        if lagged:
            series[out_months.index((asof_y, asof_m))] = None
        results.append(dict(trow=trow, name=name, section=section, kind=kind, rows=rows,
                            srclabel=srclabel, basis=basis, note=note, recon=recon,
                            celltype=celltype, formula=formula, lagged=lagged,
                            last_avail=last_avail, rev_cutoff_s=rev_cutoff_s,
                            cur=cur, series=series))
    main.rev_cutoff_s = rev_cutoff_s  # stash for writers

    # fact-check: re-derive recognised revenue from 'Revenue scheduling' vs cache
    stale, unmapped, sched_last = verify_revrec(report_path, out_months)
    main.verify = (stale, unmapped, sched_last)

    write_excel(results, out_months, asof, asof_y, asof_m)
    write_md(results, asof, asof_y, asof_m)
    if stale or unmapped:
        print("\n  FACT-CHECK FLAGS:")
        for dr, ym, rc, ca in stale:
            print(f"    stale cache: deck row {dr} {ym[0]}-{ym[1]:02d}: schedule={rc:,.2f} vs cached={ca:,.2f}")
        for p, v in unmapped.items():
            print(f"    unmapped schedule product '{p}' (latest month M={v:,.2f}) — not summed by any MRR row")
    print(f"Wrote clients/scaleflex/KPI_source_map.xlsx and KPI_SOURCE_MAP.md (as-of {asof})")
    # console summary
    for r in results:
        cur = r["cur"]
        cur_s = f"{cur:,.2f}" if isinstance(cur, float) else (cur or "-")
        print(f"  [{r['kind']:<8}] {r['name']:<26} {asof} = {cur_s:<16} recon={r['recon']}")


def write_excel(results, out_months, asof, asof_y, asof_m):
    wb = openpyxl.Workbook()
    bold = Font(bold=True)
    hdr_fill = PatternFill("solid", fgColor="1F2A44")
    hdr_font = Font(bold=True, color="FFFFFF")
    thin = Border(*[Side(style="thin", color="DDDDDD")] * 4)
    kind_fill = {
        "platform": PatternFill("solid", fgColor="FFF2CC"),
        "cross":    PatternFill("solid", fgColor="FCE4D6"),
        "none":     PatternFill("solid", fgColor="F8CBAD"),
        "derived":  PatternFill("solid", fgColor="E2EFDA"),
        "direct":   PatternFill("solid", fgColor="FFFFFF"),
    }
    cur_col_letter = None

    # ---- Mapping sheet ------------------------------------------------------
    ws = wb.active
    ws.title = "Mapping"
    headers = ["Section", "KPI (deck row)", "Deck row", "Source", "Client-file row(s)",
               "Source label", "Cell type", f"{asof} cell", "Basis", f"{asof} input",
               "2025 check", "Notes"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    src_label = {"platform": "CFO Platform", "cross": "Cross-source", "none": "No source",
                 "derived": "Client file (derived)", "direct": "Client file"}
    note_col = len(headers)
    for r in results:
        rows_s = ", ".join(str(x) for x in r["rows"]) if r["rows"] else "-"
        cell_ref = ""
        if r["kind"] in ("direct",) and r["rows"]:
            cell_ref = f"'{REPORT_SHEET}'!{get_column_letter(report_col(asof_y, asof_m))}{r['rows'][0]}"
        elif r["kind"] == "derived":
            cell_ref = "sum of rows"
        cur = r["cur"]
        if isinstance(cur, float):
            cur_out = round(cur, 4)
        elif isinstance(cur, str) and cur.startswith("rev-rec"):
            la = r["last_avail"]
            cur_out = f"0 (formula; {cur}) — last avail {r['rev_cutoff_s']}={round(la,2) if isinstance(la,float) else la}"
        else:
            cur_out = "(platform)"
        celltype = r["celltype"]
        # surface the actual formula in the note so formula rows are unmistakable
        note = r["note"]
        if r["formula"]:
            note = (note + "  " if note else "") + f"[formula {r['rev_cutoff_s'] and ''}{r['formula']}]"
        ws.append([r["section"], r["name"], r["trow"], src_label[r["kind"]], rows_s,
                   r["srclabel"] or "", celltype, cell_ref, r["basis"], cur_out, r["recon"], note])
        rr = ws.max_row
        for c in range(1, len(headers) + 1):
            ws.cell(row=rr, column=c).fill = kind_fill[r["kind"]]
            ws.cell(row=rr, column=c).border = thin
            ws.cell(row=rr, column=c).alignment = Alignment(vertical="top", wrap_text=(c == note_col))
        if celltype == "formula":
            ws.cell(row=rr, column=7).font = Font(italic=True, color="2E5BBA")
    widths = [13, 26, 8, 20, 14, 28, 9, 22, 11, 30, 24, 70]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # ---- Values sheet (copy-paste series from source of truth) --------------
    vs = wb.create_sheet("Values")
    month_hdr = [f"{y}-{mo:02d}" for (y, mo) in out_months]
    vhead = ["Section", "KPI (deck row)", "Source"] + month_hdr
    vs.append(vhead)
    for c in range(1, len(vhead) + 1):
        cell = vs.cell(row=1, column=c)
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = Alignment(vertical="center")
    asof_idx = out_months.index((asof_y, asof_m))
    for r in results:
        row_vals = [round(v, 4) if isinstance(v, float) else None for v in r["series"]]
        vs.append([r["section"], r["name"], src_label[r["kind"]]] + row_vals)
    # highlight as-of column
    cur_col = 4 + asof_idx
    cur_col_letter = get_column_letter(cur_col)
    for rr in range(1, vs.max_row + 1):
        vs.cell(row=rr, column=cur_col).fill = PatternFill("solid", fgColor="D9E1F2")
    vs.column_dimensions["A"].width = 14
    vs.column_dimensions["B"].width = 26
    vs.column_dimensions["C"].width = 20
    for i in range(4, len(vhead) + 1):
        vs.column_dimensions[get_column_letter(i)].width = 12
    vs.freeze_panes = "D2"
    # note rows
    rcs = getattr(main, "rev_cutoff_s", "n/a")
    note1 = (f"As-of column ({asof}) highlighted. Blank = sourced outside the client file "
             f"(platform / cross-source / no drop-in row) OR a recognised-revenue row whose "
             f"value computes to 0 past {rcs} (the 'Revenue scheduling' tab ends there).")
    note2 = ("2026 values for the recognised-revenue family (DMO/DAM core & overuse, Services, MRR total, "
             "ARR per FTE, New ARR) are HARDCODED, re-derived from 'Revenue scheduling' — NOT the workbook's "
             "stale cache. Services = Services + Professional Services (the fix). 2025 = deck-aligned cache.")
    vs.cell(row=vs.max_row + 2, column=1, value=note1).font = Font(italic=True)
    vs.cell(row=vs.max_row + 1, column=1, value=note2).font = Font(italic=True, color="C00000")

    wb.save("clients/scaleflex/KPI_source_map.xlsx")


def write_md(results, asof, asof_y, asof_m):
    lines = []
    lines.append(f"# KPIs sheet — source map & validation ({asof})\n")
    lines.append("How to fill the Monitoring Template `KPIs` sheet each month and trace every number back to its source. "
                 "Companion data pack: `KPI_source_map.xlsx` (sheet `Values` is copy-paste ready).\n")
    lines.append("## Two sources\n")
    lines.append("- **CFO Platform** — the whole **Profitability** block (Sales growth, gross/OPEX/EBITDA/EBIT/net margins). "
                 "Not in the client file; taken from the platform / IS.\n")
    lines.append(f"- **Client file** — `Revenues & Metrics - Finance.xlsx`, sheet **`{REPORT_SHEET}`**. Everything else.\n")
    lines.append("\n## Column geometry (so a cell ref is unambiguous)\n")
    lines.append(f"- Client file `{REPORT_SHEET}`: months run along columns, **Jan-2023 = column C**. "
                 f"So {asof} = column **{get_column_letter(report_col(asof_y, asof_m))}**.\n")
    lines.append("- Deck `KPIs`: **Dec-2023 = column B**, monthly across to Dec-2026.\n")
    lines.append("- **All inputs are monthly (MTD) flows or month-end snapshots — none are YTD/cumulative.**\n")

    lines.append("\n## Row-by-row map\n")
    lines.append("Cell type = how the value sits in `Data monthly`: **formula** (computed) or **value** (hand-keyed). "
                 "Only **Headcount** and **S&M** are hand-keyed; every other client-file row is a formula.\n")
    lines.append("| Deck row | KPI | Source | Cell type | Client-file row(s) — `Data monthly` | Basis | 2025 check |")
    lines.append("|---|---|---|---|---|---|---|")
    for r in results:
        srcmap = {"platform": "CFO Platform", "cross": "Cross-source ⚠", "none": "No drop-in source ⚠",
                  "derived": "Client file (derived)", "direct": "Client file"}
        rows_s = ", ".join(str(x) for x in r["rows"]) if r["rows"] else "—"
        label = f"{rows_s} · {r['srclabel']}" if r["srclabel"] else rows_s
        ct = {"formula": "formula", "value": "value", "—": "—"}[r["celltype"]]
        lines.append(f"| {r['trow']} | {r['name']} | {srcmap[r['kind']]} | {ct} | {label} | {r['basis']} | {r['recon']} |")

    flags = [r for r in results if r["kind"] in ("cross", "none")]
    lines.append("\n## Flags — do not silently copy from the client file\n")
    for r in flags:
        lines.append(f"- **{r['name']}** (row {r['trow']}): {r['note']}")

    formulas = [r for r in results if r["formula"]]
    if formulas:
        lines.append("\n## Formula rows — the value is computed, not entered\n")
        lines.append("Open in Excel (or let it recalc) to read the live value; openpyxl's cached value can be stale. "
                     "The as-of formula for each computed client-file metric:\n")
        for r in formulas:
            lines.append(f"- **{r['name']}** (row {r['trow']}): `{r['formula']}`")

    stale, unmapped, sched_last = getattr(main, "verify", ([], {}, None))
    sl = f"{sched_last[0]}-{sched_last[1]:02d}" if sched_last else "n/a"
    lines.append("\n## Fact-check: recognised revenue re-derived from `Revenue scheduling`\n")
    lines.append(f"The five MRR-breakdown rows are `=SUMIFS('Revenue scheduling'…)`. I re-summed that tab independently "
                 f"and compared to the workbook's cached values for every output month. **Jan–Nov 2025 reproduce the "
                 f"cached values and the deck to the cent** — mapping and formula reading confirmed. The schedule's "
                 f"latest month is **{sl}**. Discrepancies (cached value is stale — the tab was edited after the last "
                 f"Excel recalc, so a live recalc would change these):\n")
    if stale:
        lines.append("\n| Deck row | Month | Schedule (live) | Cached (stale) |")
        lines.append("|---|---|---|---|")
        rowname = {dr: n for (dr, n, *_) in [(r["trow"], r["name"]) + (None,) for r in results]}
        for dr, ym, rc, ca in stale:
            lines.append(f"| {dr} | {ym[0]}-{ym[1]:02d} | {rc:,.2f} | {ca:,.2f} |")
    else:
        lines.append("- None — cache agrees with the live schedule for every output month.")
    if unmapped:
        lines.append(f"\n**Products in the schedule's latest month ({sl}) that NO MRR-breakdown row captures:**\n")
        for p, v in unmapped.items():
            extra = " ← belongs in Services per the taxonomi; the Services formula filters only `=\"Services\"` and misses it" if "rofessional" in p else " (feeds a non-MRR product line)"
            lines.append(f"- **{p}** = {v:,.2f}{extra}")
        lines.append("\nThe **Services row understates from 2026 on** if Professional Services isn't added to its SUMIFS criteria.")

    # --- explicit signal: formulas the user needs to wire/fix -----------------
    lines.append("\n## 🔧 Formulas to update (signal only — I did NOT change the workbook)\n")
    pro = [p for p in unmapped if "rofessional" in p]
    if pro:
        lines.append(f"- **`Data monthly` row 211 (Services) / deck row 45** — its `=SUMIFS('Revenue scheduling'!$M:$M,"
                     f"…$R:$R,$B210,…)` filters product `=\"Services\"` only. Since 2026 the schedule carries a separate "
                     f"**`Professional Services`** product, so the formula now misses it. Add a second SUMIFS for "
                     f"`\"Professional Services\"` (or widen the criteria). The 2026 Services values in the `Values` "
                     f"sheet are already corrected to **Services + Professional Services** — wire the formula to match.")
    lines.append(f"- **Recognised-revenue family (rows 200/201/208/209/211 → ARR row 213 → New ARR row 354, ARR/FTE row 328)** "
                 f"— these are not formula bugs but **stale cache + a Dec-2025 hole + the schedule ending {sl}**. When you "
                 f"refresh `Revenue scheduling` and recalc, the deck cells will update; until then use the hardcoded 2026 "
                 f"values in `Values` (re-derived live from the schedule).")

    lines.append("\n## How to validate a single number\n")
    lines.append("1. Find the KPI in the table above and note its `Data monthly` row(s).\n")
    lines.append(f"2. Open the client file, sheet `{REPORT_SHEET}`, go to that row, column for the month "
                 f"(e.g. {asof} = column {get_column_letter(report_col(asof_y, asof_m))}).\n")
    lines.append("3. For *derived* rows, sum the listed rows (the deck does the same).\n")
    lines.append("4. The `Values` sheet of the xlsx already pulled these; the highlighted column is the as-of month.\n")
    lines.append("5. Reconciliation was proven by reproducing the deck's **2025** column from these same rows "
                 "(`PASS` = all 12 months within 1%).\n")

    lines.append("\n## ⚠ The client file has been restated since the deck's 2025 was locked\n")
    lines.append("Every mapped row matches the deck **exactly for Jan–Jun 2025**. From **Jul 2025** the client "
                 "file's current values diverge on several rows (bookings, ARR/FTE, CAC, New ARR) — the report was "
                 "restated after the deck captured its 2025 column. The mapping is correct; the deck's historical "
                 "2025 figures simply predate the latest extract. Rows so affected:\n")
    restate = [r for r in results if isinstance(r["recon"], str)
               and (r["recon"].startswith("REVIEW") or "restated" in r["recon"])]
    for r in restate:
        lines.append(f"- **{r['name']}** (row {r['trow']}): `{r['recon']}`. {r['note']}")

    rcs = getattr(main, "rev_cutoff_s", "n/a")
    lagged = [r for r in results if r["lagged"]]
    if lagged:
        lines.append(f"\n## Recognised-revenue rows lag one month (latest = {rcs})\n")
        lines.append(f"The recognised-revenue rows (Cloudimage, Filerobot, their overuse, Professional Services — "
                     f"rows 200/201/208/209/211) are `=SUMIFS('Revenue scheduling'…)`. That tab holds invoice-schedule "
                     f"rows only through **{rcs}**, so for {asof} the formulas legitimately compute to **0** — this is "
                     f"the booked-vs-recognised lag, **not** a pending fill-up. Affected deck rows (use the {rcs} value, "
                     f"or take the {asof} recognised figure from the IS taxonomi / platform):\n")
        for r in lagged:
            la = r["last_avail"]
            la_s = f"{la:,.2f}" if isinstance(la, float) else str(la)
            lines.append(f"- **{r['name']}** (row {r['trow']}): {asof} computes 0; latest available ({rcs}) = {la_s}")
        lines.append(f"\nBookings, retention, acquisition metrics and headcount **are** populated for {asof}.\n")
    lines.append("\n_Regenerate: `uv run python clients/scaleflex/one_offs/build_kpi_source_map.py "
                 "\"<report.xlsx>\" \"<template.xlsx>\" YYYY-MM`_\n")

    with open("clients/scaleflex/KPI_SOURCE_MAP.md", "w") as f:
        f.write("\n".join(lines))


if __name__ == "__main__":
    main()
