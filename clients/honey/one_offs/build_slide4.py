"""One-off: render slide-4 B2C charts for Honey's Q1 2026 deck.

Lives under ``clients/honey/one_offs/`` because operational subscription
data isn't yet modeled in ``core/data/``. This script will be retired
when the operational data layer lands and the slide-4 charts can be
expressed as ordinary chart specs.

Two charts, 10 months Jun-25 → Mar-26:

1. Active Subscriptions by Package — stacked bars, 5 categories
   (Subscription / Custom plan / Adopt a hive / Adopt a hive + / Support a hive).
2. Subscription Flow — diverging stacked bars (positive: New + Resumed;
   negative: Cancelled + Paused).

Methodology (validated against Dec-25 chart):
- Active by type: rows in month's B2C SUB file where col 6 ('На пауза') is
  empty, grouped by col 11 ('Тип на абонамента'). Latin AM and Cyrillic
  АМ both mean "Subscription" in source — merged.
- Flow:
    New        = ids in month T not in T-1
    Cancelled  = ids in T-1 not in T
    Paused     = paused-set in T minus paused-set in T-1
    Resumed    = paused-set in T-1 minus paused-set in T

Months for which we don't have SUB files (Jun-25 .. Oct-25) fall back to
values transcribed from the previous Dec-25 deck's slide-4 chart.

Outputs:
    clients/honey/one_offs/_out/slide4_active_b2c.png
    clients/honey/one_offs/_out/slide4_flow_b2c.png
"""

from __future__ import annotations

import datetime as dt
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
from openpyxl import load_workbook


_HERE = Path(__file__).resolve().parent
_CLIENT = _HERE.parent
_OUT = _HERE / "_out"


# Display order for the active-by-package stack (bottom → top).
_PACKAGE_ORDER = [
    "Subscription",
    "Custom plan",
    "Adopt a hive",
    "Adopt a hive +",
    "Support a hive",
]

# Map from raw SUB-file 'Тип на абонамента' codes to display labels.
_TYPE_TO_LABEL = {
    "AM":  "Subscription",     # Latin AM
    "АМ":  "Subscription",     # Cyrillic АМ — merged
    "CS":  "Custom plan",
    "ОК":  "Adopt a hive",
    "ОК+": "Adopt a hive +",
    "ПК":  "Support a hive",
}

# Brand-ish palette: green-teal family, distinct lightness ramps for each
# package band, plus dedicated colours for the flow chart.
_PACKAGE_PALETTE = {
    "Subscription":   "#1F3D3A",
    "Custom plan":    "#3B6160",
    "Adopt a hive":   "#5B8A87",
    "Adopt a hive +": "#88B0AC",
    "Support a hive": "#BCD6D2",
}
_FLOW_PALETTE = {
    "New":       "#2E7D54",
    "Resumed":   "#90C9A9",
    "Cancelled": "#C24C44",
    "Paused":    "#E89B5A",
}

# Months on the X axis (10 columns).
_MONTHS = [
    ("Jun 25", dt.date(2025, 6, 1)),
    ("Jul 25", dt.date(2025, 7, 1)),
    ("Aug 25", dt.date(2025, 8, 1)),
    ("Sep 25", dt.date(2025, 9, 1)),
    ("Oct 25", dt.date(2025, 10, 1)),
    ("Nov 25", dt.date(2025, 11, 1)),
    ("Dec 25", dt.date(2025, 12, 1)),
    ("Jan 26", dt.date(2026, 1, 1)),
    ("Feb 26", dt.date(2026, 2, 1)),
    ("Mar 26", dt.date(2026, 3, 1)),
]

# Filenames for the per-month B2C SUB files. None = no file uploaded; use
# transcribed chart values instead.
_SUB_FILES = {
    "Nov 25": "raw/sales_report/11-25/Copy of B2C SUB November 2025.xlsx",
    "Dec 25": "raw/sales_report/12-25/B2C SUB December 2025.xlsx",
    "Jan 26": "raw/sales_report/01/B2C SUB Jan 2026.xlsx",
    "Feb 26": "raw/sales_report/02/B2C SUB Feb 2026.xlsx",
    "Mar 26": "raw/sales_report/03/B2C SUB March 2026.xlsx",
}

# Transcribed from the Dec-25 deck's slide-4 chart 1 (per the user, AM Cyrillic
# and AM Latin are both "Subscription" — values are pre-merged here).
_ACTIVE_FROM_CHART = {
    "Jun 25": {"Subscription": 191, "Custom plan": 81, "Adopt a hive": 418,
               "Adopt a hive +": 70, "Support a hive": 326},
    "Jul 25": {"Subscription": 177, "Custom plan": 67, "Adopt a hive": 383,
               "Adopt a hive +": 60, "Support a hive": 295},
    "Aug 25": {"Subscription": 171, "Custom plan": 70, "Adopt a hive": 375,
               "Adopt a hive +": 58, "Support a hive": 281},
    "Sep 25": {"Subscription": 169, "Custom plan": 68, "Adopt a hive": 381,
               "Adopt a hive +": 57, "Support a hive": 276},
    "Oct 25": {"Subscription": 171, "Custom plan": 62, "Adopt a hive": 375,
               "Adopt a hive +": 60, "Support a hive": 275},
}

# Transcribed flow values from chart 2. Used for Jun-Nov 25 (no t-1 file
# available for those months to do cross-month diff). Empty = blank in
# original chart.
_FLOW_FROM_CHART = {
    "Jun 25": {"New": 22, "Resumed": 0, "Cancelled": 0, "Paused": 0},
    "Jul 25": {"New": 29, "Resumed": 0, "Cancelled": 30, "Paused": 104},
    "Aug 25": {"New": 20, "Resumed": 15, "Cancelled": 30, "Paused": 31},
    "Sep 25": {"New": 16, "Resumed": 28, "Cancelled": 23, "Paused": 26},
    "Oct 25": {"New": 63, "Resumed": 26, "Cancelled": 91, "Paused": 21},
    "Nov 25": {"New": 21, "Resumed": 17, "Cancelled": 36, "Paused": 9},
}


# ---------------------------------------------------------------------------
# Data extraction
# ---------------------------------------------------------------------------

@dataclass(frozen=True)
class Snapshot:
    """One month's per-subscription state from a B2C SUB file."""
    ids: frozenset[str]
    paused_ids: frozenset[str]
    type_by_id: dict[str, str]


def _load_snapshot(path: Path) -> Snapshot:
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    ids: list[str] = []
    paused: list[str] = []
    type_by_id: dict[str, str] = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[0] is None:
            continue
        sid = str(r[0]).strip()
        ids.append(sid)
        if r[6] is not None and str(r[6]).strip() not in ("", "None"):
            paused.append(sid)
        type_by_id[sid] = (
            str(r[11]).strip() if r[11] is not None else "-"
        )
    wb.close()
    return Snapshot(
        ids=frozenset(ids),
        paused_ids=frozenset(paused),
        type_by_id=type_by_id,
    )


def _active_by_package(snap: Snapshot) -> dict[str, int]:
    """Active = not paused, grouped and merged into display labels."""
    counts = Counter()
    for sid in snap.ids:
        if sid in snap.paused_ids:
            continue
        raw_type = snap.type_by_id.get(sid, "-")
        label = _TYPE_TO_LABEL.get(raw_type)
        if label is None:
            continue  # ignore unknown / placeholder rows
        counts[label] += 1
    return {p: counts.get(p, 0) for p in _PACKAGE_ORDER}


def _flow(prev: Snapshot, curr: Snapshot) -> dict[str, int]:
    new = curr.ids - prev.ids
    cancelled = prev.ids - curr.ids
    newly_paused = curr.paused_ids - prev.paused_ids
    newly_resumed = prev.paused_ids - curr.paused_ids
    return {
        "New":       len(new),
        "Resumed":   len(newly_resumed),
        "Cancelled": len(cancelled),
        "Paused":    len(newly_paused),
    }


def assemble_data() -> tuple[dict[str, dict[str, int]], dict[str, dict[str, int]]]:
    """Return (active_by_month, flow_by_month) — two dicts keyed by month label."""
    active: dict[str, dict[str, int]] = {}
    flow: dict[str, dict[str, int]] = {}

    snapshots: dict[str, Snapshot] = {}
    for label, rel in _SUB_FILES.items():
        snapshots[label] = _load_snapshot(_CLIENT / rel)

    # Active: derive where we have a SUB file; fall back to chart for the rest.
    for label, _ in _MONTHS:
        if label in snapshots:
            active[label] = _active_by_package(snapshots[label])
        else:
            active[label] = dict(_ACTIVE_FROM_CHART[label])

    # Flow: cross-month diff where consecutive SUB files exist.
    for i, (label, _) in enumerate(_MONTHS):
        if i == 0:
            flow[label] = dict(_FLOW_FROM_CHART.get(label, {}))
            continue
        prev_label = _MONTHS[i - 1][0]
        if label in snapshots and prev_label in snapshots:
            flow[label] = _flow(snapshots[prev_label], snapshots[label])
        else:
            flow[label] = dict(_FLOW_FROM_CHART.get(label, {
                "New": 0, "Resumed": 0, "Cancelled": 0, "Paused": 0,
            }))

    return active, flow


# ---------------------------------------------------------------------------
# Rendering
# ---------------------------------------------------------------------------

def _style() -> None:
    plt.rcParams.update({
        "font.family": "sans-serif",
        "font.sans-serif": ["Helvetica Neue", "Helvetica", "Arial", "DejaVu Sans"],
        "font.size": 10,
        "axes.spines.top": False,
        "axes.spines.right": False,
        "axes.spines.left": False,
        "axes.edgecolor": "#999999",
        "axes.labelcolor": "#444444",
        "xtick.color": "#444444",
        "ytick.color": "#444444",
        "ytick.labelsize": 9,
        "xtick.labelsize": 10,
        "axes.grid": True,
        "axes.grid.axis": "y",
        "grid.color": "#F2F2F2",
        "grid.linewidth": 0.4,
        "axes.axisbelow": True,
    })


def render_active(active: dict[str, dict[str, int]], out: Path) -> None:
    fig, ax = plt.subplots(figsize=(14, 4.2), dpi=150)

    months = [m for m, _ in _MONTHS]
    x = list(range(len(months)))
    bottom = [0] * len(months)

    for pkg in _PACKAGE_ORDER:
        values = [active[m].get(pkg, 0) for m in months]
        bars = ax.bar(x, values, bottom=bottom, label=pkg,
                      color=_PACKAGE_PALETTE[pkg], width=0.7,
                      edgecolor="white", linewidth=0.6)
        for xi, v, bot in zip(x, values, bottom):
            if v >= 30:  # only label segments that fit
                ax.text(xi, bot + v / 2, f"{v}",
                        ha="center", va="center",
                        color="white", fontsize=8, fontweight="bold")
        bottom = [b + v for b, v in zip(bottom, values)]

    # Total label above each bar
    for xi, total in zip(x, bottom):
        ax.text(xi, total + 12, f"{total}",
                ha="center", va="bottom",
                color="#1F3D3A", fontsize=10, fontweight="bold")

    ax.set_xticks(x)
    ax.set_xticklabels(months)
    ax.set_ylim(0, max(bottom) * 1.12)
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v, _: f"{int(v):,}"))
    ax.set_ylabel("Active subscriptions")

    ax.legend(loc="upper center", bbox_to_anchor=(0.5, -0.12),
              ncol=len(_PACKAGE_ORDER), frameon=False, fontsize=9)

    ax.tick_params(axis="x", length=0)
    ax.tick_params(axis="y", length=0)

    fig.tight_layout()
    out.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(out, dpi=200, bbox_inches="tight", facecolor="white")
    plt.close(fig)


def render_flow(flow: dict[str, dict[str, int]], out: Path) -> None:
    fig, ax = plt.subplots(figsize=(14, 4.2), dpi=150)

    months = [m for m, _ in _MONTHS]
    x = list(range(len(months)))

    # Above zero: New + Resumed (stacked).
    pos_bottom = [0] * len(months)
    for kind in ("New", "Resumed"):
        values = [flow[m].get(kind, 0) for m in months]
        ax.bar(x, values, bottom=pos_bottom,
               color=_FLOW_PALETTE[kind], label=kind,
               width=0.7, edgecolor="white", linewidth=0.6)
        for xi, v, bot in zip(x, values, pos_bottom):
            if v >= 5:
                ax.text(xi, bot + v / 2, f"{v}",
                        ha="center", va="center",
                        color="white", fontsize=8, fontweight="bold")
        pos_bottom = [b + v for b, v in zip(pos_bottom, values)]

    # Below zero: Cancelled + Paused (stacked downward).
    neg_top = [0] * len(months)
    for kind in ("Cancelled", "Paused"):
        values = [flow[m].get(kind, 0) for m in months]
        # Plot as negative
        neg_values = [-v for v in values]
        ax.bar(x, neg_values, bottom=neg_top,
               color=_FLOW_PALETTE[kind], label=kind,
               width=0.7, edgecolor="white", linewidth=0.6)
        for xi, v, top in zip(x, values, neg_top):
            if v >= 5:
                ax.text(xi, top - v / 2, f"{v}",
                        ha="center", va="center",
                        color="white", fontsize=8, fontweight="bold")
        neg_top = [t - v for t, v in zip(neg_top, values)]

    # Zero baseline emphasised
    ax.axhline(0, color="#222", linewidth=1.0)

    # Y-axis: show absolute values (no minus signs for the negative side)
    y_max = max(max(pos_bottom), abs(min(neg_top))) * 1.15
    ax.set_ylim(-y_max, y_max)
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v, _: f"{int(abs(v))}"))

    ax.set_xticks(x)
    ax.set_xticklabels(months)
    ax.set_ylabel("Subscriptions")

    ax.legend(loc="upper center", bbox_to_anchor=(0.5, -0.12),
              ncol=4, frameon=False, fontsize=9)

    ax.tick_params(axis="x", length=0)
    ax.tick_params(axis="y", length=0)

    fig.tight_layout()
    out.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(out, dpi=200, bbox_inches="tight", facecolor="white")
    plt.close(fig)


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main() -> None:
    _style()
    active, flow = assemble_data()

    print("Active by package (per month, totals):")
    for label, _ in _MONTHS:
        a = active[label]
        total = sum(a.values())
        print(f"  {label}: {a}  total={total}")

    print("\nFlow per month:")
    for label, _ in _MONTHS:
        print(f"  {label}: {flow[label]}")

    render_active(active, _OUT / "slide4_active_b2c.png")
    render_flow(flow, _OUT / "slide4_flow_b2c.png")
    print(f"\nWrote: {_OUT / 'slide4_active_b2c.png'}")
    print(f"Wrote: {_OUT / 'slide4_flow_b2c.png'}")


if __name__ == "__main__":
    main()
