"""Tests for core.model.recalc — LibreOffice headless recalculation."""

from __future__ import annotations

import openpyxl
import pytest

from core.model import recalc as recalc_mod
from core.model.recalc import SofficeNotFound, recalc, soffice_available


def test_missing_soffice_raises(monkeypatch) -> None:
    """When no soffice binary can be located, _find_soffice fails loud."""
    monkeypatch.setattr(recalc_mod.shutil, "which", lambda name: None)
    monkeypatch.setattr(recalc_mod, "_MAC", recalc_mod.Path("/nonexistent/soffice"))
    with pytest.raises(SofficeNotFound):
        recalc_mod._find_soffice()


@pytest.mark.skipif(not soffice_available(), reason="LibreOffice not installed")
def test_recalc_computes_formula(tmp_path) -> None:
    """A formula authored by openpyxl (no cached value) comes back computed."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"], ws["A2"] = 1, 2
    ws["A3"] = "=A1+A2"
    wb.calculation.fullCalcOnLoad = True
    path = tmp_path / "m.xlsx"
    wb.save(path)

    out = recalc(path)
    assert out.active["A3"].value == 3
