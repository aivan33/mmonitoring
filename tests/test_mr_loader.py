"""Tests for ``core/loaders/mr.py`` — extract_month."""

from __future__ import annotations

import datetime as dt
import logging
from pathlib import Path

import pytest
import yaml
from openpyxl import Workbook

from core.loaders.mr import extract_month


_REPO = Path(__file__).resolve().parent.parent
_REAL_MR = _REPO / "clients" / "farada" / "raw" / "mr_2026-03.xlsx"
_REAL_MAPPING = _REPO / "clients" / "farada" / "mapping.yaml"


def _make_mini_mr(path: Path) -> None:
    """Synthetic MR with a P&L, CF, BS sheet, header row 2, dates Jan-Mar 2026."""
    wb = Workbook()
    wb.remove(wb.active)

    pnl = wb.create_sheet("P&L")
    pnl.cell(2, 1, "Currency: EUR")
    pnl.cell(2, 2, dt.datetime(2026, 1, 1))
    pnl.cell(2, 3, dt.datetime(2026, 2, 1))
    pnl.cell(2, 4, dt.datetime(2026, 3, 1))
    pnl.cell(4, 1, "Sales ")  # subtotal — irrelevant
    pnl.cell(5, 1, "Food Logistics")  # sub-header
    pnl.cell(6, 1, "Faraday-Ox Sensors")
    pnl.cell(6, 2, 100); pnl.cell(6, 3, 200); pnl.cell(6, 4, 300)
    pnl.cell(7, 1, "Eval-Kits")
    pnl.cell(7, 2, 50); pnl.cell(7, 3, "-"); pnl.cell(7, 4, None)

    cf = wb.create_sheet("CF")
    cf.cell(2, 2, "Currency: EUR")
    cf.cell(2, 3, dt.datetime(2026, 3, 1))
    cf.cell(4, 2, "Cash received from customers ")
    cf.cell(4, 3, 999.5)

    bs = wb.create_sheet("BS")
    bs.cell(2, 2, "Currency: EUR")
    bs.cell(2, 3, dt.datetime(2026, 3, 1))
    bs.cell(6, 2, "Research and Development Asset")
    bs.cell(6, 3, 8888)

    wb.save(path)


@pytest.fixture
def mini_mr(tmp_path):
    p = tmp_path / "mini_mr.xlsx"
    _make_mini_mr(p)
    return p


@pytest.fixture
def mini_mapping():
    return {
        "mapping_is": [
            {"mr_row": 6, "mr_label": "Faraday-Ox Sensors",
             "data": "Sales", "grp": "Food Logistics", "subgroup": "Faraday-Ox Sensors"},
            {"mr_row": 7, "mr_label": "Eval-Kits",
             "data": "Sales", "grp": "Food Logistics", "subgroup": "Eval-Kits"},
            {"mr_row": None, "mr_label": None,
             "data": "Computed", "grp": "Computed", "subgroup": "Computed"},
        ],
        "mapping_cf": [
            {"mr_row": 4, "mr_label": "Cash received from customers ",
             "data": "Cash Flow from Operating Activities",
             "grp": "Cash received from customers",
             "subgroup": "Cash received from customers"},
        ],
        "mapping_bs": [
            {"mr_row": 6, "mr_label": "Research and Development Asset",
             "data": "Non-current assets",
             "grp": "Non-tangible fixed assets",
             "subgroup": "Research and Development Asset"},
        ],
    }


# Happy path -----------------------------------------------------------------

def test_extract_month_is(mini_mr, mini_mapping):
    result = extract_month(mini_mr, mini_mapping, 2026, 3, "IS")
    assert result[("Sales", "Food Logistics", "Faraday-Ox Sensors")] == 300.0
    # "-" placeholder coerces to None
    assert result[("Sales", "Food Logistics", "Eval-Kits")] is None
    # mr_row=null → None (derived/non-MR taxonomi rows)
    assert result[("Computed", "Computed", "Computed")] is None


def test_extract_month_cf(mini_mr, mini_mapping):
    result = extract_month(mini_mr, mini_mapping, 2026, 3, "CF")
    assert result[("Cash Flow from Operating Activities",
                   "Cash received from customers",
                   "Cash received from customers")] == 999.5


def test_extract_month_bs(mini_mr, mini_mapping):
    result = extract_month(mini_mr, mini_mapping, 2026, 3, "BS")
    assert result[("Non-current assets",
                   "Non-tangible fixed assets",
                   "Research and Development Asset")] == 8888.0


def test_extract_month_picks_correct_column(mini_mr, mini_mapping):
    """Verifies header-date scan resolves to the right column."""
    jan = extract_month(mini_mr, mini_mapping, 2026, 1, "IS")
    feb = extract_month(mini_mr, mini_mapping, 2026, 2, "IS")
    mar = extract_month(mini_mr, mini_mapping, 2026, 3, "IS")
    assert jan[("Sales", "Food Logistics", "Faraday-Ox Sensors")] == 100.0
    assert feb[("Sales", "Food Logistics", "Faraday-Ox Sensors")] == 200.0
    assert mar[("Sales", "Food Logistics", "Faraday-Ox Sensors")] == 300.0


# Label fallback -------------------------------------------------------------

def test_label_mismatch_falls_back_to_label_search(mini_mr, mini_mapping, caplog):
    """Mapping points at row 5 ('Food Logistics' sub-header) but mr_label is
    'Eval-Kits'. Loader should search the sheet, find Eval-Kits at row 7,
    and emit a warning naming both row indices."""
    bad = dict(mini_mapping)
    bad["mapping_is"] = [
        {"mr_row": 5, "mr_label": "Eval-Kits",
         "data": "Sales", "grp": "Food Logistics", "subgroup": "Eval-Kits"},
    ]
    with caplog.at_level(logging.WARNING):
        result = extract_month(mini_mr, bad, 2026, 3, "IS")
    # Row 7's Mar value is None ("-" placeholder) — but the lookup succeeded.
    assert result[("Sales", "Food Logistics", "Eval-Kits")] is None
    # Warning should mention both the configured row (5) and the found row (7).
    assert any("row 7" in r.getMessage() and "5" in r.getMessage()
               for r in caplog.records)


def test_label_not_found_anywhere_yields_none_with_warning(
    mini_mr, mini_mapping, caplog,
):
    bad = dict(mini_mapping)
    bad["mapping_is"] = [
        {"mr_row": 6, "mr_label": "Nonexistent Label",
         "data": "Sales", "grp": "X", "subgroup": "Y"},
    ]
    with caplog.at_level(logging.WARNING):
        result = extract_month(mini_mr, bad, 2026, 3, "IS")
    assert result[("Sales", "X", "Y")] is None
    assert any("Nonexistent Label" in r.getMessage() for r in caplog.records)


# Sign flip ------------------------------------------------------------------

def test_sign_flip_negates_value(mini_mr, mini_mapping):
    """Mapping entry with sign=-1 returns the value negated."""
    bad = dict(mini_mapping)
    bad["mapping_is"] = [
        {"mr_row": 6, "mr_label": "Faraday-Ox Sensors", "sign": -1,
         "data": "Sales", "grp": "Food Logistics",
         "subgroup": "Faraday-Ox Sensors"},
    ]
    result = extract_month(mini_mr, bad, 2026, 3, "IS")
    # Underlying MR value at row 6 col 4 (Mar) = 300; with sign=-1 → -300.
    assert result[("Sales", "Food Logistics", "Faraday-Ox Sensors")] == -300.0


def test_sign_default_is_positive(mini_mr, mini_mapping):
    """Without a sign field, value passes through unchanged."""
    result = extract_month(mini_mr, mini_mapping, 2026, 3, "IS")
    assert result[("Sales", "Food Logistics", "Faraday-Ox Sensors")] == 300.0


# Errors ---------------------------------------------------------------------

def test_unknown_statement_raises(mini_mr, mini_mapping):
    with pytest.raises(ValueError, match="statement"):
        extract_month(mini_mr, mini_mapping, 2026, 3, "BAD")


def test_period_not_in_header_raises(mini_mr, mini_mapping):
    with pytest.raises(ValueError, match="2030-01"):
        extract_month(mini_mr, mini_mapping, 2030, 1, "IS")


# Integration against the real MR file --------------------------------------

@pytest.mark.skipif(not _REAL_MR.exists(),
                    reason="real MR file not present")
def test_real_mr_is_march_2026_spot_checks():
    mapping = yaml.safe_load(_REAL_MAPPING.read_text())
    result = extract_month(_REAL_MR, mapping, 2026, 3, "IS")
    # Spot: R&D Total Payroll Germany Mar = 52545.24
    assert result[("R&D", "Total Payroll", "Germany")] == pytest.approx(52545.24)
    # Spot: G&A Accounting Mar = 4297.8
    assert result[("G&A", "External Professional Services",
                   "Accounting")] == pytest.approx(4297.8)
    # Sales Mar 2026 sums to 0 (MR P&L row 4 "Sales " = 0 for Mar)
    sales_sum = sum(
        v for (data, _, _), v in result.items()
        if data == "Sales" and v is not None
    )
    assert sales_sum == 0


@pytest.mark.skipif(not _REAL_MR.exists(),
                    reason="real MR file not present")
def test_real_mr_cf_march_2026_spot_checks():
    mapping = yaml.safe_load(_REAL_MAPPING.read_text())
    result = extract_month(_REAL_MR, mapping, 2026, 3, "CF")
    # CFI/CAPEX Mar = -52106.51
    assert result[("Cash Flow from Investing Activities",
                   "CAPEX", "CAPEX")] == pytest.approx(-52106.51)
    # Beginning Cash Mar = 1389708.27
    assert result[("Beginning Cash Balance", "Beginning Cash Balance",
                   "Beginning Cash Balance")] == pytest.approx(1389708.27)
    # Sum of activities = Excess Cash for the Period
    cfo = sum(v for (d, _, _), v in result.items()
              if d == "Cash Flow from Operating Activities" and v is not None)
    cfi = sum(v for (d, _, _), v in result.items()
              if d == "Cash Flow from Investing Activities" and v is not None)
    cff = sum(v for (d, _, _), v in result.items()
              if d == "Cash Flow from Financing Activities" and v is not None)
    # MR's Excess Cash for the Period for Mar 2026 = 340825.34
    assert cfo + cfi + cff == pytest.approx(340825.34, abs=0.5)


@pytest.mark.skipif(not _REAL_MR.exists(),
                    reason="real MR file not present")
def test_real_mr_bs_march_2026_spot_checks():
    mapping = yaml.safe_load(_REAL_MAPPING.read_text())
    result = extract_month(_REAL_MR, mapping, 2026, 3, "BS")
    # Cash & cash equivalents Mar = 1728419.46
    assert result[("Cash and cash equivalents", "Cash and cash equivalents",
                   "Cash and cash equivalents")] == pytest.approx(1728419.46)
    # Share capital Mar = 5436593.07
    assert result[("Equity", "Share capital",
                   "Share capital")] == pytest.approx(5436593.07)
    # Computed/derived rows (mr_row=null) → None
    assert result[("Working Capital", "Working Capital",
                   "Working Capital")] is None
