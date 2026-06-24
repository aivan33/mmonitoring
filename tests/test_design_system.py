"""Tests for core.model.design_system — the canonical house design system.

Locks the reconciled canon (Hybrid palette + plain number formats + Century Gothic) as constants,
and — where the gitignored reference workbooks are present — asserts those constants match what is
actually in the files, so the design system stays true rather than drifting from the docs.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from core.model import design_system as ds

FARADA = Path("clients/farada/modeling/farada_model_v8.xlsx")
CUPFFEE = Path("clients/cupffee/budget/cupfee-q2-26.xlsx")


def test_canon_constants():
    assert ds.FONT == "Century Gothic" and ds.FONT_SIZE == 10
    assert ds.PALETTE["section_band"] == "FFD8D8D8"
    assert ds.PALETTE["active"] == "FFDDFBFF"        # Hybrid: Farada cyan
    assert ds.PALETTE["input_value"] == "FFFEF2CB"
    assert ds.PALETTE["statement_banner"] == "FFD5EBF4"
    assert ds.SCENARIOS == ["Realistic", "Optimistic", "Pessimistic"]   # cols L/M/N
    assert ds.SCENARIO_COLS == ["L", "M", "N"]
    # only Realistic (L) is required; Optimistic/Pessimistic (M/N) are optional, may stay empty
    assert ds.SCENARIO_COLS_REQUIRED == ["L"]
    # plain number formats (not accounting)
    assert ds.NUMBER_FORMATS["int"] == "#,##0" and ds.NUMBER_FORMATS["pct"] == "0.0%"


def test_helpers_build_openpyxl_objects():
    f = ds.font(bold=True)
    assert f.name == "Century Gothic" and f.size == 10 and f.bold
    fl = ds.fill("active")
    assert fl.fgColor.rgb == "FFDDFBFF" and fl.patternType == "solid"
    assert ds.active_formula(16) == "=OFFSET(K16,0,$D$2)"


def _fill(cell):
    fl = cell.fill
    return fl.fgColor.rgb if fl and fl.patternType and isinstance(fl.fgColor.rgb, str) else None


@pytest.mark.skipif(not FARADA.exists(), reason="Farada model gitignored / absent")
def test_canon_matches_farada():
    # located by content (not hardcoded cells) so it survives the I–V Inputs reflow
    wb = openpyxl.load_workbook(FARADA)
    inp, isx = wb[" Inputs"], wb["IS"]
    band = next(inp.cell(r, 3) for r in range(1, inp.max_row + 1)
                if isinstance(inp.cell(r, 1).value, str) and inp.cell(r, 1).value.strip().endswith("."))
    assert band.font.name == ds.FONT and _fill(band) == ds.PALETTE["section_band"]
    act = next(r for r in range(1, inp.max_row + 1)
               if isinstance(inp.cell(r, 10).value, str) and "OFFSET" in inp.cell(r, 10).value)
    assert _fill(inp.cell(act, 10)) == ds.PALETTE["active"]
    assert _fill(inp.cell(act, 12)) == ds.PALETTE["input_value"]
    assert _fill(isx["A1"]) == ds.PALETTE["statement_banner"]


@pytest.mark.skipif(not CUPFFEE.exists(), reason="Cupffee budget gitignored / absent")
def test_section_band_shared_with_cupffee():
    wb = openpyxl.load_workbook(CUPFFEE)
    assert _fill(wb[" Inputs"]["C7"]) == ds.PALETTE["section_band"]   # the one role all share
