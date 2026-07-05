"""Tests for core.model.integrity — gate checks over a recomputed workbook.

These operate on plain openpyxl workbooks with literal values (standing in for
LibreOffice-recomputed numbers), so they need no soffice.
"""

from __future__ import annotations

import openpyxl
import pytest

from core.model.integrity import run_all


def _wb():
    wb = openpyxl.Workbook()
    bs = wb.active
    bs.title = "BS"
    cf = wb.create_sheet("CF")
    # BS row 34 = check (0 across 2 months); row 10 = cash.
    for c in (3, 4):
        bs.cell(34, c).value = 0.0
        bs.cell(10, c).value = 500.0
        cf.cell(24, c).value = 500.0          # CF ending cash == BS cash
        # subtotal: row 20 == 15 + 16
        cf.cell(15, c).value = 100.0
        cf.cell(16, c).value = 40.0
        cf.cell(20, c).value = 140.0
    return wb


GATES = {
    "error_scan": {"sheets": "all"},
    "balance_checks": [
        {"name": "BS check", "sheet": "BS", "row": 34,
         "cols": {"start": 3, "end": 4}, "tolerance": 1.0}],
    "ties": [
        {"name": "CF ending == BS cash", "a": {"sheet": "CF", "row": 24},
         "b": {"sheet": "BS", "row": 10}, "cols": {"start": 3, "end": 4}}],
    "subtotals": [
        {"name": "CFF", "sheet": "CF", "parent": 20, "children": [15, 16],
         "cols": {"start": 3, "end": 4}}],
}


def test_clean_model_has_no_violations() -> None:
    assert run_all(_wb(), GATES) == []


def test_error_cell_is_flagged() -> None:
    wb = _wb()
    wb["CF"].cell(24, 3).value = "#REF!"
    v = run_all(wb, GATES)
    assert any(x.check == "error_cell" and "#REF!" in x.detail for x in v)


def test_unbalanced_bs_is_flagged() -> None:
    wb = _wb()
    wb["BS"].cell(34, 4).value = 100.0
    v = run_all(wb, GATES)
    assert any(x.check == "bs_balance" and x.cell.endswith("34") for x in v)


def test_broken_tie_is_flagged() -> None:
    wb = _wb()
    wb["CF"].cell(24, 3).value = 999.0
    v = run_all(wb, GATES)
    assert any(x.check == "statements_tie" for x in v)


def test_subtotal_that_does_not_foot_is_flagged() -> None:
    wb = _wb()
    wb["CF"].cell(20, 4).value = 999.0
    v = run_all(wb, GATES)
    assert any(x.check == "subtotals_foot" for x in v)


def test_error_scan_skips_configured_rows() -> None:
    # A documented skip (e.g. an unbuilt section) suppresses only those rows.
    wb = _wb()
    wb["CF"].cell(24, 3).value = "#REF!"
    gates = {**GATES, "ties": [],
             "error_scan": {"sheets": "all", "skip": {"CF": [[24, 24]]}}}
    v = run_all(wb, gates)
    assert not any(x.check == "error_cell" for x in v)
