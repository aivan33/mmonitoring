"""Tests for ``core/data/loaders/kpis.py`` — load_kpi_wide_xlsx."""

from __future__ import annotations

import datetime as dt
from pathlib import Path

import pytest
from openpyxl import Workbook

from core.data.loaders.kpis import KPIRow, load_kpi_wide_xlsx


def _make_kpi_xlsx(path: Path, header_row, data_rows) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("KPIs")
    ws.append(header_row)
    for row in data_rows:
        ws.append(row)
    wb.save(path)


# Header parsing -----------------------------------------------------------

def test_month_year_short_header(tmp_path):
    p = tmp_path / "x.xlsx"
    _make_kpi_xlsx(p,
        ["Month", "Jan 26", "Feb 26", "Mar 26"],
        [
            ["GMV", 100, 200, 300],
            ["# Invoices", 10, 20, 30],
        ],
    )
    rows = list(load_kpi_wide_xlsx(p, year=2026, entity="e1"))
    assert len(rows) == 6
    by_period = {(r.period_date, r.kpi): r.value for r in rows}
    assert by_period[(dt.date(2026, 1, 1), "GMV")] == 100
    assert by_period[(dt.date(2026, 3, 1), "# Invoices")] == 30


def test_full_month_name_header(tmp_path):
    p = tmp_path / "x.xlsx"
    _make_kpi_xlsx(p,
        ["", "January", "February", "March"],
        [["GMV", 100, 200, 300]],
    )
    rows = list(load_kpi_wide_xlsx(p, year=2026, entity="e1"))
    # Default year=2026 supplies the year for bare month names.
    assert {r.period_date for r in rows} == {
        dt.date(2026, 1, 1), dt.date(2026, 2, 1), dt.date(2026, 3, 1)
    }


def test_dt_date_header(tmp_path):
    p = tmp_path / "x.xlsx"
    _make_kpi_xlsx(p,
        ["", dt.datetime(2026, 1, 1), dt.datetime(2026, 2, 1)],
        [["GMV", 100, 200]],
    )
    rows = list(load_kpi_wide_xlsx(p, year=2026, entity="e1"))
    assert {r.period_date for r in rows} == {dt.date(2026, 1, 1), dt.date(2026, 2, 1)}


def test_non_month_header_columns_skipped(tmp_path):
    """Columns whose header doesn't parse as a month are silently dropped."""
    p = tmp_path / "x.xlsx"
    _make_kpi_xlsx(p,
        ["Month", "Jan 26", "Notes", "Feb 26"],
        [["GMV", 100, "some text", 200]],
    )
    rows = list(load_kpi_wide_xlsx(p, year=2026, entity="e1"))
    # Only the two month columns produce rows.
    assert len(rows) == 2
    assert {r.period_date for r in rows} == {dt.date(2026, 1, 1), dt.date(2026, 2, 1)}


# Value handling ------------------------------------------------------------

def test_null_cells_skipped(tmp_path):
    p = tmp_path / "x.xlsx"
    _make_kpi_xlsx(p,
        ["", "Jan 26", "Feb 26", "Mar 26"],
        [["GMV", 100, None, 300]],
    )
    rows = list(load_kpi_wide_xlsx(p, year=2026, entity="e1"))
    # Feb is null → no row emitted.
    assert len(rows) == 2
    months = sorted(r.period_date.month for r in rows)
    assert months == [1, 3]


def test_empty_kpi_label_row_skipped(tmp_path):
    p = tmp_path / "x.xlsx"
    _make_kpi_xlsx(p,
        ["", "Jan 26"],
        [
            ["GMV", 100],
            [None, 999],   # no label → skipped
            ["", 999],     # empty string → skipped
            ["# Boxes", 50],
        ],
    )
    rows = list(load_kpi_wide_xlsx(p, year=2026, entity="e1"))
    assert {r.kpi for r in rows} == {"GMV", "# Boxes"}


def test_dash_treated_as_null(tmp_path):
    p = tmp_path / "x.xlsx"
    _make_kpi_xlsx(p,
        ["", "Jan 26"],
        [["GMV", "-"]],
    )
    rows = list(load_kpi_wide_xlsx(p, year=2026, entity="e1"))
    assert rows == []


def test_string_numeric_values_skipped(tmp_path):
    """If a cell is a non-numeric string (not a known null sentinel), skip
    rather than fail."""
    p = tmp_path / "x.xlsx"
    _make_kpi_xlsx(p,
        ["", "Jan 26", "Feb 26"],
        [["GMV", "lol", 200]],
    )
    rows = list(load_kpi_wide_xlsx(p, year=2026, entity="e1"))
    assert len(rows) == 1
    assert rows[0].period_date == dt.date(2026, 2, 1)


# Currency conversion -------------------------------------------------------

def test_currency_eur_no_conversion(tmp_path):
    p = tmp_path / "x.xlsx"
    _make_kpi_xlsx(p, ["", "Jan 26"], [["GMV", 100]])
    rows = list(load_kpi_wide_xlsx(p, year=2026, entity="e1", currency="EUR"))
    assert rows[0].value == 100.0


def test_currency_usd_divides_by_rate(tmp_path):
    """Convention: fx_rate = source units per 1 EUR. Divide source by rate."""
    p = tmp_path / "x.xlsx"
    _make_kpi_xlsx(p, ["", "Jan 26"], [["GMV", 1087]])
    # USD rate 1.087 → $1087 / 1.087 = €1000.
    rows = list(load_kpi_wide_xlsx(
        p, year=2026, entity="e1", currency="USD", fx_rate=1.087,
    ))
    assert rows[0].value == pytest.approx(1000.0, abs=0.01)


def test_currency_non_eur_without_rate_raises(tmp_path):
    p = tmp_path / "x.xlsx"
    _make_kpi_xlsx(p, ["", "Jan 26"], [["GMV", 100]])
    with pytest.raises(ValueError, match="fx_rate"):
        list(load_kpi_wide_xlsx(p, year=2026, entity="e1", currency="USD"))


# Real almacena file --------------------------------------------------------

_REPO = Path(__file__).resolve().parent.parent
_ALM = _REPO / "clients" / "almacena" / "raw" / "profitability_q1.xlsx"


@pytest.mark.skipif(not _ALM.exists(),
                    reason="real almacena profitability file not present")
def test_real_almacena_profitability_loads():
    """Smoke test on the real almacena file: 3 months × 8 KPIs = 24 rows
    (assuming all cells have values, which they do for Q1 2026)."""
    rows = list(load_kpi_wide_xlsx(
        _ALM, year=2026, entity="ap_foundation",
        currency="USD", fx_rate=1.087,
    ))
    assert len(rows) >= 18  # tolerant lower bound (some cells may be null)
    kpis = {r.kpi for r in rows}
    assert "GMV" in kpis
    assert "Funded Amount" in kpis
    assert "Average Days Outstanding" in kpis
    # Currency conversion sanity: GMV Jan was $17,044,519 USD;
    # /1.087 ≈ €15,680,330.
    gmv_jan = next(
        r for r in rows
        if r.kpi == "GMV" and r.period_date == dt.date(2026, 1, 1)
    )
    assert gmv_jan.value == pytest.approx(15_680_330, abs=1000)


def test_dimensionless_kpis_skip_currency_conversion(tmp_path):
    """KPIs listed as dimensionless pass through unconverted even when the
    source is USD."""
    from openpyxl import Workbook
    p = tmp_path / "x.xlsx"
    wb = Workbook(); wb.remove(wb.active)
    ws = wb.create_sheet("KPIs")
    ws.append(["", "Jan 26"])
    ws.append(["GMV", 1087])             # USD → /1.087 = €1000
    ws.append(["# Invoices", 69])        # dimensionless → stays 69
    ws.append(["Cash Drag %", 0.15])     # dimensionless → stays 0.15
    wb.save(p)
    rows = list(load_kpi_wide_xlsx(
        p, year=2026, entity="e1",
        currency="USD", fx_rate=1.087,
        dimensionless_kpis=["# Invoices", "Cash Drag %"],
    ))
    by_kpi = {r.kpi: r.value for r in rows}
    assert by_kpi["GMV"] == pytest.approx(1000.0, abs=0.01)
    assert by_kpi["# Invoices"] == 69.0
    assert by_kpi["Cash Drag %"] == pytest.approx(0.15)

