"""Tests for core.data.integrity — the R3/R4/R5 checker engine.

R3: rows whose label looks like a Total/Subtotal/Net but aren't registered.
R4: registered aggregate must equal Σ(leaf × sign) within tolerance.
R5: registered aggregate's source cell must be an Excel formula, not hardcoded.
"""

from __future__ import annotations

import sqlite3
from pathlib import Path

import pytest
from openpyxl import Workbook

from core.data.aggregate_formulas import AggregateFormula, FormulaLeaf
from core.data.integrity import (
    Finding,
    IntegrityReport,
    check_integrity,
)
from core.data.schema import wipe_and_create


def _insert(
    conn: sqlite3.Connection,
    *,
    data: str,
    grp: str,
    subgroup: str,
    value: float,
    period: str = "2025-01-01",
    scenario: str = "actual",
    entity: str = "demo",
    statement: str = "IS",
    is_aggregate: int = 0,
    display_order: int = 0,
) -> None:
    conn.execute(
        "INSERT INTO financials VALUES (?,?,?,?,?,?,?,?,?,?)",
        (period, entity, scenario, statement, data, grp, subgroup,
         display_order, value, is_aggregate),
    )


@pytest.fixture
def db(tmp_path: Path) -> sqlite3.Connection:
    path = tmp_path / "t.db"
    wipe_and_create(path)
    conn = sqlite3.connect(path)
    yield conn
    conn.close()


# ---------------------------------------------------------------------------
# Empty / trivial cases
# ---------------------------------------------------------------------------

class TestEmpty:
    def test_empty_registry_empty_db_yields_no_findings(
        self, db: sqlite3.Connection,
    ) -> None:
        report = check_integrity(db, registry={})
        assert report.findings == ()
        assert not report.has_failures()


# ---------------------------------------------------------------------------
# R4: registered aggregate recompute
# ---------------------------------------------------------------------------

class TestR4Recompute:
    def _setup_burn(self, db: sqlite3.Connection, *, parsed_burn: float) -> None:
        # Two leaves: -100 and -50. Expected gross_burn = -150.
        _insert(db, data="Cost of Sales", grp="Materials", subgroup="x", value=-100)
        _insert(db, data="S&M",           grp="Events",    subgroup="y", value=-50)
        _insert(db, data="KPI", grp="Burn", subgroup="Gross",
                value=parsed_burn, is_aggregate=1)
        db.commit()

    def _registry(self) -> dict[str, AggregateFormula]:
        return {"gross_burn": AggregateFormula(
            name="gross_burn",
            data="KPI", grp="Burn", subgroup="Gross",
            leaves=(
                FormulaLeaf(data="Cost of Sales", sign=1),
                FormulaLeaf(data="S&M",           sign=1),
            ),
        )}

    def test_match_yields_no_finding(self, db: sqlite3.Connection) -> None:
        self._setup_burn(db, parsed_burn=-150)
        report = check_integrity(db, registry=self._registry())
        assert report.failures == ()

    def test_mismatch_above_tolerance_fails(self, db: sqlite3.Connection) -> None:
        # Parsed -200, recomputed -150 → delta 50, well above €1 tolerance.
        self._setup_burn(db, parsed_burn=-200)
        report = check_integrity(db, registry=self._registry())
        assert len(report.failures) == 1
        f = report.failures[0]
        assert f.rule == "R4"
        assert f.name == "gross_burn"
        assert "delta" in f.message.lower()

    def test_mismatch_below_tolerance_passes(self, db: sqlite3.Connection) -> None:
        # 0.50 EUR delta vs default 1.00 tolerance.
        self._setup_burn(db, parsed_burn=-150.5)
        report = check_integrity(db, registry=self._registry())
        assert report.failures == ()

    def test_explicit_triplet_leaf_works(self, db: sqlite3.Connection) -> None:
        _insert(db, data="Sales", grp="Distributors", subgroup="220", value=100)
        _insert(db, data="Sales", grp="Distributors", subgroup="110", value=50)
        _insert(db, data="Sales", grp="Distributors", subgroup="Total",
                value=100, is_aggregate=1)  # wrong: hardcoded 100, should be 150
        db.commit()
        registry = {"distrib_total": AggregateFormula(
            name="distrib_total",
            data="Sales", grp="Distributors", subgroup="Total",
            leaves=(
                FormulaLeaf(data="Sales", grp="Distributors", subgroup="220", sign=1),
                FormulaLeaf(data="Sales", grp="Distributors", subgroup="110", sign=1),
            ),
        )}
        report = check_integrity(db, registry=registry)
        assert len(report.failures) == 1
        assert report.failures[0].rule == "R4"

    def test_mixed_signs(self, db: sqlite3.Connection) -> None:
        # Net Income = Revenue - Costs. Stored: 100 - 30 = 70. Pass.
        _insert(db, data="Sales", grp="g", subgroup="x", value=100)
        _insert(db, data="Cost of Sales", grp="m", subgroup="y", value=30)
        _insert(db, data="KPI", grp="Margin", subgroup="Net Income",
                value=70, is_aggregate=1)
        db.commit()
        registry = {"net_income": AggregateFormula(
            name="net_income",
            data="KPI", grp="Margin", subgroup="Net Income",
            leaves=(
                FormulaLeaf(data="Sales",         sign=1),
                FormulaLeaf(data="Cost of Sales", sign=-1),
            ),
        )}
        report = check_integrity(db, registry=registry)
        assert report.failures == ()


# ---------------------------------------------------------------------------
# R3: total-lookalike rows that aren't registered
# ---------------------------------------------------------------------------

class TestR3LookalikePattern:
    def test_total_subgroup_unregistered_warns(
        self, db: sqlite3.Connection,
    ) -> None:
        _insert(db, data="Sales", grp="g", subgroup="leaf", value=100)
        _insert(db, data="Sales", grp="g", subgroup="Total Sales", value=100)
        db.commit()
        report = check_integrity(db, registry={})
        r3 = [f for f in report.findings if f.rule == "R3"]
        assert len(r3) == 1
        assert r3[0].severity == "warn"
        assert "Total Sales" in r3[0].message

    def test_subtotal_grp_unregistered_warns(
        self, db: sqlite3.Connection,
    ) -> None:
        _insert(db, data="Sales", grp="Subtotal", subgroup="x", value=200)
        db.commit()
        report = check_integrity(db, registry={})
        r3 = [f for f in report.findings if f.rule == "R3"]
        assert len(r3) == 1

    def test_registered_total_not_warned(
        self, db: sqlite3.Connection,
    ) -> None:
        _insert(db, data="Sales", grp="g", subgroup="leaf", value=100)
        _insert(db, data="Sales", grp="g", subgroup="Total", value=100,
                is_aggregate=1)
        db.commit()
        registry = {"total": AggregateFormula(
            name="total",
            data="Sales", grp="g", subgroup="Total",
            leaves=(FormulaLeaf(data="Sales", grp="g", subgroup="leaf"),),
        )}
        report = check_integrity(db, registry=registry)
        r3 = [f for f in report.findings if f.rule == "R3"]
        assert r3 == []

    def test_non_total_label_not_warned(self, db: sqlite3.Connection) -> None:
        _insert(db, data="Sales", grp="Distributors", subgroup="220", value=100)
        db.commit()
        report = check_integrity(db, registry={})
        assert [f for f in report.findings if f.rule == "R3"] == []


# ---------------------------------------------------------------------------
# R5: source-cell type audit
# ---------------------------------------------------------------------------

class TestR5CellTypeAudit:
    def _make_workbook(self, path: Path, *, gross_value, gross_is_formula: bool) -> None:
        wb = Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet("IS")
        ws["A1"] = 100
        ws["A2"] = 50
        if gross_is_formula:
            ws["A47"] = "=A1+A2"
        else:
            ws["A47"] = gross_value
        wb.save(path)

    def _registry_with_cell(self) -> dict[str, AggregateFormula]:
        return {"gross_burn": AggregateFormula(
            name="gross_burn",
            data="KPI", grp="Burn", subgroup="Gross",
            leaves=(FormulaLeaf(data="Sales", sign=1),),
            source_cell="IS!A47",
        )}

    def test_formula_cell_no_warn(
        self, db: sqlite3.Connection, tmp_path: Path,
    ) -> None:
        wb = tmp_path / "src.xlsx"
        self._make_workbook(wb, gross_value=150, gross_is_formula=True)
        # Need an aggregate row in DB for R4 to even consider the formula,
        # but R5 should run regardless. Add a leaf and aggregate matching.
        _insert(db, data="Sales", grp="g", subgroup="x", value=150)
        _insert(db, data="KPI", grp="Burn", subgroup="Gross", value=150, is_aggregate=1)
        db.commit()
        report = check_integrity(
            db, registry=self._registry_with_cell(),
            workbook_paths=[wb],
        )
        assert [f for f in report.findings if f.rule == "R5"] == []

    def test_hardcoded_cell_warns(
        self, db: sqlite3.Connection, tmp_path: Path,
    ) -> None:
        wb = tmp_path / "src.xlsx"
        self._make_workbook(wb, gross_value=150, gross_is_formula=False)
        _insert(db, data="Sales", grp="g", subgroup="x", value=150)
        _insert(db, data="KPI", grp="Burn", subgroup="Gross", value=150, is_aggregate=1)
        db.commit()
        report = check_integrity(
            db, registry=self._registry_with_cell(),
            workbook_paths=[wb],
        )
        r5 = [f for f in report.findings if f.rule == "R5"]
        assert len(r5) == 1
        assert r5[0].severity == "warn"
        assert "hardcoded" in r5[0].message.lower()

    def test_no_source_cell_no_finding(
        self, db: sqlite3.Connection, tmp_path: Path,
    ) -> None:
        # Registry entry has no source_cell — R5 simply skips it.
        registry = {"agg": AggregateFormula(
            name="agg",
            data="KPI", grp="Burn", subgroup="Gross",
            leaves=(FormulaLeaf(data="Sales", sign=1),),
            source_cell=None,
        )}
        _insert(db, data="Sales", grp="g", subgroup="x", value=100)
        _insert(db, data="KPI", grp="Burn", subgroup="Gross", value=100, is_aggregate=1)
        db.commit()
        report = check_integrity(db, registry=registry, workbook_paths=[])
        assert [f for f in report.findings if f.rule == "R5"] == []


# ---------------------------------------------------------------------------
# IntegrityReport surface
# ---------------------------------------------------------------------------

class TestReportSurface:
    def test_failures_warnings_partition(self) -> None:
        rep = IntegrityReport(findings=(
            Finding(rule="R4", severity="fail", name="a", message="m1"),
            Finding(rule="R3", severity="warn", name="b", message="m2"),
        ))
        assert len(rep.failures) == 1
        assert len(rep.warnings) == 1
        assert rep.has_failures() is True

    def test_no_failures_no_has_failures(self) -> None:
        rep = IntegrityReport(findings=(
            Finding(rule="R3", severity="warn", name="x", message="m"),
        ))
        assert not rep.has_failures()
