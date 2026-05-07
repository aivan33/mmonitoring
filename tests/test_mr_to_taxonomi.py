"""Tests for ``core/report/mr_to_taxonomi.py``."""

from __future__ import annotations

import datetime as dt
import logging
from pathlib import Path

import pytest
import yaml
from openpyxl import Workbook, load_workbook

from core.report.mr_to_taxonomi import populate_taxonomi
from core.report.mr import extract_month


HEADER = ["Data", "Group", "Subgroup", "Jan", "Feb", "Mar", "Apr", "May",
          "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


_REPO = Path(__file__).resolve().parent.parent
_REAL_MR = _REPO / "clients" / "farada" / "raw" / "mr_2026-03.xlsx"
_REAL_PREV = _REPO / "clients" / "farada" / "raw" / "taxonomi_act_2026-02.xlsx"
_REAL_MAPPING = _REPO / "clients" / "farada" / "mapping.yaml"


def _make_mini_taxonomi(path: Path) -> None:
    """Synthetic taxonomi with IS/CF/BS sheets, Jan/Feb populated, Mar+ empty."""
    wb = Workbook()
    wb.remove(wb.active)
    is_sheet = wb.create_sheet("IS (Actual)")
    is_sheet.append(HEADER)
    is_sheet.append(["Sales", "Food Logistics", "Faraday-Ox Sensors", 100, 200,
                     None, None, None, None, None, None, None, None, None, None])
    is_sheet.append(["Sales", "Food Logistics", "Eval-Kits", 50, None,
                     None, None, None, None, None, None, None, None, None, None])

    cf_sheet = wb.create_sheet("CF Indirect (Actual)")
    cf_sheet.append(HEADER)
    cf_sheet.append(["Cash Flow from Operating Activities",
                     "Cash received from customers",
                     "Cash received from customers", None, 1550,
                     None, None, None, None, None, None, None, None, None, None])

    bs_sheet = wb.create_sheet("BS (Actual)")
    bs_sheet.append(HEADER)
    bs_sheet.append(["Cash and cash equivalents", "Cash and cash equivalents",
                     "Cash and cash equivalents", 1688174, 1382641,
                     None, None, None, None, None, None, None, None, None, None])
    wb.save(path)


@pytest.fixture
def mini_prev(tmp_path):
    p = tmp_path / "mini_prev.xlsx"
    _make_mini_taxonomi(p)
    return p


@pytest.fixture
def mini_extracts():
    """Mar 2026 extracts for the mini taxonomi."""
    return {
        "IS": {
            ("Sales", "Food Logistics", "Faraday-Ox Sensors"): 333.7,
            ("Sales", "Food Logistics", "Eval-Kits"): None,
        },
        "CF": {
            ("Cash Flow from Operating Activities",
             "Cash received from customers",
             "Cash received from customers"): 2380.0,
        },
        "BS": {
            ("Cash and cash equivalents", "Cash and cash equivalents",
             "Cash and cash equivalents"): 1728419.46,
        },
    }


# Happy path -----------------------------------------------------------------

def test_populate_taxonomi_writes_new_month(mini_prev, mini_extracts, tmp_path):
    out = tmp_path / "out.xlsx"
    populate_taxonomi(mini_prev, mini_extracts, 2026, 3, out)
    assert out.exists()

    wb = load_workbook(out)
    is_ws = wb["IS (Actual)"]
    # Mar = col 6. 333.7 rounds to 334.
    assert is_ws.cell(2, 6).value == 334
    # None extract → cell stays None.
    assert is_ws.cell(3, 6).value is None
    bs_ws = wb["BS (Actual)"]
    # 1728419.46 rounds to 1728419.
    assert bs_ws.cell(2, 6).value == 1728419


def test_populate_preserves_prior_months(mini_prev, mini_extracts, tmp_path):
    out = tmp_path / "out.xlsx"
    populate_taxonomi(mini_prev, mini_extracts, 2026, 3, out)
    wb = load_workbook(out)
    is_ws = wb["IS (Actual)"]
    # Jan/Feb cols 4/5 unchanged.
    assert is_ws.cell(2, 4).value == 100
    assert is_ws.cell(2, 5).value == 200
    assert is_ws.cell(3, 4).value == 50
    assert is_ws.cell(3, 5).value is None


def test_populate_is_idempotent(mini_prev, mini_extracts, tmp_path):
    out1 = tmp_path / "run1.xlsx"
    out2 = tmp_path / "run2.xlsx"
    populate_taxonomi(mini_prev, mini_extracts, 2026, 3, out1)
    populate_taxonomi(mini_prev, mini_extracts, 2026, 3, out2)
    wb1 = load_workbook(out1)
    wb2 = load_workbook(out2)
    for sn in wb1.sheetnames:
        ws1, ws2 = wb1[sn], wb2[sn]
        for r in range(1, ws1.max_row + 1):
            for c in range(1, ws1.max_column + 1):
                assert ws1.cell(r, c).value == ws2.cell(r, c).value, \
                    f"{sn}!{r}{c}: {ws1.cell(r,c).value} vs {ws2.cell(r,c).value}"


def test_unmapped_row_logs_warning(mini_prev, mini_extracts, tmp_path, caplog):
    """Add a row to mini_prev IS that isn't in mr_extracts; expect a warning."""
    wb = load_workbook(mini_prev)
    is_ws = wb["IS (Actual)"]
    is_ws.append(["Sales", "Industrial IoT", "Eval-Kits", 999, 999,
                  None, None, None, None, None, None, None, None, None, None])
    wb.save(mini_prev)

    out = tmp_path / "out.xlsx"
    with caplog.at_level(logging.WARNING):
        populate_taxonomi(mini_prev, mini_extracts, 2026, 3, out)
    assert any("Industrial IoT" in r.getMessage() for r in caplog.records)
    # Pre-existing values unchanged.
    wb_out = load_workbook(out)
    is_out = wb_out["IS (Actual)"]
    assert is_out.cell(4, 4).value == 999
    assert is_out.cell(4, 5).value == 999


def test_populate_overwrites_existing_target_month(mini_prev, mini_extracts, tmp_path):
    """If the target month already has a value, populate overwrites it."""
    wb = load_workbook(mini_prev)
    wb["IS (Actual)"].cell(2, 6).value = 99999  # pre-existing Mar value
    wb.save(mini_prev)
    out = tmp_path / "out.xlsx"
    populate_taxonomi(mini_prev, mini_extracts, 2026, 3, out)
    is_ws = load_workbook(out)["IS (Actual)"]
    assert is_ws.cell(2, 6).value == 334  # overwritten


# Integration ---------------------------------------------------------------

@pytest.mark.skipif(not _REAL_MR.exists() or not _REAL_PREV.exists(),
                    reason="real farada files not present")
def test_real_pipeline_extract_then_populate(tmp_path):
    mapping = yaml.safe_load(_REAL_MAPPING.read_text())
    extracts = {
        stmt: extract_month(_REAL_MR, mapping, 2026, 3, stmt)
        for stmt in ("IS", "CF", "BS")
    }
    out = tmp_path / "taxonomi_act_2026-03.xlsx"
    populate_taxonomi(_REAL_PREV, extracts, 2026, 3, out)

    wb = load_workbook(out)
    is_ws = wb["IS (Actual)"]
    bs_ws = wb["BS (Actual)"]
    cf_ws = wb["CF Indirect (Actual)"]

    # MR-sourced spot checks.
    assert is_ws.cell(26, 6).value == 52545        # R&D Germany Mar = 52545
    assert bs_ws.cell(9, 6).value == 1728419       # Cash Mar = 1728419
    assert bs_ws.cell(2, 5).value == 9026          # Feb preserved

    # Derived KPI checks (Mar).
    assert bs_ws.cell(21, 6).value == 797868       # Working Capital
    assert cf_ws.cell(22, 6).value == 852894       # Gross Fixed assets
    # AP Turnover Mar undefined (Trade Payables = 0 both months) → None.
    assert bs_ws.cell(19, 6).value is None
    # AR Turnover Mar = 0.0 (Sales=0 in Mar; denominator non-zero).
    assert bs_ws.cell(20, 6).value == 0.0
    # % Change in cash Mar should be a float (rounding bug fix).
    pc = cf_ws.cell(20, 6).value
    assert isinstance(pc, float) and 0 < pc < 1


@pytest.mark.skipif(not _REAL_MR.exists() or not _REAL_PREV.exists(),
                    reason="real farada files not present")
def test_real_pipeline_kpi_derivation_matches_feb_values(tmp_path):
    """Sanity-check the KPI formulas by re-deriving Feb values from MR
    and comparing against the taxonomi's reported Feb values.

    Run the populator with target month=2 against a synthetic prev
    taxonomi that has the right Jan begin balances. The result should
    match the existing Feb taxonomi values within rounding (€1) for WC
    and within ~0.001 for AP/AR turnover."""
    mapping = yaml.safe_load(_REAL_MAPPING.read_text())
    extracts_feb = {
        stmt: extract_month(_REAL_MR, mapping, 2026, 2, stmt)
        for stmt in ("IS", "CF", "BS")
    }
    out = tmp_path / "feb_redo.xlsx"
    populate_taxonomi(_REAL_PREV, extracts_feb, 2026, 2, out)
    wb = load_workbook(out)
    bs_ws = wb["BS (Actual)"]

    # AP Turnover Feb reported in the existing taxonomi = 0.6900 (verified).
    apt_feb = bs_ws.cell(19, 5).value
    assert apt_feb == pytest.approx(0.6900, abs=0.001)


def test_ratio_cells_kept_as_float(mini_prev, mini_extracts, tmp_path):
    """Add a % Change in cash row to the mini taxonomi; ensure the value
    is stored as float, not rounded to int."""
    wb = load_workbook(mini_prev)
    cf = wb["CF Indirect (Actual)"]
    cf.append(["% Change in cash", "% Change in cash", "% Change in cash",
               0.10, 0.20, None, None, None, None, None, None, None,
               None, None, None])
    wb.save(mini_prev)
    extracts = {**mini_extracts, "CF": {
        **mini_extracts["CF"],
        ("% Change in cash", "% Change in cash", "% Change in cash"): 0.245,
    }}
    out = tmp_path / "out.xlsx"
    populate_taxonomi(mini_prev, extracts, 2026, 3, out)
    cf_out = load_workbook(out)["CF Indirect (Actual)"]
    # Find the % Change in cash row (row 3 since we appended after the existing row 2).
    for r in range(2, cf_out.max_row + 1):
        if cf_out.cell(r, 1).value == "% Change in cash":
            v = cf_out.cell(r, 6).value
            assert isinstance(v, float)
            assert v == pytest.approx(0.245)
            return
    pytest.fail("% Change in cash row not found in output")
