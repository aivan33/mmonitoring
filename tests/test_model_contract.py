"""Tests for core.model.contract — the structure layer of the model parser.

A general classifier engine turns sheet names into typed SheetInfo(entity, role,
statement) driven by a small per-client Rules config; ModelContract groups them
and exposes the actuals/budget/driver seams plus the taxonomi month-axis.
"""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from core.model.contract import (
    ModelContract,
    Rules,
    TaxonomiAxis,
    classify_sheet,
    load_rules,
    read_contract,
)

ALMACENA = Path("clients/almacena/budget/Almacena-26_AprActuals.xlsx")


@pytest.fixture
def rules():
    return Rules(
        entity_patterns={"foundation": ["found"], "bg": ["bg"], "consolidated": ["cons"]},
        default_entity="consolidated",
        role_overrides={
            " Inputs": "engine",
            "Inputs_Foundation": "engine",
            "KPIs": "driver",
        },
        separator_marker=">>>",
        taxonomi_axis=TaxonomiAxis(header_row=1, first_month_col="D", months=12, year=2026),
    )


# ---- classify_sheet (the general engine) ----

def test_separator_sheet(rules):
    info = classify_sheet("Consolidated >>>", rules)
    assert info.role == "separator"
    assert info.entity is None and info.statement is None


def test_consolidated_taxonomi(rules):
    info = classify_sheet("is_cons_taxonomi", rules)
    assert (info.role, info.entity, info.statement) == ("taxonomi", "consolidated", "IS")


def test_foundation_taxonomi(rules):
    info = classify_sheet("cf_found_taxonomi", rules)
    assert (info.role, info.entity, info.statement) == ("taxonomi", "foundation", "CF")


def test_yearly_defaults_to_consolidated(rules):
    info = classify_sheet("IS_Yearly", rules)
    assert (info.role, info.entity, info.statement) == ("yearly", "consolidated", "IS")


def test_actuals_sheet_has_no_statement_letter(rules):
    info = classify_sheet("actuals_found", rules)
    assert (info.role, info.entity, info.statement) == ("actuals", "foundation", None)


def test_underscore_act_is_actuals(rules):
    assert classify_sheet("BV_act", rules).role == "actuals"


def test_bare_statement_defaults_to_consolidated(rules):
    info = classify_sheet("IS", rules)
    assert (info.role, info.entity, info.statement) == ("statement", "consolidated", "IS")


def test_bg_statement(rules):
    info = classify_sheet("IS_BG", rules)
    assert (info.role, info.entity, info.statement) == ("statement", "bg", "IS")


def test_engine_override_is_shared_when_no_entity_marker(rules):
    info = classify_sheet(" Inputs", rules)
    assert info.role == "engine" and info.entity is None


def test_engine_override_keeps_entity_marker(rules):
    info = classify_sheet("Inputs_Foundation", rules)
    assert (info.role, info.entity) == ("engine", "foundation")


def test_driver_override_is_shared(rules):
    info = classify_sheet("KPIs", rules)
    assert info.role == "driver" and info.entity is None


def test_unknown_sheet_is_other_not_dropped(rules):
    info = classify_sheet("Random Sheet", rules)
    assert info.role == "other"


# ---- ModelContract grouping + seams ----

@pytest.fixture
def contract(tmp_path, rules):
    wb = Workbook()
    wb.remove(wb.active)
    for name in ["Consolidated >>>", "IS", "is_cons_taxonomi", "Consolidated Actuals",
                 "is_found_taxonomi", "actuals_found", " Inputs", "KPIs"]:
        ws = wb.create_sheet(title=name)
        if "taxonomi" in name:
            ws["A1"] = "Data"
            for i, m in enumerate(["Jan", "Feb", "Mar", "Apr"]):
                ws.cell(row=1, column=4 + i, value=m)
            ws.cell(row=2, column=4, value=100)  # Jan populated
            ws.cell(row=2, column=5, value=200)  # Feb populated
    path = tmp_path / "model.xlsx"
    wb.save(path)
    return read_contract(path, rules)


def test_lists_entities_present(contract):
    assert contract.entities() == ["consolidated", "foundation"]


def test_groups_by_role(contract):
    names = {s.name for s in contract.by_role("taxonomi")}
    assert names == {"is_cons_taxonomi", "is_found_taxonomi"}


def test_seams_pair_budget_and_actuals_per_entity(contract):
    seams = contract.seams()
    assert seams["consolidated"]["budget"] == ["is_cons_taxonomi"]
    assert seams["consolidated"]["actuals"] == ["Consolidated Actuals"]
    assert seams["foundation"]["actuals"] == ["actuals_found"]


def test_month_axis_maps_columns_to_periods(contract):
    axis = contract.month_axis()
    assert axis["D"] == "2026-01"
    assert axis["O"] == "2026-12"


def test_last_populated_month_reads_values(contract):
    # synthetic taxonomi has Jan+Feb filled
    assert contract.last_populated_month("is_cons_taxonomi") == "2026-02"


# ---- validation against the real Almacena workbook ----

@pytest.mark.skipif(not ALMACENA.exists(), reason="gitignored client model absent")
def test_almacena_classifies_with_real_rules():
    rules = load_rules(Path("clients/almacena/model_rules.yaml"))
    contract = read_contract(ALMACENA, rules)
    assert set(contract.entities()) >= {"consolidated", "foundation", "bg", "bvnv"}
    by_name = {s.name: s for s in contract.sheets}
    assert (by_name["is_cons_taxonomi"].role, by_name["is_cons_taxonomi"].entity) == ("taxonomi", "consolidated")
    assert by_name[" Inputs"].role == "engine"
    assert by_name["KPIs"].role == "driver"
    # every sheet is classified (no role is None/empty)
    assert all(s.role for s in contract.sheets)
