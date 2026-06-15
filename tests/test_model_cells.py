"""Tests for core.model.cells — the cell layer of the model parser.

Each cell exposes its cached value, its raw formula string (if any), its
number format, and a semantic dtype. Fixtures are authored with openpyxl, so
formula cells carry NO cached value (openpyxl does not calculate) — that is the
documented behaviour; real Excel-saved workbooks carry the cache.
"""

from __future__ import annotations

import datetime as dt

import pytest
from openpyxl import Workbook

from core.model.cells import read_cells


@pytest.fixture
def workbook(tmp_path):
    """A small workbook covering each cell kind."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = 42                      # number literal
    ws["A2"] = "hello"                 # text
    ws["A3"] = "=A1+A1"               # formula
    ws["A4"] = True                    # bool
    ws["A5"] = dt.date(2026, 4, 30)    # date
    ws["B1"] = 1234.5
    ws["B1"].number_format = "#,##0.00"
    wb.create_sheet("Other")
    path = tmp_path / "wb.xlsx"
    wb.save(path)
    return read_cells(path)


def test_reads_a_number_literal(workbook):
    c = workbook.cell("Sheet1", "A1")
    assert c.value == 42
    assert c.formula is None
    assert c.dtype == "number"
    assert c.sheet == "Sheet1"
    assert c.coord == "A1"


def test_reads_a_text_cell(workbook):
    c = workbook.cell("Sheet1", "A2")
    assert c.value == "hello"
    assert c.dtype == "text"
    assert c.formula is None


def test_captures_the_formula_string(workbook):
    c = workbook.cell("Sheet1", "A3")
    assert c.formula == "=A1+A1"
    assert c.dtype == "formula"
    assert c.is_formula is True
    # openpyxl-authored fixtures have no cached result
    assert c.value is None


def test_reads_a_boolean_cell(workbook):
    c = workbook.cell("Sheet1", "A4")
    assert c.value is True
    assert c.dtype == "bool"


def test_reads_a_date_cell(workbook):
    c = workbook.cell("Sheet1", "A5")
    assert c.dtype == "date"
    assert c.value == dt.datetime(2026, 4, 30)


def test_empty_cell_is_dtype_empty(workbook):
    c = workbook.cell("Sheet1", "Z99")
    assert c.value is None
    assert c.formula is None
    assert c.dtype == "empty"


def test_preserves_number_format(workbook):
    assert workbook.cell("Sheet1", "B1").number_format == "#,##0.00"


def test_lists_sheets(workbook):
    assert workbook.sheets() == ["Sheet1", "Other"]


def test_unknown_sheet_raises(workbook):
    with pytest.raises(KeyError):
        workbook.cell("Nope", "A1")
