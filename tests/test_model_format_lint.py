"""Tests for core.model.format_lint — design-system conformance (warn-only)."""

from __future__ import annotations

import openpyxl

from core.model import design_system as ds
from core.model.format_lint import font_conformance, inputs_grammar, lint


def test_font_flags_non_century_gothic() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "ok"
    ws["A1"].font = ds.font()          # Century Gothic
    ws["A2"] = "bad"                    # default Calibri
    v = font_conformance(wb)
    assert any(x.cell == "A2" for x in v)
    assert not any(x.cell == "A1" for x in v)


def test_inputs_grammar_flags_broken_selector() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = " Inputs"
    ws["J5"] = "=OFFSET(K5,0,$D$2)"     # correct scenario selector
    ws["J6"] = "=K6"                    # broken — not the OFFSET selector
    v = inputs_grammar(wb, " Inputs")
    assert any(x.cell == "J6" for x in v)
    assert not any(x.cell == "J5" for x in v)


def test_lint_combines_and_is_clean_on_conformant_sheet() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = " Inputs"
    ws["C4"] = "Revenue days"
    ws["C4"].font = ds.font()
    ws["J4"] = "=OFFSET(K4,0,$D$2)"
    ws["J4"].font = ds.font()
    assert lint(wb, inputs_sheet=" Inputs") == []
