"""Write a new month's column into a copy of the prior taxonomi-actual xlsx.

Bridges ``core/report/mr.py`` (MR → keyed values) and the canonical taxonomi
format. Uses ``openpyxl`` load-modify-save so cell formatting, column widths,
and other workbook metadata from the prior month carry over untouched.

Two writes happen:
    1. MR-sourced cells (per ``mapping_is/cf/bs``) get the values from
       ``extract_month``.
    2. Derived KPI rows are computed from the per-client ``kpi_derivations:``
       block in mapping.yaml. Four formula types are supported today:

       - ``sum``: Σ over a list of source rows.
       - ``working_capital``: Σ(current_assets) − Σ(current_liabilities).
       - ``turnover_ratio``: numerator / avg(begin_balance, end_balance).
       - ``ap_turnover``: (Σ cost_data − Σ personnel) / avg(begin, end).

Per-row floats are declared via ``store_as: float`` on either a mapping
entry or a kpi_derivations entry. Without it, numeric cells round to int.
"""

from __future__ import annotations

import logging
from pathlib import Path

from openpyxl import load_workbook

logger = logging.getLogger(__name__)


# Canonical taxonomi sheet name per statement code. Format-defined, not
# client-configurable.
_SHEET_FOR: dict[str, str] = {
    "IS": "IS (Actual)",
    "CF": "CF Indirect (Actual)",
    "BS": "BS (Actual)",
}


def populate_taxonomi(
    prev_taxonomi: str | Path,
    mr_extracts: dict[str, dict[tuple[str, str, str], float | None]],
    year: int,
    month: int,
    out_path: str | Path,
    *,
    mapping: dict | None = None,
) -> None:
    """Copy ``prev_taxonomi`` and overwrite the new month's column with
    MR-sourced values plus derived KPIs declared in ``mapping['kpi_derivations']``.

    Args:
        prev_taxonomi: path to prior month's taxonomi-actual xlsx.
        mr_extracts: ``{statement_code: {(data, grp, subgroup): value}}``
            from ``core/report/mr.extract_month``. Statement codes:
            ``'IS' | 'CF' | 'BS'``.
        year, month: target period.
        out_path: where to save the populated xlsx.
        mapping: parsed mapping.yaml. Optional; without it no derivations
            run and no per-row float overrides apply.

    Behavior:
        - Target column = ``3 + month`` (Jan=4, ..., Dec=15).
        - MR-sourced numeric cells round to integers; cells flagged
          ``store_as: float`` (in mapping entries or kpi_derivations entries)
          are stored as floats.
        - Idempotent: same inputs produce the same output content.
    """
    out_path = Path(out_path)
    target_col = 3 + month  # Jan=4, ..., Dec=15

    norm_extracts: dict[str, dict[tuple[str, str, str], float | None]] = {
        stmt: {tuple(_strip(p) for p in key): val for key, val in d.items()}
        for stmt, d in mr_extracts.items()
    }
    for stmt in _SHEET_FOR:
        norm_extracts.setdefault(stmt, {})

    float_keys = _collect_float_keys(mapping)

    wb = load_workbook(prev_taxonomi)
    try:
        # Step 1: derive KPIs from MR values + prior-month begin balances.
        for entry in (mapping or {}).get("kpi_derivations") or []:
            stmt_for_target = entry["statement_for_target"]
            target_key = tuple(entry["target"])
            value = _evaluate_formula(entry, norm_extracts, wb, month)
            norm_extracts.setdefault(stmt_for_target, {})[target_key] = value

        # Step 2: write per-sheet.
        for stmt_code, sheet_name in _SHEET_FOR.items():
            if sheet_name not in wb.sheetnames:
                logger.warning(
                    "prev taxonomi has no sheet %r — skipping %s.",
                    sheet_name, stmt_code,
                )
                continue
            ws = wb[sheet_name]
            extracts = norm_extracts.get(stmt_code, {})
            _populate_sheet(ws, extracts, target_col, sheet_name, float_keys)

        out_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(out_path)
    finally:
        wb.close()


# --- KPI formula dispatch --------------------------------------------------

def _evaluate_formula(
    entry: dict, extracts: dict, wb, month: int,
) -> float | None:
    formula = entry.get("formula")
    if formula == "sum":
        return _formula_sum(entry, extracts)
    if formula == "working_capital":
        return _formula_working_capital(entry, extracts)
    if formula == "turnover_ratio":
        return _formula_turnover_ratio(entry, extracts, wb, month)
    if formula == "ap_turnover":
        return _formula_ap_turnover(entry, extracts, wb, month)
    raise ValueError(
        f"kpi_derivations: unknown formula type {formula!r}; "
        f"expected one of: sum, working_capital, turnover_ratio, ap_turnover"
    )


def _formula_sum(entry: dict, extracts: dict) -> float | None:
    """Σ over a list of source rows. Returns None (and logs a warning naming the
    missing key(s)) if any component is absent from the extract or present with
    a None/non-numeric value — a dropped extract must fail loud, not silently
    count as 0 and yield a plausible-but-wrong total."""
    total, missing = _sum_sources(entry.get("sources") or [], extracts)
    if missing:
        logger.warning("kpi %s: missing component(s) %s → value is None",
                       entry.get("target"), missing)
        return None
    return total


def _formula_working_capital(entry: dict, extracts: dict) -> float | None:
    """Working Capital = Σ(current_assets) − Σ(current_liabilities). Returns None
    (with a warning naming the missing key(s)) if any component is absent or
    None, rather than treating a dropped BS line as 0."""
    ca, miss_a = _sum_sources(entry.get("current_assets") or [], extracts)
    cl, miss_l = _sum_sources(entry.get("current_liabilities") or [], extracts)
    missing = miss_a + miss_l
    if missing:
        logger.warning("kpi %s: missing component(s) %s → value is None",
                       entry.get("target"), missing)
        return None
    return ca - cl


def _sum_sources(
    sources: list, extracts: dict,
) -> tuple[float, list[tuple[str, str, str]]]:
    """Sum source values, collecting the keys of any that are missing. A key
    that is absent and one present with value None are both 'missing'."""
    total = 0.0
    missing: list[tuple[str, str, str]] = []
    for src in sources:
        v = _get_source_opt(extracts, src)
        if v is None:
            missing.append(tuple(src["key"]))
        else:
            total += v
    return total, missing


def _formula_turnover_ratio(
    entry: dict, extracts: dict, wb, month: int,
) -> float | None:
    """Turnover ratio = numerator / avg(begin, end). Returns None when the
    average balance is 0 or when begin balance is unavailable."""
    numerator = _resolve_numerator(entry["numerator"], extracts)
    avg = _avg_balance(entry["avg_balance_key"], extracts, wb, month)
    if avg is None or avg == 0:
        return None
    return numerator / avg


def _formula_ap_turnover(
    entry: dict, extracts: dict, wb, month: int,
) -> float | None:
    """AP Turnover = (Σ cost_data − Σ personnel_keys) / avg(begin, end).

    cost_data_classes lists statement-level data classes whose extracts are
    summed (across all statements/scenarios). personnel_keys lists specific
    rows to subtract from that total.
    """
    cost_classes = set(entry.get("cost_data_classes") or [])
    personnel_keys = {tuple(k) for k in entry.get("personnel_keys") or []}

    cost_total = 0.0
    personnel_total = 0.0
    for ext in extracts.values():
        for k, v in ext.items():
            if not isinstance(v, (int, float)):
                continue
            if k[0] in cost_classes:
                cost_total += v
            if k in personnel_keys:
                personnel_total += v

    avg = _avg_balance(entry["avg_balance_key"], extracts, wb, month)
    if avg is None or avg == 0:
        return None
    return (cost_total - personnel_total) / avg


def _resolve_numerator(spec: dict, extracts: dict) -> float:
    kind = spec["type"]
    if kind == "data_aggregate":
        stmt = spec["source"]
        data = spec["data"]
        return sum(
            v for k, v in extracts.get(stmt, {}).items()
            if k[0] == data and isinstance(v, (int, float))
        )
    if kind == "source_value":
        return _get_source(extracts, {"statement": spec["source"], "key": spec["key"]})
    raise ValueError(f"unknown numerator type {kind!r}")


def _avg_balance(
    spec: dict, extracts: dict, wb, month: int,
) -> float | None:
    """avg(begin, end). begin = prior-month BS value; end = current extract."""
    end = _get_source(extracts, spec)
    sheet = _SHEET_FOR[spec["statement"]]
    begin = _read_prior_month(wb, sheet, tuple(spec["key"]), month)
    if begin is None:
        return None
    return (begin + end) / 2


def _get_source(extracts: dict, src: dict) -> float:
    """Extract one source value; missing/None counts as 0.0."""
    v = extracts.get(src["statement"], {}).get(tuple(src["key"]))
    return float(v) if isinstance(v, (int, float)) else 0.0


def _get_source_opt(extracts: dict, src: dict) -> float | None:
    """Like ``_get_source`` but distinguishes missing from zero: returns None
    when the statement/key is absent, or present with a None/non-numeric value."""
    v = extracts.get(src["statement"], {}).get(tuple(src["key"]))
    return float(v) if isinstance(v, (int, float)) else None


def _read_prior_month(
    wb, sheet_name: str, key: tuple[str, str, str], month: int,
) -> float | None:
    """Read the value at ``key`` from the prior month's column of
    ``sheet_name``. Returns ``None`` if month==1, the sheet is missing,
    the row isn't found, or the cell is empty."""
    if month <= 1:
        return None
    if sheet_name not in wb.sheetnames:
        return None
    ws = wb[sheet_name]
    prior_col = 3 + (month - 1)
    for r in range(2, ws.max_row + 1):
        d = _strip(ws.cell(r, 1).value)
        g = _strip(ws.cell(r, 2).value)
        s = _strip(ws.cell(r, 3).value)
        if (d, g, s) == key:
            v = ws.cell(r, prior_col).value
            return float(v) if isinstance(v, (int, float)) else None
    return None


# --- per-sheet writer ------------------------------------------------------

def _collect_float_keys(
    mapping: dict | None,
) -> set[tuple[str, str, str]]:
    """Build the set of triplets that should be stored as float, not rounded.
    Reads ``store_as: float`` from per-row mapping entries and from
    kpi_derivations entries."""
    if mapping is None:
        return set()
    keys: set[tuple[str, str, str]] = set()
    for stmt_key in ("mapping_is", "mapping_cf", "mapping_bs"):
        for entry in mapping.get(stmt_key) or []:
            if entry.get("store_as") == "float":
                keys.add((entry["data"], entry["grp"], entry["subgroup"]))
    for entry in mapping.get("kpi_derivations") or []:
        if entry.get("store_as") == "float":
            keys.add(tuple(entry["target"]))
    return keys


def _populate_sheet(
    ws,
    extracts: dict[tuple[str, str, str], float | None],
    target_col: int,
    sheet_name: str,
    float_keys: set[tuple[str, str, str]],
) -> None:
    for r in range(2, ws.max_row + 1):
        d = ws.cell(r, 1).value
        g = ws.cell(r, 2).value
        s = ws.cell(r, 3).value
        if d is None and g is None and s is None:
            continue
        key = (_strip(d), _strip(g), _strip(s))
        if key not in extracts:
            logger.warning(
                "taxonomi %s row %d %r has no MR mapping — leaving cell unchanged.",
                sheet_name, r, key,
            )
            continue
        value = extracts[key]
        if value is None:
            ws.cell(r, target_col).value = None
        elif key in float_keys:
            ws.cell(r, target_col).value = float(value)
        else:
            ws.cell(r, target_col).value = round(value)


def _strip(v):
    if isinstance(v, str):
        return v.strip()
    return v
