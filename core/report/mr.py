"""Loader for management-reporting (MR) workbooks.

The MR is a per-client master accounting workbook (DATEV-derived for
FaradaIC, in-house spreadsheets for others). Layouts vary across clients,
so per-statement layout settings live in each client's ``mapping.yaml``
under an ``mr_layout:`` block. Output feeds ``core/report/mr_to_taxonomi.py``
which writes a new month's column into a copy of the prior taxonomi-actual
xlsx.

A client without an ``mr_layout:`` block falls back to FaradaIC-style
defaults: sheets ``P&L`` / ``CF`` / ``BS``, header row 2, ``dt.date`` cells
in the header for column matching.
"""

from __future__ import annotations

import datetime as dt
import logging
from pathlib import Path

from openpyxl import load_workbook

logger = logging.getLogger(__name__)


# Defaults preserve FaradaIC behavior when ``mr_layout`` is absent from
# mapping.yaml — existing clients keep working without config changes.
_DEFAULT_LAYOUT: dict[str, dict] = {
    "IS": {"sheet": "P&L", "header_row": 2, "label_col": 1, "period_format": "date_cell"},
    "CF": {"sheet": "CF",  "header_row": 2, "label_col": 2, "period_format": "date_cell"},
    "BS": {"sheet": "BS",  "header_row": 2, "label_col": 2, "period_format": "date_cell"},
}

_VALID_PERIOD_FORMATS = ("date_cell", "month_name")

_MAPPING_KEY: dict[str, str] = {
    "IS": "mapping_is",
    "CF": "mapping_cf",
    "BS": "mapping_bs",
}

# Cell strings treated as "no value".
_BLANK_STRINGS = frozenset({"", "-", "—", "n/a", "N/A"})

# For ``period_format='month_name'``: English month names → 1..12.
_MONTH_NAMES: dict[str, int] = {
    name.lower(): idx
    for idx, name in enumerate(
        ["January", "February", "March", "April", "May", "June",
         "July", "August", "September", "October", "November", "December"],
        start=1,
    )
}


def extract_month(
    mr_path: str | Path,
    mapping: dict,
    year: int,
    month: int,
    statement: str,
) -> dict[tuple[str, str, str], float | None]:
    """Extract one month's values from an MR sheet, mapped onto taxonomi keys.

    Args:
        mr_path: path to the MR workbook.
        mapping: parsed ``mapping.yaml`` dict; must contain the per-statement
            list keyed by ``mapping_is`` / ``mapping_cf`` / ``mapping_bs``.
            May contain an ``mr_layout`` block overriding the per-statement
            sheet name, header row, label column, and period format.
        year, month: the target period.
        statement: ``'IS' | 'CF' | 'BS'``.

    Returns:
        ``dict`` mapping ``(data, grp, subgroup)`` → ``float`` or ``None``.
        ``None`` is emitted when the cell is empty / a dash placeholder, or
        when the mapping entry has ``mr_row=null`` (derived/non-MR rows).

    Behavior:
        - Resolves the target column by scanning ``layout['header_row']``
          for a header that matches ``(year, month)`` according to
          ``layout['period_format']``. Raises ``ValueError`` if absent.
        - For each entry: reads the configured row's label; if it matches
          ``mr_label``, uses the configured row. Otherwise searches the
          label column for ``mr_label`` and uses that row instead, emitting
          a ``logging.warning`` with both row indices. If the label can't
          be found anywhere, the entry's value is ``None`` and a warning
          is emitted.
    """
    if statement not in _DEFAULT_LAYOUT:
        raise ValueError(
            f"statement={statement!r}; expected one of {list(_DEFAULT_LAYOUT)}"
        )

    layout = _resolve_layout(mapping, statement)
    entries = mapping[_MAPPING_KEY[statement]]

    wb = load_workbook(mr_path, data_only=True)
    try:
        ws = wb[layout["sheet"]]
        target_col = _find_period_column(ws, year, month, layout, statement)
        return _extract(
            ws, entries,
            label_col=layout["label_col"],
            header_row=layout["header_row"],
            target_col=target_col,
            statement=statement,
        )
    finally:
        wb.close()


def _resolve_layout(mapping: dict, statement: str) -> dict:
    """Merge per-client ``mr_layout`` over the per-statement defaults."""
    layout = dict(_DEFAULT_LAYOUT[statement])
    user_layout = (mapping.get("mr_layout") or {}).get(statement) or {}
    layout.update(user_layout)
    if layout["period_format"] not in _VALID_PERIOD_FORMATS:
        raise ValueError(
            f"mr_layout[{statement}].period_format={layout['period_format']!r}; "
            f"expected one of {_VALID_PERIOD_FORMATS}"
        )
    return layout


def _find_period_column(
    ws, year: int, month: int, layout: dict, statement: str,
) -> int:
    header_row = layout["header_row"]
    fmt = layout["period_format"]
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if fmt == "date_cell":
            if isinstance(v, dt.date) and v.year == year and v.month == month:
                return c
        elif fmt == "month_name":
            if isinstance(v, str) and _MONTH_NAMES.get(v.strip().lower()) == month:
                return c
    raise ValueError(
        f"{statement} sheet {layout['sheet']!r}: no header column for "
        f"{year}-{month:02d} in row {header_row} "
        f"(period_format={fmt!r})"
    )


def _extract(
    ws, entries, *, label_col: int, header_row: int, target_col: int,
    statement: str,
) -> dict[tuple[str, str, str], float | None]:
    out: dict[tuple[str, str, str], float | None] = {}
    for entry in entries:
        key = (entry["data"], entry["grp"], entry["subgroup"])
        mr_row = entry["mr_row"]
        if mr_row is None:
            out[key] = None
            continue
        row = _resolve_row(
            ws, mr_row, entry["mr_label"], label_col, header_row, statement,
        )
        if row is None:
            out[key] = None
            continue
        value = _coerce_value(ws.cell(row, target_col).value)
        if value is not None:
            sign = entry.get("sign", 1)
            value = value * sign
        out[key] = value
    return out


def _resolve_row(
    ws, mr_row: int, mr_label: str, label_col: int, header_row: int,
    statement: str,
) -> int | None:
    actual = ws.cell(mr_row, label_col).value
    if actual == mr_label:
        return mr_row
    for r in range(header_row + 1, ws.max_row + 1):
        if ws.cell(r, label_col).value == mr_label:
            logger.warning(
                "%s: configured row %d has label %r, but %r found at row %d "
                "— using row %d. Update mapping.yaml.",
                statement, mr_row, actual, mr_label, r, r,
            )
            return r
    logger.warning(
        "%s: configured row %d has label %r, expected %r — label not found "
        "elsewhere in sheet. Leaving cell null.",
        statement, mr_row, actual, mr_label,
    )
    return None


def _coerce_value(raw):
    if raw is None:
        return None
    if isinstance(raw, str):
        if raw.strip() in _BLANK_STRINGS:
            return None
        try:
            return float(raw)
        except ValueError:
            return None
    return float(raw)
