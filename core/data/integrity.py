"""Integrity engine — runs R3, R4, R5 against a loaded financials DB.

R3 — total-lookalike scan
    Rows whose grp/subgroup label matches a Total/Subtotal/Net pattern but
    aren't registered as aggregates. Warn-only: hint to register them or
    rename the row.

R4 — registered aggregate recompute
    For every registry entry, recompute Σ(leaf × sign) and compare to the
    parsed aggregate value. Fail above tolerance — this is the canonical
    "hardcoded value drifted from its formula" trip-wire.

R5 — source-cell type audit
    For every registry entry with ``source_cell`` set, open the workbook
    and verify the cell is an Excel formula (data_type='f'). Hardcoded
    values warn — they're correct today but will silently drift.
"""

from __future__ import annotations

import re
import sqlite3
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from core.data.aggregate_formulas import AggregateFormula


@dataclass(frozen=True)
class Finding:
    rule: str       # "R3" | "R4" | "R5" | "R6"
    severity: str   # "fail" | "warn"
    name: str       # aggregate name or row identifier
    message: str


@dataclass(frozen=True)
class IntegrityReport:
    findings: tuple[Finding, ...] = ()

    @property
    def failures(self) -> tuple[Finding, ...]:
        return tuple(f for f in self.findings if f.severity == "fail")

    @property
    def warnings(self) -> tuple[Finding, ...]:
        return tuple(f for f in self.findings if f.severity == "warn")

    def has_failures(self) -> bool:
        return any(f.severity == "fail" for f in self.findings)


# Matches "Total", "Subtotal", "Sum of …", "Net Income / Net Burn / Net …".
# Plain "Net" (no following word) is NOT flagged — too generic.
_TOTAL_PATTERN = re.compile(
    r"^\s*(total|subtotal|sum|net\s+\S)", re.IGNORECASE,
)


def check_integrity(
    conn: sqlite3.Connection,
    registry: dict[str, AggregateFormula],
    *,
    workbook_paths: Iterable[Path] | None = None,
    tolerance: float = 1.0,
) -> IntegrityReport:
    findings: list[Finding] = []
    findings.extend(_run_r4(conn, registry, tolerance))
    findings.extend(_run_r3(conn, registry))
    findings.extend(_run_r6(conn, tolerance))
    findings.extend(_run_r7(conn))
    findings.extend(_run_r9(conn))
    if workbook_paths is not None:
        findings.extend(_run_r5(registry, list(workbook_paths)))
    return IntegrityReport(findings=tuple(findings))


# ---------------------------------------------------------------------------
# R4
# ---------------------------------------------------------------------------

def _run_r4(
    conn: sqlite3.Connection,
    registry: dict[str, AggregateFormula],
    tolerance: float,
) -> Iterable[Finding]:
    for name, formula in registry.items():
        rows = conn.execute(
            "SELECT period_date, scenario, entity, value FROM financials "
            "WHERE data=? AND grp=? AND subgroup=? AND value IS NOT NULL",
            (formula.data, formula.grp, formula.subgroup),
        ).fetchall()
        for period, scenario, entity, parsed in rows:
            recomputed = _recompute(conn, formula, period, scenario, entity)
            if recomputed is None:
                continue
            delta = parsed - recomputed
            if abs(delta) > tolerance:
                yield Finding(
                    rule="R4",
                    severity="fail",
                    name=name,
                    message=(
                        f"{name} at {period} ({scenario}/{entity}): "
                        f"parsed {parsed:.2f}, recomputed {recomputed:.2f}, "
                        f"delta {delta:+.2f} (tolerance {tolerance:.2f})"
                    ),
                )


def _recompute(
    conn: sqlite3.Connection,
    formula: AggregateFormula,
    period: str,
    scenario: str,
    entity: str,
) -> float | None:
    """Return Σ(leaf × sign) over leaves only (is_aggregate=0). Returns None
    if no leaf data was found at all — caller skips rather than fails."""
    total = 0.0
    saw_any = False
    for leaf in formula.leaves:
        sql = (
            "SELECT SUM(value) FROM financials "
            "WHERE data=? AND period_date=? AND scenario=? AND entity=? "
            "AND is_aggregate=0"
        )
        args: list = [leaf.data, period, scenario, entity]
        if leaf.grp is not None:
            sql += " AND grp=?"
            args.append(leaf.grp)
        if leaf.subgroup is not None:
            sql += " AND subgroup=?"
            args.append(leaf.subgroup)
        row = conn.execute(sql, args).fetchone()
        v = row[0] if row else None
        if v is None:
            continue
        saw_any = True
        total += leaf.sign * v
    return total if saw_any else None


# ---------------------------------------------------------------------------
# R3
# ---------------------------------------------------------------------------

def _run_r3(
    conn: sqlite3.Connection,
    registry: dict[str, AggregateFormula],
) -> Iterable[Finding]:
    registered = {(f.data, f.grp, f.subgroup) for f in registry.values()}
    rows = conn.execute(
        "SELECT DISTINCT data, grp, subgroup FROM financials "
        "WHERE is_aggregate=0"
    ).fetchall()
    for data, grp, subgroup in rows:
        if (data, grp, subgroup) in registered:
            continue  # registered → already covered by R4
        for label in (subgroup, grp):
            if label and _TOTAL_PATTERN.match(str(label)):
                yield Finding(
                    rule="R3",
                    severity="warn",
                    name=f"{data} / {grp} / {subgroup}",
                    message=(
                        f"row labelled like a total but not registered: "
                        f"({data!r}, {grp!r}, {subgroup!r})"
                    ),
                )
                break  # one finding per row, even if both grp+subgroup match


# ---------------------------------------------------------------------------
# R6 — three-statement cash coherence
# ---------------------------------------------------------------------------

# Standard taxonomi names. If a client uses different labels for these rows
# the check silently skips — better than false positives.
_CASH_DATA = "Cash and cash equivalents"
_CFO_DATA = "Cash Flow from Operating Activities"
_CFI_DATA = "Cash Flow from Investing Activities"
_CFF_DATA = "Cash Flow from Financing Activities"


def _run_r6(
    conn: sqlite3.Connection,
    tolerance: float,
) -> Iterable[Finding]:
    """For every (scenario, entity) pair that has BOTH BS and CF data, walk
    consecutive periods and verify ΔCash[t] ≈ CFO[t] + CFI[t] + CFF[t].
    Pairs lacking either statement are skipped — incomplete data isn't a
    coherence violation."""
    combos = conn.execute(
        "SELECT DISTINCT bs.scenario, bs.entity FROM financials bs "
        "WHERE bs.statement='BS' "
        "AND EXISTS (SELECT 1 FROM financials cf "
        "            WHERE cf.statement='CF' "
        "            AND cf.scenario=bs.scenario AND cf.entity=bs.entity)"
    ).fetchall()
    for scenario, entity in combos:
        periods = [
            row[0] for row in conn.execute(
                "SELECT DISTINCT period_date FROM financials "
                "WHERE statement='BS' AND scenario=? AND entity=? "
                "AND data=? "
                "ORDER BY period_date",
                (scenario, entity, _CASH_DATA),
            ).fetchall()
        ]
        if len(periods) < 2:
            continue
        for prev, curr in zip(periods, periods[1:]):
            cash_prev = _scalar(conn,
                "SELECT SUM(value) FROM financials "
                "WHERE statement='BS' AND scenario=? AND entity=? "
                "AND data=? AND period_date=? AND is_aggregate=0",
                (scenario, entity, _CASH_DATA, prev),
            )
            cash_curr = _scalar(conn,
                "SELECT SUM(value) FROM financials "
                "WHERE statement='BS' AND scenario=? AND entity=? "
                "AND data=? AND period_date=? AND is_aggregate=0",
                (scenario, entity, _CASH_DATA, curr),
            )
            if cash_prev is None or cash_curr is None:
                continue
            cf_total = sum(
                _scalar(conn,
                    "SELECT SUM(value) FROM financials "
                    "WHERE statement='CF' AND scenario=? AND entity=? "
                    "AND data=? AND period_date=? AND is_aggregate=0",
                    (scenario, entity, data, curr),
                ) or 0.0
                for data in (_CFO_DATA, _CFI_DATA, _CFF_DATA)
            )
            delta_cash = cash_curr - cash_prev
            mismatch = delta_cash - cf_total
            if abs(mismatch) > tolerance:
                # warn, not fail: cash-bridge math accumulates rounding
                # across three statements. R4 (registered formulas) is
                # where exact-match failures belong.
                yield Finding(
                    rule="R6",
                    severity="warn",
                    name=f"cash-flow-coherence/{scenario}/{entity}",
                    message=(
                        f"{curr} ({scenario}/{entity}): ΔCash {delta_cash:+.2f} "
                        f"vs CFO+CFI+CFF {cf_total:+.2f}, "
                        f"mismatch {mismatch:+.2f} (tolerance {tolerance:.2f})"
                    ),
                )


def _scalar(
    conn: sqlite3.Connection, sql: str, args: tuple,
) -> float | None:
    row = conn.execute(sql, args).fetchone()
    return None if row is None or row[0] is None else float(row[0])


# ---------------------------------------------------------------------------
# R7 — period continuity
# ---------------------------------------------------------------------------

def _run_r7(conn: sqlite3.Connection) -> Iterable[Finding]:
    """For each (scenario, entity, statement), every month between the
    earliest and latest loaded period must be present. Range start ≠ Jan 1
    is fine; gaps inside the range are not."""
    combos = conn.execute(
        "SELECT DISTINCT scenario, entity, statement FROM financials"
    ).fetchall()
    for scenario, entity, statement in combos:
        periods = sorted(
            row[0] for row in conn.execute(
                "SELECT DISTINCT period_date FROM financials "
                "WHERE scenario=? AND entity=? AND statement=?",
                (scenario, entity, statement),
            ).fetchall()
        )
        if len(periods) < 2:
            continue
        present = set(periods)
        expected = _months_between(periods[0], periods[-1])
        missing = sorted(expected - present)
        if missing:
            yield Finding(
                rule="R7",
                severity="fail",
                name=f"period-continuity/{scenario}/{entity}/{statement}",
                message=(
                    f"{scenario}/{entity}/{statement}: missing months in range "
                    f"{periods[0]}..{periods[-1]}: {missing}"
                ),
            )


def _months_between(start: str, end: str) -> set[str]:
    """All YYYY-MM-01 ISO dates from ``start`` through ``end`` inclusive."""
    import datetime as dt
    s = dt.date.fromisoformat(start)
    e = dt.date.fromisoformat(end)
    out: set[str] = set()
    cur = s
    while cur <= e:
        out.add(cur.isoformat())
        # Step to next first-of-month.
        if cur.month == 12:
            cur = dt.date(cur.year + 1, 1, 1)
        else:
            cur = dt.date(cur.year, cur.month + 1, 1)
    return out


# ---------------------------------------------------------------------------
# R9 — scenario coverage parity
# ---------------------------------------------------------------------------

def _run_r9(conn: sqlite3.Connection) -> Iterable[Finding]:
    """For each (entity, statement), warn if scenarios cover materially
    different period ranges. Common case: 'realistic' budget shorter than
    'actual' history — useful to know but not necessarily wrong."""
    combos = conn.execute(
        "SELECT DISTINCT entity, statement FROM financials"
    ).fetchall()
    for entity, statement in combos:
        rows = conn.execute(
            "SELECT scenario, COUNT(DISTINCT period_date) "
            "FROM financials WHERE entity=? AND statement=? "
            "GROUP BY scenario",
            (entity, statement),
        ).fetchall()
        if len(rows) < 2:
            continue
        max_count = max(c for _, c in rows)
        for scenario, count in rows:
            if count < max_count:
                yield Finding(
                    rule="R9",
                    severity="warn",
                    name=f"scenario-coverage/{entity}/{statement}/{scenario}",
                    message=(
                        f"{scenario}/{entity}/{statement}: covers {count} "
                        f"period(s); other scenarios cover up to {max_count}"
                    ),
                )


# ---------------------------------------------------------------------------
# R5
# ---------------------------------------------------------------------------

def _run_r5(
    registry: dict[str, AggregateFormula],
    workbook_paths: list[Path],
) -> Iterable[Finding]:
    cache: dict[Path, "Workbook"] = {}  # noqa: F821
    try:
        for name, formula in registry.items():
            yield from _audit_one_cell(name, formula, workbook_paths, cache)
    finally:
        for wb in cache.values():
            wb.close()


def _audit_one_cell(
    name: str,
    formula: AggregateFormula,
    workbook_paths: list[Path],
    cache: dict,
) -> Iterable[Finding]:
    if formula.source_cell is None:
        return
    if "!" not in formula.source_cell:
        yield Finding(
            rule="R5", severity="warn", name=name,
            message=(
                f"{name}: source_cell={formula.source_cell!r} is not in "
                f"'Sheet!Cell' form; skipping cell-type audit"
            ),
        )
        return
    sheet_name, cell_ref = formula.source_cell.split("!", 1)

    for path in workbook_paths:
        path = Path(path)
        if path not in cache:
            cache[path] = load_workbook(path, data_only=False, read_only=False)
        wb = cache[path]
        if sheet_name not in wb.sheetnames:
            continue
        try:
            cell = wb[sheet_name][cell_ref]
        except (ValueError, KeyError) as exc:
            yield Finding(
                rule="R5", severity="warn", name=name,
                message=(
                    f"{name}: could not read cell {formula.source_cell!r} "
                    f"in {path.name}: {exc}"
                ),
            )
            return
        if cell.data_type != "f":
            yield Finding(
                rule="R5", severity="warn", name=name,
                message=(
                    f"{name}: source_cell {formula.source_cell!r} is "
                    f"hardcoded (data_type={cell.data_type!r}), expected formula. "
                    f"Will silently drift when upstream cells change."
                ),
            )
        return  # found the sheet; done

    yield Finding(
        rule="R5", severity="warn", name=name,
        message=(
            f"{name}: sheet {sheet_name!r} (from source_cell="
            f"{formula.source_cell!r}) not found in any provided workbook"
        ),
    )
