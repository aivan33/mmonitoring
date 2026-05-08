"""Loader for the canonical taxonomy-format xlsx files.

Sheet name pattern: ``<STMT>( Indirect)? (<SCEN>)`` — e.g. ``IS (Actual)``,
``CF Indirect (Realistic)``. Layout is fixed: row 1 is the header
``Data | Group | Subgroup | Jan | ... | Dec``; row 2+ are data rows.
"""

from __future__ import annotations

import datetime as dt
import re
from pathlib import Path
from typing import Iterable, NamedTuple

from openpyxl import load_workbook


class FinancialRow(NamedTuple):
    period_date: dt.date
    entity: str
    scenario: str
    statement: str
    data: str
    grp: str
    subgroup: str
    display_order: int
    value: float | None
    is_aggregate: int = 0


VALID_SCENARIOS = ("actual", "pessimistic", "realistic", "optimistic")
VALID_STATEMENTS = ("IS", "CF", "BS")

# Matches "IS (Actual)", "CF Indirect (Realistic)", "BS (Pessimistic)", etc.
_SHEET_NAME_RE = re.compile(
    r"^\s*(IS|CF|BS)(?:\s+Indirect)?\s*\(\s*(\w+)\s*\)\s*$"
)

# Strings treated as NULL when they appear in monthly cells. ``-`` and ``—``
# are common Excel placeholders for "no value" in financial spreadsheets.
_NULL_STRINGS = {"", "-", "—", "n/a", "N/A"}


def _is_null_cell(raw: object) -> bool:
    if raw is None:
        return True
    if isinstance(raw, str) and raw.strip() in _NULL_STRINGS:
        return True
    return False


def _parse_sheet_name(name: str) -> tuple[str, str] | None:
    """Return ``(statement, scenario)`` if the sheet matches the taxonomy
    pattern, else ``None`` (skip silently). Raises ``ValueError`` if a
    recognized statement carries an unknown scenario name."""
    m = _SHEET_NAME_RE.match(name)
    if not m:
        return None
    statement = m.group(1)
    scenario = m.group(2).lower()
    if scenario not in VALID_SCENARIOS:
        raise ValueError(
            f"Sheet {name!r}: unknown scenario {scenario!r}; "
            f"expected one of {VALID_SCENARIOS}"
        )
    return statement, scenario


def load_taxonomy_xlsx(
    path: str | Path,
    *,
    year: int,
    entity: str,
    currency: str = "EUR",
    fx_rate: float | None = None,
    emit_null_cells: bool = False,
) -> Iterable[FinancialRow]:
    """Iterate every (statement, scenario) sheet in ``path`` and yield rows.

    Args:
        path: xlsx file in canonical taxonomy format.
        year: anchors month columns to ISO dates (Jan col → ``year``-01-01).
        entity: stamped on every yielded row.
        currency: source currency. ``EUR`` = pass-through; otherwise
            every numeric value is divided by ``fx_rate``.
        fx_rate: required when ``currency != 'EUR'``.
        emit_null_cells: when ``True``, emit one row per (data, grp, subgroup,
            month) with ``value=None`` for null cells. Default skips them.

    Sheets that don't match the taxonomy name pattern are skipped silently
    (e.g. ``Summary``, ``Notes``). Empty sheets and rows where all monthly
    cells are null are skipped.
    """
    if currency != "EUR" and fx_rate is None:
        raise ValueError(
            f"currency={currency!r} requires fx_rate to be set "
            f"(generic non-EUR conversion factor)"
        )
    return _iter_rows(
        Path(path),
        year=year,
        entity=entity,
        currency=currency,
        fx_rate=fx_rate,
        emit_null_cells=emit_null_cells,
    )


def _iter_rows(
    path: Path,
    *,
    year: int,
    entity: str,
    currency: str,
    fx_rate: float | None,
    emit_null_cells: bool,
) -> Iterable[FinancialRow]:
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        for sheet_name in wb.sheetnames:
            parsed = _parse_sheet_name(sheet_name)
            if parsed is None:
                continue
            statement, scenario = parsed
            ws = wb[sheet_name]
            yield from _iter_sheet(
                ws, statement, scenario,
                year=year, entity=entity,
                currency=currency, fx_rate=fx_rate,
                emit_null_cells=emit_null_cells,
            )
    finally:
        wb.close()


def _iter_sheet(
    ws,
    statement: str,
    scenario: str,
    *,
    year: int,
    entity: str,
    currency: str,
    fx_rate: float | None,
    emit_null_cells: bool,
) -> Iterable[FinancialRow]:
    rows_iter = ws.iter_rows(values_only=True)
    try:
        next(rows_iter)  # consume header
    except StopIteration:
        return

    # Buffer eligible rows so we can detect "no eligible rows" → skip the
    # whole sheet (matches the spec's "empty sheet" rule).
    eligible: list[tuple[int, str, str, str, tuple]] = []
    display_order = 0
    for row in rows_iter:
        if row is None or len(row) < 15:
            continue
        data, grp, subgroup = row[0], row[1], row[2]
        monthly = row[3:15]
        if data is None and grp is None and subgroup is None:
            continue
        if all(_is_null_cell(v) for v in monthly):
            continue
        if data is None or grp is None or subgroup is None:
            # Spec: data/grp/subgroup always non-null on a data row.
            continue
        eligible.append((
            display_order,
            str(data).strip(),
            str(grp).strip(),
            str(subgroup).strip(),
            tuple(monthly),
        ))
        display_order += 1

    if not eligible:
        return

    for order, data, grp, subgroup, monthly in eligible:
        for month_idx, raw in enumerate(monthly, start=1):
            if _is_null_cell(raw):
                if not emit_null_cells:
                    continue
                value: float | None = None
            else:
                value = float(raw)
                if currency != "EUR":
                    assert fx_rate is not None  # validated at entry
                    value = value / fx_rate
            yield FinancialRow(
                period_date=dt.date(year, month_idx, 1),
                entity=entity,
                scenario=scenario,
                statement=statement,
                data=data,
                grp=grp,
                subgroup=subgroup,
                display_order=order,
                value=value,
            )
