"""Loader for operational KPI xlsx files.

These are platform metrics that don't fit IS/CF/BS — GMV, # invoices,
days outstanding, active subscriptions, etc. Stored long-form in the
``operational_kpis`` table.

Two source formats are supported today:

- ``kpi_wide``: row 1 is a header with month labels (``Jan 26``,
  ``January``, ``2026-01``, etc.). Column 1 is a header label
  (typically ``Month`` or empty). Rows 2+ have the KPI name in
  column 1 and monthly values in subsequent columns.

- (``kpi_long`` is reserved for the future: one row per
  (period, kpi, value) triple.)

The loader normalizes month headers into ``YYYY-MM-01`` ISO dates by
matching against English month names with optional 2- or 4-digit year
suffix; the ``year`` parameter resolves headers without a year.
"""

from __future__ import annotations

import datetime as dt
import re
from pathlib import Path
from typing import Iterable, NamedTuple

from openpyxl import load_workbook


class KPIRow(NamedTuple):
    period_date: dt.date
    entity: str
    kpi: str
    value: float | None


_NULL_STRINGS = frozenset({"", "-", "—", "n/a", "N/A"})

_MONTH_NAMES = {
    name.lower(): idx
    for idx, name in enumerate(
        ["January", "February", "March", "April", "May", "June",
         "July", "August", "September", "October", "November", "December"],
        start=1,
    )
}
_MONTH_NAMES.update({
    name.lower(): idx
    for idx, name in enumerate(
        ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
         "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"],
        start=1,
    )
})

# Matches "Jan 26", "Jan 2026", "January 2026", "Jan-26", "January", or just "Jan".
_MONTH_HEADER_RE = re.compile(
    r"^\s*([A-Za-z]+)(?:\s*[-/\s]?\s*(\d{2,4}))?\s*$"
)


def _is_null_cell(raw) -> bool:
    if raw is None:
        return True
    if isinstance(raw, str) and raw.strip() in _NULL_STRINGS:
        return True
    return False


def _parse_month_header(cell, default_year: int) -> dt.date | None:
    """Return ``YYYY-MM-01`` for a header cell, or None if it isn't a
    recognizable month label.

    Recognized: bare month name (``Jan``, ``January``); month + year
    (``Jan 26``, ``January 2026``, ``Jan-26``); or a ``dt.date`` cell.
    """
    if isinstance(cell, dt.date):
        return dt.date(cell.year, cell.month, 1)
    if not isinstance(cell, str):
        return None
    m = _MONTH_HEADER_RE.match(cell)
    if not m:
        return None
    month = _MONTH_NAMES.get(m.group(1).lower())
    if month is None:
        return None
    year_token = m.group(2)
    if year_token is None:
        year = default_year
    else:
        y = int(year_token)
        year = 2000 + y if y < 100 else y
    return dt.date(year, month, 1)


def load_kpi_wide_xlsx(
    path: str | Path,
    *,
    year: int,
    entity: str,
    currency: str = "EUR",
    fx_rate: float | None = None,
) -> Iterable[KPIRow]:
    """Iterate rows of a ``kpi_wide`` xlsx, yielding one KPIRow per
    (period, kpi) pair.

    Args:
        path: xlsx file in kpi_wide format.
        year: default year for headers without a year (e.g. ``Jan``).
        entity: stamped on every yielded row.
        currency: source currency. ``EUR`` = pass-through; otherwise
            every numeric value is divided by ``fx_rate``.
        fx_rate: required when ``currency != 'EUR'``.

    Behavior:
        - Reads the first sheet of the workbook.
        - Header (row 1) is scanned for month labels; non-month columns
          (e.g. column 1's "Month" header) are skipped.
        - Rows where column 1 is null are skipped.
        - Cells with no value are skipped (no row emitted).
        - Numeric values are stored as floats; non-numeric strings are skipped.
    """
    if currency != "EUR" and fx_rate is None:
        raise ValueError(
            f"currency={currency!r} requires fx_rate to be set"
        )
    wb = load_workbook(Path(path), data_only=True, read_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration:
            return

        # Map each column index → period_date. Skip columns whose header
        # doesn't parse as a month.
        col_to_date: dict[int, dt.date] = {}
        for col_idx, cell in enumerate(header):
            if col_idx == 0:
                continue  # column 1 is the KPI label column
            d = _parse_month_header(cell, year)
            if d is not None:
                col_to_date[col_idx] = d

        for row in rows_iter:
            if not row or row[0] is None:
                continue
            kpi = str(row[0]).strip()
            if not kpi:
                continue
            for col_idx, period_date in col_to_date.items():
                if col_idx >= len(row):
                    continue
                raw = row[col_idx]
                if _is_null_cell(raw):
                    continue
                try:
                    value = float(raw)
                except (TypeError, ValueError):
                    continue
                if currency != "EUR":
                    assert fx_rate is not None
                    # Convention matches core/data/loaders/financials.py:
                    # fx_rate = source units per 1 EUR. Divide source by rate
                    # to get EUR (e.g. BGN 1.95583 means $195.583 BGN ÷ 1.95583
                    # = €100; USD ~1.087 means $100 USD ÷ 1.087 ≈ €92).
                    value = value / fx_rate
                yield KPIRow(
                    period_date=period_date,
                    entity=entity,
                    kpi=kpi,
                    value=value,
                )
    finally:
        wb.close()
