"""Integrity gates over a recomputed model workbook.

Each check reads the LibreOffice-recomputed values (see ``core.model.recalc``) and
returns a list of :class:`Violation` — it never raises, so one broken check doesn't
mask the rest. Gate targets (which rows are the BS check, the statement ties, the
subtotals) come from a per-client ``model_gates.yaml``; the checks themselves are
generic.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import yaml
from openpyxl.utils import get_column_letter as gc

# Excel error sentinels openpyxl surfaces as strings on a recomputed workbook.
_ERROR_STRINGS = frozenset(
    {"#REF!", "#NAME?", "#VALUE!", "#DIV/0!", "#N/A", "#NULL!", "#NUM!"}
)


@dataclass(frozen=True)
class Violation:
    check: str
    sheet: str
    cell: str
    detail: str


def _num(v):
    return v if isinstance(v, (int, float)) else None


def _cols(spec: dict) -> range:
    c = spec["cols"]
    return range(c["start"], c["end"] + 1)


def error_cells(wb, sheets, skip: dict | None = None) -> list[Violation]:
    """Any cell that recomputed to an Excel error is a hard failure.

    ``skip`` maps a sheet name to a list of ``[lo, hi]`` row ranges to ignore —
    used to scope out an explicitly-unbuilt section (documented in the gates yaml).
    """
    skip = skip or {}
    out: list[Violation] = []
    for sh in sheets:
        if sh not in wb.sheetnames:
            continue
        ws = wb[sh]
        ranges = skip.get(sh) or []
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value in _ERROR_STRINGS:
                    if any(lo <= cell.row <= hi for lo, hi in ranges):
                        continue
                    out.append(Violation("error_cell", sh, cell.coordinate, cell.value))
    return out


def bs_balances(wb, checks) -> list[Violation]:
    """Each balance-check row must be ~0 across its month columns."""
    out: list[Violation] = []
    for chk in checks:
        ws = wb[chk["sheet"]]
        r, tol = chk["row"], chk.get("tolerance", 1.0)
        for c in _cols(chk):
            v = _num(ws.cell(r, c).value)
            if v is not None and abs(v) > tol:
                out.append(Violation(
                    "bs_balance", chk["sheet"], f"{gc(c)}{r}",
                    f"{chk['name']}: {v:.2f} exceeds |{tol}|"))
    return out


def statements_tie(wb, ties) -> list[Violation]:
    """Two rows (e.g. CF ending cash vs BS cash) must be equal per column."""
    out: list[Violation] = []
    for tie in ties:
        a, b = tie["a"], tie["b"]
        wsa, wsb = wb[a["sheet"]], wb[b["sheet"]]
        tol = tie.get("tolerance", 1.0)
        for c in _cols(tie):
            va, vb = _num(wsa.cell(a["row"], c).value), _num(wsb.cell(b["row"], c).value)
            if va is None or vb is None or abs(va - vb) > tol:
                out.append(Violation(
                    "statements_tie", a["sheet"], f"{gc(c)}{a['row']}",
                    f"{tie['name']}: {va} vs {vb} (col {gc(c)})"))
    return out


def subtotals_foot(wb, subtotals) -> list[Violation]:
    """A parent row must equal the sum of its children per column."""
    out: list[Violation] = []
    for sub in subtotals:
        ws = wb[sub["sheet"]]
        parent, children = sub["parent"], sub["children"]
        tol = sub.get("tolerance", 1.0)
        for c in _cols(sub):
            p = _num(ws.cell(parent, c).value) or 0.0
            kids = sum(_num(ws.cell(k, c).value) or 0.0 for k in children)
            if abs(p - kids) > tol:
                out.append(Violation(
                    "subtotals_foot", sub["sheet"], f"{gc(c)}{parent}",
                    f"{sub['name']}: parent {p:.2f} != Σchildren {kids:.2f}"))
    return out


def run_all(wb, gates: dict) -> list[Violation]:
    """Run every configured gate and return the combined violations."""
    error_scan = gates.get("error_scan") or {}
    scan = error_scan.get("sheets")
    sheets = wb.sheetnames if scan in (None, "all") else scan
    return (
        error_cells(wb, sheets, error_scan.get("skip"))
        + bs_balances(wb, gates.get("balance_checks") or [])
        + statements_tie(wb, gates.get("ties") or [])
        + subtotals_foot(wb, gates.get("subtotals") or [])
    )


def load_gates(client: str, root: Path | None = None) -> dict:
    """Read ``clients/<client>/model_gates.yaml`` (the gate target manifest)."""
    base = root or Path(__file__).resolve().parents[2]
    path = base / "clients" / client / "model_gates.yaml"
    if not path.exists():
        raise FileNotFoundError(f"no model_gates.yaml for client {client!r} at {path}")
    return yaml.safe_load(path.read_text()) or {}
