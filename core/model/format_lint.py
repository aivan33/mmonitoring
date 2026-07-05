"""Design-system conformance lint for a model workbook.

Checks the machine-verifiable house-standard rules from ``core.model.design_system``
and reports :class:`FormatViolation`s — it never auto-fixes (format is a given; a
human decides). Intended to run **warn-only** in the gate until the existing client
models are brought fully to canon.

Two rules today (the highest-signal ones):
  * **font** — every value cell must be Century Gothic (Calibri is the outlier the
    canon calls out; a default-styled "naked" cell trips this too).
  * **inputs grammar** — on the Inputs sheet, an active-column (J) formula must be
    the scenario selector ``=OFFSET(K{r},0,$D$2)``.
"""

from __future__ import annotations

import re
from dataclasses import dataclass

from openpyxl.utils import column_index_from_string

from core.model import design_system as ds


@dataclass(frozen=True)
class FormatViolation:
    rule: str
    sheet: str
    cell: str
    detail: str


def font_conformance(wb) -> list[FormatViolation]:
    """Every cell that carries a value must use the house font."""
    out: list[FormatViolation] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                name = cell.font.name if cell.font is not None else None
                if name is not None and name != ds.FONT:
                    out.append(FormatViolation(
                        "font", ws.title, cell.coordinate,
                        f"font {name!r} != {ds.FONT!r}"))
    return out


def inputs_grammar(wb, sheet: str) -> list[FormatViolation]:
    """Active-column (J) formulas on the Inputs sheet must be the OFFSET selector."""
    out: list[FormatViolation] = []
    if sheet not in wb.sheetnames:
        return out
    ws = wb[sheet]
    jcol = column_index_from_string(ds.INPUT_COLUMNS["active"])
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, jcol).value
        if isinstance(v, str) and v.startswith("="):
            expected = ds.active_formula(r).replace(" ", "")
            if v.replace(" ", "") != expected:
                out.append(FormatViolation(
                    "inputs_grammar", sheet, f"{ds.INPUT_COLUMNS['active']}{r}",
                    f"active cell not {expected}: {v!r}"))
    return out


def lint(wb, *, inputs_sheet: str = " Inputs") -> list[FormatViolation]:
    """Run all design-system checks and return the combined violations."""
    return font_conformance(wb) + inputs_grammar(wb, inputs_sheet)
