"""The canonical house design system for client financial models — the single source of truth.

Reconciled from the reference budgets (Cupffee / Almacena / Farada), which had diverged. Decisions
(2026-06-22): **Hybrid palette** — Farada's cyan active cell + light-blue statement banners + grey
section bands + cream inputs, with Cupffee's **three scenario columns L/M/N** (Realistic / Optimistic
/ Pessimistic) as standard; **plain number formats**; **Century Gothic 10pt** (Almacena's Calibri was
the outlier). Builders import these constants so formatting is uniform; the model-building skill's
`references/design-system.md` documents the same canon in prose.

A model is ONE scenario-switchable workbook: cell ``D2`` selects the scenario, and each input's
active cell ``J{r} = OFFSET(K{r},0,$D$2)`` reads the chosen column among K/L/M/N.
"""
from __future__ import annotations

from openpyxl.styles import Font, PatternFill

FONT = "Century Gothic"
FONT_SIZE = 10
RIBBON_SIZE = 9

# fill palette: role -> ARGB hex
PALETTE = {
    "section_band": "FFD8D8D8",       # grey — Inputs section headers (I./II./…)
    "input_value": "FFFEF2CB",        # cream — editable scenario value cells (L/M/N)
    "active": "FFDDFBFF",             # cyan — the OFFSET active cell (col J)
    "statement_banner": "FFD5EBF4",   # light blue — statement titles / section bands / line headers
    "ribbon": "FFFFFFFF",             # white — the month date ribbon
}

# plain number formats (NOT accounting) — explicit constants; never harvested
NUMBER_FORMATS = {
    "int": "#,##0",
    "eur": "€#,##0.00",
    "pct": "0.0%",
    "num2": "0.00",
    "date": "[$-409]mmm\\-yy",
}

# scenarios live in columns K(anchor)/L/M/N; D2 selects via OFFSET.
# The three columns are part of the standard layout, but only Realistic (L) is REQUIRED —
# Optimistic / Pessimistic (M/N) are optional and may stay empty.
SCENARIOS = ["Realistic", "Optimistic", "Pessimistic"]
SCENARIO_COLS = ["L", "M", "N"]
SCENARIO_COLS_REQUIRED = ["L"]
SELECTOR_CELL = "D2"

# Inputs sheet column grammar: role -> column letter
INPUT_COLUMNS = {
    "section": "A",      # Roman section number (I., II., …)
    "subnumber": "B",    # sub-number (1.1, 3.1.1)
    "label": "C",
    "unit": "D",
    "threshold": "F",    # ladder threshold, if any
    "start": "G", "end": "H",   # date window, if any
    "active": "J",       # =OFFSET(K{r},0,$D$2)
    "anchor": "K",       # OFFSET anchor (blank)
    "scenario_first": "L",
    "notes": "O",
}


def font(bold: bool = False, size: int = FONT_SIZE) -> Font:
    return Font(name=FONT, size=size, bold=bold)


def fill(role: str) -> PatternFill:
    return PatternFill("solid", fgColor=PALETTE[role])


def active_formula(row: int) -> str:
    """The scenario-active formula for an input on ``row`` (reads K/L/M/N via D2)."""
    return f"=OFFSET(K{row},0,$D$2)"
