"""The structure layer of the model parser.

A general classifier engine turns sheet names into typed :class:`SheetInfo`
(entity, role, statement), driven by a small per-client :class:`Rules` config.
:class:`ModelContract` groups the sheets and exposes the actuals/budget/driver
seams plus the taxonomi month-axis.

The engine knows the *generic* financial conventions (taxonomi / yearly /
actuals / IS-CF-BS statements); the client config carries the specifics — entity
name patterns, exact-name role overrides, the separator marker, and the taxonomi
month-axis. See ``clients/almacena/model_rules.yaml`` and
``clients/almacena/MODEL_CONTRACT.md``.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

import yaml
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

# A statement sheet name starts with IS/CF/BS as a delimited token.
_STATEMENT_RE = re.compile(r"^\s*(IS|CF|BS)(?=[_ ]|$)", re.IGNORECASE)
# An "actuals" sheet either says "actual" or carries an "act" token (e.g. BV_act).
_ACT_TOKEN_RE = re.compile(r"(^|[_ ])act([_ ]|$)", re.IGNORECASE)

_STATEMENT_ROLES = {"statement", "taxonomi", "yearly"}
_ENTITY_DEFAULT_ROLES = {"statement", "taxonomi", "yearly", "actuals"}

# Engine/driver sheets that recur across clients (cupffee/honey/farada/almacena)
# under the same names — recognised by the engine, not per-client config. Matched
# on the whitespace-normalised name so "HR ", " Inputs", "Pro  Forma" all hit.
_GENERIC_ENGINE = {"inputs", "pro forma"}
_GENERIC_DRIVER = {"hr", "kpis"}


def _norm(name: str) -> str:
    """Collapse internal/edge whitespace and lowercase, for tolerant matching."""
    return re.sub(r"\s+", " ", name.strip()).lower()


def _generic_role(name: str) -> str | None:
    """Engine/driver role for sheets that recur across clients, else ``None``."""
    n = _norm(name)
    if n in _GENERIC_ENGINE or n.startswith(("inputs_", "inputs ")):
        return "engine"  # incl. entity-suffixed engine inputs, e.g. "Inputs_Foundation"
    if n in _GENERIC_DRIVER:
        return "driver"
    return None


@dataclass(frozen=True)
class TaxonomiAxis:
    """Where the months live on the budget taxonomi tabs.

    Two modes. *Positional* (``first_month_col`` + ``months`` + ``year``) keys
    months by column position — robust to mislabeled header years (Almacena,
    cupffee). *Date-header* (``header_dates=True``) reads the real ISO dates out
    of ``header_row`` — for clients with no single-year taxonomi tab whose bare
    IS/CF/BS span several years (honey, farada)."""

    header_row: int
    first_month_col: str | None = None
    months: int | None = None
    year: int | None = None
    header_dates: bool = False


@dataclass
class Rules:
    """Per-client classification config (the engine is general)."""

    entity_patterns: dict[str, list[str]]
    default_entity: str = "consolidated"
    role_overrides: dict[str, str] = field(default_factory=dict)
    separator_marker: str = ">>>"
    taxonomi_axis: TaxonomiAxis | None = None


@dataclass(frozen=True)
class SheetInfo:
    """A classified sheet."""

    name: str
    entity: str | None
    role: str  # separator | taxonomi | yearly | actuals | statement | engine | driver | other
    statement: str | None  # IS | CF | BS | None


def _detect_entity(name: str, rules: Rules) -> str | None:
    low = name.lower()
    for entity, subs in rules.entity_patterns.items():
        if any(s.lower() in low for s in subs):
            return entity
    return None


def _detect_statement(name: str) -> str | None:
    m = _STATEMENT_RE.match(name.strip())
    return m.group(1).upper() if m else None


def _detect_role(name: str, rules: Rules) -> str:
    if rules.separator_marker and rules.separator_marker in name:
        return "separator"
    if name in rules.role_overrides:
        return rules.role_overrides[name]
    low = name.lower()
    if "taxonomi" in low:
        return "taxonomi"
    if "yearly" in low:
        return "yearly"
    if "actual" in low or _ACT_TOKEN_RE.search(name):
        return "actuals"
    if _STATEMENT_RE.match(name.strip()):
        return "statement"
    generic = _generic_role(name)
    if generic is not None:
        return generic
    return "other"


def classify_sheet(name: str, rules: Rules) -> SheetInfo:
    """Classify one sheet name into a :class:`SheetInfo`."""
    role = _detect_role(name, rules)
    if role == "separator":
        return SheetInfo(name, None, "separator", None)
    entity = _detect_entity(name, rules)
    if entity is None and role in _ENTITY_DEFAULT_ROLES:
        entity = rules.default_entity
    statement = _detect_statement(name) if role in _STATEMENT_ROLES else None
    return SheetInfo(name, entity, role, statement)


class ModelContract:
    """The classified sheets of one model workbook, grouped and queryable."""

    def __init__(self, sheets: list[SheetInfo], rules: Rules, path: Path) -> None:
        self.sheets = sheets
        self.rules = rules
        self._path = path

    def entities(self) -> list[str]:
        return sorted({s.entity for s in self.sheets if s.entity})

    def by_role(self, role: str) -> list[SheetInfo]:
        return [s for s in self.sheets if s.role == role]

    def by_entity(self, entity: str) -> list[SheetInfo]:
        return [s for s in self.sheets if s.entity == entity]

    def taxonomi(self, entity: str | None = None) -> list[SheetInfo]:
        return [s for s in self.sheets if s.role == "taxonomi" and (entity is None or s.entity == entity)]

    def budget(self, entity: str | None = None) -> list[SheetInfo]:
        """Budget-side sheets: the taxonomi tabs, falling back to the bare
        ``IS``/``CF``/``BS`` statement sheet for any statement that has no
        taxonomi tab (some clients keep no taxonomi, or only an IS one). Only the
        plainly-named statement falls back — ``IS_platform`` variants do not."""
        taxes = self.taxonomi(entity)
        covered = {s.statement for s in taxes}
        fallback = [
            s
            for s in self.sheets
            if s.role == "statement"
            and (entity is None or s.entity == entity)
            and s.statement is not None
            and s.statement not in covered
            and _norm(s.name) == s.statement.lower()
        ]
        return taxes + fallback

    def actuals(self, entity: str | None = None) -> list[SheetInfo]:
        return [s for s in self.sheets if s.role == "actuals" and (entity is None or s.entity == entity)]

    def drivers(self) -> list[SheetInfo]:
        return self.by_role("driver")

    def engine(self) -> list[SheetInfo]:
        return self.by_role("engine")

    def month_axis(self) -> dict[str, str]:
        """Map each taxonomi month column letter to its ``YYYY-MM`` period."""
        ax = self.rules.taxonomi_axis
        if ax is None:
            return {}
        if ax.header_dates:
            return self._date_header_axis(ax)
        start = column_index_from_string(ax.first_month_col)
        return {
            get_column_letter(start + i): f"{ax.year}-{i + 1:02d}"
            for i in range(ax.months)
        }

    def _date_header_axis(self, ax: TaxonomiAxis) -> dict[str, str]:
        """Read column->period from the real dates in ``ax.header_row`` of the
        first budget sheet (all budget sheets share the same header layout)."""
        budget = self.budget()
        if not budget:
            return {}
        wb = load_workbook(self._path, read_only=True, data_only=True)
        try:
            ws = wb[budget[0].name]
            header = next(ws.iter_rows(min_row=ax.header_row, max_row=ax.header_row))
        finally:
            wb.close()
        return {
            get_column_letter(cell.column): f"{cell.value.year}-{cell.value.month:02d}"
            for cell in header
            if isinstance(cell.value, datetime)
        }

    def seams(self) -> dict[str, dict[str, list[str]]]:
        """Per entity, the budget (taxonomi) and actuals sheet names."""
        return {
            ent: {
                "budget": [s.name for s in self.budget(ent)],
                "actuals": [s.name for s in self.actuals(ent)],
            }
            for ent in self.entities()
        }

    def last_populated_month(self, sheet_name: str) -> str | None:
        """The last month column on a taxonomi sheet that holds any value.

        Axis-agnostic: works off :meth:`month_axis` (column->period), so it
        serves both the positional and date-header axes."""
        ax = self.rules.taxonomi_axis
        if ax is None:
            return None
        axis = self.month_axis()
        if not axis:
            return None
        cols = list(axis)
        periods = list(axis.values())
        idx = [column_index_from_string(c) for c in cols]
        base = min(idx)
        wb = load_workbook(self._path, read_only=True, data_only=True)
        try:
            ws = wb[sheet_name]
            rows = list(
                ws.iter_rows(
                    min_row=ax.header_row + 1,
                    min_col=base,
                    max_col=max(idx),
                    values_only=True,
                )
            )
        finally:
            wb.close()
        last = None
        for k, ci in enumerate(idx):
            if any(row[ci - base] is not None for row in rows):
                last = k
        return None if last is None else periods[last]


def read_contract(path: str | Path, rules: Rules) -> ModelContract:
    """Classify every sheet in the workbook at ``path``."""
    path = Path(path)
    wb = load_workbook(path, read_only=True)
    try:
        names = list(wb.sheetnames)
    finally:
        wb.close()
    sheets = [classify_sheet(n, rules) for n in names]
    return ModelContract(sheets, rules, path)


def load_rules(path: str | Path) -> Rules:
    """Load a per-client :class:`Rules` config from YAML."""
    data = yaml.safe_load(Path(path).read_text())
    m = data["model"]
    ax = m.get("taxonomi_axis")
    axis = (
        TaxonomiAxis(
            header_row=ax["header_row"],
            first_month_col=ax.get("first_month_col"),
            months=ax.get("months"),
            year=ax.get("year"),
            header_dates=ax.get("header_dates", False),
        )
        if ax
        else None
    )
    return Rules(
        entity_patterns=m["entity_patterns"],
        default_entity=m.get("default_entity", "consolidated"),
        role_overrides=m.get("role_overrides", {}),
        separator_marker=m.get("separator_marker", ">>>"),
        taxonomi_axis=axis,
    )
