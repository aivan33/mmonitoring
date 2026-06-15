"""The cell layer of the model parser.

``read_cells(path)`` loads a workbook twice with openpyxl — once with
``data_only=True`` for the cached computed values and once with
``data_only=False`` for the raw formula strings — and exposes every cell as a
:class:`Cell` (value, formula, number_format, dtype, sheet, coord).

Note on cached values: openpyxl never *calculates* formulas. Workbooks last
saved by Excel/LibreOffice carry cached results, so ``Cell.value`` is the
computed number; workbooks authored by openpyxl (e.g. test fixtures) do not, so
a formula cell's ``value`` is ``None`` while its ``formula`` string is present.
"""

from __future__ import annotations

import datetime as dt
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


@dataclass(frozen=True)
class Cell:
    """A single parsed cell."""

    sheet: str
    coord: str
    value: object | None
    formula: str | None
    number_format: str
    dtype: str  # one of: empty | number | text | bool | date | formula

    @property
    def is_formula(self) -> bool:
        return self.formula is not None


def _classify(value: object | None, formula: str | None) -> str:
    if formula is not None:
        return "formula"
    if value is None:
        return "empty"
    if isinstance(value, bool):
        return "bool"
    if isinstance(value, (int, float)):
        return "number"
    if isinstance(value, (dt.datetime, dt.date)):
        return "date"
    return "text"


def _formula_string(raw: object) -> str:
    """Coerce an openpyxl formula value to its string form.

    Plain formulas come back as ``"=A1+B1"``; array formulas come back as an
    ``ArrayFormula`` object carrying the text under ``.text``.
    """
    if isinstance(raw, str):
        return raw
    return getattr(raw, "text", str(raw))


class Cells:
    """Parsed cells of one workbook, addressable by ``(sheet, coord)``."""

    def __init__(self, values_wb, formulas_wb) -> None:
        self._values = values_wb
        self._formulas = formulas_wb

    def sheets(self) -> list[str]:
        return list(self._formulas.sheetnames)

    def cell(self, sheet: str, coord: str) -> Cell:
        if sheet not in self._formulas.sheetnames:
            raise KeyError(f"no such sheet: {sheet!r}")
        fcell = self._formulas[sheet][coord]
        formula = _formula_string(fcell.value) if fcell.data_type == "f" else None
        value = self._values[sheet][coord].value
        return Cell(
            sheet=sheet,
            coord=coord.upper(),
            value=value,
            formula=formula,
            number_format=fcell.number_format,
            dtype=_classify(value, formula),
        )


def read_cells(path: str | Path) -> Cells:
    """Parse the workbook at ``path`` into an addressable :class:`Cells`."""
    path = Path(path)
    values_wb = load_workbook(path, data_only=True)
    formulas_wb = load_workbook(path, data_only=False)
    return Cells(values_wb, formulas_wb)
