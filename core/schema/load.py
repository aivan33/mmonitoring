"""Load an Excel model into the schema (role B — query/validate over parsed Excel).

Reads a workbook and populates `core.schema`'s tables: the ` Inputs` sheet → sections/groups/
inputs/input_value (Pillar 1); ProForma → proforma lines; IS/CF/BS(+_Y) → statement lines
(Pillars 2-3). Dependencies (the lineage edges) are extracted by parsing each line's formula with
`core.model.parse_refs` and resolving refs to an input (via the input rows' J/F/G/H cells) or
another line (by sheet+row). Values stay in Excel; only the Realistic scenario value is captured.

Heuristic parser, tuned to the standardized layout (Farada/Cupffee): section headers are
UPPER-case rows; groups are short sub-labels / sub-numbered rows; inputs are labelled rows with a
Realistic value (col L) or an `=OFFSET` active cell (col J).

Run/import:  from core.schema.load import load_model
"""
from __future__ import annotations

import datetime as dt

from openpyxl import load_workbook
from openpyxl.utils import coordinate_to_tuple
from openpyxl.worksheet.formula import ArrayFormula

from core.model import parse_refs
from core.schema import create_db

# sheet name -> pillar, for the standardized model layout
PILLARS = {" Inputs": "input", "ProForma": "proforma",
           "IS": "statement", "CF": "statement", "BS": "statement",
           "IS_Y": "statement", "CF_Y": "statement", "BS_Y": "statement"}


def _ft(cell):
    v = cell.value
    return v.text if isinstance(v, ArrayFormula) else (v if isinstance(v, str) and v.startswith("=") else None)


def _num(v):
    return float(v) if isinstance(v, (int, float)) and not isinstance(v, bool) else None


def _iso(v):
    return v.date().isoformat() if isinstance(v, dt.datetime) else (v.isoformat() if isinstance(v, dt.date) else None)


def _dtype(unit):
    u = (unit or "").lower()
    if "%" in u:
        return "pct"
    if "eur" in u:
        return "eur"
    if "date" in u:
        return "date"
    if any(k in u for k in ("day", "month", "year", "#", "pcs", "sensor", "meas")):
        return "count"
    return "num"


def _upperish(s: str) -> bool:
    letters = [c for c in s if c.isalpha()]
    return bool(letters) and sum(c.isupper() for c in letters) / len(letters) > 0.6


def _is_note(c: str) -> bool:
    c = c.strip()
    return c.startswith(("-", "<", ">")) or (len(c) > 34 and not _upperish(c)) or (c[:2].rstrip(")").isdigit() and ")" in c[:3])


def load_model(db_path: str, xlsx_path: str, client_name: str, model_name: str,
               base_ccy: str = "EUR", start_date: str | None = None, horizon: int = 60):
    conn = create_db(db_path)
    wbf = load_workbook(xlsx_path)
    wbv = load_workbook(xlsx_path, data_only=True)
    conn.execute("INSERT INTO client VALUES (1, ?)", (client_name,))
    conn.execute("INSERT INTO model VALUES (1, 1, ?, ?, ?, ?)", (model_name, base_ccy, start_date, horizon))
    conn.execute("INSERT INTO scenario VALUES (1, 1, 'Realistic', 1)")

    cell2input: dict[tuple[str, str], int] = {}   # (sheet, COORD J/F/G/H) -> input_id
    row2line: dict[tuple[str, int], int] = {}      # (sheet, row) -> line_id
    n = {"section": 0, "grp": 0, "input": 0, "line": 0}

    def nid(k):
        n[k] += 1
        return n[k]

    # ---- Pillar 1: inputs ----------------------------------------------
    isheet = " Inputs"
    wf, wv = wbf[isheet], wbv[isheet]
    cur_sec = cur_grp = None
    so = go = io = 0
    for r in range(1, wf.max_row + 1):
        c = wf.cell(r, 3).value
        if c is None or not str(c).strip():
            continue
        c = str(c).strip()
        l = wv.cell(r, 12).value
        jf = _ft(wf.cell(r, 10))
        is_input = (_num(l) is not None) or (jf is not None and "OFFSET" in jf)
        if is_input:
            if cur_grp is None:                 # ensure a group exists
                if cur_sec is None:
                    cur_sec = nid("section"); so += 1
                    conn.execute("INSERT INTO section VALUES (?,1,'input',NULL,'(unsectioned)',?)", (cur_sec, so))
                cur_grp = nid("grp"); go += 1
                conn.execute("INSERT INTO grp VALUES (?,?,NULL,'(general)',?)", (cur_grp, cur_sec, go))
            iid = nid("input"); io += 1
            unit = wf.cell(r, 4).value
            conn.execute("INSERT INTO input VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                         (iid, cur_grp, c, str(unit) if unit else None, _dtype(unit),
                          _num(wf.cell(r, 6).value), _iso(wf.cell(r, 7).value), _iso(wf.cell(r, 8).value),
                          (lambda o: str(o) if o else None)(wf.cell(r, 15).value), io, f"{isheet}!J{r}"))
            conn.execute("INSERT INTO input_value VALUES (?,1,?)", (iid, _num(l)))
            for col in ("J", "F", "G", "H"):
                cell2input[(isheet, f"{col}{r}")] = iid
        elif _is_note(c):
            continue
        elif _upperish(c):                       # section header
            cur_sec = nid("section"); so += 1; go = 0; cur_grp = None
            conn.execute("INSERT INTO section VALUES (?,1,'input',?,?,?)", (cur_sec, str(wf.cell(r, 2).value or "")[:8] or None, c, so))
        elif len(c) <= 34:                       # short label = group/sub-header
            if cur_sec is None:
                cur_sec = nid("section"); so += 1
                conn.execute("INSERT INTO section VALUES (?,1,'input',NULL,'(unsectioned)',?)", (cur_sec, so))
            cur_grp = nid("grp"); go += 1; io = 0
            conn.execute("INSERT INTO grp VALUES (?,?,?,?,?)", (cur_grp, cur_sec, str(wf.cell(r, 2).value or "")[:8] or None, c, go))

    # ---- Pillars 2-3: lines (one section per proforma/statement sheet) --
    for sheet, pillar in PILLARS.items():
        if pillar == "input" or sheet not in wbf.sheetnames:
            continue
        wf = wbf[sheet]
        sec = nid("section"); so += 1
        conn.execute("INSERT INTO section VALUES (?,1,?,NULL,?,?)", (sec, pillar, sheet, so))
        lo = 0
        for r in range(1, wf.max_row + 1):
            a = wf.cell(r, 1).value
            if a is None or not str(a).strip():
                continue
            formula = next((_ft(wf.cell(r, col)) for col in range(2, 8) if _ft(wf.cell(r, col))), None)
            broken = next((f for col in range(2, 63) for f in [_ft(wf.cell(r, col))] if f and "#REF!" in f), None)
            lo += 1
            lid = nid("line")
            conn.execute("INSERT INTO line VALUES (?,?,?,?,NULL,?,?)",
                         (lid, sec, str(a).strip(), "header" if formula is None else "leaf", lo, f"{sheet}!{r}"))
            stored = broken or formula                     # prefer a broken-ref formula so it surfaces
            if stored:
                conn.execute("INSERT INTO line_formula VALUES (?,?)", (lid, stored[:480]))
            row2line[(sheet, r)] = lid

    # ---- dependencies (lineage edges) from formula refs ----------------
    edges = set()
    for (sheet, r), lid in row2line.items():
        formula = next((_ft(wbf[sheet].cell(r, col)) for col in range(2, 8) if _ft(wbf[sheet].cell(r, col))), None)
        if not formula:
            continue
        for ref in parse_refs(formula, sheet).refs:
            for rsheet, rcoord in _cells_of(ref):
                if (rsheet, rcoord) in cell2input:
                    edges.add((lid, "input", cell2input[(rsheet, rcoord)]))
                else:
                    row = coordinate_to_tuple(rcoord)[0]
                    tgt = row2line.get((rsheet, row))
                    if tgt and tgt != lid:
                        edges.add((lid, "line", tgt))
    conn.executemany("INSERT OR IGNORE INTO line_dependency VALUES (?,?,?)", edges)
    conn.commit()
    return conn


def _cells_of(ref):
    """(sheet, coord) per ref cell — expand modest ranges (≤200 cells), else just the start."""
    if not ref.is_range:
        yield (ref.sheet, ref.start)
        return
    try:
        r1, c1 = coordinate_to_tuple(ref.start)
        r2, c2 = coordinate_to_tuple(ref.end)
        if (abs(r2 - r1) + 1) * (abs(c2 - c1) + 1) <= 200:
            yield from ref.cells()
            return
    except Exception:
        pass
    yield (ref.sheet, ref.start)
